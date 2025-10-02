import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class ZitounaReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVÉ ZITOUNA vers Excel")
        self.root.geometry("600x500")
        
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"RELEVE_ZITOUNA_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        self.setup_ui()

    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur RELEVÉ ZITOUNA",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVÉ ZITOUNA:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.browse_pdf, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                              command=self.convertir, 
                              font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF ZITOUNA", filetypes=[["PDF", "*.pdf"]])
        if path:
            self.pdf_path.set(path)
            base = os.path.splitext(os.path.basename(path))[0]
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.excel_name.set(f"RELEVE_ZITOUNA_{base}_{ts}")

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF ZITOUNA.")
            return
        self.progress['value'] = 10; self.root.update_idletasks()
        try:
            print(f"DEBUG ZITOUNA - PDF sélectionné: {path}")
        except Exception:
            pass
        rows = self.parse_pdf(path)
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire des transactions du relevé ZITOUNA.")
            return
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'RELEVE_ZITOUNA'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        self.progress['value'] = 100
        # Message de succès plus positif
        success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n\n"
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        messagebox.showinfo("✅ Conversion réussie", success_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def parse_pdf(self, pdf_path: str):
        """Parser robuste pour ZITOUNA: gère index avant date, 2 dates/ligne, libellé multi-lignes
        et montants situés après la date valeur (débit/crédit)."""
        date_line_re = re.compile(r"^\s*(?:\d+\s+)?(\d{2}/\d{2}/\d{4})")
        date_any_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*(?:[.,]\d+)?|-?\d+[.,]\d+")

        def to_float(s: str):
            if not s:
                return None
            s = s.replace('\u00a0', ' ').strip()
            neg = s.startswith('-')
            s = s.lstrip('-').replace(' ', '')
            
            # Handle different number formats for Zitouna:
            # Format 1: 3,500,000 (thousands separator with comma)
            # Format 2: 2,850 (decimal with comma)
            # Format 3: 0,350 (decimal with comma)
            # Format 4: 3500,000 (decimal with comma, 3 digits after)
            
            if ',' in s and '.' in s:
                # Both comma and dot present - assume dot is thousands separator, comma is decimal
                s = s.replace('.', '').replace(',', '.')
            elif ',' in s:
                # Only comma present - check if it's decimal or thousands separator
                parts = s.split(',')
                if len(parts) == 2 and len(parts[1]) <= 3:
                    # Likely decimal separator (e.g., 2,850 or 0,350 or 3500,000)
                    s = s.replace(',', '.')
                else:
                    # Likely thousands separator (e.g., 3,500,000)
                    s = s.replace(',', '')
            else:
                # No comma, keep as is
                pass
                
            try:
                val = float(s)
                return -val if neg else val
            except Exception:
                return None

        def _clean_zitouna_amount(amount_str: str) -> str:
            """
            Nettoie et formate un montant ZITOUNA selon le format standard
            Format: 3500,000 (virgule pour décimales, 3 chiffres après virgule)
            """
            if not amount_str:
                return None
            
            # Nettoyer le montant - garder le signe négatif et gérer les formats spéciaux
            # Exemple: "3500,000" -> "3500,000", "0500)" -> "500,000"
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            # Gérer les montants avec parenthèses (ex: "0500)" -> "500")
            if cleaned.endswith(')'):
                cleaned = cleaned[:-1]  # Enlever la parenthèse fermante
            
            # Gérer les montants avec des zéros en début (ex: "0500" -> "500")
            if cleaned.startswith('0') and len(cleaned) > 1:
                cleaned = cleaned.lstrip('0')
                if not cleaned:  # Si tout était des zéros
                    cleaned = '0'
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]  # Enlever le signe négatif
            
            # Filtrer les montants à zéro et les montants dégradés
            if (cleaned == '0' or cleaned == '0,000' or cleaned == '0.000' or 
                cleaned == '000' or cleaned == '000,000' or cleaned == '000.000' or
                cleaned == '00' or cleaned == '00,000' or cleaned == '00.000'):
                return None
            
            try:
                # Traitement spécial pour préserver le format original Zitouna
                original_cleaned = cleaned
                
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 12.409,000 (point = milliers, virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 3500,000 (virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est un format avec décimales ou milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales: 5.950 -> 5,950
                        cleaned = cleaned.replace('.', ',')
                    else:
                        # Probablement des milliers: 12.409 -> 12409,000
                        cleaned = cleaned.replace('.', '') + ',000'
                
                # Si pas de virgule, ajouter ,000
                if ',' not in cleaned:
                    cleaned = cleaned + ',000'
                
                # S'assurer qu'il y a exactement 3 chiffres après la virgule
                if ',' in cleaned:
                    parts = cleaned.split(',')
                    integer_part = parts[0]
                    decimal_part = parts[1] if len(parts) > 1 else "000"
                    
                    # Garder exactement 3 chiffres après la virgule
                    if len(decimal_part) >= 3:
                        formatted_decimal = decimal_part[:3]
                    else:
                        formatted_decimal = decimal_part.ljust(3, '0')
                    
                    # Retourner le format Zitouna standard
                    result = f"{integer_part},{formatted_decimal}"
                    return f"-{result}" if is_negative else result
                else:
                    # Pas de virgule trouvée, ajouter ,000
                    result = f"{cleaned},000"
                    return f"-{result}" if is_negative else result
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not text.strip():
                    # Fallback OCR pour les PDFs scannés
                    ocr_text = self._extract_text_via_ocr(pdf_path)
                    if ocr_text.strip():
                        text = ocr_text
                        break  # Utiliser le texte OCR pour toutes les pages
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                i = 0
                while i < len(lines):
                    line = lines[i]
                    m = date_line_re.match(line)
                    if not m:
                        i += 1
                        continue
                    op_date = m.group(1)

                    # Accumuler libellé multi-lignes jusqu'à la prochaine date
                    j = i + 1
                    desc = []
                    while j < len(lines) and not date_line_re.match(lines[j]):
                        desc.append(lines[j])
                        j += 1

                    combined = (line + ' ' + ' '.join(desc)).strip()
                    # Après 2ème date (date valeur), se trouvent débit/crédit
                    dates = list(date_any_re.finditer(combined))
                    tail_start = dates[1].end() if len(dates) >= 2 else m.end()
                    tail = combined[tail_start:]

                    raw_amounts = amount_re.findall(tail)
                    filtered = []
                    for a in raw_amounts:
                        if re.search(r"\d{2}\.\d{4}", a):  # ex: 05.2025 (référence)
                            continue
                        if re.fullmatch(r"\d{7,}", a.replace(' ', '')):
                            continue
                        filtered.append(a)
                    # Utiliser la nouvelle fonction de formatage Zitouna
                    amount_formatted = _clean_zitouna_amount(filtered[0]) if filtered else None
                    amount = to_float(filtered[0]) if filtered else None

                    lib = combined[m.end():tail_start].strip()
                    lib = re.sub(r"\s+", " ", lib)
                    # Supprimer la date valeur du libellé (format DD/MM/YYYY)
                    lib = re.sub(r"\b\d{2}/\d{2}/\d{4}\b", "", lib).strip()

                    debit = credit = None
                    if amount is not None and amount_formatted is not None:
                        # Classification basée sur les mots-clés Zitouna spécifiques
                        upper_lib = lib.upper()
                        
                        # Mots-clés pour les crédits Zitouna
                        credit_keywords = [
                            'VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT', 'CREDIT',
                            'VIREMENT RECU', 'ENCAISSEMENT EFFET'
                        ]
                        
                        # Mots-clés pour les débits Zitouna
                        debit_keywords = [
                            'PAIEMENT', 'COMMISSION', 'FRAIS', 'TVA', 'RETRAIT', 'DEBIT',
                            'COMM REGLEMENT EFFET', 'COMMISSION RETOUR EFFET', 'FRAIS PDL',
                            'PAIEMENT EFFET', 'PAIEMENT PRINCIPAL', 'PAIEMENT PROFIT',
                            'PRIME TAKAFUL', 'DROIT DE TIMBRE', 'COMMISSIONS PAYEES',
                            'PAIEMENT PRELEVEMENT', 'COM ENC CHEQUE','COMM REMISE EFFET'
                        ]
                        
                        # Vérifier les crédits en premier
                        is_credit = any(kw in upper_lib for kw in credit_keywords)
                        is_debit = any(kw in upper_lib for kw in debit_keywords)
                        
                        if is_credit and not is_debit:
                            # Transaction de crédit - utiliser le formatage Zitouna
                            credit = amount_formatted
                        elif is_debit and not is_credit:
                            # Transaction de débit - utiliser le formatage Zitouna (sans signe négatif)
                            debit = amount_formatted.lstrip('-') if amount_formatted else None
                        else:
                            # Par défaut, utiliser le signe du montant
                            if filtered[0].strip().startswith('-') or amount < 0:
                                debit = amount_formatted.lstrip('-') if amount_formatted else None
                            else:
                                credit = amount_formatted

                    # Filtrer les lignes vides et non-transactionnelles
                    if (not re.search(r"(?i)solde|totaux|report", lib) and 
                        lib.strip() and len(lib.strip()) > 2 and
                        (debit is not None or credit is not None)):
                        
                        # Éviter d'avoir le même montant en débit et crédit
                        if debit and credit and debit == credit:
                            # Si les montants sont identiques, garder seulement le crédit (versements)
                            if any(kw in lib.upper() for kw in ["VERSEMENT", "DEPOT", "ENCAISSEMENT", "REM COM", "AV.TPE", "VIREMENT RECU"]):
                                debit = None
                            else:
                                credit = None
                        
                        results.append({"date": op_date, "libelle": lib, "debit": debit, "credit": credit})

                    i = j
        return results

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "J03"
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        # Number format 3 decimals
        max_row = ws.max_row
        pattern = '#\u00A0##0,000'
        for r in range(2, max_row + 1):
            ws[f'C{r}'].number_format = pattern
            ws[f'D{r}'].number_format = pattern
        # Borders
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

def main():
    root = tk.Tk(); ZitounaReleveConverter(root); root.mainloop()

if __name__ == '__main__':
    main()

