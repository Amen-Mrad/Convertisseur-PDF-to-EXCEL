import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
import fitz  # PyMuPDF
import cv2
import numpy as np
from PIL import Image
import io
import subprocess
import sys
try:
    import pytesseract  # type: ignore
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False
    print("DEBUG BTK RELEVÉ - pytesseract non disponible, OCR désactivé")

class BTKReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVÉ BTK vers Excel")
        self.root.geometry("600x500")
        self.root.configure(bg='#f0f0f0')
        
        # Variables
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar()
        self.excel_name.set(f"RELEVE_BTK_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        
        self.create_widgets()
    
    def create_widgets(self):
        # Titre principal
        title_frame = tk.Frame(self.root, bg='#f0f0f0')
        title_frame.pack(pady=20)
        
        title_label = tk.Label(title_frame, text="Convertisseur RELEVÉ BTK"
                              font=('Arial', 18, 'bold'), bg='#f0f0f0')
        title_label.pack(pady=(0,4))
        subtitle_label = tk.Label(title_frame, text="Conversion PDF vers Excel"
                                 font=('Arial', 12), fg='#666', bg='#f0f0f0')
        subtitle_label.pack()
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Sélection du fichier PDF
        pdf_frame = tk.Frame(main_frame, bg='#f0f0f0')
        pdf_frame.pack(fill='x', pady=10)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVÉ BTK:"
                font=('Arial', 12, 'bold'), bg='#f0f0f0').pack(anchor='w')
        
        pdf_input_frame = tk.Frame(pdf_frame, bg='#f0f0f0')
        pdf_input_frame.pack(fill='x', pady=5)
        
        pdf_entry = tk.Entry(pdf_input_frame, textvariable=self.pdf_path
                            font=('Arial', 10), width=50)
        pdf_entry.pack(side='left', fill='x', expand=True)
        
        browse_button = tk.Button(pdf_input_frame, text="Parcourir"
                                 command=self.browse_pdf, 
                                 font=('Arial', 10, 'bold'), padx=20)
        browse_button.pack(side='right', padx=(10, 0))
        
        # Nom du fichier Excel
        excel_frame = tk.Frame(main_frame, bg='#f0f0f0')
        excel_frame.pack(fill='x', pady=10)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:"
                font=('Arial', 12, 'bold'), bg='#f0f0f0').pack(anchor='w')
        
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name
                              font=('Arial', 10), width=50)
        excel_entry.pack(fill='x', pady=5)
        
        # Boutons
        buttons_frame = tk.Frame(main_frame, bg='#f0f0f0')
        buttons_frame.pack(pady=20)
        
        self.convert_button = tk.Button(buttons_frame, text="Convertir en Excel"
                                       command=self.convert_to_excel, bg='#A3A3A3', fg='#0088FF'
                                       font=('Arial', 14, 'bold'), padx=30, pady=10)
        self.convert_button.pack(side='left', padx=5)
        
        self.retour_button = tk.Button(buttons_frame, text="Retour page d'accueil"
                                      command=self.retour_accueil, 
                                      font=('Arial', 12, 'bold'), padx=20, pady=10)
        self.retour_button.pack(side='left', padx=5)
        
        # Barre de progression

        # Label de statut
        , bg='#f0f0f0')
        self.status_label.pack()
        
        # Informations
        info_frame = tk.Frame(main_frame, bg='#f0f0f0')
        info_frame.pack(fill='x', pady=20)
        
        info_text = "Format supporté : RELEVÉ BTK avec dates DD/MM/YYYY"
        tk.Label(info_frame, text=info_text, font=('Arial', 9), bg='#f0f0f0').pack()
        
        info_text2 = "Les fichiers Excel seront créés dans le dossier Téléchargements"
        tk.Label(info_frame, text=info_text2, font=('Arial', 9), bg='#f0f0f0').pack()
    
    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF RELEVÉ BTK"
            filetypes=[("Fichiers PDF", "*.pdf")]
        )
        if filename:
            self.pdf_path.set(filename)
    
    def is_btk_releve_pdf(self, pdf_path):
        """Vérifie si le PDF est un RELEVÉ BTK"""
        try:
            # Vérifier d'abord le nom du fichier
            filename_ok = self._check_btk_filename(pdf_path)
            print(f"DEBUG BTK RELEVÉ - Nom de fichier: {filename_ok}")
            
            # Vérifier les mots-clés textuels
            text_keywords_ok = self._check_btk_text_keywords(pdf_path)
            
            # Vérifier la structure du tableau
            table_structure_ok = self._check_btk_table_structure(pdf_path)
            
            # Vérifier la présence du logo (optionnel)
            logo_ok = self._check_btk_logo_presence(pdf_path)
            
            print(f"DEBUG BTK RELEVÉ - Mots-clés: {text_keywords_ok}, Structure: {table_structure_ok}, Logo: {logo_ok}")
            
            # Accepter si le nom de fichier est OK OU si on a les autres critères
            return filename_ok or (text_keywords_ok and table_structure_ok)
            
        except Exception as e:
            print(f"Erreur vérification BTK RELEVÉ: {e}")
            return False
    
    def _check_btk_filename(self, pdf_path):
        """Vérifie si le nom du fichier suggère un RELEVÉ BTK"""
        try:
            filename = os.path.basename(pdf_path).upper()
            print(f"DEBUG BTK RELEVÉ - Nom de fichier: {filename}")
            
            # Mots-clés dans le nom de fichier
            filename_keywords = ["BTK", "RELEVE", "RELEVÉ", "TUNISO", "KOWEITIENNE"]
            
            found_keywords = []
            for keyword in filename_keywords:
                if keyword in filename:
                    found_keywords.append(keyword)
                    print(f"DEBUG BTK RELEVÉ - Mot-clé trouvé dans le nom: {keyword}")
            
            # Accepter si on trouve au moins 2 mots-clés dans le nom
            return len(found_keywords) >= 2
            
        except Exception as e:
            print(f"Erreur vérification nom de fichier BTK RELEVÉ: {e}")
            return False
    
    def _check_btk_text_keywords(self, pdf_path):
        """Vérifie les mots-clés BTK RELEVÉ dans le texte"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages[:3]:  # Vérifier les 3 premières pages
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text.upper()
                
                print(f"DEBUG BTK RELEVÉ - Texte extrait (premiers 500 caractères): {text[:500]}")
                
                # Mots-clés spécifiques à BTK RELEVÉ - ÉTENDUS
                btk_keywords = [
                    "BTK", "BANQUE TUNISO-KOWEITIENNE", "RELEVÉ DE COMPTE"
                    "TUNISO-KOWEITIENNE", "TUNISO KOWEITIENNE", "BTKOTNTT"
                    "RELEVE", "RELEVÉ", "TUNISO KOWEITIENNE", "BANQUE TUNISO KOWEITIENNE"
                ]
                
                # Mots-clés essentiels - PLUS FLEXIBLES
                essential_keywords = ["BTK", "RELEVÉ", "RELEVE", "TND", "DINAR TUNISIEN", "TUNISO"]
                
                found_keywords = []
                essential_found = []
                
                for keyword in btk_keywords:
                    if keyword in text:
                        found_keywords.append(keyword)
                        print(f"DEBUG BTK RELEVÉ - Mot-clé trouvé: {keyword}")
                
                for keyword in essential_keywords:
                    if keyword in text:
                        essential_found.append(keyword)
                        print(f"DEBUG BTK RELEVÉ - Mot-clé essentiel trouvé: {keyword}")
                
                # Patterns spécifiques BTK RELEVÉ - PLUS FLEXIBLES
                btk_patterns = [
                    r"RELEVÉ DE COMPTE.*BTK"
                    r"BTK.*RELEVÉ DE COMPTE"
                    r"RELEVE DE COMPTE.*BTK"
                    r"BTK.*RELEVE DE COMPTE"
                    r"PÉRIODE DU.*AU.*\d{4}"
                    r"SOLDE AU.*\d{4}"
                    r"TUNISO.*KOWEITIENNE"
                    r"BTK.*BANQUE"
                ]
                
                pattern_found = False
                for pattern in btk_patterns:
                    if re.search(pattern, text):
                        pattern_found = True
                        print(f"DEBUG BTK RELEVÉ - Pattern BTK trouvé: {pattern}")
                        break
                
                # Vérifier aussi la présence de dates au format DD/MM/YYYY
                date_pattern = r'\d{1,2}/\d{1,2}/\d{4}'
                dates_found = re.findall(date_pattern, text)
                print(f"DEBUG BTK RELEVÉ - Dates trouvées: {dates_found[:5]}")  # Afficher les 5 premières
                
                print(f"DEBUG BTK RELEVÉ - Mots-clés trouvés: {found_keywords}")
                print(f"DEBUG BTK RELEVÉ - Mots-clés essentiels trouvés: {essential_found}")
                
                # Logique de détection plus permissive
                # Accepter si on trouve au moins 1 mot-clé essentiel OU 2 mots-clés généraux OU un pattern OU des dates
                detection_ok = (len(essential_found) >= 1 or 
                              len(found_keywords) >= 2 or 
                              pattern_found or 
                              len(dates_found) >= 3)
                
                print(f"DEBUG BTK RELEVÉ - Détection OK: {detection_ok}")
                return detection_ok
                
        except Exception as e:
            print(f"Erreur vérification mots-clés BTK RELEVÉ: {e}")
            return False
    
    def _check_btk_table_structure(self, pdf_path):
        """Vérifie la structure du tableau BTK RELEVÉ"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages[:3]):
                    tables = page.extract_tables()
                    if tables:
                        print(f"DEBUG BTK RELEVÉ - {len(tables)} tableaux trouvés sur la page {page_num + 1}")
                        
                        for table_idx, table in enumerate(tables):
                            if len(table) > 0:
                                # Vérifier les en-têtes de colonnes
                                header_row = table[0]
                                header_text = " ".join([str(cell) for cell in header_row if cell])
                                
                                print(f"DEBUG BTK RELEVÉ - Header du tableau: {header_text}")
                                
                                # Vérifier la présence de colonnes typiques BTK RELEVÉ - PLUS FLEXIBLE
                                if any(keyword in header_text.upper() for keyword in 
                                      ["DATE", "VALEUR", "LIBELLÉ", "LIBELLE", "DÉBIT", "DEBIT", "CRÉDIT", "CREDIT", "SOLDE"]):
                                    print("DEBUG BTK RELEVÉ - Structure de tableau BTK RELEVÉ détectée")
                                    return True
                    else:
                        # Si aucun tableau trouvé, vérifier le texte brut
                        text = page.extract_text()
                        if text:
                            # Chercher des patterns de transactions BTK RELEVÉ dans le texte - PLUS FLEXIBLE
                            if re.search(r'\d{1,2}/\d{1,2}/\d{4}\s+[A-Za-z\s]+\s+[\d\s.,]+', text):
                                print("DEBUG BTK RELEVÉ - Pattern de transaction BTK RELEVÉ détecté dans le texte")
                                return True
                            # Chercher aussi des patterns plus simples
                            if re.search(r'\d{1,2}/\d{1,2}/\d{4}', text):
                                print("DEBUG BTK RELEVÉ - Pattern de date BTK RELEVÉ détecté dans le texte")
                                return True
                            # Chercher des montants avec espaces (format BTK)
                            if re.search(r'\d{1,3}\s+\d{3},\d{3}', text):
                                print("DEBUG BTK RELEVÉ - Pattern de montant BTK RELEVÉ détecté dans le texte")
                                return True
            print("DEBUG BTK RELEVÉ - Aucune structure de tableau BTK détectée")
            return False
        except Exception as e:
            print(f"Erreur vérification structure BTK RELEVÉ: {e}")
            return False
    
    def _check_btk_logo_presence(self, pdf_path):
        """Vérifie la présence du logo BTK"""
        try:
            # Vérifier si le fichier logo existe
            logo_path = "logo/btk.png"
            if not os.path.exists(logo_path):
                print(f"DEBUG BTK RELEVÉ - Logo non trouvé à {logo_path}, détection basée sur le texte")
                return True  # Fallback sur la détection textuelle
            
            # Extraire les images du PDF
            doc = fitz.open(pdf_path)
            logo_found = False
            
            for page_num in range(min(3, len(doc))):  # Vérifier les 3 premières pages
                page = doc[page_num]
                image_list = page.get_images()
                
                if image_list:
                    print(f"DEBUG BTK RELEVÉ - {len(image_list)} images trouvées sur la page {page_num + 1}")
                    # Pour l'instant, on considère que la présence d'images indique un logo
                    logo_found = True
                    break
            
            doc.close()
            
            if logo_found:
                print("DEBUG BTK RELEVÉ - Images détectées dans le PDF (logo probable)")
            else:
                print("DEBUG BTK RELEVÉ - Aucune image détectée, mais on continue avec la détection textuelle")
            
            # Toujours retourner True car la détection textuelle est plus fiable
            return True
            
        except Exception as e:
            print(f"Erreur vérification logo BTK RELEVÉ: {e}")
            return True  # Fallback sur la détection textuelle
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF BTK RELEVÉ"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages[:3]):  # Chercher dans les 3 premières pages
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines[:30]:  # Chercher dans les 30 premières lignes
                            # Chercher des patterns d'année spécifiques au format BTK RELEVÉ
                            year_patterns = [
                                r'PÉRIODE DU.*?(\d{4})',  # "Période du 01/06/2025 Au 30/06/2025"
                                r'AU.*?(\d{4})',  # "Au 30/06/2025"
                                r'(\d{4})',  # Année simple
                                r'(\d{1,2}/\d{1,2}/(\d{4}))',  # Date avec année
                            ]
                            
                            for pattern in year_patterns:
                                matches = re.findall(pattern, line)
                                for match in matches:
                                    if isinstance(match, tuple):
                                        year = match[1] if len(match) > 1 else match[0]
                                    else:
                                        year = match
                                    
                                    year_int = int(year)
                                    
                                    if 2020 <= year_int <= 2030:  # Années plausibles
                                        print(f"DEBUG BTK RELEVÉ - Année détectée: {year_int} dans la ligne: {line}")
                                        return year_int
            
            print("DEBUG BTK RELEVÉ - Aucune année détectée, utilisation de l'année actuelle")
            return datetime.now().year
            
        except Exception as e:
            print(f"Erreur détection année BTK RELEVÉ: {e}")
            return datetime.now().year
    
    def is_scanned_pdf(self, pdf_path):
        """Détecter si le PDF est scanné (images) ou contient du texte extractible"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Vérifier la première page
                first_page = pdf.pages[0]
                text = first_page.extract_text()
                
                # Si très peu de texte ou pas de texte, c'est probablement scanné
                if not text or len(text.strip()) < 50:
                    print("DEBUG BTK RELEVÉ - PDF détecté comme scanné (peu de texte extractible)")
                    return True
                
                # Vérifier si le texte contient des mots-clés BTK
                text_upper = text.upper()
                btk_keywords = ["BTK", "RELEVE", "COMPTE", "SOLDE", "DEBIT", "CREDIT"]
                keyword_count = sum(1 for keyword in btk_keywords if keyword in text_upper)
                
                if keyword_count < 2:
                    print("DEBUG BTK RELEVÉ - PDF détecté comme scanné (pas de mots-clés BTK)")
                    return True
                
                print("DEBUG BTK RELEVÉ - PDF détecté comme texte extractible")
                return False
                
        except Exception as e:
            print(f"DEBUG BTK RELEVÉ - Erreur détection PDF: {e}")
            return True  # En cas d'erreur, supposer que c'est scanné

    def extract_text_with_ocr(self, pdf_path):
        """Extraire le texte d'un PDF scanné avec OCR"""
        if not TESSERACT_AVAILABLE:
            print("DEBUG BTK RELEVÉ - OCR non disponible, impossible de traiter le PDF scanné")
            return ""
        
        try:
            print("DEBUG BTK RELEVÉ - Début extraction OCR...")
            
            # Convertir PDF en images
            doc = fitz.open(pdf_path)
            all_text = ""
            
            for page_num in range(len(doc)):
                print(f"DEBUG BTK RELEVÉ - OCR page {page_num + 1}/{len(doc)}")
                
                # Convertir la page en image
                page = doc[page_num]
                mat = fitz.Matrix(2.0, 2.0)  # Augmenter la résolution
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                # Convertir en image PIL
                image = Image.open(io.BytesIO(img_data))
                
                # Préprocesser l'image pour améliorer l'OCR
                img_array = np.array(image)
                
                # Convertir en niveaux de gris
                if len(img_array.shape) == 3:
                    gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
                else:
                    gray = img_array
                
                # Améliorer le contraste
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                enhanced = clahe.apply(gray)
                
                # Appliquer un seuil pour binariser
                _, binary = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                
                # OCR avec Tesseract
                custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,/-: '
                page_text = pytesseract.image_to_string(binary, config=custom_config, lang='fra')
                
                all_text += page_text + "\n"
                print(f"DEBUG BTK RELEVÉ - Texte OCR extrait page {page_num + 1}: {len(page_text)} caractères")
            
            doc.close()
            print(f"DEBUG BTK RELEVÉ - Extraction OCR terminée: {len(all_text)} caractères total")
            return all_text
            
        except Exception as e:
            print(f"DEBUG BTK RELEVÉ - Erreur OCR: {e}")
            return ""

    def extract_table_data(self, pdf_path):
        """Extrait les données du tableau des transactions BTK RELEVÉ (texte ou scanné)"""
        try:
            print("DEBUG BTK RELEVÉ - Début extraction données, fichier:", pdf_path)
            # Détecter si le PDF est scanné
            if self.is_scanned_pdf(pdf_path):
                print("DEBUG BTK RELEVÉ - PDF scanné détecté, utilisation OCR")
                # Extraire le texte avec OCR
                ocr_text = self.extract_text_with_ocr(pdf_path)
                if ocr_text:
                    # Parser le texte OCR comme du texte normal
                    year = self.detect_year_from_pdf(pdf_path)
                    print(f"DEBUG BTK RELEVÉ - Année détectée: {year}")
                    transactions = self.parse_btk_transactions_from_text(ocr_text, year)
                    # Si toujours vide, parser de façon ultra-simple
                    if not transactions:
                        transactions = self.parse_btk_text_simple(ocr_text, year)
                    print("DEBUG BTK RELEVÉ - Aperçu OCR (400 char):", ocr_text[:400].replace('\n',' '))
                    print(f"DEBUG BTK RELEVÉ - Transactions extraites de l'OCR: {len(transactions)}")
                    return transactions
                else:
                    print("DEBUG BTK RELEVÉ - Échec extraction OCR")
                    return []
            else:
                print("DEBUG BTK RELEVÉ - PDF texte détecté, extraction normale")
                # Logique normale pour PDF avec texte extractible
                with pdfplumber.open(pdf_path) as pdf:
                    all_transactions = []
                
                    for page_num, page in enumerate(pdf.pages):
                        print(f"DEBUG BTK RELEVÉ - Traitement page {page_num + 1}")
                        
                        # Essayer d'extraire les tableaux
                        tables = page.extract_tables()
                        
                        if tables:
                            print(f"DEBUG BTK RELEVÉ - {len(tables)} tableaux trouvés sur la page {page_num + 1}")
                            table_transactions_found = False
                            for table_idx, table in enumerate(tables):
                                print(f"DEBUG BTK RELEVÉ - Traitement tableau {table_idx + 1}")
                                transactions = self.parse_btk_table(table)
                                if transactions:  # Si on trouve des transactions dans le tableau
                                    all_transactions.extend(transactions)
                                    table_transactions_found = True
                                    print(f"DEBUG BTK RELEVÉ - {len(transactions)} transactions extraites du tableau")
                            
                            # Si aucun tableau n'a donné de transactions, essayer l'extraction de texte
                            if not table_transactions_found:
                                print(f"DEBUG BTK RELEVÉ - Aucune transaction trouvée dans les tableaux, extraction du texte")
                                text = page.extract_text()
                                if text:
                                    print(f"DEBUG BTK RELEVÉ - Texte extrait (premiers 200 caractères): {text[:200]}")
                                    year = self.detect_year_from_pdf(pdf_path)
                                    print(f"DEBUG BTK RELEVÉ - Année détectée: {year}")
                                    transactions = self.parse_btk_transactions_from_text(text, year)
                                    if not transactions:
                                        transactions = self.parse_btk_text_simple(text, year)
                                    print(f"DEBUG BTK RELEVÉ - Transactions extraites du texte: {len(transactions)}")
                                    all_transactions.extend(transactions)
                                else:
                                    print("DEBUG BTK RELEVÉ - Aucun texte extrait de la page")
                        else:
                            # Si aucun tableau trouvé, essayer d'extraire le texte brut directement
                            print(f"DEBUG BTK RELEVÉ - Aucun tableau trouvé sur la page {page_num + 1}, extraction du texte")
                            text = page.extract_text()
                            if text:
                                print(f"DEBUG BTK RELEVÉ - Texte extrait (premiers 200 caractères): {text[:200]}")
                                year = self.detect_year_from_pdf(pdf_path)
                                transactions = self.parse_btk_transactions_from_text(text, year)
                                if not transactions:
                                    transactions = self.parse_btk_text_simple(text, year)
                                print(f"DEBUG BTK RELEVÉ - Transactions extraites du texte: {len(transactions)}")
                                all_transactions.extend(transactions)
                            else:
                                print("DEBUG BTK RELEVÉ - Aucun texte extrait de la page")
                    
                    print(f"DEBUG BTK RELEVÉ - Total transactions extraites: {len(all_transactions)}")
                    # Fallback global: si aucune transaction trouvée, concaténer tout le texte et parser une fois
                    if not all_transactions:
                        print("DEBUG BTK RELEVÉ - Fallback global: concaténation de tout le texte du PDF")
                        try:
                            full_text = "\n".join([p.extract_text() or "" for p in pdf.pages])
                            year = self.detect_year_from_pdf(pdf_path)
                            more = self.parse_btk_transactions_from_text(full_text, year)
                            if not more:
                                more = self.parse_btk_text_simple(full_text, year)
                            print(f"DEBUG BTK RELEVÉ - Transactions extraites via fallback global: {len(more)}")
                            all_transactions.extend(more)
                        except Exception as e:
                            print(f"DEBUG BTK RELEVÉ - Erreur fallback global: {e}")
                # Dernier recours: OCR même si document détecté texte
                if not all_transactions:
                    print("DEBUG BTK RELEVÉ - Dernier recours: lancer OCR")
                    ocr_text = self.extract_text_with_ocr(pdf_path)
                    if ocr_text:
                        year = self.detect_year_from_pdf(pdf_path)
                        ocr_tx = self.parse_btk_transactions_from_text(ocr_text, year)
                        if not ocr_tx:
                            ocr_tx = self.parse_btk_text_simple(ocr_text, year)
                        print(f"DEBUG BTK RELEVÉ - OCR fallback a extrait {len(ocr_tx)} transactions")
                        all_transactions.extend(ocr_tx)
                    return all_transactions
                
        except Exception as e:
            print(f"Erreur extraction données BTK RELEVÉ: {e}")
            return []

    def parse_btk_text_simple(self, text, year):
        """Parser très tolérant pour BTK (scanné):
        - Chaque transaction commence par une date (DD/MM/YYYY)
        - Montants à la fin: Débit puis Crédit OU un seul montant
        - On ignore la colonne Solde
        """
        transactions = []
        try:
            import re
            lines = [ln.strip() for ln in text.split('\n') if ln.strip()]
            date_re = re.compile(r"^(\d{2}/\d{2}/\d{4})")
            amount_re = re.compile(r"-?[\d ]+[.,]\d{3}")
            current = None
            for line in lines:
                m = date_re.match(line)
                if m:
                    # Finaliser la précédente
                    if current and (current.get('debit') or current.get('credit')):
                        transactions.append({
                            'Date': current['date']
                            'Libellé': current['libelle'].strip()
                            'Débit': current.get('debit') or ''
                            'Crédit': current.get('credit') or ''
                        })
                    current = {'date': m.group(1), 'libelle': line[m.end():].strip(), 'debit': '', 'credit': ''}
                else:
                    if current:
                        # continuation libellé
                        current['libelle'] = (current['libelle'] + ' ' + line).strip()
                # essayer d'extraire les montants à la fin de la ligne
                amts = list(amount_re.finditer(line))
                if current and amts:
                    # prendre les 1 ou 2 derniers montants
                    nums = [line[m.start():m.end()] for m in amts[-2:]]
                    def fmt(s):
                        v = s.replace(' ', '').replace('.', '').replace(',', '.')
                        try:
                            f = float(v)
                        except Exception:
                            return ''
                        return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
                    if len(nums) == 2:
                        current['debit'] = fmt(nums[0])
                        current['credit'] = fmt(nums[1])
                    else:
                        # un seul montant -> utiliser heuristique: si libellé contient mots débit
                        lib_up = current['libelle'].upper()
                        debit_hint = any(k in lib_up for k in ['FRAIS', 'AGIOS', 'RET', 'PRELEV', 'PAIEMENT'])
                        if debit_hint:
                            current['debit'] = fmt(nums[0])
                        else:
                            current['credit'] = fmt(nums[0])
            if current and (current.get('debit') or current.get('credit')):
                transactions.append({
                    'Date': current['date']
                    'Libellé': current['libelle'].strip()
                    'Débit': current.get('debit') or ''
                    'Crédit': current.get('credit') or ''
                })
        except Exception as e:
            print(f"DEBUG BTK SIMPLE PARSE ERROR: {e}")
        return transactions
    
    def extract_text_with_ocr(self, page):
        """Extrait le texte d'une page PDF avec OCR"""
        try:
            if not TESSERACT_AVAILABLE:
                return None
            
            # Convertir la page en image
            image = page.to_image(resolution=300)
            pil_image = image.original
            
            # Convertir en format OpenCV
            opencv_image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
            
            # Améliorer l'image pour l'OCR
            gray = cv2.cvtColor(opencv_image, cv2.COLOR_BGR2GRAY)
            denoised = cv2.medianBlur(gray, 3)
            
            # Appliquer un seuil pour améliorer la lisibilité
            _, thresh = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Convertir back en PIL pour tesseract
            pil_thresh = Image.fromarray(thresh)
            
            # Extraire le texte avec tesseract
            text = pytesseract.image_to_string(pil_thresh, lang='fra')
            
            return text
            
        except Exception as e:
            print(f"Erreur OCR BTK RELEVÉ: {e}")
            return None
    
    def parse_btk_table(self, table):
        """Parse un tableau BTK RELEVÉ extrait par pdfplumber"""
        transactions = []
        try:
            print(f"DEBUG BTK RELEVÉ - Tableau reçu avec {len(table)} lignes")
            
            if len(table) < 2:
                print("DEBUG BTK RELEVÉ - Tableau trop petit")
                return transactions
            
            # Afficher la structure du tableau
            for i, row in enumerate(table[:3]):  # Afficher les 3 premières lignes
                print(f"DEBUG BTK RELEVÉ - Ligne {i}: {row}")
            
            # Détecter l'année depuis le contenu du tableau
            year = None
            for row in table:
                for cell in row:
                    if cell and isinstance(cell, str):
                        year_match = re.search(r'(\d{4})', cell)
                        if year_match:
                            year_int = int(year_match.group(1))
                            if 2020 <= year_int <= 2030:
                                year = year_int
                                print(f"DEBUG BTK RELEVÉ - Année détectée dans le tableau: {year}")
                                break
                if year:
                    break
            
            if not year:
                year = datetime.now().year
                print(f"DEBUG BTK RELEVÉ - Année par défaut: {year}")
            
            # Parser les lignes de données
            for i, row in enumerate(table[1:], 1):  # Ignorer la première ligne (header)
                if row and any(cell for cell in row if cell):
                    print(f"DEBUG BTK RELEVÉ - Traitement ligne {i}: {row}")
                    
                    # Vérifier si c'est une ligne de total ou d'en-tête
                    row_text = " ".join([str(cell) for cell in row if cell])
                    if any(keyword in row_text.upper() for keyword in [
                        "TOTAL", "SOLDE", "REPORT", "PAGE", "TITULAIRE", "COMPTE"
                    ]):
                        print(f"DEBUG BTK RELEVÉ - Ligne ignorée (total/en-tête): {row_text}")
                        continue
                    
                    transaction = self.parse_btk_transaction_row(row, year)
                    if transaction:
                        transactions.append(transaction)
                        print(f"DEBUG BTK RELEVÉ - Transaction ajoutée: {transaction}")
                    else:
                        print(f"DEBUG BTK RELEVÉ - Ligne ignorée (pas de transaction valide)")
            
        except Exception as e:
            print(f"Erreur parsing tableau BTK RELEVÉ: {e}")
        
        print(f"DEBUG BTK RELEVÉ - Total transactions extraites du tableau: {len(transactions)}")
        return transactions
    
    def parse_btk_transaction_row(self, row, year):
        """Parse une ligne de transaction BTK RELEVÉ avec colonnes séparées"""
        try:
            # Nettoyer la ligne - garder toutes les colonnes même vides
            clean_row = [str(cell).strip() if cell else "" for cell in row]
            
            print(f"DEBUG BTK RELEVÉ - Ligne brute: {row}")
            print(f"DEBUG BTK RELEVÉ - Ligne nettoyée: {clean_row}")
            print(f"DEBUG BTK RELEVÉ - Nombre de colonnes: {len(clean_row)}")
            
            # Structure BTK RELEVÉ: Date, Valeur, Libellé, Débit, Crédit, Solde
            if len(clean_row) < 4:
                print(f"DEBUG BTK RELEVÉ - Pas assez de colonnes: {len(clean_row)}")
                return None
            
            date_ope = clean_row[0]
            date_valeur = clean_row[1] if len(clean_row) > 1 else ""
            libelle = clean_row[2] if len(clean_row) > 2 else ""
            debit = clean_row[3] if len(clean_row) > 3 else ""
            credit = clean_row[4] if len(clean_row) > 4 else ""
            
            # Si la date est vide, essayer de la récupérer depuis la colonne suivante
            if not date_ope and len(clean_row) > 1:
                date_ope = clean_row[1]
                date_valeur = clean_row[2] if len(clean_row) > 2 else ""
                libelle = clean_row[3] if len(clean_row) > 3 else ""
                debit = clean_row[4] if len(clean_row) > 4 else ""
                credit = clean_row[5] if len(clean_row) > 5 else ""
            
            print(f"DEBUG BTK RELEVÉ - Date opé: '{date_ope}', Libellé: '{libelle[:50]}...', Débit: '{debit}', Crédit: '{credit}'")
            
            # Vérifier si c'est une date valide (format DD/MM/YYYY)
            if not self.is_date_btk_releve(date_ope):
                print(f"DEBUG BTK RELEVÉ - Date invalide: '{date_ope}'")
                return None
            
            # Formater la date
            date_formatted = self.format_date_btk_releve(date_ope)
            print(f"DEBUG BTK RELEVÉ - Date formatée: '{date_formatted}'")
            
            # Déterminer si c'est un débit ou crédit
            montant_debit = None
            montant_credit = None
            
            if debit and debit != '' and debit != '0':
                # Nettoyer le montant débit
                debit_clean = debit.replace(' ', '').replace(',', '.')
                try:
                    montant_debit = float(debit_clean)
                    print(f"DEBUG BTK RELEVÉ - Montant débit: {montant_debit}")
                except:
                    print(f"DEBUG BTK RELEVÉ - Erreur parsing débit: '{debit}'")
            
            if credit and credit != '' and credit != '0':
                # Nettoyer le montant crédit
                credit_clean = credit.replace(' ', '').replace(',', '.')
                try:
                    montant_credit = float(credit_clean)
                    print(f"DEBUG BTK RELEVÉ - Montant crédit: {montant_credit}")
                except:
                    print(f"DEBUG BTK RELEVÉ - Erreur parsing crédit: '{credit}'")
            
            # Vérifier qu'on a au moins un montant
            if not montant_debit and not montant_credit:
                print(f"DEBUG BTK RELEVÉ - Aucun montant valide trouvé")
                return None
            
            return {
                'date': date_formatted
                'libelle': libelle
                'debit': montant_debit
                'credit': montant_credit
            }
            
        except Exception as e:
            print(f"Erreur parsing ligne BTK RELEVÉ: {e}")
            return None
    
    def is_date_btk_releve(self, date_str):
        """Vérifie si la chaîne est une date BTK RELEVÉ valide (DD/MM/YYYY)"""
        if not date_str:
            return False
        # Pattern pour DD/MM/YYYY
        pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'
        return bool(re.match(pattern, date_str.strip()))
    
    def format_date_btk_releve(self, date_str):
        """Formate une date BTK RELEVÉ (DD/MM/YYYY -> DD/MM/YYYY)"""
        try:
            # La date est déjà au bon format DD/MM/YYYY
            return date_str.strip()
        except:
            return "01/01/2025"
    
    def clean_ocr_date(self, date_str):
        """Nettoie une date extraite par OCR"""
        try:
            # Remplacer les caractères OCR erronés
            cleaned = date_str.replace(':', '/').replace('.', '/')
            
            # Extraire les parties de la date
            parts = cleaned.split('/')
            if len(parts) == 3:
                day = parts[0].strip()
                month = parts[1].strip()
                year = parts[2].strip()
                
                # Nettoyer les caractères OCR
                day = re.sub(r'[^\d]', '', day)
                month = re.sub(r'[^\d]', '', month)
                year = re.sub(r'[^\d]', '', year)
                
                # Valider et formater
                if len(day) == 1:
                    day = '0' + day
                if len(month) == 1:
                    month = '0' + month
                if len(year) == 2:
                    year = '20' + year
                
                return f"{day}/{month}/{year}"
            
            return date_str
        except:
            return date_str
    
    def clean_ocr_description(self, description):
        """Nettoie une description extraite par OCR"""
        try:
            # Remplacer les caractères OCR bizarres par des caractères corrects
            replacements = {
                'mVIREMENTEMISMEM': 'VIREMENT EMIS MEM'
                'wmCOMXTVAVIR.EMIS': 'COM & TVA VIR. EMIS'
                'WVIREMENTEMISAUTBQ': 'VIREMENT EMIS AUT BQ'
                'TUNISIANADU': 'TUNISIANA DU'
                'EBRETRAITESP.DEPLACE': 'RETRAIT ESP. DEPLACE'
                'PRELEV.RECUAUTBQE': 'PRELEV. RECU AUT BQE'
                'AGIOSDU30/09/2024AU': 'AGIOS DU 30/09/2024 AU'
                'RETRAII': 'RETRAIT'
                'REMÆEFF.ES': 'REM.EFF.ES'
                'INT/ESCEFF BC SP': 'INT/ESC EFF BQ SP'
                'COM&TVA AUTORISATI': 'COM & TVA AUTORISATI'
                'VIREMENTEMISAUTBQ': 'VIREMENT EMIS AUT BQ'
                'AGIOS OU': 'AGIOS DU'
            }
            
            # Appliquer les remplacements
            for wrong, correct in replacements.items():
                description = description.replace(wrong, correct)
            
            # Nettoyer les caractères bizarres restants
            description = re.sub(r'[^\w\s&/.-]', ' ', description)
            description = re.sub(r'\s+', ' ', description).strip()
            
            return description
        except:
            return description

    def is_solde_amount(self, amount, line, all_amounts):
        """Détecte si un montant est un montant de solde (à filtrer).
        Ne filtre que si la ligne contient explicitement des mots-clés de solde/total
        """
        try:
            line_upper = line.upper()

            # Filtrer uniquement si la ligne contient des indicateurs explicites de solde/total
            if re.search(r"\b(SOLDE|BALANCE|TOTAL)\b", line_upper):
                return True

            return False
        except Exception:
            return False
    
    def extract_btk_description(self, line, date_str, amounts):
        """Extrait la description d'une ligne BTK RELEVÉ"""
        try:
            # Patterns de libellés typiques BTK RELEVÉ - plus flexibles pour OCR
            libelle_patterns = [
                r'RETRAIT ESPECE \d+'
                r'RETRAII ESPECE \d+',  # Variante OCR
                r'PAIEMENT EFFET \d+'
                r'COM & TVA PAI\. EFFET \d+'
                r'COM&TVAPAIEFFET \d+',  # Variante OCR
                r'PRELEV\. RECU AUT BQE \d+'
                r'REM\.EFF\.ES \d+'
                r'REMÆEFF\.ES \d+',  # Variante OCR
                r'INT/ESC EFF BQ SP \d+'
                r'INT/ESCEFF BC SP \d+',  # Variante OCR
                r'COM & TVA AUTORISATI \d+'
                r'COM&TVA AUTORISATI \d+',  # Variante OCR
                r'VIREMENT EMIS AUT BQ'
                r'VIREMENTEMISAUTBQ',  # Variante OCR
                r'AGIOS DU \d{2}/\d{2}/\d{2} AU \d{2}/\d{2}/\d{2}'
                r'AGIOS OU \d{2}/\d{2}/\d{2} AU \d{2}/\d{2}/\d{2}',  # Variante OCR
                r'Commission Virement'
                r'TVA/Com Virement'
                r'En faveur \w+'
            ]
            
            # Chercher les patterns de libellé dans la ligne
            for pattern in libelle_patterns:
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    description = match.group(0)
                    # Nettoyer la description des caractères OCR bizarres
                    description = self.clean_ocr_description(description)
                    return description
            
            # Si aucun pattern trouvé, essayer d'extraire manuellement
            # Enlever la date et les montants de la ligne
            description = line
            
            # Enlever la date
            description = description.replace(date_str, '', 1)
            
            # Enlever les montants (en commençant par la fin pour éviter les conflits)
            for amount in reversed(amounts):
                if amount in description:
                    description = description.replace(amount, '', 1)
            
            # Enlever les caractères de séparation OCR
            description = re.sub(r'[|;:\'"]+', ' ', description)
            description = re.sub(r'\s+', ' ', description).strip()
            
            # Si la description est trop courte ou contient des caractères suspects, essayer une approche différente
            if len(description) < 5 or re.search(r'[^\w\s/.-]', description):
                # Chercher des mots-clés connus
                keywords = ['RETRAIT', 'PAIEMENT', 'PRELEV', 'COM', 'TVA', 'VIREMENT', 'AGIOS', 'EFFET', 'ESPECE']
                for keyword in keywords:
                    if keyword in line.upper():
                        # Extraire autour du mot-clé
                        start = line.upper().find(keyword)
                        if start >= 0:
                            # Prendre 50 caractères autour du mot-clé
                            start = max(0, start - 10)
                            end = min(len(line), start + 60)
                            description = line[start:end]
                            description = re.sub(r'[|;:\'"]+', ' ', description)
                            description = re.sub(r'\s+', ' ', description).strip()
                            break
            
            return description
            
        except Exception as e:
            print(f"Erreur extraction description BTK RELEVÉ: {e}")
            return ""
    
    def parse_btk_transactions_from_text(self, text, year):
        """Parse les transactions BTK RELEVÉ depuis le texte brut"""
        transactions = []
        seen_transactions = set()
        
        try:
            print(f"DEBUG BTK RELEVÉ - Parsing du texte brut (premiers 1000 caractères): {text[:1000]}")
            
            # Pattern flexible pour BTK RELEVÉ - dates avec erreurs OCR
            btk_date_pattern = re.compile(r'(\d{1,2}[:\/\.]\d{1,2}[:\/\.]\d{4})')
            
            # Pattern pour extraire les montants BTK RELEVÉ (format avec espaces)
            btk_amount_pattern = re.compile(r'(\d{1,3}\s+\d{3},\d{3}|\d+,\d{3}|\d+\.\d{3})')
            
            # Pour BTK RELEVÉ, utiliser les lignes brutes directement
            lines = [ln.strip() for ln in text.split('\n')]
            print(f"DEBUG BTK RELEVÉ - {len(lines)} lignes de texte à analyser")
            
            for line_num, line in enumerate(lines):
                line = line.strip()
                if not line or len(line) < 10:
                    continue
                
                print(f"DEBUG BTK RELEVÉ - Ligne {line_num}: {line}")
                
                # Ignorer seulement les en-têtes vraiment non-transactionnels
                if any(header in line.upper() for header in [
                    'DATE VALEUR', 'LIBELLÉ', 'DÉBIT CRÉDIT', 'TOTAUX', 'SOLDE'
                    'BTK RELEVÉ', 'COMPTE', 'RIB:', 'TITULAIRE', 'CLIENT'
                    'PÉRIODE DU', 'NUMRODE', 'DEVISE:', 'CATEGORIE', 'SOLDEDEPART'
                    'PAGE', 'SAUFERREUR', 'TOTAL DES OPÉRATIONS', 'NOUVEAU SOLDE'
                    'SOLDE AU', 'IDENTIFIANT NATIONAL', 'BANQUE AGENCE', 'IBAN BIC'
                    'MONTPLAISIR', 'TUNIS', 'TUNISIE', 'CITÉ EL KHADHRA'
                ]):
                    print(f"DEBUG BTK RELEVÉ - Ligne ignorée (en-tête): {line}")
                    continue
                
                # Fonction helper pour nettoyer les montants
                def clean_amount(amount_str):
                    """Nettoie et convertit un montant en float"""
                    if not amount_str or amount_str.strip() == '':
                        return None
                    try:
                        # Remplacer virgule par point et enlever espaces
                        clean_amount = amount_str.replace(',', '.').replace(' ', '')
                        if clean_amount == '':
                            return None
                        
                        # Gérer les montants négatifs
                        is_negative = False
                        if clean_amount.startswith('-'):
                            is_negative = True
                            clean_amount = clean_amount[1:]
                        
                        amount = float(clean_amount)
                        if is_negative:
                            amount = -amount
                        
                        return amount
                    except ValueError:
                        return None
                
                # Chercher une date dans la ligne
                date_match = btk_date_pattern.search(line)
                if not date_match:
                    # Essayer de chercher des dates plus flexibles
                    flexible_date_pattern = re.compile(r'(\d{1,2}[:\/\.]\d{1,2}[:\/\.]\d{2,4})')
                    date_match = flexible_date_pattern.search(line)
                    if not date_match:
                        print(f"DEBUG BTK RELEVÉ - Aucune date trouvée dans la ligne")
                        continue
                
                date_str = date_match.group(1)
                # Nettoyer la date OCR
                date_str = self.clean_ocr_date(date_str)
                print(f"DEBUG BTK RELEVÉ - Date trouvée et nettoyée: {date_str}")
                
                # Analyser la structure de la ligne BTK RELEVÉ
                # Format typique: Date Date_Valeur Libellé Débit Crédit Solde
                parts = line.split()
                
                # Chercher tous les montants dans la ligne
                amounts = btk_amount_pattern.findall(line)
                print(f"DEBUG BTK RELEVÉ - Montants trouvés: {amounts}")
                
                # Extraire la description de manière plus intelligente
                description = self.extract_btk_description(line, date_str, amounts)
                
                print(f"DEBUG BTK RELEVÉ - Description extraite: '{description}'")
                
                # Si pas de description, essayer d'extraire manuellement
                if not description or len(description.strip()) < 3:
                    # Enlever la date et les montants pour isoler la description
                    temp_line = line
                    temp_line = temp_line.replace(date_str, '', 1)
                    for amount in amounts:
                        temp_line = temp_line.replace(amount, '', 1)
                    
                    # Nettoyer et extraire la description
                    description = re.sub(r'[^\w\s/.-]', ' ', temp_line)
                    description = re.sub(r'\s+', ' ', description).strip()
                    
                    print(f"DEBUG BTK RELEVÉ - Description extraite manuellement: '{description}'")

                if not description or len(description.strip()) < 2:
                    print(f"DEBUG BTK RELEVÉ - Description vide, ligne ignorée")
                    continue
                
                # Ignorer les lignes contenant "Report du" (plus flexible)
                if re.search(r'(?i)report\s+du|report\s+au|report\s+le', description):
                    print(f"DEBUG BTK RELEVÉ - Ligne 'Report' ignorée: {description}")
                    continue
                
                # Supprimer les dates au début du libellé uniquement si elles sont suivies de texte (évite de supprimer une vraie date)
                description_before = description
                description = re.sub(r'^\d{1,2}[\s/.-]\d{1,2}[\s/.-]\d{2,4}\s+(?=[A-Za-zÀ-ÿ])', '', description).strip()
                if description != description_before:
                    print(f"DEBUG BTK RELEVÉ - Date supprimée du libellé: '{description_before}' -> '{description}'")
                
                # Supprimer aussi les variantes si suivies de texte
                description_before = description
                description = re.sub(r'^\d{1,2}[\s/.-]\d{1,2}[\s/.-]\d{2,4}\s*(?=[A-Za-zÀ-ÿ])', '', description).strip()
                if description != description_before:
                    print(f"DEBUG BTK RELEVÉ - Date alternative supprimée du libellé: '{description_before}' -> '{description}'")
                
                # Supprimer les dates partielles (ex: "19/11/2 -" -> "-")
                description_before = description
                description = re.sub(r'^\d{1,2}[\s/.-]\d{1,2}[\s/.-]\d{1}\s*-\s*$', '-', description).strip()
                if description != description_before:
                    print(f"DEBUG BTK RELEVÉ - Date partielle supprimée du libellé: '{description_before}' -> '{description}'")
                
                # Nettoyer les libellés qui ne contiennent que des montants ou des chiffres
                if re.match(r'^[\d\s,.-]+$', description) and len(description) > 10:
                    print(f"DEBUG BTK RELEVÉ - Libellé contenant seulement des chiffres ignoré: {description}")
                    continue

                # Séparer les montants du libellé pour éviter les mélanges
                # Chercher des montants à la fin du libellé et les extraire
                libelle_clean = description
                montant_at_end_pattern = r'(.+?)\s+([\d\s,.-]+)$'
                libelle_match = re.match(montant_at_end_pattern, libelle_clean)
                
                if libelle_match:
                    libelle_text = libelle_match.group(1).strip()
                    potential_amount = libelle_match.group(2).strip()
                    
                    # Vérifier si c'est un vrai montant (avec virgule ou point décimal)
                    if re.search(r'[\d]+[,.]\d+', potential_amount) and len(potential_amount) > 3:
                        # C'est un vrai montant, le garder dans le libellé mais ne pas l'utiliser comme montant de transaction
                        print(f"DEBUG BTK RELEVÉ - Montant détecté dans libellé (gardé): {potential_amount}")
                        libelle_clean = libelle_text
                    else:
                        # C'est probablement un numéro de référence, le garder dans le libellé
                        print(f"DEBUG BTK RELEVÉ - Numéro de référence dans libellé (gardé): {potential_amount}")
                
                description = libelle_clean
                
                # Traiter les montants - filtrer les montants de solde
                valid_amounts = []
                for amount_str in amounts:
                    amount = self.clean_amount(amount_str)
                    if amount is not None and amount > 0 and amount < 1000000:  # Éviter les montants aberrants
                        # Filtrer les montants de solde (généralement très grands et croissants/décroissants)
                        if not self.is_solde_amount(amount, line, amounts):
                            valid_amounts.append(amount)
                        else:
                            print(f"DEBUG BTK RELEVÉ - Montant de solde filtré: {amount}")
                
                print(f"DEBUG BTK RELEVÉ - Montants valides: {valid_amounts}")
                
                # Si pas de montant valide, essayer de trouver des montants dans la ligne complète
                if not valid_amounts:
                    # Chercher des montants dans la ligne complète (pas seulement la description)
                    line_amounts = re.findall(r'(\d{1,3}\s+\d{3},\d{3}|\d+,\d{3})', line)
                    for line_amount in line_amounts:
                        amount = self.clean_amount(line_amount)
                        if amount is not None and amount > 0 and amount < 1000000:
                            # Filtrer les montants de solde
                            if not self.is_solde_amount(amount, line, line_amounts):
                                valid_amounts.append(amount)
                            else:
                                print(f"DEBUG BTK RELEVÉ - Montant de solde filtré (ligne): {amount}")
                
                if not valid_amounts:
                    # Essayer de trouver des montants plus simples
                    simple_amounts = re.findall(r'(\d+[.,]\d+)', line)
                    for simple_amount in simple_amounts:
                        amount = self.clean_amount(simple_amount)
                        if amount is not None and amount > 0 and amount < 10000000:
                            # Filtrer les montants de solde
                            if not self.is_solde_amount(amount, line, simple_amounts):
                                valid_amounts.append(amount)
                            else:
                                print(f"DEBUG BTK RELEVÉ - Montant de solde filtré (simple): {amount}")
                    
                    if not valid_amounts:
                        print(f"DEBUG BTK RELEVÉ - Aucun montant valide trouvé")
                        continue
                
                # Prendre le montant le plus probable (généralement le plus grand et raisonnable)
                # Filtrer les montants trop petits ou trop grands
                reasonable_amounts = [a for a in valid_amounts if 1 <= a <= 10000000]
                if reasonable_amounts:
                    amount = max(reasonable_amounts)
                else:
                    amount = max(valid_amounts)
                
                # Classification débit/crédit corrigée pour BTK RELEVE
                desc_upper = description.upper()
                
                # Mots-clés qui indiquent un DÉBIT (sortie d'argent) - BTK RELEVE
                debit_keywords = [
                    "COMM", "RETRAIT", "PAIEMENT", "PRELEV", "COM", "TVA", "AGIOS"
                    "VIREMENT EMIS", "DEBIT", "FRAIS", "INT/ESC", "REM.EFF.ES"
                    "COMMISSION", "PRELEVEMENT", "RETRAIT ESPECE", "PAIEMENT EFFET"
                    "COM & TVA", "VIREMENT EMIS AUT BQ", "AGIOS DU", "REGLEMENT INTERET"
                    "INTERET DE RETARD", "INTERET EN IMP", "RECOUVREMENT IMPAYES"
                ]
                
                # Mots-clés qui indiquent un CRÉDIT (entrée d'argent) - BTK RELEVE
                credit_keywords = [
                    "VIREMENT RECU", "CREDIT", "REMISE", "DEPOT", "ENCAISSEMENT", "TRS TPE"
                    "VIREMENT RECU AUT BQ", "REMISE EFFET", "ENCAISSEMENT EFFET"
                    "VERST ESP DEP"
                ]
                
                # Classification basée sur les mots-clés avec priorité
                is_debit = any(keyword in desc_upper for keyword in debit_keywords)
                is_credit = any(keyword in desc_upper for keyword in credit_keywords)
                
                # Priorité aux mots-clés spécifiques
                if is_debit and not is_credit:
                    debit, credit = amount, None
                    print(f"DEBUG BTK RELEVÉ - Classification DÉBIT: {description[:50]}...")
                elif is_credit and not is_debit:
                    debit, credit = None, amount
                    print(f"DEBUG BTK RELEVÉ - Classification CRÉDIT: {description[:50]}...")
                elif is_debit and is_credit:
                    # Conflit : priorité aux mots-clés les plus spécifiques
                    if any(keyword in desc_upper for keyword in ["REGLEMENT INTERET", "INTERET DE RETARD", "INTERET EN IMP", "COM & TVA", "RECOUVREMENT IMPAYES"]):
                        debit, credit = amount, None
                        print(f"DEBUG BTK RELEVÉ - Classification DÉBIT (mots-clés spécifiques): {description[:50]}...")
                    else:
                        debit, credit = None, amount
                        print(f"DEBUG BTK RELEVÉ - Classification CRÉDIT (conflit résolu): {description[:50]}...")
                else:
                    # Fallback: analyser le montant et le contexte
                    if amount < 100:
                        debit, credit = amount, None
                        print(f"DEBUG BTK RELEVÉ - Classification DÉBIT par défaut (montant petit): {amount}")
                    else:
                        debit, credit = None, amount
                        print(f"DEBUG BTK RELEVÉ - Classification CRÉDIT par défaut (montant grand): {amount}")
                
                # Créer une clé unique pour éviter les doublons
                transaction_key = f"{date_str}|{description}|{debit}|{credit}"
                
                if transaction_key not in seen_transactions:
                    seen_transactions.add(transaction_key)
                    transaction = {
                        "date": date_str
                        "libelle": description
                        "debit": debit
                        "credit": credit
                    }
                    transactions.append(transaction)
                    print(f"DEBUG BTK RELEVÉ - Transaction ajoutée: {transaction}")
                else:
                    print(f"DEBUG BTK RELEVÉ - Transaction en doublon ignorée")
            
            print(f"DEBUG BTK RELEVÉ - Total transactions extraites du texte: {len(transactions)}")
            return transactions
            
        except Exception as e:
            print(f"Erreur parsing texte BTK RELEVÉ: {e}")
            return []

def convert_to_excel(self):
        """Convertit le PDF en Excel"""
        if not self.pdf_path.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
            return
        
        if not self.excel_name.get():
            messagebox.showerror("Erreur", "Veuillez entrer un nom pour le fichier Excel")
            return
        
        try:
            
            self.convert_button.config(state='disabled')
            
            self.root.update()
            
            # Vérifier que c'est un PDF BTK RELEVÉ (avec option de bypass)
            is_btk_releve = self.is_btk_releve_pdf(self.pdf_path.get())
            print(f"DEBUG BTK RELEVÉ - Détection: {is_btk_releve}")
            
            if not is_btk_releve:
                # Demander à l'utilisateur s'il veut continuer malgré la détection échouée
                result = messagebox.askyesno("Détection échouée"
                    "Le fichier n'a pas été détecté comme un RELEVÉ BTK valide.\n"
                    "Voulez-vous continuer la conversion quand même ?\n\n"
                    "Cela peut fonctionner si le fichier est bien un RELEVÉ BTK.")
                if not result:
                    return
                else:
                    print("DEBUG BTK RELEVÉ - L'utilisateur a choisi de continuer malgré la détection échouée")
            
            # Extraire les données
            
            self.root.update()
            
            print("DEBUG BTK RELEVÉ - Début extraction des données...")
            transactions = self.extract_table_data(self.pdf_path.get())
            print(f"DEBUG BTK RELEVÉ - Résultat extraction: {len(transactions)} transactions trouvées")
            
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF")
                return
            
            # Créer le DataFrame
            
            self.root.update()
            
            df = pd.DataFrame(transactions)
            
            # Réorganiser les colonnes
            df = df[['date', 'libelle', 'debit', 'credit']]
            
            # Créer le fichier Excel dans le dossier Téléchargements
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = os.path.join(downloads_path, f"{self.excel_name.get()}.xlsx")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='J03', index=False)
                
                # Formater le fichier Excel
                workbook = writer.book
                worksheet = writer.sheets['J03']
                
                # Mettre les en-têtes en jaune
                from openpyxl.styles import PatternFill, Border, Side
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin')
                    right=Side(style='thin')
                    top=Side(style='thin')
                    bottom=Side(style='thin')
                )
                
                # Appliquer le formatage aux en-têtes
                for col in range(1, 5):  # 4 colonnes
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = yellow_fill
                    cell.border = thin_border
                
                # Appliquer les bordures à toutes les cellules
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border

            self.convert_button.config(state='normal')

            messagebox.showinfo("Succès", f"Fichier Excel créé avec succès!\nEmplacement: {excel_path}")
            
        except Exception as e:
            
            self.convert_button.config(state='normal')
            
            messagebox.showerror("Erreur", f"Erreur lors de la conversion: {str(e)}")

    def clean_amount(self, amount_str):
        """Convertit une chaîne de montant en float. Gère virgules, espaces et signe."""
        if amount_str is None:
            return None
        s = str(amount_str).strip()
        if s == "" or s == "-":
            return None
        try:
            # Normaliser: espaces -> rien, virgule -> point
            s_norm = s.replace(" ", "").replace(",", ".")
            # Gérer signe négatif éventuel en tête
            negative = False
            if s_norm.startswith("-"):
                negative = True
                s_norm = s_norm[1:]
            value = float(s_norm)
            return -value if negative else value
        except Exception:
            return None

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

def main():
    root = tk.Tk()
    app = BTKReleveConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
