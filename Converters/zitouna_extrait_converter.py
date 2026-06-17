import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class ZitounaExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT ZITOUNA vers Excel")
        self.root.geometry("600x500")
        
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"EXTRAT_ZITOUNA_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        self.setup_ui()

    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur EXTRAT ZITOUNA",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAT ZITOUNA:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.browse_pdf, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                              command=self.convertir, 
                              font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF ZITOUNA EXTRAT", filetypes=[["PDF", "*.pdf"]])
        if path:
            # Détection automatique du type de document
            try:
                detector = PDFBankDetector()
                detection_result = detector.detect_document_type(path)
                
                # Vérifier si c'est un extrait Zitouna Bank
                if detection_result['type'] == 'extrait_zitouna':
                    self.pdf_path.set(path)
                    base = os.path.splitext(os.path.basename(path))[0]
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    self.excel_name.set(f"EXTRAT_ZITOUNA_{base}_{ts}")
                    
                    # Message de confirmation discret
                    print(f"✅ Document détecté: EXTRAT ZITOUNA BANK (Confiance: {detection_result['confidence']:.1%})")
                    
                else:
                    # Afficher le type détecté et demander confirmation
                    summary = detector.get_detection_summary(detection_result)
                    response = messagebox.askyesno(
                        "⚠️ Type de document détecté", 
                        f"Document détecté: {summary}\n\n"
                        f"Ce convertisseur est conçu pour les extraits ZITOUNA BANK.\n"
                        f"Voulez-vous continuer quand même ?"
                    )
                    
                    if response:
                        self.pdf_path.set(path)
                        base = os.path.splitext(os.path.basename(path))[0]
                        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                        self.excel_name.set(f"EXTRAT_ZITOUNA_{base}_{ts}")
                    else:
                        return  # Ne pas sélectionner le fichier
                        
            except Exception as e:
                # En cas d'erreur de détection, continuer quand même
                print(f"⚠️ Erreur de détection: {str(e)} - Continuation avec le fichier sélectionné")
                self.pdf_path.set(path)
                base = os.path.splitext(os.path.basename(path))[0]
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                self.excel_name.set(f"EXTRAT_ZITOUNA_{base}_{ts}")

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            print("⚠️ Attention: PDF manquant - Veuillez choisir un fichier PDF ZITOUNA EXTRAT.")
            return
        self.progress['value'] = 10; self.root.update_idletasks()
        try:
            print(f"DEBUG ZITOUNA EXTRAT - PDF sélectionné: {path}")
        except Exception:
            pass
        
        # Utiliser le parser layout amélioré pour capturer toutes les lignes
        rows = self._parse_pdf_layout_improved(path)
        if not rows:
            # Fallback sur le parser complet
            rows = self._parse_pdf_comprehensive(path)
        if not rows:
            # Fallback sur le parser standard
            rows = self.parse_pdf(path)
        if not rows:
            print("❌ Erreur: Aucune transaction - Impossible d'extraire des transactions de l'extrait ZITOUNA.")
            return
        
        # Trier par date décroissante (DD/MM/YYYY) - du plus récent au plus ancien comme le PDF
        def _parse_date_safe(s: str):
            try:
                return datetime.strptime(s, "%d/%m/%Y")
            except Exception:
                return None
        rows.sort(key=lambda r: (_parse_date_safe(r.get('date') or '') or datetime.min), reverse=True)
        
        # Créer le DataFrame et formater les montants
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        
        # Convertir les montants en float pour le formatage Excel
        for col in ['debit', 'credit']:
            df[col] = df[col].apply(lambda x: float(x.replace('.', '').replace(',', '.')) if pd.notna(x) and x and x != '' else None)
        
        # Formater les montants avec espaces comme séparateurs de milliers
        for col in ['debit', 'credit']:
            df[col] = df[col].apply(lambda x: self._format_amount_with_spaces(x) if pd.notna(x) and x else None)
        
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'EXTRAT_ZITOUNA'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        self.progress['value'] = 100
        # Message de succès plus positif
        success_msg = f"✅ Conversion EXTRAT terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n\n"
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        print("✅ Succès: Conversion réussie - " + success_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            self.root.destroy()
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            print("❌ Erreur: Impossible de retourner à la page d'accueil: " + str(e))

    def parse_pdf(self, pdf_path: str):
        """Parser pour les extraits ZITOUNA BANK"""
        # Expressions régulières
        date_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*[.,]\d{2,3}|-?\d+[.,]\d{2,3}")
        
        def clean_zitouna_extrait_amount(amount_str: str) -> str:
            """
            Nettoie et formate un montant ZITOUNA selon le format standard
            Format: 3500,000 (virgule pour décimales, 3 chiffres après virgule)
            """
            if not amount_str:
                return None
            
            # CORRECTION: Supprimer seulement les chiffres isolés qui sont clairement des jours de dates
            # (ex: "5 404,000" -> "404,000") mais pas "3.814.040" -> "814.040"
            # Ne supprimer que si c'est suivi d'un espace ET que le reste ne ressemble pas à un montant
            if re.match(r'^([1-9]|[12][0-9]|3[01])\s+', amount_str) and not re.search(r'[.,]\d', amount_str):
                amount_str = re.sub(r'^([1-9]|[12][0-9]|3[01])\s*', '', amount_str)
            
            # Nettoyer le montant - garder le signe négatif et gérer les formats spéciaux
            # Exemple: "3500,000" -> "3500,000", "0500)" -> "500,000"
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            # Gérer les montants avec parenthèses (ex: "0500)" -> "500")
            if cleaned.endswith(')'):
                cleaned = cleaned[:-1]  # Enlever la parenthèse fermante
            
            # Gérer les montants avec des zéros en début (ex: "0500" -> "500")
            if cleaned.startswith('0') and len(cleaned) > 1:
                cleaned = cleaned.lstrip('0')
                if not cleaned:  # Si tout était des zéros
                    cleaned = '0'
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]  # Enlever le signe négatif
            
            # Filtrer les montants à zéro et les montants dégradés
            if (cleaned == '0' or cleaned == '0,000' or cleaned == '0.000' or 
                cleaned == '000' or cleaned == '000,000' or cleaned == '000.000' or
                cleaned == '00' or cleaned == '00,000' or cleaned == '00.000'):
                return None
            
            try:
                # Traitement spécial pour préserver le format original Zitouna
                original_cleaned = cleaned
                
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 12.409,000 (point = milliers, virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 3500,000 (virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est un format avec décimales ou milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales: 5.950 -> 5,950
                        cleaned = cleaned.replace('.', ',')
                    else:
                        # Probablement des milliers: 12.409 -> 12409,000
                        cleaned = cleaned.replace('.', '') + ',000'
                
                # Si pas de virgule, ajouter ,000
                if ',' not in cleaned:
                    cleaned = cleaned + ',000'
                
                # S'assurer qu'il y a exactement 3 chiffres après la virgule
                if ',' in cleaned:
                    parts = cleaned.split(',')
                    integer_part = parts[0]
                    decimal_part = parts[1] if len(parts) > 1 else "000"
                    
                    # Garder exactement 3 chiffres après la virgule
                    if len(decimal_part) >= 3:
                        formatted_decimal = decimal_part[:3]
                    else:
                        formatted_decimal = decimal_part.ljust(3, '0')
                    
                    # Retourner le format Zitouna standard
                    result = f"{integer_part},{formatted_decimal}"
                    return f"-{result}" if is_negative else result
                else:
                    # Pas de virgule trouvée, ajouter ,000
                    result = f"{cleaned},000"
                    return f"-{result}" if is_negative else result
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not text.strip():
                    # Fallback OCR pour les PDFs scannés
                    ocr_text = self._extract_text_via_ocr(pdf_path)
                    if ocr_text.strip():
                        text = ocr_text
                        break  # Utiliser le texte OCR pour toutes les pages
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                
                # Chercher les lignes de transaction
                print(f"DEBUG: Page avec {len(lines)} lignes")
                for i, line in enumerate(lines):
                    # Vérifier si la ligne contient une date
                    dates = date_re.findall(line)
                    
                    # CORRECTION: Nettoyer la ligne avant d'extraire les montants
                    # pour éviter le mélange avec les dates, libellés et références
                    line_clean = line
                    
                    # Supprimer toutes les dates (DD/MM/YYYY) de la ligne
                    line_clean = re.sub(r'\d{1,2}/\d{1,2}/\d{2,4}', '', line_clean)
                    
                    # Supprimer les années isolées (2025, 2024, etc.)
                    line_clean = re.sub(r'\b(19|20)\d{2}\b', '', line_clean)
                    
                    # Supprimer les nombres isolés qui pourraient être des jours de dates
                    line_clean = re.sub(r'\b([1-9]|[12][0-9]|3[01])\b', '', line_clean)
                    
                    # Supprimer les mois isolés (1-12)
                    line_clean = re.sub(r'\b(1[0-2]|[1-9])\b', '', line_clean)
                    
                    # Maintenant extraire les montants de la ligne nettoyée
                    amounts = amount_re.findall(line_clean)
                    
                    print(f"DEBUG: Ligne {i+1}: {line[:50]}... - Dates: {len(dates)}, Montants: {len(amounts)}")
                    
                    if len(dates) >= 1:  # Au moins 1 date (plus flexible)
                        # Extraire les montants
                        
                        if len(amounts) >= 1:
                            # Nettoyer la ligne
                            parts = line.split()
                            
                            # Extraire les dates
                            date_operation = dates[0] if len(dates) > 0 else ""
                            date_valeur = dates[1] if len(dates) > 1 else date_operation  # Utiliser la même date si pas de date valeur
                            
                            # Extraire le libellé et la référence
                            libelle_parts = []
                            reference = ""
                            amount_started = False
                            
                            for part in parts:
                                if amount_re.match(part):
                                    amount_started = True
                                elif not amount_started and part not in dates:
                                    # Chercher des patterns de référence (codes alphanumériques)
                                    if re.match(r'^[A-Z0-9]{8,}$', part) and not reference:
                                        reference = part
                                    else:
                                        libelle_parts.append(part)
                            
                            libelle = ' '.join(libelle_parts)
                            
                            # Traiter les montants
                            debit = credit = None
                            for amount in amounts:
                                cleaned_amount = clean_zitouna_extrait_amount(amount)
                                if cleaned_amount:
                                    # Déterminer si c'est un débit ou crédit
                                    if amount.startswith('-'):
                                        # Montant négatif = débit
                                        debit = cleaned_amount.lstrip('-')
                                    else:
                                        # Montant positif = crédit
                                        credit = cleaned_amount
                            
                            # Si on a un seul montant, déterminer le type par le libellé
                            if (debit and not credit) or (credit and not debit):
                                if any(keyword in libelle.upper() for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT']):
                                    if debit:
                                        credit = debit
                                        debit = None
                                elif any(keyword in libelle.upper() for keyword in ['PAIEMENT', 'COMMISSION', 'FRAIS', 'TVA', 'RETRAIT']):
                                    if credit:
                                        debit = credit
                                        credit = None
                            
                            # Ajouter la transaction si elle est valide
                            if libelle.strip() and (debit or credit):
                                print(f"DEBUG: Transaction trouvée - Date: {date_operation}, Libellé: {libelle[:30]}..., Débit: {debit}, Crédit: {credit}")
                                results.append({
                                    "date": date_operation,  # Date d'opération du PDF
                                    "libelle": libelle,      # Opération du PDF
                                    "debit": debit,          # Débit du PDF
                                    "credit": credit         # Crédit du PDF
                                })
        
        # Si aucun résultat trouvé, essayer une approche plus permissive
        if not results:
            print("DEBUG: Aucune transaction trouvée avec le parser standard, tentative avec approche permissive...")
            results = self._parse_pdf_permissive(pdf_path)
        
        print(f"DEBUG: Total de {len(results)} transactions trouvées")
        return results
    
    def _parse_pdf_permissive(self, pdf_path: str):
        """Parser permissif pour les extraits ZITOUNA BANK"""
        date_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*[.,]\d{2,3}|-?\d+[.,]\d{2,3}")
        
        def clean_zitouna_extrait_amount(amount_str: str) -> str:
            """
            Nettoie et formate un montant ZITOUNA selon le format standard
            Format: 3500,000 (virgule pour décimales, 3 chiffres après virgule)
            """
            if not amount_str:
                return None
            
            # CORRECTION: Supprimer seulement les chiffres isolés qui sont clairement des jours de dates
            # (ex: "5 404,000" -> "404,000") mais pas "3.814.040" -> "814.040"
            # Ne supprimer que si c'est suivi d'un espace ET que le reste ne ressemble pas à un montant
            if re.match(r'^([1-9]|[12][0-9]|3[01])\s+', amount_str) and not re.search(r'[.,]\d', amount_str):
                amount_str = re.sub(r'^([1-9]|[12][0-9]|3[01])\s*', '', amount_str)
            
            # Nettoyer le montant - garder le signe négatif et gérer les formats spéciaux
            # Exemple: "3500,000" -> "3500,000", "0500)" -> "500,000"
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            # Gérer les montants avec parenthèses (ex: "0500)" -> "500")
            if cleaned.endswith(')'):
                cleaned = cleaned[:-1]  # Enlever la parenthèse fermante
            
            # Gérer les montants avec des zéros en début (ex: "0500" -> "500")
            if cleaned.startswith('0') and len(cleaned) > 1:
                cleaned = cleaned.lstrip('0')
                if not cleaned:  # Si tout était des zéros
                    cleaned = '0'
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]  # Enlever le signe négatif
            
            # Filtrer les montants à zéro et les montants dégradés
            if (cleaned == '0' or cleaned == '0,000' or cleaned == '0.000' or 
                cleaned == '000' or cleaned == '000,000' or cleaned == '000.000' or
                cleaned == '00' or cleaned == '00,000' or cleaned == '00.000'):
                return None
            
            try:
                # Traitement spécial pour préserver le format original Zitouna
                original_cleaned = cleaned
                
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 12.409,000 (point = milliers, virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 3500,000 (virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est un format avec décimales ou milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales: 5.950 -> 5,950
                        cleaned = cleaned.replace('.', ',')
                    else:
                        # Probablement des milliers: 12.409 -> 12409,000
                        cleaned = cleaned.replace('.', '') + ',000'
                
                # Si pas de virgule, ajouter ,000
                if ',' not in cleaned:
                    cleaned = cleaned + ',000'
                
                # S'assurer qu'il y a exactement 3 chiffres après la virgule
                if ',' in cleaned:
                    parts = cleaned.split(',')
                    integer_part = parts[0]
                    decimal_part = parts[1] if len(parts) > 1 else "000"
                    
                    # Garder exactement 3 chiffres après la virgule
                    if len(decimal_part) >= 3:
                        formatted_decimal = decimal_part[:3]
                    else:
                        formatted_decimal = decimal_part.ljust(3, '0')
                    
                    # Retourner le format Zitouna standard
                    result = f"{integer_part},{formatted_decimal}"
                    return f"-{result}" if is_negative else result
                else:
                    # Pas de virgule trouvée, ajouter ,000
                    result = f"{cleaned},000"
                    return f"-{result}" if is_negative else result
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                
                # Approche plus permissive : chercher toute ligne avec des montants
                print(f"DEBUG PERMISSIF: Page avec {len(lines)} lignes")
                for i, line in enumerate(lines):
                    dates = date_re.findall(line)
                    
                    # CORRECTION: Nettoyer la ligne avant d'extraire les montants
                    # pour éviter le mélange avec les dates, libellés et références
                    line_clean = line
                    
                    # Supprimer toutes les dates (DD/MM/YYYY) de la ligne
                    line_clean = re.sub(r'\d{1,2}/\d{1,2}/\d{2,4}', '', line_clean)
                    
                    # Supprimer les années isolées (2025, 2024, etc.)
                    line_clean = re.sub(r'\b(19|20)\d{2}\b', '', line_clean)
                    
                    # Supprimer les nombres isolés qui pourraient être des jours de dates
                    line_clean = re.sub(r'\b([1-9]|[12][0-9]|3[01])\b', '', line_clean)
                    
                    # Supprimer les mois isolés (1-12)
                    line_clean = re.sub(r'\b(1[0-2]|[1-9])\b', '', line_clean)
                    
                    # Maintenant extraire les montants de la ligne nettoyée
                    amounts = amount_re.findall(line_clean)
                    
                    print(f"DEBUG PERMISSIF: Ligne {i+1}: {line[:50]}... - Dates: {len(dates)}, Montants: {len(amounts)}")
                    
                    if len(amounts) >= 1:  # Au moins un montant
                        # Extraire les dates
                        date_operation = dates[0] if len(dates) > 0 else ""
                        date_valeur = dates[1] if len(dates) > 1 else date_operation
                        
                        # Extraire le libellé (tout ce qui n'est pas une date ou un montant)
                        parts = line.split()
                        libelle_parts = []
                        reference = ""
                        
                        for part in parts:
                            if not date_re.match(part) and not amount_re.match(part):
                                # Chercher des patterns de référence
                                if re.match(r'^[A-Z0-9]{8,}$', part) and not reference:
                                    reference = part
                                else:
                                    libelle_parts.append(part)
                        
                        libelle = ' '.join(libelle_parts)
                        
                        # Traiter les montants
                        debit = credit = None
                        for amount in amounts:
                            cleaned_amount = clean_zitouna_extrait_amount(amount)
                            if cleaned_amount:
                                # Déterminer si c'est un débit ou crédit
                                if amount.startswith('-'):
                                    # Montant négatif = débit
                                    debit = cleaned_amount.lstrip('-')
                                else:
                                    # Montant positif = crédit
                                    credit = cleaned_amount
                        
                        # Si on a un seul montant, déterminer le type par le libellé
                        if (debit and not credit) or (credit and not debit):
                            if any(keyword in libelle.upper() for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT']):
                                if debit:
                                    credit = debit
                                    debit = None
                            elif any(keyword in libelle.upper() for keyword in ['PAIEMENT', 'COMMISSION', 'FRAIS', 'TVA']):
                                if credit:
                                    debit = credit
                                    credit = None
                        
                        # Ajouter la transaction si elle est valide
                        if libelle.strip() and (debit or credit):
                            print(f"DEBUG PERMISSIF: Transaction trouvée - Date: {date_operation}, Libellé: {libelle[:30]}..., Débit: {debit}, Crédit: {credit}")
                            results.append({
                                "date": date_operation,  # Date d'opération du PDF
                                "libelle": libelle,      # Opération du PDF
                                "debit": debit,          # Débit du PDF
                                "credit": credit         # Crédit du PDF
                            })
        
        return results

    def _parse_pdf_by_layout(self, pdf_path: str):
        """Extraction basée sur positions x/y: Date, Libellé, Débit, Crédit."""
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*[.,]\d{2,3}|-?\d+[.,]\d{2,3}")
        header_stop = [
            'extrait', 'compte', 'liste', 'transactions', 'agence', 'solde', 'actuel', 'date début', 'date fin'
            'type d\'opération', 'type d’opération', 'devise', 'banque', 'zitouna'
        ]
        rows = []
        def clean_amount(a: str):
            if not a:
                return None
            a = a.replace(' ', '')
            # format 1.234,567 or 123,456 or 123.456
            neg = a.startswith('-')
            if neg:
                a = a[1:]
            if ',' in a and '.' in a:
                a = a.replace('.', '').replace(',', '.')
            elif ',' in a:
                a = a.replace(',', '.')
            elif '.' in a and a.count('.') > 1:
                a = a.replace('.', '')
            try:
                val = float(a)
            except Exception:
                return None
            # Retourner chaine formattée fr (.)
            if val == int(val):
                s = f"{int(abs(val)):,}".replace(',', '.')
            else:
                s = f"{abs(val):,.3f}".replace(',', '.').replace('.', ',', 1)
            return f"-{s}" if neg else s
        
        with pdfplumber.open(pdf_path) as pdf:
            # Détecter colonnes sur 1ère page
            first = pdf.pages[0]
            header_words = first.extract_words() or []
            # bornes par défaut
            x_date_max = 130
            x_lib_min, x_lib_max = 135, 640
            x_debit_min, x_debit_max = 640, 760
            x_credit_min = 760
            try:
                for w in header_words:
                    t = w['text'].lower()
                    if 'date opération' in t or t.startswith('date'):
                        x_date_max = max(x_date_max, int(w['x1']) + 8)
                    elif 'opération' in t or 'operation' in t or 'libell' in t:
                        x_lib_min = min(x_lib_min, int(w['x0']))
                        x_lib_max = max(x_lib_max, int(w['x1']) + 8)
                    elif 'débit' in t or 'debit' in t:
                        x_debit_min = min(x_debit_min, int(w['x0']) - 6)
                        x_debit_max = max(x_debit_max, int(w['x1']) + 6)
                    elif 'crédit' in t or 'credit' in t:
                        x_credit_min = min(x_credit_min, int(w['x0']) - 6)
            except Exception:
                pass
            
            for page in pdf.pages:
                words = page.extract_words() or []
                # grouper par y
                lines = {}
                for w in words:
                    yk = round(w['top'] / 3)
                    lines.setdefault(yk, []).append(w)
                for yk in sorted(lines.keys()):
                    ws = sorted(lines[yk], key=lambda m: m['x0'])
                    date_text = ''
                    lib_parts = []
                    debit_text = ''
                    credit_text = ''
                    for w in ws:
                        x0, x1 = w['x0'], w['x1']
                        t = w['text']
                        tl = t.lower()
                        # filtrer têtes/pieds
                        if any(sw in tl for sw in header_stop):
                            lib_parts = []
                            date_text = ''
                            debit_text = credit_text = ''
                            break
                        if x1 <= x_date_max and re.match(r"^\d{2}/\d{2}/\d{4}$", t):
                            date_text = t
                        elif x0 >= x_credit_min and amount_re.match(t):
                            credit_text = t
                        elif x0 >= x_debit_min and x1 <= x_debit_max and amount_re.match(t):
                            debit_text = t
                        elif x0 >= x_lib_min and x1 <= x_lib_max:
                            lib_parts.append(t)
                    libelle = ' '.join(lib_parts).strip()
                    if not libelle and not (debit_text or credit_text):
                        continue
                    # Nettoyage montants
                    debit_fmt = clean_amount(debit_text)
                    credit_fmt = clean_amount(credit_text)
                    if (debit_fmt or credit_fmt) and libelle:
                        rows.append({
                            'date': date_text,
                            'libelle': libelle,
                            'debit': debit_fmt,
                            'credit': credit_fmt
                        })
        return rows

    def _parse_pdf_layout_improved(self, pdf_path: str):
        """Parser amélioré basé sur les coordonnées pour capturer toutes les lignes"""
        date_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*[.,]\d{2,3}|-?\d+[.,]\d{2,3}")
        
        def clean_zitouna_extrait_amount(amount_str: str) -> str:
            """
            Nettoie et formate un montant ZITOUNA selon le format standard
            Format: 3500,000 (virgule pour décimales, 3 chiffres après virgule)
            """
            if not amount_str:
                return None
            
            # CORRECTION: Supprimer seulement les chiffres isolés qui sont clairement des jours de dates
            # (ex: "5 404,000" -> "404,000") mais pas "3.814.040" -> "814.040"
            # Ne supprimer que si c'est suivi d'un espace ET que le reste ne ressemble pas à un montant
            if re.match(r'^([1-9]|[12][0-9]|3[01])\s+', amount_str) and not re.search(r'[.,]\d', amount_str):
                amount_str = re.sub(r'^([1-9]|[12][0-9]|3[01])\s*', '', amount_str)
            
            # Nettoyer le montant - garder le signe négatif et gérer les formats spéciaux
            # Exemple: "3500,000" -> "3500,000", "0500)" -> "500,000"
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            # Gérer les montants avec parenthèses (ex: "0500)" -> "500")
            if cleaned.endswith(')'):
                cleaned = cleaned[:-1]  # Enlever la parenthèse fermante
            
            # Gérer les montants avec des zéros en début (ex: "0500" -> "500")
            if cleaned.startswith('0') and len(cleaned) > 1:
                cleaned = cleaned.lstrip('0')
                if not cleaned:  # Si tout était des zéros
                    cleaned = '0'
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]  # Enlever le signe négatif
            
            # Filtrer les montants à zéro et les montants dégradés
            if (cleaned == '0' or cleaned == '0,000' or cleaned == '0.000' or 
                cleaned == '000' or cleaned == '000,000' or cleaned == '000.000' or
                cleaned == '00' or cleaned == '00,000' or cleaned == '00.000'):
                return None
            
            try:
                # Traitement spécial pour préserver le format original Zitouna
                original_cleaned = cleaned
                
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 12.409,000 (point = milliers, virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 3500,000 (virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est un format avec décimales ou milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales: 5.950 -> 5,950
                        cleaned = cleaned.replace('.', ',')
                    else:
                        # Probablement des milliers: 12.409 -> 12409,000
                        cleaned = cleaned.replace('.', '') + ',000'
                
                # Si pas de virgule, ajouter ,000
                if ',' not in cleaned:
                    cleaned = cleaned + ',000'
                
                # S'assurer qu'il y a exactement 3 chiffres après la virgule
                if ',' in cleaned:
                    parts = cleaned.split(',')
                    integer_part = parts[0]
                    decimal_part = parts[1] if len(parts) > 1 else "000"
                    
                    # Garder exactement 3 chiffres après la virgule
                    if len(decimal_part) >= 3:
                        formatted_decimal = decimal_part[:3]
                    else:
                        formatted_decimal = decimal_part.ljust(3, '0')
                    
                    # Retourner le format Zitouna standard
                    result = f"{integer_part},{formatted_decimal}"
                    return f"-{result}" if is_negative else result
                else:
                    # Pas de virgule trouvée, ajouter ,000
                    result = f"{cleaned},000"
                    return f"-{result}" if is_negative else result
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            print(f"DEBUG: PDF contient {len(pdf.pages)} pages")
            for page_num, page in enumerate(pdf.pages, 1):
                # Extraire les mots avec leurs coordonnées
                words = page.extract_words()
                if not words:
                    print(f"DEBUG: Page {page_num} - Aucun mot extrait")
                    continue
                
                print(f"DEBUG: Page {page_num} avec {len(words)} mots extraits")
                
                # Grouper les mots par ligne (même y approximatif)
                lines = {}
                for word in words:
                    y = round(word['top'], 1)  # Arrondir pour grouper
                    if y not in lines:
                        lines[y] = []
                    lines[y].append(word)
                
                # Trier les lignes par position y
                sorted_lines = sorted(lines.items())
                print(f"DEBUG: Page {page_num} - {len(sorted_lines)} lignes détectées")
                
                filtered_count = 0
                for y_pos, words_in_line in sorted_lines:
                    # Trier les mots par position x
                    words_in_line.sort(key=lambda w: w['x0'])
                    
                    # Reconstituer le texte de la ligne
                    line_text = ' '.join([w['text'] for w in words_in_line])
                    
                    # Filtrer les lignes non-transactionnelles
                    if self._filter_line(line_text):
                        filtered_count += 1
                        continue
                    
                    # Chercher les dates et montants
                    dates = date_re.findall(line_text)
                    amounts = amount_re.findall(line_text)
                    
                    print(f"DEBUG: Ligne: {line_text[:60]}... - Dates: {len(dates)}, Montants: {len(amounts)}")
                    
                    # Accepter les lignes avec au moins une date et un montant (être plus permissif)
                    if len(dates) >= 1 and len(amounts) >= 1:
                        # Extraire les données
                        date_operation = dates[0]
                        date_valeur = dates[1] if len(dates) > 1 else date_operation
                        
                        # Extraire le libellé complet (tout ce qui n'est pas date ou montant)
                        libelle_parts = []
                        for word in words_in_line:
                            word_text = word['text']
                            if not date_re.match(word_text) and not amount_re.match(word_text):
                                libelle_parts.append(word_text)
                        
                        libelle = ' '.join(libelle_parts).strip()
                        
                        # Si le libellé est vide, essayer de le reconstituer différemment
                        if not libelle:
                            # Prendre tous les mots entre la première date et le premier montant
                            date_found = False
                            amount_found = False
                            for word in words_in_line:
                                word_text = word['text']
                                if date_re.match(word_text):
                                    date_found = True
                                elif amount_re.match(word_text):
                                    amount_found = True
                                elif date_found and not amount_found:
                                    libelle_parts.append(word_text)
                            libelle = ' '.join(libelle_parts).strip()
                        
                        # Traiter les montants - Classification intelligente débit/crédit
                        debit = credit = None
                        
                        # 1. D'abord, classifier par le signe du montant (méthode principale)
                        for amount in amounts:
                            cleaned_amount = clean_zitouna_extrait_amount(amount)
                            if cleaned_amount:
                                if amount.startswith('-'):
                                    # Montant négatif = débit (enlever le signe pour l'affichage)
                                    debit = cleaned_amount.lstrip('-')
                                else:
                                    # Montant positif = crédit
                                    credit = cleaned_amount
                                break  # Prendre seulement le premier montant valide
                        
                        # 2. Si classification par signe échoue, utiliser les mots-clés du libellé
                        if not debit and not credit:
                            # Aucun montant trouvé, essayer de le reconstituer
                            for amount in amounts:
                                cleaned_amount = clean_zitouna_extrait_amount(amount)
                                if cleaned_amount:
                                    # Classification par mots-clés dans le libellé
                                    libelle_upper = libelle.upper()
                                    
                                    # Mots-clés pour crédit (entrées d'argent)
                                    credit_keywords = ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT', 'VIREMENT RECU', 'AV TPE']
                                    
                                    # Mots-clés pour débit (sorties d'argent)
                                    debit_keywords = ['PAIEMENT', 'COMMISSION', 'FRAIS', 'TVA', 'RETRAIT', 'PRELEVEMENT', 'EFFET', 'COMM REGLEMENT', 'COMM REMISE', 'DROIT DE TIMBRE', 'PRIME TAKAFUL', 'PROFIT IJARA', 'PRINCIPAL IJARA']
                                    
                                    if any(keyword in libelle_upper for keyword in credit_keywords):
                                        credit = cleaned_amount
                                    elif any(keyword in libelle_upper for keyword in debit_keywords):
                                        debit = cleaned_amount
                                    else:
                                        # Par défaut, si le montant est négatif dans le PDF, c'est un débit
                                        if amount.startswith('-'):
                                            debit = cleaned_amount.lstrip('-')
                                        else:
                                            credit = cleaned_amount
                                    break
                        
                        # 3. Validation finale : s'assurer qu'on a soit débit soit crédit, pas les deux
                        if debit and credit:
                            # Si on a les deux, garder seulement celui qui correspond le mieux au libellé
                            libelle_upper = libelle.upper()
                            if any(keyword in libelle_upper for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT']):
                                debit = None  # Garder seulement le crédit
                            else:
                                credit = None  # Garder seulement le débit
                        
                        # Ajouter la transaction si on a un montant (être plus permissif sur le libellé)
                        if (debit or credit):
                            # Si le libellé est vide, utiliser une description par défaut
                            if not libelle or libelle == "Transaction":
                                libelle = "Transaction bancaire"
                            
                            print(f"DEBUG: Transaction ajoutée - Date: {date_operation}, Libellé: '{libelle}', Débit: {debit}, Crédit: {credit}")
                            results.append({
                                "date": date_operation,
                                "libelle": libelle,
                                "debit": debit,
                                "credit": credit
                            })
                
                print(f"DEBUG: Page {page_num} - {filtered_count} lignes filtrées, {len([r for r in results if r.get('page') == page_num])} transactions ajoutées")
        
        print(f"DEBUG: Total de {len(results)} transactions trouvées avec le parser amélioré")
        return results

    def _parse_pdf_comprehensive(self, pdf_path: str):
        """Parser complet qui capture toutes les lignes de transaction"""
        date_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*[.,]\d{2,3}|-?\d+[.,]\d{2,3}")
        
        def clean_zitouna_extrait_amount(amount_str: str) -> str:
            """
            Nettoie et formate un montant ZITOUNA selon le format standard
            Format: 3500,000 (virgule pour décimales, 3 chiffres après virgule)
            """
            if not amount_str:
                return None
            
            # CORRECTION: Supprimer seulement les chiffres isolés qui sont clairement des jours de dates
            # (ex: "5 404,000" -> "404,000") mais pas "3.814.040" -> "814.040"
            # Ne supprimer que si c'est suivi d'un espace ET que le reste ne ressemble pas à un montant
            if re.match(r'^([1-9]|[12][0-9]|3[01])\s+', amount_str) and not re.search(r'[.,]\d', amount_str):
                amount_str = re.sub(r'^([1-9]|[12][0-9]|3[01])\s*', '', amount_str)
            
            # Nettoyer le montant - garder le signe négatif et gérer les formats spéciaux
            # Exemple: "3500,000" -> "3500,000", "0500)" -> "500,000"
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            # Gérer les montants avec parenthèses (ex: "0500)" -> "500")
            if cleaned.endswith(')'):
                cleaned = cleaned[:-1]  # Enlever la parenthèse fermante
            
            # Gérer les montants avec des zéros en début (ex: "0500" -> "500")
            if cleaned.startswith('0') and len(cleaned) > 1:
                cleaned = cleaned.lstrip('0')
                if not cleaned:  # Si tout était des zéros
                    cleaned = '0'
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]  # Enlever le signe négatif
            
            # Filtrer les montants à zéro et les montants dégradés
            if (cleaned == '0' or cleaned == '0,000' or cleaned == '0.000' or 
                cleaned == '000' or cleaned == '000,000' or cleaned == '000.000' or
                cleaned == '00' or cleaned == '00,000' or cleaned == '00.000'):
                return None
            
            try:
                # Traitement spécial pour préserver le format original Zitouna
                original_cleaned = cleaned
                
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 12.409,000 (point = milliers, virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 3500,000 (virgule = décimales)
                    # Exemple: "3500,000" -> "3500,000" (préserver tel quel)
                    pass  # Garder le format original
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est un format avec décimales ou milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales: 5.950 -> 5,950
                        cleaned = cleaned.replace('.', ',')
                    else:
                        # Probablement des milliers: 12.409 -> 12409,000
                        cleaned = cleaned.replace('.', '') + ',000'
                
                # Si pas de virgule, ajouter ,000
                if ',' not in cleaned:
                    cleaned = cleaned + ',000'
                
                # S'assurer qu'il y a exactement 3 chiffres après la virgule
                if ',' in cleaned:
                    parts = cleaned.split(',')
                    integer_part = parts[0]
                    decimal_part = parts[1] if len(parts) > 1 else "000"
                    
                    # Garder exactement 3 chiffres après la virgule
                    if len(decimal_part) >= 3:
                        formatted_decimal = decimal_part[:3]
                    else:
                        formatted_decimal = decimal_part.ljust(3, '0')
                    
                    # Retourner le format Zitouna standard
                    result = f"{integer_part},{formatted_decimal}"
                    return f"-{result}" if is_negative else result
                else:
                    # Pas de virgule trouvée, ajouter ,000
                    result = f"{cleaned},000"
                    return f"-{result}" if is_negative else result
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # Essayer d'abord l'extraction de tableau
                tables = page.extract_tables()
                if tables:
                    print(f"DEBUG: Tableau trouvé avec {len(tables)} tables")
                    for table in tables:
                        for row in table:
                            if row and len(row) >= 3:  # Au moins 3 colonnes
                                row_text = ' '.join([str(cell) for cell in row if cell])
                                if self._filter_line(row_text):
                                    continue
                                
                                dates = date_re.findall(row_text)
                                amounts = amount_re.findall(row_text)
                                
                                if len(dates) >= 1 and len(amounts) >= 1:
                                    date_operation = dates[0]
                                    
                                    # Extraire le libellé complet (colonnes du milieu)
                                    libelle_parts = []
                                    for i, cell in enumerate(row[1:-1]):  # Exclure première et dernière colonne
                                        if cell and not date_re.match(str(cell)) and not amount_re.match(str(cell)):
                                            libelle_parts.append(str(cell))
                                    libelle = ' '.join(libelle_parts).strip()
                                    
                                    # Traiter les montants - Classification intelligente débit/crédit
                                    debit = credit = None
                                    print(f"DEBUG: Montants trouvés: {amounts}")
                                    
                                    # 1. Classification par signe du montant (méthode principale)
                                    for amount in amounts:
                                        # Exclure les dates (format DD/MM/YYYY ou DD)
                                        if re.match(r'^\d{1,2}$', amount) or re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', amount):
                                            print(f"DEBUG: Ignoré (date): '{amount}'")
                                            continue
                                        
                                        # Exclure les codes de référence (longs codes alphanumériques)
                                        if re.match(r'^[A-Z0-9]{8,}$', amount):
                                            print(f"DEBUG: Ignoré (référence): '{amount}'")
                                            continue
                                        
                                        # Prendre seulement les montants avec virgule ou point (vrais montants)
                                        if ',' in amount or ('.' in amount and not amount.startswith('.')):
                                            cleaned_amount = clean_zitouna_extrait_amount(amount)
                                            print(f"DEBUG: Montant original: '{amount}' -> Nettoyé: '{cleaned_amount}'")
                                            if cleaned_amount:
                                                if amount.startswith('-'):
                                                    debit = cleaned_amount.lstrip('-')
                                                    print(f"DEBUG: Débit assigné (signe négatif): {debit}")
                                                else:
                                                    credit = cleaned_amount
                                                    print(f"DEBUG: Crédit assigné (signe positif): {credit}")
                                                break
                                    
                                    # 2. Si pas de classification par signe, utiliser les mots-clés
                                    if not debit and not credit:
                                        for amount in amounts:
                                            if ',' in amount or ('.' in amount and not amount.startswith('.')):
                                                cleaned_amount = clean_zitouna_extrait_amount(amount)
                                                if cleaned_amount:
                                                    libelle_upper = libelle.upper()
                                                    
                                                    # Mots-clés pour crédit (entrées d'argent)
                                                    credit_keywords = ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT', 'VIREMENT RECU', 'AV TPE']
                                                    
                                                    # Mots-clés pour débit (sorties d'argent)
                                                    debit_keywords = ['PAIEMENT', 'COMMISSION', 'FRAIS', 'TVA', 'RETRAIT', 'PRELEVEMENT', 'EFFET', 'COMM REGLEMENT', 'COMM REMISE', 'DROIT DE TIMBRE', 'PRIME TAKAFUL', 'PROFIT IJARA', 'PRINCIPAL IJARA']
                                                    
                                                    if any(keyword in libelle_upper for keyword in credit_keywords):
                                                        credit = cleaned_amount
                                                        print(f"DEBUG: Crédit assigné (mots-clés): {credit}")
                                                    elif any(keyword in libelle_upper for keyword in debit_keywords):
                                                        debit = cleaned_amount
                                                        print(f"DEBUG: Débit assigné (mots-clés): {debit}")
                                                    else:
                                                        # Par défaut selon le signe
                                                        if amount.startswith('-'):
                                                            debit = cleaned_amount.lstrip('-')
                                                            print(f"DEBUG: Débit assigné (défaut): {debit}")
                                                        else:
                                                            credit = cleaned_amount
                                                            print(f"DEBUG: Crédit assigné (défaut): {credit}")
                                                    break
                                    
                                    # 3. Validation finale : s'assurer qu'on a soit débit soit crédit, pas les deux
                                    if debit and credit:
                                        libelle_upper = libelle.upper()
                                        if any(keyword in libelle_upper for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT']):
                                            debit = None  # Garder seulement le crédit
                                            print(f"DEBUG: Débit supprimé, garde crédit: {credit}")
                                        else:
                                            credit = None  # Garder seulement le débit
                                            print(f"DEBUG: Crédit supprimé, garde débit: {debit}")
                                    
                                    if (debit or credit):
                                        # Si le libellé est vide, utiliser une description par défaut
                                        if not libelle or libelle == "Transaction":
                                            libelle = "Transaction bancaire"
                                        
                                        results.append({
                                            "date": date_operation,
                                            "libelle": libelle,
                                            "debit": debit,
                                            "credit": credit
                                        })
                
                # Si pas de tableau, utiliser l'extraction de texte
                if not results:
                    text = page.extract_text()
                    if text:
                        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                        
                        for line in lines:
                            if self._filter_line(line):
                                continue
                            
                            dates = date_re.findall(line)
                            amounts = amount_re.findall(line)
                            
                            if len(dates) >= 1 and len(amounts) >= 1:
                                date_operation = dates[0]
                                
                                # Extraire le libellé complet
                                parts = line.split()
                                libelle_parts = []
                                for part in parts:
                                    if not date_re.match(part) and not amount_re.match(part):
                                        libelle_parts.append(part)
                                libelle = ' '.join(libelle_parts).strip()
                                
                                # Traiter les montants - Classification intelligente débit/crédit
                                debit = credit = None
                                print(f"DEBUG: Montants trouvés: {amounts}")
                                
                                # 1. Classification par signe du montant (méthode principale)
                                for amount in amounts:
                                    # Exclure les dates (format DD/MM/YYYY ou DD)
                                    if re.match(r'^\d{1,2}$', amount) or re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', amount):
                                        print(f"DEBUG: Ignoré (date): '{amount}'")
                                        continue
                                    
                                    # Exclure les codes de référence (longs codes alphanumériques)
                                    if re.match(r'^[A-Z0-9]{8,}$', amount):
                                        print(f"DEBUG: Ignoré (référence): '{amount}'")
                                        continue
                                    
                                    # Prendre seulement les montants avec virgule ou point (vrais montants)
                                    if ',' in amount or ('.' in amount and not amount.startswith('.')):
                                        cleaned_amount = clean_zitouna_extrait_amount(amount)
                                        print(f"DEBUG: Montant original: '{amount}' -> Nettoyé: '{cleaned_amount}'")
                                        if cleaned_amount:
                                            if amount.startswith('-'):
                                                debit = cleaned_amount.lstrip('-')
                                                print(f"DEBUG: Débit assigné (signe négatif): {debit}")
                                            else:
                                                credit = cleaned_amount
                                                print(f"DEBUG: Crédit assigné (signe positif): {credit}")
                                            break
                                
                                # 2. Si pas de classification par signe, utiliser les mots-clés
                                if not debit and not credit:
                                    for amount in amounts:
                                        if ',' in amount or ('.' in amount and not amount.startswith('.')):
                                            cleaned_amount = clean_zitouna_extrait_amount(amount)
                                            if cleaned_amount:
                                                libelle_upper = libelle.upper()
                                                
                                                # Mots-clés pour crédit (entrées d'argent)
                                                credit_keywords = ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT', 'VIREMENT RECU', 'AV TPE']
                                                
                                                # Mots-clés pour débit (sorties d'argent)
                                                debit_keywords = ['PAIEMENT', 'COMMISSION', 'FRAIS', 'TVA', 'RETRAIT', 'PRELEVEMENT', 'EFFET', 'COMM REGLEMENT', 'COMM REMISE', 'DROIT DE TIMBRE', 'PRIME TAKAFUL', 'PROFIT IJARA', 'PRINCIPAL IJARA']
                                                
                                                if any(keyword in libelle_upper for keyword in credit_keywords):
                                                    credit = cleaned_amount
                                                    print(f"DEBUG: Crédit assigné (mots-clés): {credit}")
                                                elif any(keyword in libelle_upper for keyword in debit_keywords):
                                                    debit = cleaned_amount
                                                    print(f"DEBUG: Débit assigné (mots-clés): {debit}")
                                                else:
                                                    # Par défaut selon le signe
                                                    if amount.startswith('-'):
                                                        debit = cleaned_amount.lstrip('-')
                                                        print(f"DEBUG: Débit assigné (défaut): {debit}")
                                                    else:
                                                        credit = cleaned_amount
                                                        print(f"DEBUG: Crédit assigné (défaut): {credit}")
                                                break
                                
                                # 3. Validation finale : s'assurer qu'on a soit débit soit crédit, pas les deux
                                if debit and credit:
                                    libelle_upper = libelle.upper()
                                    if any(keyword in libelle_upper for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT']):
                                        debit = None  # Garder seulement le crédit
                                        print(f"DEBUG: Débit supprimé, garde crédit: {credit}")
                                    else:
                                        credit = None  # Garder seulement le débit
                                        print(f"DEBUG: Crédit supprimé, garde débit: {debit}")
                                
                                if (debit or credit):
                                    # Si le libellé est vide, utiliser une description par défaut
                                    if not libelle or libelle == "Transaction":
                                        libelle = "Transaction bancaire"
                                    
                                    results.append({
                                        "date": date_operation,
                                        "libelle": libelle,
                                        "debit": debit,
                                        "credit": credit
                                    })
        
        print(f"DEBUG: Total de {len(results)} transactions trouvées avec le parser complet")
        return results

    def _format_amount_with_spaces(self, amount_float):
        """Formate un montant avec espaces comme séparateurs de milliers et virgule comme décimal"""
        if amount_float is None:
            return None
        
        # Convertir en string avec 3 décimales
        amount_str = f"{amount_float:,.3f}"
        
        # Remplacer les virgules par des espaces (séparateurs de milliers)
        # et le point par une virgule (séparateur décimal)
        formatted = amount_str.replace(',', ' ').replace('.', ',', 1)
        
        return formatted

    def _filter_line(self, line_text: str) -> bool:
        """Filtre les lignes non-transactionnelles"""
        line_upper = line_text.upper()
        
        # Exclure les lignes qui sont clairement des en-têtes/pieds de page
        # (mais pas les lignes de transaction qui contiennent ces mots)
        
        # Lignes qui sont uniquement des en-têtes (pas de montants)
        header_only_patterns = [
            'EXTRAT', 'COMPTE', 'AGENCE', 'PAGE', 'NOM', 'R.S.', 'SOLDE', 'VEILLE'
            'DU', 'AU', 'EXTXML', 'BANQUE', 'ATTIJARI', 'ZITOUNA', 'BANK'
            'DATE', 'OPERATION', 'LIBELLE', 'DEBIT', 'CREDIT', 'REFERENCE'
            'MONTANT', 'TOTAL', 'SOUS TOTAL', 'RELEVE', 'PERIODE'
        ]
        
        # Vérifier si c'est une ligne d'en-tête pure (sans montants)
        has_amount = bool(re.search(r'\d+[.,]\d+', line_text))  # Contient un montant
        has_date = bool(re.search(r'\d{2}/\d{2}/\d{4}', line_text))  # Contient une date
        
        # Si c'est une ligne d'en-tête pure (contient des mots-clés mais pas de montants/dates)
        if not has_amount and not has_date:
            for keyword in header_only_patterns:
                if keyword in line_upper:
                    return True
        
        # Exclure les lignes trop courtes (probablement des titres)
        if len(line_text.strip()) < 10:
            return True
            
        return False

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "J03"
        
        # Styles
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        
        # En-têtes
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        
        # Largeurs des colonnes
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 50  # Libellé
        ws.column_dimensions['C'].width = 16  # Débit
        ws.column_dimensions['D'].width = 16  # Crédit
        
        # Format des nombres avec espaces comme séparateurs de milliers et virgule comme décimal
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            # Format: espace pour milliers, virgule pour décimales, 3 décimales exactement
            ws[f'C{r}'].number_format = '# ##0,000'  # Débit
            ws[f'D{r}'].number_format = '# ##0,000'  # Crédit
        
        # Bordures
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        
        wb.save(path)

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

def main():
    root = tk.Tk()
    ZitounaExtraitConverter(root)
    root.mainloop()

if __name__ == '__main__':
    main()
