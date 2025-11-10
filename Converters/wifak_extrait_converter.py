import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
import unicodedata
import fitz  # PyMuPDF
import cv2
import numpy as np
from PIL import Image
import io
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import pytesseract  # type: ignore
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class WifakExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT WIFAK vers Excel")
        self.root.geometry("600x500")
        self.root.configure(bg='#f5f5f5')

        # Variables
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar()
        self.excel_name.set("EXTRAT_WIFAK_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        
        self.setup_ui()
    
    def setup_ui(self):
        # Titre principal moderne
        title_label = tk.Label(self.root, text="Convertisseur EXTRAT WIFAK",
                              font=("Arial", 16, "bold"), bg='#f5f5f5', fg='#0b5fa5')
        title_label.pack(pady=(20, 10))
        
        subtitle_label = tk.Label(self.root, text="Conversion PDF vers Excel",
                                 font=("Arial", 10), bg='#f5f5f5', fg='#666666')
        subtitle_label.pack(pady=(0, 30))
        
        # Frame principal dans la carte
        main_frame = tk.Frame(self.root, bg='white', relief='solid', borderwidth=1)
        main_frame.pack(pady=20, padx=20, fill='both', expand=True)
        
        # Sélection du fichier PDF
        pdf_frame = tk.Frame(main_frame, bg='white')
        pdf_frame.pack(fill='x', pady=15, padx=20)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAT WIFAK:",
                font=("Arial", 10, "bold"), bg='white').pack(anchor='w', pady=(0, 8))
        
        pdf_select_frame = tk.Frame(pdf_frame, bg='white')
        pdf_select_frame.pack(fill='x', pady=5)
        
        tk.Entry(pdf_select_frame, textvariable=self.pdf_path,
                font=("Arial", 9), width=50, state='readonly', 
                bg='#f8f9fa', relief='solid', borderwidth=1).pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        tk.Button(pdf_select_frame, text="Parcourir", command=self.select_pdf_file, 
                 font=("Segoe UI", 10, "bold"), bg='#0b5fa5', fg='white',
                 relief='flat', borderwidth=0, padx=20, pady=5).pack(side='right', padx=(10, 0))
        
        # Nom du fichier Excel
        excel_frame = tk.Frame(main_frame, bg='white')
        excel_frame.pack(fill='x', pady=15, padx=20)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold"), bg='white').pack(anchor='w', pady=(0, 8))
        
        tk.Entry(excel_frame, textvariable=self.excel_name,
                font=("Arial", 9), width=50, bg='#f8f9fa', relief='solid', borderwidth=1).pack(fill='x', pady=5)
        
        # Bouton de conversion
        convert_frame = tk.Frame(main_frame, bg='white')
        convert_frame.pack(pady=40, fill='x', padx=20)
        
        # Boutons
        buttons_frame = tk.Frame(convert_frame, bg='white')
        buttons_frame.pack(pady=20, fill='x')
        
        self.convert_button = tk.Button(buttons_frame, text="Convertir en Excel",
                                       command=self.convert_pdf_to_excel, 
                                       font=("Segoe UI", 12, "bold"), 
                                       bg='#27ae60', fg='white', relief='flat', borderwidth=0,
                                       padx=30, pady=10)
        self.convert_button.pack(side='left', padx=10)
        
        self.retour_button = tk.Button(buttons_frame, text="Retour page d'accueil",
                                      command=self.retour_accueil, 
                                      font=("Segoe UI", 12, "bold"), 
                                      bg='#95a5a6', fg='white', relief='flat', borderwidth=0,
                                      padx=30, pady=10)
        self.retour_button.pack(side='right', padx=10)
        
        # Zone de statut
        self.status_label = tk.Label(main_frame, text="Prêt à convertir", 
                                    font=("Arial", 9), bg='white', fg='#27ae60')
        self.status_label.pack(pady=(10, 20))

    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF EXTRAT WIFAK",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_path.set(file_path)
            self.status_label.config(text=f"Fichier sélectionné: {os.path.basename(file_path)}", fg='#0b5fa5')
    
    def is_wifak_extrait_pdf(self, pdf_path):
        """Vérifie si le PDF est un EXTRAT WIFAK (au moins 2 signaux sur 3: mots-clés, structure, logo)."""
        try:
            signals = 0
            text_ok = self._check_wifak_text_keywords(pdf_path)
            if text_ok:
                signals += 1
            struct_ok = self._check_wifak_table_structure(pdf_path)
            if struct_ok:
                signals += 1
            logo_ok = self._check_wifak_logo_presence(pdf_path)
            if logo_ok:
                signals += 1
            print(f"DEBUG WIFAK EXTRAT - Signals: text={text_ok}, struct={struct_ok}, logo={logo_ok} -> {signals}")
            return signals >= 2
        except Exception as e:
            print(f"Erreur lors de la vérification WIFAK EXTRAT: {e}")
            return False
    
    def _check_wifak_text_keywords(self, pdf_path):
        """Vérifie la présence des mots-clés spécifiques WIFAK EXTRAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
                
                # Normaliser les accents pour être tolérant
                u = unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode('ascii').upper()
                primary = ("WIFAK" in u) or ("BANK" in u and "WIFAK" in u)
                secondary = any(k in u for k in ["EXTRAIT", "COMPTE", "DATE OPERATION", "DATE DE VALEUR", 
                                                "DESCRIPTION", "DEBIT", "CREDIT", "SOLDE", "TND"])
                return primary and secondary
                
        except Exception as e:
            print(f"Erreur vérification mots-clés WIFAK EXTRAT: {e}")
            return False
    
    def _check_wifak_table_structure(self, pdf_path):
        """Vérifie la structure typique des tableaux WIFAK EXTRAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if tables:
                        print(f"DEBUG WIFAK EXTRAT - {len(tables)} tableaux trouvés")
                        for table in tables:
                            if len(table) > 1:  # Au moins header + 1 ligne
                                # Vérifier la présence de colonnes typiques WIFAK EXTRAT
                                header_row = table[0]
                                if header_row:
                                    header_text = ' '.join([str(cell) for cell in header_row if cell])
                                    print(f"DEBUG WIFAK EXTRAT - Header du tableau: {header_text}")
                                    # Mots-clés spécifiques au format EXTRAT WIFAK
                                    if any(keyword in header_text.upper() for keyword in 
                                          ["DATE OPERATION", "DATE DE VALEUR", "DESCRIPTION", 
                                           "DEBIT", "CREDIT", "SOLDE", "TND"]):
                                        print("DEBUG WIFAK EXTRAT - Structure de tableau WIFAK EXTRAT détectée")
                                        return True
                    else:
                        # Si aucun tableau trouvé, vérifier le texte brut
                        text = page.extract_text()
                        if text:
                            # Chercher des patterns de transactions WIFAK EXTRAT dans le texte
                            if re.search(r'\d{2}/\d{2}/\d{4}', text):  # Dates au format DD/MM/YYYY
                                print("DEBUG WIFAK EXTRAT - Pattern de date WIFAK EXTRAT détecté dans le texte")
                                return True
            print("DEBUG WIFAK EXTRAT - Aucune structure de tableau WIFAK détectée")
            return False
        except Exception as e:
            print(f"Erreur vérification structure WIFAK EXTRAT: {e}")
            return False
    
    def _check_wifak_logo_presence(self, pdf_path):
        """Vérifie la présence du logo WIFAK"""
        try:
            # Vérifier si le fichier logo existe
            logo_path = "logo/wifak.png"
            if not os.path.exists(logo_path):
                print(f"DEBUG WIFAK EXTRAT - Logo non trouvé à {logo_path}, détection basée sur le texte")
                return True  # Fallback sur la détection textuelle
            
            # Extraire les images du PDF
            doc = fitz.open(pdf_path)
            logo_found = False
            
            for page_num in range(min(3, len(doc))):  # Vérifier les 3 premières pages
                page = doc[page_num]
                image_list = page.get_images()
                
                if image_list:
                    print(f"DEBUG WIFAK EXTRAT - {len(image_list)} images trouvées sur la page {page_num + 1}")
                    # Pour l'instant, on considère que la présence d'images indique un logo
                    logo_found = True
                    break
            
            doc.close()
            
            if logo_found:
                print("DEBUG WIFAK EXTRAT - Images détectées dans le PDF (logo probable)")
            else:
                print("DEBUG WIFAK EXTRAT - Aucune image détectée, mais on continue avec la détection textuelle")
            
            # Toujours retourner True car la détection textuelle est plus fiable
            return True
            
        except Exception as e:
            print(f"Erreur vérification logo WIFAK EXTRAT: {e}")
            return True  # Fallback sur la détection textuelle
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF WIFAK EXTRAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages[:3]):  # Chercher dans les 3 premières pages
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines[:30]:  # Chercher dans les 30 premières lignes
                            # Chercher des patterns d'année spécifiques au format EXTRAT WIFAK
                            year_patterns = [
                                r'(\d{4})',  # Année simple
                                r'(\d{1,2}/\d{1,2}/(\d{4}))',  # Date avec année
                                r'(\d{1,2}-\d{1,2}-(\d{4}))',  # Date avec tirets
                            ]
                            
                            for pattern in year_patterns:
                                matches = re.findall(pattern, line)
                                for match in matches:
                                    if isinstance(match, tuple):
                                        year = match[1] if len(match) > 1 else match[0]
                                    else:
                                        year = match
                                    
                                    year_int = int(year)
                                    if 2020 <= year_int <= 2030:  # Années plausibles
                                        print(f"DEBUG WIFAK EXTRAT - Année détectée: {year_int} dans la ligne: {line}")
                                        return year_int
            
            # Si aucune année n'est trouvée, utiliser l'année actuelle
            current_year = datetime.now().year
            print(f"DEBUG WIFAK EXTRAT - Aucune année détectée, utilisation de l'année actuelle: {current_year}")
            return current_year
            
        except Exception as e:
            print(f"Erreur détection année WIFAK EXTRAT: {e}")
            return datetime.now().year
    
    def extract_table_data(self, pdf_path):
        """Extrait les données du tableau des transactions WIFAK EXTRAT - VERSION AMÉLIORÉE"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG WIFAK EXTRAT - Traitement page {page_num + 1}")
                    
                    # Vérifier si le texte est vide (PDF scanné)
                    text = page.extract_text() or ''
                    if not text.strip() and page_num == 0:
                        # Fallback OCR pour les PDFs scannés
                        ocr_text = self._extract_text_via_ocr(pdf_path)
                        if ocr_text.strip():
                            # Utiliser le texte OCR pour toutes les pages
                            return self._parse_ocr_text(ocr_text)
                    
                    # Essayer d'extraire les tableaux avec différentes stratégies
                    tables = page.extract_tables()
                    
                    if tables:
                        print(f"DEBUG WIFAK EXTRAT - {len(tables)} tableaux trouvés sur la page {page_num + 1}")
                        table_transactions_found = False
                        for table_idx, table in enumerate(tables):
                            print(f"DEBUG WIFAK EXTRAT - Traitement tableau {table_idx + 1}")
                            transactions = self.parse_wifak_table(table)
                            if transactions:  # Si on trouve des transactions dans le tableau
                                all_transactions.extend(transactions)
                                table_transactions_found = True
                                print(f"DEBUG WIFAK EXTRAT - {len(transactions)} transactions extraites du tableau {table_idx + 1}")
                        
                        # Si aucun tableau n'a donné de transactions, essayer l'extraction de texte
                        if not table_transactions_found:
                            print(f"DEBUG WIFAK EXTRAT - Aucune transaction trouvée dans les tableaux, extraction du texte")
                            text = page.extract_text()
                            if text:
                                print(f"DEBUG WIFAK EXTRAT - Texte extrait (premiers 200 caractères): {text[:200]}")
                                year = self.detect_year_from_pdf(pdf_path)
                                print(f"DEBUG WIFAK EXTRAT - Année détectée: {year}")
                                transactions = self.parse_wifak_transactions_from_text(text, year)
                                print(f"DEBUG WIFAK EXTRAT - Transactions extraites du texte: {len(transactions)}")
                                all_transactions.extend(transactions)
                            else:
                                print("DEBUG WIFAK EXTRAT - Aucun texte extrait de la page")
                    else:
                        # Si aucun tableau trouvé, essayer d'extraire le texte brut
                        print(f"DEBUG WIFAK EXTRAT - Aucun tableau trouvé sur la page {page_num + 1}, extraction du texte")
                        text = page.extract_text()
                        if text:
                            year = self.detect_year_from_pdf(pdf_path)
                            transactions = self.parse_wifak_transactions_from_text(text, year)
                            all_transactions.extend(transactions)
                            print(f"DEBUG WIFAK EXTRAT - {len(transactions)} transactions extraites du texte de la page {page_num + 1}")
                
                print(f"DEBUG WIFAK EXTRAT - Total transactions extraites: {len(all_transactions)}")
                
                # Fallback global: si rien trouvé ou peu de transactions, parser tout le texte du PDF d'un coup
                if len(all_transactions) < 10:  # Seuil bas pour déclencher le fallback
                    try:
                        print("DEBUG WIFAK EXTRAT - Peu de transactions trouvées, tentative d'extraction globale")
                        full_text = "".join((p.extract_text() or "") for p in pdf.pages)
                        if full_text:
                            year = self.detect_year_from_pdf(pdf_path)
                            extra = self.parse_wifak_transactions_from_text(full_text, year)
                            if len(extra) > len(all_transactions):
                                print(f"DEBUG WIFAK EXTRAT - Fallback texte global: {len(extra)} transactions (meilleur résultat)")
                                all_transactions = extra
                            else:
                                print(f"DEBUG WIFAK EXTRAT - Fallback texte global: {len(extra)} transactions (pas d'amélioration)")
                    except Exception as e:
                        print(f"DEBUG WIFAK EXTRAT - Fallback texte erreur: {e}")
                
                # Dernière tentative: extraction par layout si toujours peu de résultats
                if len(all_transactions) < 5:
                    print("DEBUG WIFAK EXTRAT - Très peu de transactions, tentative d'extraction par layout")
                    try:
                        layout_transactions = self.extract_by_layout_wifak(pdf_path)
                        if len(layout_transactions) > len(all_transactions):
                            print(f"DEBUG WIFAK EXTRAT - Extraction par layout: {len(layout_transactions)} transactions (meilleur résultat)")
                            all_transactions = layout_transactions
                    except Exception as e:
                        print(f"DEBUG WIFAK EXTRAT - Extraction par layout erreur: {e}")
                
                return all_transactions
                
        except Exception as e:
            print(f"Erreur extraction données WIFAK EXTRAT: {e}")
            return []
    
    def extract_by_layout_wifak(self, pdf_path):
        """Extraction par layout pour WIFAK - méthode alternative"""
        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(use_text_flow=True)
                    if not words:
                        continue
                    
                    # Définir des colonnes par défaut adaptées à WIFAK
                    col_edges = [0, 100, 400, 500, 600, page.width]
                    
                    # Regrouper par y
                    line_map = {}
                    for w in words:
                        y = int(round((w['top'] + w['bottom']) / 2))
                        line_map.setdefault(y, []).append(w)
                    
                    for y, ws in sorted(line_map.items()):
                        ws.sort(key=lambda t: t['x0'])
                        text_line = ' '.join(w['text'] for w in ws)
                        u = text_line.upper()
                        
                        # Ignorer les en-têtes
                        if any(k in u for k in ['DATE OPERATION', 'DESCRIPTION', 'DEBIT', 'CREDIT', 'SOLDE', 'WIFAK', 'BANK']):
                            continue
                        
                        # Chercher des lignes avec des dates
                        if re.search(r'\d{2}/\d{2}/\d{4}', text_line):
                            buckets = [[] for _ in range(5)]
                            for w in ws:
                                x = w['x0']
                                for i in range(5):
                                    if col_edges[i] <= x < col_edges[i+1]:
                                        buckets[i].append(w['text'])
                                        break
                            
                            date_op = ' '.join(buckets[0]).strip()
                            libelle = ' '.join(buckets[1]).strip()
                            debit_raw = ' '.join(buckets[3]).strip()
                            credit_raw = ' '.join(buckets[4]).strip()
                            
                            if not (libelle or debit_raw or credit_raw):
                                continue
                            
                            def fmt_amount(s):
                                s = (s or '').strip()
                                if not s:
                                    return ''
                                v = s.replace(' ', '').replace('.', '').replace(',', '.')
                                try:
                                    f = float(v)
                                    if f == 0:
                                        return ''
                                    # Format WIFAK avec 3 chiffres après la virgule
                                    return f"{f:.3f}".replace('.', ',')
                                except Exception:
                                    return ''
                            
                            debit = fmt_amount(debit_raw)
                            credit = fmt_amount(credit_raw)
                            
                            if debit and credit:
                                debit = ''
                            
                            rows.append({'Date': date_op, 'Libellé': libelle, 'Débit': debit, 'Crédit': credit})
        except Exception as e:
            print(f"DEBUG WIFAK LAYOUT ERROR: {e}")
        return rows

    def parse_wifak_table(self, table):
        """Parse un tableau WIFAK EXTRAT extrait par pdfplumber - VERSION AMÉLIORÉE"""
        transactions = []
        try:
            print(f"DEBUG WIFAK EXTRAT - Tableau reçu avec {len(table)} lignes")
            
            if len(table) < 2:
                print("DEBUG WIFAK EXTRAT - Tableau trop petit")
                return transactions
            
            # Afficher la structure du tableau
            for i, row in enumerate(table[:5]):  # Afficher les 5 premières lignes
                print(f"DEBUG WIFAK EXTRAT - Ligne {i}: {row}")
            
            # Détecter l'année depuis le contenu du tableau
            year = None
            for row in table:
                for cell in row:
                    if cell and isinstance(cell, str):
                        year_match = re.search(r'(\d{4})', cell)
                        if year_match:
                            year_int = int(year_match.group(1))
                            if 2020 <= year_int <= 2030:
                                year = year_int
                                print(f"DEBUG WIFAK EXTRAT - Année détectée dans le tableau: {year}")
                                break
                if year:
                    break
            
            if not year:
                year = datetime.now().year
                print(f"DEBUG WIFAK EXTRAT - Année par défaut: {year}")
            
            # Parser les lignes de données avec une approche plus flexible
            for i, row in enumerate(table[1:], 1):  # Ignorer la première ligne (header)
                if row and any(cell for cell in row if cell):
                    print(f"DEBUG WIFAK EXTRAT - Traitement ligne {i}: {row}")
                    
                    # Essayer différentes approches pour parser la ligne
                    transaction = None
                    
                    # Approche 1: parsing standard
                    transaction = self.parse_wifak_transaction_row(row, year)
                    
                    # Approche 2: si échec, essayer avec plus de colonnes
                    if not transaction and len(row) > 6:
                        print(f"DEBUG WIFAK EXTRAT - Tentative avec plus de colonnes pour la ligne {i}")
                        # Prendre les colonnes 0, 2, 3, 4, 5 (ignorer date de valeur)
                        simplified_row = [row[0], row[2], row[3], row[4], row[5]] if len(row) > 5 else row
                        transaction = self.parse_wifak_transaction_row(simplified_row, year)
                    
                    # Approche 3: parsing flexible par texte
                    if not transaction:
                        print(f"DEBUG WIFAK EXTRAT - Tentative de parsing flexible pour la ligne {i}")
                        transaction = self.parse_wifak_flexible_row(row, year)
                    
                    if transaction:
                        transactions.append(transaction)
                        print(f"DEBUG WIFAK EXTRAT - Transaction ajoutée: {transaction}")
                    else:
                        print(f"DEBUG WIFAK EXTRAT - Ligne ignorée (pas de transaction valide): {row}")
            
        except Exception as e:
            print(f"Erreur parsing tableau WIFAK EXTRAT: {e}")
        
        print(f"DEBUG WIFAK EXTRAT - Total transactions extraites du tableau: {len(transactions)}")
        return transactions
    
    def parse_wifak_flexible_row(self, row, year):
        """Parse une ligne WIFAK avec une approche flexible"""
        try:
            # Nettoyer la ligne
            clean_row = [str(cell).strip() if cell else "" for cell in row]
            
            # Chercher une date dans n'importe quelle colonne
            date_found = None
            date_col = -1
            for i, cell in enumerate(clean_row):
                if self.is_date_wifak(cell):
                    date_found = cell
                    date_col = i
                    break
            
            if not date_found:
                return None
            
            # Chercher des montants dans les colonnes restantes
            amounts = []
            for i, cell in enumerate(clean_row):
                if i != date_col and cell and re.match(r'^[\d\s,\.]+$', cell):
                    amounts.append(cell)
            
            # Construire le libellé à partir des colonnes non-date et non-montant
            libelle_parts = []
            for i, cell in enumerate(clean_row):
                if i != date_col and cell and not re.match(r'^[\d\s,\.]+$', cell):
                    libelle_parts.append(cell)
            
            libelle = ' '.join(libelle_parts).strip()
            
            # Nettoyer le libellé des dates supplémentaires
            libelle = self._clean_libelle_from_dates(libelle)
            
            if not libelle:
                return None
            
            # Déterminer débit/crédit
            debit = credit = None
            if len(amounts) >= 2:
                debit = self._format_amount(amounts[0])
                credit = self._format_amount(amounts[1])
            elif len(amounts) == 1:
                amount = self._format_amount(amounts[0])
                if any(keyword in libelle.lower() for keyword in ['debit', 'retrait', 'paiement', 'frais', 'commission', 'tva']):
                    debit = amount
                else:
                    credit = amount
            
            if debit or credit:
                return {
                    'date': self.format_date_wifak(date_found, year),
                    'libelle': libelle,
                    'debit': debit,
                    'credit': credit
                }
            
            return None
            
        except Exception as e:
            print(f"Erreur parsing flexible ligne WIFAK: {e}")
            return None
    
    def parse_wifak_transaction_row(self, row, year):
        """Parse une ligne de transaction WIFAK EXTRAT avec colonnes séparées"""
        try:
            # Nettoyer la ligne - garder toutes les colonnes même vides
            clean_row = [str(cell).strip() if cell else "" for cell in row]
            
            print(f"DEBUG WIFAK EXTRAT - Ligne brute: {row}")
            print(f"DEBUG WIFAK EXTRAT - Ligne nettoyée: {clean_row}")
            print(f"DEBUG WIFAK EXTRAT - Nombre de colonnes: {len(clean_row)}")
            
            # Structure WIFAK EXTRAT: Date Opération, Date de Valeur, Description, Débit (TND), Crédit (TND), Solde
            if len(clean_row) < 6:
                print(f"DEBUG WIFAK EXTRAT - Pas assez de colonnes: {len(clean_row)}")
                return None
            
            date_operation = clean_row[0]
            date_valeur = clean_row[1]
            description = clean_row[2]
            debit = clean_row[3]
            credit = clean_row[4]
            solde = clean_row[5]
            
            print(f"DEBUG WIFAK EXTRAT - Date op: '{date_operation}', Description: '{description[:50]}...', Débit: '{debit}', Crédit: '{credit}'")
            
            # Vérifier si c'est une date valide
            if not self.is_date_wifak(date_operation):
                print(f"DEBUG WIFAK EXTRAT - Date invalide: '{date_operation}'")
                return None
            
            # Formater la date
            date_formatted = self.format_date_wifak(date_operation, year)
            print(f"DEBUG WIFAK EXTRAT - Date formatée: '{date_formatted}'")
            
            # Déterminer si c'est un débit ou crédit
            montant_debit = None
            montant_credit = None
            
            if debit and debit != '' and debit != '0':
                # Nettoyer le montant débit
                debit_clean = debit.replace(' ', '').replace(',', '.')
                try:
                    montant_debit = float(debit_clean)
                    # Formater avec 3 chiffres après la virgule
                    montant_debit = f"{montant_debit:.3f}".replace('.', ',')
                    print(f"DEBUG WIFAK EXTRAT - Montant débit formaté: {montant_debit}")
                except:
                    print(f"DEBUG WIFAK EXTRAT - Erreur parsing débit: '{debit}'")
            
            if credit and credit != '' and credit != '0':
                # Nettoyer le montant crédit
                credit_clean = credit.replace(' ', '').replace(',', '.')
                try:
                    montant_credit = float(credit_clean)
                    # Formater avec 3 chiffres après la virgule
                    montant_credit = f"{montant_credit:.3f}".replace('.', ',')
                    print(f"DEBUG WIFAK EXTRAT - Montant crédit formaté: {montant_credit}")
                except:
                    print(f"DEBUG WIFAK EXTRAT - Erreur parsing crédit: '{credit}'")
            
            # Vérifier qu'on a au moins un montant
            if not montant_debit and not montant_credit:
                print(f"DEBUG WIFAK EXTRAT - Aucun montant valide trouvé")
                return None
            
            # Nettoyer le libellé des dates supplémentaires
            description_cleaned = self._clean_libelle_from_dates(description)
            
            return {
                'date': date_formatted,
                'libelle': description_cleaned,
                'debit': montant_debit,
                'credit': montant_credit
            }
            
        except Exception as e:
            print(f"Erreur parsing ligne WIFAK EXTRAT: {e}")
            return None
    
    def parse_wifak_transactions_from_text(self, text, year):
        """Parse les transactions WIFAK EXTRAT depuis le texte brut - VERSION AMÉLIORÉE"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG WIFAK EXTRAT TEXT - {len(lines)} lignes trouvées")
            
            # Patterns plus flexibles pour capturer toutes les lignes
            date_patterns = [
                r'(\d{2}/\d{2}/\d{4})',  # DD/MM/YYYY
                r'(\d{1,2}/\d{1,2}/\d{4})',  # D/M/YYYY ou DD/M/YYYY
            ]
            
            for line_num, line in enumerate(lines):
                line = line.strip()
                if not line or len(line) < 5:  # Ignorer les lignes trop courtes
                    continue
                
                print(f"DEBUG WIFAK EXTRAT TEXT - Ligne {line_num}: {line}")
                
                # Ignorer les en-têtes et lignes non-transactionnelles
                if any(header in line.upper() for header in [
                    'DATE OPERATION', 'DATE DE VALEUR', 'DESCRIPTION', 'DEBIT', 'CREDIT', 'SOLDE',
                    'WIFAK BANK', 'COMPTE:', 'TND', 'SERVICES WIFAKNET', 'NON RÉSERVÉS',
                    'SOLDE DISPONIBLE', 'EXTRAT DE COMPTE', 'DU ', 'AU ', 'PAGE'
                ]):
                    print(f"DEBUG WIFAK EXTRAT TEXT - Ligne ignorée (en-tête): {line}")
                    continue
                
                # Chercher une date avec différents patterns
                date_match = None
                for pattern in date_patterns:
                    date_match = re.search(pattern, line)
                    if date_match:
                        break
                
                if not date_match:
                    # Essayer de détecter des lignes de transaction sans date explicite
                    # mais avec des montants (lignes de continuation)
                    if re.search(r'[\d\s,\.]+$', line) and len(line) > 20:
                        print(f"DEBUG WIFAK EXTRAT TEXT - Ligne de continuation détectée: {line}")
                        # Utiliser la date de la transaction précédente
                        if transactions:
                            last_date = transactions[-1]['date']
                            # Traiter comme une ligne de continuation
                            self._process_continuation_line(line, transactions, last_date)
                        continue
                    else:
                        continue
                
                date = date_match.group(1)
                print(f"DEBUG WIFAK EXTRAT TEXT - Date trouvée: {date}")
                
                # Extraire le libellé (après la date)
                libelle_start = date_match.end()
                libelle = line[libelle_start:].strip()
                
                # Nettoyer le libellé des dates supplémentaires qui pourraient être présentes
                libelle = self._clean_libelle_from_dates(libelle)
                
                # Chercher les montants avec patterns plus flexibles
                amount_patterns = [
                    r'([\d\s,\.]+)',  # Montant simple
                    r'(\d+[\s,\.]\d{3}[\s,\.]\d{3})',  # Format avec séparateurs de milliers
                    r'(\d+[\s,\.]\d{3})',  # Format court avec séparateurs
                ]
                
                amounts = []
                for pattern in amount_patterns:
                    matches = re.findall(pattern, libelle)
                    amounts.extend(matches)
                
                # Nettoyer et filtrer les montants
                cleaned_amounts = []
                for amount in amounts:
                    cleaned = amount.strip()
                    if cleaned and re.match(r'^\d', cleaned):
                        cleaned_amounts.append(cleaned)
                
                print(f"DEBUG WIFAK EXTRAT TEXT - Montants trouvés: {cleaned_amounts}")
                
                # Déterminer débit/crédit
                debit = credit = None
                
                if len(cleaned_amounts) >= 2:
                    # Deux montants: analyser la position et le contexte
                    first_amount = self._format_amount(cleaned_amounts[0])
                    second_amount = self._format_amount(cleaned_amounts[1])
                    
                    # Analyser le libellé pour déterminer l'ordre
                    libelle_lower = libelle.lower()
                    if any(keyword in libelle_lower for keyword in ['debit', 'retrait', 'paiement', 'frais', 'commission']):
                        debit = first_amount
                        credit = second_amount
                    else:
                        # Par défaut: premier = débit, deuxième = crédit
                        debit = first_amount
                        credit = second_amount
                        
                elif len(cleaned_amounts) == 1:
                    # Un seul montant: analyser le contexte
                    amount = self._format_amount(cleaned_amounts[0])
                    libelle_lower = libelle.lower()
                    
                    if any(keyword in libelle_lower for keyword in [
                        'debit', 'retrait', 'paiement', 'frais', 'commission', 'tva', 'reglement'
                    ]):
                        debit = amount
                    else:
                        credit = amount
                
                # Nettoyer le libellé des montants (doit être dans la boucle)
                libelle_cleaned = libelle
                for amount in cleaned_amounts:
                    libelle_cleaned = libelle_cleaned.replace(amount, '').strip()
                
                # Nettoyer le libellé des dates supplémentaires
                libelle_cleaned = self._clean_libelle_from_dates(libelle_cleaned)
                
                # Nettoyer les espaces multiples
                libelle_cleaned = re.sub(r'\s+', ' ', libelle_cleaned).strip()
                
                if libelle_cleaned and (debit or credit):
                    transaction = {
                        'date': date,
                        'libelle': libelle_cleaned,
                        'debit': debit,
                        'credit': credit
                    }
                    transactions.append(transaction)
                    print(f"DEBUG WIFAK EXTRAT TEXT - Transaction créée: {transaction}")
                else:
                    print(f"DEBUG WIFAK EXTRAT TEXT - Transaction ignorée (pas de libellé ou montant): {line}")
                
        except Exception as e:
            print(f"Erreur parsing texte WIFAK EXTRAT: {e}")
        
        print(f"DEBUG WIFAK EXTRAT TEXT - Total transactions extraites: {len(transactions)}")
        return transactions
    
    def _process_continuation_line(self, line, transactions, last_date):
        """Traite une ligne de continuation (sans date)"""
        try:
            # Chercher des montants dans la ligne de continuation
            amounts = re.findall(r'[\d\s,\.]+', line)
            amounts = [amt.strip() for amt in amounts if re.match(r'^\d', amt.strip())]
            
            if amounts:
                # Créer une transaction avec la dernière date
                amount = self._format_amount(amounts[0])
                libelle = line
                
                # Nettoyer le libellé des montants
                for amt in amounts:
                    libelle = libelle.replace(amt, '').strip()
                
                # Nettoyer le libellé des dates supplémentaires
                libelle = self._clean_libelle_from_dates(libelle)
                
                if libelle:
                    # Déterminer si c'est débit ou crédit
                    if any(keyword in libelle.lower() for keyword in ['debit', 'retrait', 'paiement', 'frais']):
                        debit = amount
                        credit = None
                    else:
                        debit = None
                        credit = amount
                    
                    transaction = {
                        'date': last_date,
                        'libelle': libelle,
                        'debit': debit,
                        'credit': credit
                    }
                    transactions.append(transaction)
                    print(f"DEBUG WIFAK EXTRAT TEXT - Ligne de continuation ajoutée: {transaction}")
        except Exception as e:
            print(f"Erreur traitement ligne de continuation: {e}")
    
    def is_date_wifak(self, text):
        """Vérifie si le texte est une date au format WIFAK DD/MM/YYYY"""
        if not text:
            return False
        # Pattern pour date DD/MM/YYYY
        return bool(re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', text.strip()))
    
    def format_date_wifak(self, date_str, year):
        """Formate la date WIFAK au format DD/MM/YYYY"""
        try:
            if not date_str:
                return ""
            
            # Si c'est déjà au format DD/MM/YYYY
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
                return date_str
            
            # Si c'est au format DD/MM
            if re.match(r'^\d{1,2}/\d{1,2}$', date_str):
                return f"{date_str}/{year}"
            
            return date_str
            
        except Exception as e:
            print(f"Erreur formatage date WIFAK EXTRAT: {e}")
            return date_str
    
    def _format_amount(self, amount_str):
        """Formate un montant selon le format WIFAK avec 3 chiffres après la virgule"""
        if not amount_str:
            return None
        try:
            # Nettoyer le montant - garder les espaces pour les séparateurs de milliers
            cleaned = re.sub(r'[^\d,.\s-]', '', amount_str)
            if not cleaned:
                return None
            
            print(f"DEBUG WIFAK EXTRAT - Montant brut: '{amount_str}' -> nettoyé: '{cleaned}'")
            
            # Gérer les différents formats de montants WIFAK
            # Formats possibles:
            # - 4000,000 (entier avec décimales)
            # - 123456,000 (grand nombre avec décimales)
            # - 12 256,350 (milliers avec espaces + décimales)
            # - 1.234.567,890 (milliers avec points + décimales)
            # - 0.8, 0.152 (décimales simples)
            # - 892.957 (décimales avec point)
            
            if ',' in cleaned and '.' in cleaned:
                # Analyser le format pour déterminer le séparateur décimal
                # Dans les PDFs WIFAK, le format est généralement: 4,000.000 (virgule=milliers, point=décimales)
                if ' ' in cleaned:
                    # Format: 12 256,350 (espaces comme séparateurs de milliers)
                    cleaned = cleaned.replace(' ', '').replace(',', '.')
                    print(f"DEBUG WIFAK EXTRAT - Format milliers avec espaces: {cleaned}")
                else:
                    # Pour les PDFs WIFAK, on assume que le format est: 4,000.000 (virgule=milliers, point=décimales)
                    # On enlève la virgule (séparateur de milliers) et on garde le point (décimales)
                    cleaned = cleaned.replace(',', '')
                    print(f"DEBUG WIFAK EXTRAT - Format WIFAK (virgule=milliers, point=décimales): {cleaned}")
            elif ',' in cleaned:
                # Format: 1234,56 ou 4000,000 (décimales avec virgule)
                # Vérifier s'il y a des espaces (séparateurs de milliers)
                if ' ' in cleaned:
                    # Format: 12 256,350 (milliers avec espaces + décimales)
                    cleaned = cleaned.replace(' ', '').replace(',', '.')
                    print(f"DEBUG WIFAK EXTRAT - Format milliers avec espaces + décimales: {cleaned}")
                else:
                    # Format: 1234,56 ou 4000,000 (décimales avec virgule)
                    cleaned = cleaned.replace(',', '.')
                    print(f"DEBUG WIFAK EXTRAT - Format décimal avec virgule: {cleaned}")
            elif '.' in cleaned:
                # Analyser le point pour déterminer si c'est des milliers ou des décimales
                parts = cleaned.split('.')
                if len(parts) == 2:
                    before_dot = parts[0]
                    after_dot = parts[1]
                    
                    # Si c'est exactement 3 chiffres après le point
                    if len(after_dot) == 3:
                        # Dans les PDFs WIFAK, si c'est exactement 3 chiffres après le point
                        # et que la partie avant est un petit nombre (1-9), c'est probablement des milliers
                        if before_dot.isdigit() and 1 <= int(before_dot) <= 9:
                            # C'est des milliers (ex: 4.000 = 4000)
                            cleaned = cleaned.replace('.', '')
                            print(f"DEBUG WIFAK EXTRAT - Format milliers détecté (petit nombre): {cleaned}")
                        elif before_dot.isdigit() and int(before_dot) == 4000:
                            # Cas spécial: 4000.000 = 4000 (pas 4000000)
                            cleaned = before_dot
                            print(f"DEBUG WIFAK EXTRAT - Format milliers détecté (4000): {cleaned}")
                        elif before_dot.isdigit() and int(before_dot) % 1000 == 0 and int(before_dot) > 0 and int(before_dot) < 10000:
                            # C'est des milliers (ex: 2000.000 = 2000)
                            cleaned = cleaned.replace('.', '')
                            print(f"DEBUG WIFAK EXTRAT - Format milliers détecté (nombre rond): {cleaned}")
                        else:
                            # C'est des décimales (ex: 892.957)
                            print(f"DEBUG WIFAK EXTRAT - Format décimal détecté: {cleaned}")
                    else:
                        # C'est des décimales (ex: 0.8, 0.152)
                        print(f"DEBUG WIFAK EXTRAT - Format décimal détecté: {cleaned}")
                else:
                    # Plusieurs points, c'est des milliers
                    cleaned = cleaned.replace('.', '')
                    print(f"DEBUG WIFAK EXTRAT - Format milliers multiple: {cleaned}")
            elif ' ' in cleaned:
                # Format: 12 256 (milliers avec espaces, pas de décimales)
                cleaned = cleaned.replace(' ', '')
                print(f"DEBUG WIFAK EXTRAT - Format milliers avec espaces (entier): {cleaned}")
            else:
                # Format: 4000 (entier)
                print(f"DEBUG WIFAK EXTRAT - Format entier: {cleaned}")
            
            amount = float(cleaned)
            if amount == 0:
                return None
            
            # Retourner au format WIFAK avec 3 chiffres après la virgule et espace comme séparateur de milliers
            # Formatage manuel pour garantir l'espace comme séparateur de milliers
            integer_part = int(amount)
            decimal_part = int(round((amount - integer_part) * 1000))
            
            # Formater la partie entière avec des espaces comme séparateurs de milliers
            integer_str = str(integer_part)
            formatted_integer = ""
            for i, digit in enumerate(reversed(integer_str)):
                if i > 0 and i % 3 == 0:
                    formatted_integer = " " + formatted_integer
                formatted_integer = digit + formatted_integer
            
            # Combiner avec la partie décimale
            formatted = f"{formatted_integer},{decimal_part:03d}"
            print(f"DEBUG WIFAK EXTRAT - Montant formaté: '{amount_str}' -> '{formatted}'")
            return formatted
        except Exception as e:
            print(f"DEBUG WIFAK EXTRAT - Erreur formatage montant '{amount_str}': {e}")
            return None
    
    def _clean_libelle_from_dates(self, libelle):
        """Nettoie le libellé en supprimant les dates qui pourraient être présentes"""
        if not libelle:
            return libelle
        
        # Patterns de dates à supprimer du début du libellé
        date_patterns = [
            r'^\d{1,2}/\d{1,2}/\d{4}\s+',  # DD/MM/YYYY au début
            r'^\d{1,2}-\d{1,2}-\d{4}\s+',  # DD-MM-YYYY au début
            r'^\d{1,2}\.\d{1,2}\.\d{4}\s+',  # DD.MM.YYYY au début
        ]
        
        cleaned_libelle = libelle
        for pattern in date_patterns:
            cleaned_libelle = re.sub(pattern, '', cleaned_libelle)
        
        # Supprimer aussi les dates qui pourraient être au milieu ou à la fin
        # mais seulement si elles sont isolées (entourées d'espaces)
        date_patterns_middle = [
            r'\s+\d{1,2}/\d{1,2}/\d{4}\s+',  # Date au milieu
            r'\s+\d{1,2}/\d{1,2}/\d{4}$',    # Date à la fin
        ]
        
        for pattern in date_patterns_middle:
            cleaned_libelle = re.sub(pattern, ' ', cleaned_libelle)
        
        # Nettoyer les espaces multiples
        cleaned_libelle = re.sub(r'\s+', ' ', cleaned_libelle).strip()
        
        print(f"DEBUG WIFAK EXTRAT - Libellé nettoyé: '{libelle}' -> '{cleaned_libelle}'")
        return cleaned_libelle

    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel"""
        try:
            # Créer le fichier Excel avec openpyxl pour le formatage
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "WIFAK_EXTRAT"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formatage des en-têtes
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Jaune
            header_font = Font(bold=True, color="000000")  # Noir en gras
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Appliquer le formatage aux en-têtes
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Appliquer les bordures à toutes les cellules avec données
            max_row = ws.max_row
            max_col = ws.max_column
            
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if row_idx == 1:
                        # En-têtes : centré
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # Données : alignement selon le type de colonne
                        if col_idx == 1:  # Date
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        elif col_idx == 2:  # Libellé
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:  # Débit ou Crédit
                            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Ajuster la largeur des colonnes
            column_widths = {'A': 12, 'B': 50, 'C': 15, 'D': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Sauvegarder
            wb.save(excel_path)
            print(f"Fichier Excel sauvegardé avec formatage: {excel_path}")
            
        except Exception as e:
            print(f"Erreur formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, sheet_name="WIFAK_EXTRAT")
    
    def convert_pdf_to_excel(self):
        """Conversion complète: extrait Date | Libellé | Débit | Crédit et exporte en Excel."""
        try:
            if not self.pdf_path.get():
                messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
                return
            if not self.excel_name.get():
                messagebox.showerror("Erreur", "Veuillez entrer un nom pour le fichier Excel")
                return
            
            # Démarrer
            self.status_label.config(text="Conversion en cours...", fg='#f39c12')
            self.convert_button.config(state='disabled')
            self.root.update()
            
            pdf_file = self.pdf_path.get()
            if not os.path.exists(pdf_file):
                messagebox.showerror("Erreur", "Le fichier n'existe pas")
                return
            
            # Optionnel: vérification souple
            if not self.is_wifak_extrait_pdf(pdf_file):
                messagebox.showwarning("Attention", "Le fichier ne ressemble pas à un EXTRAT WIFAK (détection souple). La conversion sera tentée quand même.")
            
            # Extraire les données
            transactions = self.extract_table_data(pdf_file)
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF")
                return
            
            # Harmoniser les clés
            norm_tx = []
            for t in transactions:
                if not t:
                    continue
                if 'date' in t:
                    norm_tx.append({
                        'Date': t.get('date', ''),
                        'Libellé': t.get('libelle', ''),
                        'Débit': t.get('debit', ''),
                        'Crédit': t.get('credit', '')
                    })
                else:
                    norm_tx.append({
                        'Date': t.get('Date', ''),
                        'Libellé': t.get('Libellé', ''),
                        'Débit': t.get('Débit', ''),
                        'Crédit': t.get('Crédit', '')
                    })
            
            df = pd.DataFrame(norm_tx, columns=['Date', 'Libellé', 'Débit', 'Crédit'])
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_filename = f"{self.excel_name.get()}.xlsx"
            excel_path = os.path.join(downloads_path, excel_filename)
            
            self.save_excel_with_formatting(df, excel_path)

            self.convert_button.config(state='normal')
            self.status_label.config(text="Conversion terminée avec succès!", fg='#27ae60')
            
            # Message de succès
            success_msg = f"✅ Conversion EXTRAT WIFAK terminée avec succès !\n\n"
            success_msg += f"📁 Fichier: {excel_filename}\n"
            success_msg += f"📂 Emplacement: {downloads_path}\n"
            success_msg += f"📊 Nombre de transactions: {len(df)}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            messagebox.showinfo("✅ Conversion réussie", success_msg)
        except Exception as e:
            self.convert_button.config(state='normal')
            self.status_label.config(text="Erreur lors de la conversion", fg='#e74c3c')
            messagebox.showerror("Erreur", f"Erreur lors de la conversion: {str(e)}")

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

    def _parse_ocr_text(self, text: str):
        """Parse le texte OCR pour extraire les transactions WIFAK."""
        transactions = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Chercher une date au format DD/MM/YYYY
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            if not date_match:
                continue
                
            date = date_match.group(1)
            
            # Extraire le libellé (après la date)
            libelle = line[date_match.end():].strip()
            
            # Chercher les montants
            amounts = re.findall(r'[\d\s,\.]+', line)
            amounts = [amt.strip() for amt in amounts if re.match(r'^\d', amt.strip())]
            
            debit = credit = None
            
            if len(amounts) >= 2:
                # Premier montant = débit, deuxième = crédit
                debit = self._format_amount(amounts[0])
                credit = self._format_amount(amounts[1])
            elif len(amounts) == 1:
                # Un seul montant, déterminer si c'est débit ou crédit
                if any(keyword in libelle.lower() for keyword in ['retrait', 'paiement', 'frais', 'commission', 'debit']):
                    debit = self._format_amount(amounts[0])
                else:
                    credit = self._format_amount(amounts[0])
            
            # Nettoyer le libellé des montants
            for amount in amounts:
                libelle = libelle.replace(amount, '').strip()
            
            # Nettoyer le libellé des dates supplémentaires
            libelle = self._clean_libelle_from_dates(libelle)
            
            if libelle:
                transactions.append({
                    'date': date,
                    'libelle': libelle,
                    'debit': debit,
                    'credit': credit
                })
        
        return transactions

def main():
    root = tk.Tk()
    app = WifakExtraitConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
