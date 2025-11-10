#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convertisseur pour les extraits de compte UBCI Bank
Détecte automatiquement les documents UBCI et extrait les transactions
"""

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from datetime import datetime
import os
import sys
import argparse
import tkinter as tk
from tkinter import messagebox, filedialog, ttk

class UBCIExtraitConverter:
    def __init__(self, pdf_path, excel_path=None, silent=False):
        self.pdf_path = pdf_path
        self.excel_path = excel_path
        self.silent = silent
        self.df = None
        self.bank_name = "UBCI"
        self.document_type = "Extrait de Compte"

    def _log(self, message):
        if not self.silent:
            print(message)

    def detect_ubci_extrait(self, text_content):
        """
        Détecte si le PDF est un extrait de compte UBCI
        """
        if not text_content or len(text_content.strip()) < 10:
            self._log("Texte trop court ou vide pour la détection")
            return True  # Accepter même si le texte est court
        
        normalized_text = text_content.lower().replace('\n', ' ').replace('\r', ' ')

        # Mots-clés UBCI étendus
        ubci_keywords = [
            r'ubci',
            r'union\s+bancaire\s+pour\s+le\s+commerce\s+et\s+l\'industrie',
            r'الإتحاد\s+البنكي\s+للتجارة\s+والصناعة',
            r'ubci\s+-\s+société\s+anonyme',
            r'swift:\s+u\.b\.c\.i\s+tntt',
            r'union\s+bancaire',
            r'commerce\s+industrie'
        ]
        
        # Mots-clés pour extrait étendus
        extrait_keywords = [
            r'extrait\s+de\s+compte',
            r'كشف\s+حساب',
            r'extrait\s+compte',
            r'extrait',
            r'compte',
            r'statement',
            r'relevé'
        ]
        
        # Mots-clés pour les transactions étendus
        transaction_keywords = [
            r'date\s+opération',
            r'natures\s+des\s+opérations',
            r'débit',
            r'crédit',
            r'date\s+valeur',
            r'ref\s+banque',
            r'opération',
            r'transaction',
            r'montant',
            r'solde'
        ]

        ubci_score = sum(1 for keyword in ubci_keywords if re.search(keyword, normalized_text, re.IGNORECASE))
        extrait_score = sum(1 for keyword in extrait_keywords if re.search(keyword, normalized_text, re.IGNORECASE))
        transaction_score = sum(1 for keyword in transaction_keywords if re.search(keyword, normalized_text, re.IGNORECASE))

        # Calcul de la confiance
        total_keywords = len(ubci_keywords) + len(extrait_keywords) + len(transaction_keywords)
        confidence = (ubci_score + extrait_score + transaction_score) / total_keywords

        self._log(f"Détection UBCI - Score UBCI: {ubci_score}, Score Extrait: {extrait_score}, Score Transactions: {transaction_score}")
        self._log(f"Confiance totale: {confidence:.2f}")

        # Seuil de confiance très permissif
        if confidence > 0.05:  # Seuil encore plus bas
            return True
        
        # Détection simple par mots-clés essentiels (très permissive)
        simple_keywords = ['ubci', 'extrait', 'compte', 'débit', 'crédit', 'opération', 'transaction', 'banque', 'date', 'montant']
        simple_score = sum(1 for keyword in simple_keywords if keyword in normalized_text)
        
        self._log(f"Détection simple - Score: {simple_score}")
        
        # Si on trouve au moins 2 mots-clés, on accepte
        if simple_score >= 2:
            return True
            
        # Dernière chance : accepter si on trouve des dates et des montants
        date_pattern = r'\d{1,2}/\d{1,2}/\d{2,4}'
        amount_pattern = r'\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?'
        
        has_dates = bool(re.search(date_pattern, normalized_text))
        has_amounts = bool(re.search(amount_pattern, normalized_text))
        
        if has_dates and has_amounts:
            self._log("Détection par dates et montants - Accepté")
            return True
            
        # En dernier recours, accepter tout document qui contient "extrait" ou "compte"
        if 'extrait' in normalized_text or 'compte' in normalized_text:
            self._log("Détection par mots-clés de base - Accepté")
            return True
            
        return False

    def extract_transactions(self):
        """
        Extrait les transactions de l'extrait UBCI
        """
        all_data = []

        with pdfplumber.open(self.pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                self._log(f"Traitement de la page {page_num + 1}")

                # Extraire le texte pour détecter l'année
                text = page.extract_text()
                if not text:
                    continue

                # Détecter l'année
                year_match = re.search(r'(\d{4})', text)
                current_year = year_match.group(1) if year_match else str(datetime.now().year)

                # Paramètres de table optimisés pour UBCI
                table_settings = {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 5,  # Augmenté pour plus de flexibilité
                    "join_tolerance": 5,  # Augmenté pour plus de flexibilité
                    "edge_min_length": 2,  # Réduit pour capturer plus de lignes
                    "min_words_vertical": 1,
                    "min_words_horizontal": 1,
                    "text_tolerance": 5,  # Augmenté
                    "text_x_tolerance": 5,  # Augmenté
                    "text_y_tolerance": 5,  # Augmenté
                    "intersection_tolerance": 5,  # Augmenté
                }

                tables = page.extract_tables(table_settings)

                for table in tables:
                    if not table or len(table) < 2:
                        continue

                    # Trouver la ligne d'en-tête avec plus de flexibilité
                    header_row_idx = None
                    for idx, row in enumerate(table):
                        if row and any(col for col in row if col and (
                            "date opération" in col.lower() or 
                            "date operation" in col.lower() or
                            "date" in col.lower() and "opération" in col.lower() or
                            "date" in col.lower() and "operation" in col.lower()
                        )):
                            header_row_idx = idx
                            break
                    
                    if header_row_idx is None:
                        self._log(f"En-tête de table non trouvé sur la page {page_num + 1}")
                        continue

                    header_row = table[header_row_idx]
                    self._log(f"En-tête trouvé: {header_row}")

                    # Identifier les indices des colonnes
                    col_indices = {
                        'date_op': -1,
                        'libelle': -1,
                        'debit': -1,
                        'credit': -1,
                        'date_valeur': -1,
                        'ref_banque': -1
                    }
                    
                    for i, header in enumerate(header_row):
                        if header:
                            header_lower = header.lower().strip()
                            # Détection plus flexible pour la colonne date
                            if ('date' in header_lower and 'opération' in header_lower) or \
                               ('date' in header_lower and 'operation' in header_lower) or \
                               ('date' in header_lower and 'op' in header_lower) or \
                               header_lower == 'date':
                                col_indices['date_op'] = i
                                self._log(f"Colonne Date trouvée à l'index {i}: '{header}'")
                            elif 'natures des opérations' in header_lower or 'natures des operations' in header_lower:
                                col_indices['libelle'] = i
                            elif 'débit' in header_lower or 'debit' in header_lower:
                                col_indices['debit'] = i
                            elif 'crédit' in header_lower or 'credit' in header_lower:
                                col_indices['credit'] = i
                            elif 'date valeur' in header_lower:
                                col_indices['date_valeur'] = i
                            elif 'ref banque' in header_lower:
                                col_indices['ref_banque'] = i

                    self._log(f"Indices des colonnes: {col_indices}")

                    # Extraire les données des lignes suivantes
                    for row_idx in range(header_row_idx + 1, len(table)):
                        row = table[row_idx]
                        if not row or all(col is None or col == '' for col in row):
                            continue

                        # Extraire les données selon les indices
                        date_op = row[col_indices['date_op']] if col_indices['date_op'] != -1 and col_indices['date_op'] < len(row) else None
                        libelle = row[col_indices['libelle']] if col_indices['libelle'] != -1 and col_indices['libelle'] < len(row) else None
                        debit_str = row[col_indices['debit']] if col_indices['debit'] != -1 and col_indices['debit'] < len(row) else None
                        credit_str = row[col_indices['credit']] if col_indices['credit'] != -1 and col_indices['credit'] < len(row) else None

                        # Nettoyer et convertir les données
                        date_formatted = self._format_date(date_op, current_year)
                        
                        # Fallback: si pas de date dans la colonne "Date opération", essayer "Date valeur"
                        if not date_formatted and col_indices['date_valeur'] != -1 and col_indices['date_valeur'] < len(row):
                            date_valeur = row[col_indices['date_valeur']]
                            date_formatted = self._format_date(date_valeur, current_year)
                            if date_formatted:
                                self._log(f"Date extraite depuis 'Date valeur': {date_formatted}")
                        
                        # Fallback: chercher une date dans le libellé
                        if not date_formatted and libelle:
                            date_match_in_libelle = re.search(r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})', str(libelle))
                            if date_match_in_libelle:
                                date_formatted = self._format_date(date_match_in_libelle.group(1), current_year)
                                if date_formatted:
                                    self._log(f"Date extraite depuis le libellé: {date_formatted}")
                        
                        libelle_cleaned = self._clean_libelle(libelle)
                        debit_amount = self._parse_amount(debit_str)
                        credit_amount = self._parse_amount(credit_str)

                        # Ajouter seulement si on a des données valides et un libellé non vide
                        if (date_formatted or libelle_cleaned or debit_amount is not None or credit_amount is not None) and libelle_cleaned and libelle_cleaned.strip():
                            all_data.append({
                                'Date': date_formatted,
                                'Libellé': libelle_cleaned,
                                'Débit': debit_amount,
                                'Crédit': credit_amount
                            })
                
                # Si aucune table n'a été trouvée, essayer l'extraction par texte
                if not all_data:
                    self._log(f"Aucune table trouvée sur la page {page_num + 1}, tentative d'extraction par texte")
                    self._extract_from_text(text, current_year, all_data)
                
                # Si toujours rien, essayer une extraction plus agressive
                if not all_data:
                    self._log(f"Tentative d'extraction agressive sur la page {page_num + 1}")
                    self._extract_aggressive(text, current_year, all_data)
                
                # Dernière tentative avec extraction par regex avancée
                if not all_data:
                    self._log(f"Tentative d'extraction par regex avancée sur la page {page_num + 1}")
                    self._extract_with_regex(text, current_year, all_data)

        self.df = pd.DataFrame(all_data)
        
        # Supprimer les lignes avec des libellés vides
        if not self.df.empty:
            # Supprimer les lignes où le libellé est vide, None ou ne contient que des espaces
            initial_count = len(self.df)
            self.df = self.df.dropna(subset=['Libellé'])  # Supprimer les lignes avec libellé None
            self.df = self.df[self.df['Libellé'].str.strip() != '']  # Supprimer les libellés vides
            self.df = self.df.reset_index(drop=True)  # Réinitialiser les index
            
            removed_count = initial_count - len(self.df)
            if removed_count > 0:
                self._log(f"Suppression de {removed_count} lignes avec libellés vides")
        
        self._log(f"Extraction terminée: {len(self.df)} transactions trouvées")
        return self.df

    def _parse_amount(self, amount_str):
        """
        Parse un montant en gérant les virgules comme séparateur décimal
        """
        if not amount_str or amount_str == '':
            return None
        
        amount_str = str(amount_str).strip()
        if not amount_str:
            return None

        # Nettoyer les espaces et caractères indésirables
        amount_str = re.sub(r'[^\d.,\-]', '', amount_str)
        
        # Gérer les cas spéciaux
        if amount_str == '-' or amount_str == '':
            return None

        # Remplacer les points par des virgules pour les milliers, puis virgule par point pour décimal
        # Format UBCI: 1.234,567 -> 1234.567
        if ',' in amount_str and '.' in amount_str:
            # Format avec milliers et décimales
            amount_str = amount_str.replace('.', '').replace(',', '.')
        elif ',' in amount_str:
            # Format avec virgule comme décimal
            amount_str = amount_str.replace(',', '.')
        
        try:
            return float(amount_str)
        except ValueError:
            # Essayer de nettoyer davantage
            amount_str = re.sub(r'[^\d.]', '', amount_str)
            try:
                return float(amount_str)
            except ValueError:
                return None

    def _format_date(self, date_str, current_year):
        """
        Formate une chaîne de date en DD/MM/YYYY
        """
        if not date_str or date_str == '':
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None

        # Nettoyer la date des caractères indésirables
        date_str = re.sub(r'[^\d/\-\.]', '', date_str)
        
        # Essayer plusieurs formats de date
        date_formats = [
            '%d/%m/%Y',  # DD/MM/YYYY
            '%d/%m/%y',  # DD/MM/YY
            '%d-%m-%Y',  # DD-MM-YYYY
            '%d-%m-%y',  # DD-MM-YY
            '%d.%m.%Y',  # DD.MM.YYYY
            '%d.%m.%y',  # DD.MM.YY
        ]
        
        for fmt in date_formats:
            try:
                dt_obj = datetime.strptime(date_str, fmt)
                # Si l'année est à 2 chiffres, ajuster
                if '%y' in fmt:
                    if dt_obj.year > datetime.now().year:
                        dt_obj = dt_obj.replace(year=dt_obj.year - 100)
                return dt_obj.strftime('%d/%m/%Y')
            except ValueError:
                continue
        
        # Si aucun format ne fonctionne, essayer d'ajouter l'année courante
        if len(date_str) <= 5 and ('/' in date_str or '-' in date_str or '.' in date_str):
            try:
                # Essayer DD/MM
                dt_obj = datetime.strptime(f"{date_str}/{current_year}", '%d/%m/%Y')
                return dt_obj.strftime('%d/%m/%Y')
            except ValueError:
                pass
            try:
                # Essayer DD-MM
                dt_obj = datetime.strptime(f"{date_str}/{current_year}", '%d-%m-%Y')
                return dt_obj.strftime('%d/%m/%Y')
            except ValueError:
                pass
            try:
                # Essayer DD.MM
                dt_obj = datetime.strptime(f"{date_str}/{current_year}", '%d.%m.%Y')
                return dt_obj.strftime('%d/%m/%Y')
            except ValueError:
                pass

        self._log(f"Impossible de formater la date: {date_str}")
        return None

    def _clean_libelle(self, libelle):
        """
        Nettoie le libellé
        """
        if not libelle:
            return None
        
        libelle = str(libelle).strip()
        if not libelle:
            return None
        
        # Remplacer les retours à la ligne par des espaces
        libelle = libelle.replace('\n', ' ').replace('\r', ' ')
        # Nettoyer les espaces multiples
        libelle = re.sub(r'\s+', ' ', libelle)
        
        return libelle

    def _extract_from_text(self, text, current_year, all_data):
        """
        Extraction par regex si l'extraction de table échoue
        """
        # Patterns pour détecter les transactions dans le texte
        date_pattern = r'(\d{1,2}/\d{1,2}/\d{2,4})'
        amount_pattern = r'(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?)'
        
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Chercher une date dans la ligne
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
                
            # Chercher des montants dans la ligne
            amounts = re.findall(amount_pattern, line)
            if len(amounts) >= 1:
                # Extraire le libellé (tout sauf la date et les montants)
                libelle = line
                for amount in amounts:
                    libelle = libelle.replace(amount, '')
                libelle = re.sub(date_pattern, '', libelle).strip()
                libelle = re.sub(r'\s+', ' ', libelle)
                
                # Déterminer si c'est un débit ou crédit
                debit_amount = None
                credit_amount = None
                
                if len(amounts) == 1:
                    # Un seul montant - déterminer le type par contexte
                    if any(word in line.lower() for word in ['débit', 'debit', 'retrait', 'virement sortant']):
                        debit_amount = self._parse_amount(amounts[0])
                    else:
                        credit_amount = self._parse_amount(amounts[0])
                elif len(amounts) == 2:
                    # Deux montants - premier débit, second crédit
                    debit_amount = self._parse_amount(amounts[0])
                    credit_amount = self._parse_amount(amounts[1])
                
                if (debit_amount is not None or credit_amount is not None) and libelle and libelle.strip():
                    all_data.append({
                        'Date': self._format_date(date_match.group(1), current_year),
                        'Libellé': libelle,
                        'Débit': debit_amount,
                        'Crédit': credit_amount
                    })

    def _extract_aggressive(self, text, current_year, all_data):
        """
        Extraction très agressive pour capturer tout ce qui ressemble à une transaction
        """
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if not line or len(line) < 5:
                continue
                
            # Chercher des patterns de transaction plus larges
            # Pattern: date + texte + montant
            date_pattern = r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})'
            amount_pattern = r'(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?)'
            
            # Chercher une date dans la ligne
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
                
            # Chercher tous les montants dans la ligne
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
                
            # Extraire le libellé (tout sauf la date et les montants)
            libelle = line
            for amount in amounts:
                libelle = libelle.replace(amount, '')
            libelle = re.sub(date_pattern, '', libelle).strip()
            libelle = re.sub(r'\s+', ' ', libelle)
            
            # Nettoyer le libellé
            if not libelle or len(libelle) < 3:
                libelle = "Transaction"
            
            # Déterminer le type de transaction
            debit_amount = None
            credit_amount = None
            
            if len(amounts) == 1:
                # Un seul montant - essayer de déterminer le type
                amount = self._parse_amount(amounts[0])
                if amount is not None:
                    # Par défaut, considérer comme crédit si pas d'indication
                    credit_amount = amount
            elif len(amounts) == 2:
                # Deux montants
                debit_amount = self._parse_amount(amounts[0])
                credit_amount = self._parse_amount(amounts[1])
            else:
                # Plusieurs montants - prendre le plus grand
                max_amount = 0
                for amount_str in amounts:
                    amount = self._parse_amount(amount_str)
                    if amount and amount > max_amount:
                        max_amount = amount
                if max_amount > 0:
                    credit_amount = max_amount
            
            # Ajouter la transaction si on a des données valides et un libellé non vide
            if (debit_amount is not None or credit_amount is not None) and libelle and libelle.strip():
                all_data.append({
                    'Date': self._format_date(date_match.group(1), current_year),
                    'Libellé': libelle,
                    'Débit': debit_amount,
                    'Crédit': credit_amount
                })

    def _extract_with_regex(self, text, current_year, all_data):
        """
        Extraction par regex avancée pour capturer toutes les transactions
        """
        # Pattern pour détecter les lignes de transaction
        # Format: Date + Libellé + Montant(s)
        transaction_pattern = r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})\s+(.+?)\s+(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?)(?:\s+(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?))?'
        
        matches = re.finditer(transaction_pattern, text, re.MULTILINE | re.DOTALL)
        
        for match in matches:
            date_str = match.group(1)
            libelle = match.group(2).strip()
            amount1 = match.group(3)
            amount2 = match.group(4) if match.group(4) else None
            
            # Nettoyer le libellé
            libelle = re.sub(r'\s+', ' ', libelle)
            if not libelle or len(libelle) < 3:
                libelle = "Transaction"
            
            # Parser les montants
            debit_amount = None
            credit_amount = None
            
            if amount2:
                # Deux montants - premier débit, second crédit
                debit_amount = self._parse_amount(amount1)
                credit_amount = self._parse_amount(amount2)
            else:
                # Un seul montant - déterminer le type par le contexte
                amount = self._parse_amount(amount1)
                if amount is not None:
                    # Analyser le libellé pour déterminer le type
                    libelle_lower = libelle.lower()
                    if any(word in libelle_lower for word in ['débit', 'debit', 'retrait', 'virement sortant', 'prélèvement']):
                        debit_amount = amount
                    else:
                        credit_amount = amount
            
            # Ajouter la transaction si on a des données valides et un libellé non vide
            if (debit_amount is not None or credit_amount is not None) and libelle and libelle.strip():
                all_data.append({
                    'Date': self._format_date(date_str, current_year),
                    'Libellé': libelle,
                    'Débit': debit_amount,
                    'Crédit': credit_amount
                })

    def _format_amount_for_excel(self, amount):
        """
        Formate un montant numérique en chaîne avec espace comme séparateur de milliers
        et virgule comme séparateur décimal avec 3 décimales (ex: "3 333,000")
        """
        if amount is None or pd.isna(amount):
            return ''
        
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return str(amount)
        
        # Toujours formater avec 3 décimales
        amount_str = f"{amount:.3f}"
        integer_part, decimal_part = amount_str.split('.')
        
        # Formater la partie entière avec espaces comme séparateur de milliers
        formatted_parts = []
        for i in range(len(integer_part) - 1, -1, -3):
            start = max(0, i - 2)
            formatted_parts.insert(0, integer_part[start:i+1])
        
        return f"{' '.join(formatted_parts)},{decimal_part}"
    
    def to_excel(self, output_path=None):
        """
        Convertit le DataFrame en fichier Excel avec style UBCI
        """
        if self.df is None or self.df.empty:
            self._log("Aucune donnée à exporter en Excel.")
            return None

        if output_path is None:
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"UBCI_Extrait_{timestamp}.xlsx"
            output_path = os.path.join(downloads_path, output_filename)

        self._log(f"Exportation vers Excel: {output_path}")

        try:
            # Créer une copie du DataFrame pour le formatage
            df_formatted = self.df.copy()
            
            # Formater les colonnes Débit et Crédit
            if 'Débit' in df_formatted.columns:
                df_formatted['Débit'] = df_formatted['Débit'].apply(self._format_amount_for_excel)
            if 'Crédit' in df_formatted.columns:
                df_formatted['Crédit'] = df_formatted['Crédit'].apply(self._format_amount_for_excel)
            
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            df_formatted.to_excel(writer, index=False, sheet_name='Transactions')
            workbook = writer.book
            sheet = writer.sheets['Transactions']

            # Styles UBCI (orange)
            header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Appliquer les styles aux en-têtes
            for col_num, column_title in enumerate(df_formatted.columns, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = alignment
                # Ajuster la largeur des colonnes
                if col_num == 1:  # Date
                    sheet.column_dimensions[chr(64 + col_num)].width = 12
                elif col_num == 2:  # Libellé
                    sheet.column_dimensions[chr(64 + col_num)].width = 50
                else:  # Débit/Crédit
                    sheet.column_dimensions[chr(64 + col_num)].width = 15

            # Appliquer les bordures aux cellules de données
            for row_num in range(2, sheet.max_row + 1):
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.border = border

            writer.close()
            self._log("Fichier Excel créé avec succès.")
            return output_path
        except Exception as e:
            self._log(f"Erreur lors de la création du fichier Excel: {e}")
            return None

    def convert(self):
        """
        Méthode principale de conversion
        """
        if not os.path.exists(self.pdf_path):
            self._log(f"Le fichier PDF n'existe pas: {self.pdf_path}")
            return False

        try:
            # Détection
            with pdfplumber.open(self.pdf_path) as pdf:
                first_page_text = ""
                if pdf.pages:
                    first_page_text = pdf.pages[0].extract_text()
                    self._log(f"Texte de la première page extrait: {len(first_page_text)} caractères")
                
                if not self.detect_ubci_extrait(first_page_text):
                    self._log("Document non reconnu comme un extrait UBCI. Tentative d'extraction quand même.")
                else:
                    self._log("Document reconnu comme extrait UBCI")
            
            # Extraction
            self.extract_transactions()
            
            if self.df is None or self.df.empty:
                self._log("Aucune transaction extraite par les méthodes standard.")
                # Essayer l'extraction par texte sur tout le document
                self._log("Tentative d'extraction par texte sur tout le document...")
                all_data = []
                with pdfplumber.open(self.pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        if text:
                            year_match = re.search(r'(\d{4})', text)
                            current_year = year_match.group(1) if year_match else str(datetime.now().year)
                            
                            # Essayer toutes les méthodes d'extraction
                            self._extract_from_text(text, current_year, all_data)
                            if not all_data:
                                self._extract_aggressive(text, current_year, all_data)
                            if not all_data:
                                self._extract_with_regex(text, current_year, all_data)
                
                if all_data:
                    self.df = pd.DataFrame(all_data)
                    self._log(f"Extraction par texte réussie: {len(self.df)} transactions trouvées")
                else:
                    self._log("Aucune transaction trouvée même avec toutes les méthodes d'extraction")
                    return False
            
            # Conversion Excel
            output_path = self.to_excel(self.excel_path)
            return output_path is not None
            
        except Exception as e:
            self._log(f"Erreur lors de la conversion: {e}")
            import traceback
            self._log(f"Traceback: {traceback.format_exc()}")
            return False

class UBCIExtraitConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur UBCI Extrait")
        self.root.geometry("600x400")
        self.root.configure(bg='#f5f5f5')

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"UBCI_Extrait_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        self.status_message = tk.StringVar()

        self._create_widgets()

    def _create_widgets(self):
        # Frame pour le titre
        title_frame = tk.Frame(self.root, bg='#FF6B35', pady=10)
        title_frame.pack(fill='x')
        title_label = tk.Label(title_frame, text="Convertisseur UBCI Extrait de Compte", bg='#FF6B35', fg='white', font=('Arial', 16, 'bold'))
        title_label.pack()

        # Frame pour la sélection du fichier PDF
        pdf_frame = tk.LabelFrame(self.root, text="Fichier PDF", bg='#f5f5f5', padx=10, pady=10)
        pdf_frame.pack(pady=10, padx=20, fill='x')

        tk.Entry(pdf_frame, textvariable=self.pdf_path, width=50, state='readonly').pack(side='left', padx=5, fill='x', expand=True)
        tk.Button(pdf_frame, text="Parcourir", command=self._browse_pdf, bg='#FF6B35', fg='white').pack(side='left', padx=5)

        # Frame pour le nom du fichier Excel
        excel_frame = tk.LabelFrame(self.root, text="Nom du fichier Excel de sortie", bg='#f5f5f5', padx=10, pady=10)
        excel_frame.pack(pady=5, padx=20, fill='x')

        tk.Entry(excel_frame, textvariable=self.excel_name, width=50).pack(side='left', padx=5, fill='x', expand=True)
        tk.Label(excel_frame, text=".xlsx", bg='#f5f5f5').pack(side='left')

        # Bouton de conversion
        convert_button = tk.Button(self.root, text="Démarrer la conversion", command=self.convertir, bg='#FF6B35', fg='white', font=('Arial', 12, 'bold'), height=2)
        convert_button.pack(pady=20, padx=20, fill='x')

        # Message de statut
        status_label = tk.Label(self.root, textvariable=self.status_message, bg='#f5f5f5', fg='blue', font=('Arial', 10))
        status_label.pack(pady=5)

    def _browse_pdf(self):
        file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if file_path:
            self.pdf_path.set(file_path)
            # Proposer un nom d'Excel basé sur le nom du PDF
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            self.excel_name.set(f"{base_name}_UBCI_Extrait_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

    def convertir(self):
        pdf_file = self.pdf_path.get()
        if not pdf_file:
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier PDF.")
            return

        excel_output_name = self.excel_name.get()
        if not excel_output_name:
            messagebox.showwarning("Attention", "Veuillez spécifier un nom pour le fichier Excel.")
            return

        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        excel_full_path = os.path.join(downloads_path, f"{excel_output_name}.xlsx")

        self.status_message.set("Conversion en cours...")
        self.root.update_idletasks()

        # Créer un convertisseur en mode debug pour avoir plus d'informations
        converter = UBCIExtraitConverter(pdf_file, excel_full_path, silent=False)
        success = converter.convert()

        if success:
            self.status_message.set(f"Conversion réussie ! Fichier enregistré dans {excel_full_path}")
            messagebox.showinfo("Succès", f"Le fichier Excel a été créé avec succès :\n{excel_full_path}")
        else:
            self.status_message.set("Échec de la conversion. Vérifiez que le PDF est un extrait UBCI valide.")
            messagebox.showerror("Erreur", "Échec de la conversion. Vérifiez que le PDF est un extrait UBCI valide ou consultez la console pour plus de détails.")

def main():
    if len(sys.argv) > 1:
        # Mode ligne de commande
        parser = argparse.ArgumentParser(description="Convertisseur UBCI Extrait PDF vers Excel")
        parser.add_argument('pdf_path', nargs='?', help='Chemin du fichier PDF UBCI')
        parser.add_argument('--output', '-o', help='Chemin du fichier Excel de sortie')
        parser.add_argument('--silent', action='store_true', help='Mode silencieux')
        
        args = parser.parse_args()
        
        if args.pdf_path:
            converter = UBCIExtraitConverter(args.pdf_path, args.output, args.silent)
            success = converter.convert()
            
            if success:
                print("Conversion réussie!")
                sys.exit(0)
            else:
                print("Échec de la conversion")
                sys.exit(1)
        else:
            # Si aucun argument n'est fourni, lancer l'interface graphique
            root = tk.Tk()
            app = UBCIExtraitConverterGUI(root)
            root.mainloop()
    else:
        # Lancer l'interface graphique par défaut
        root = tk.Tk()
        app = UBCIExtraitConverterGUI(root)
        root.mainloop()

if __name__ == "__main__":
    main()
