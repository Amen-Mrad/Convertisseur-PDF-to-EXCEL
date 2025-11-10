import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
from PIL import Image, ImageTk
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime as dt
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class STBReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVE STB Bank PDF vers Excel")
        self.root.geometry("600x500")

        # Variables
        self.pdf_file_path = tk.StringVar()
        self.excel_filename = tk.StringVar()
        self.excel_filename.set("releve_stb_" + datetime.now().strftime("%d%m%Y_%H%M"))
        
        self.setup_ui()
        
    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(self.root, text="Convertisseur RELEVE STB BANK",
                              font=("Arial", 16, "bold"), fg="#1E3A8A")  # Couleur bleue STB
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(self.root, text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame(self.root)
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section sélection fichier PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVE STB Bank:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_file_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.select_pdf_file, 
                              font=("Segoe UI", 10, "bold"), bg="#1E3A8A", fg="white")
        browse_btn.pack(pady=5)
        
        # Section nom fichier Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_filename, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section conversion
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        # Boutons
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                               command=self.convert_pdf_to_excel, 
                               font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Zone de résultats
        self.result_text = tk.Text(main_frame, height=8, width=70, font=("Arial", 8))
        self.result_text.pack(fill='both', expand=True, pady=10)
        
    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF RELEVE STB Bank",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_file_path.set(file_path)
            
    def is_stb_releve_pdf(self, pdf_path):
        """Vérifie si le PDF est un RELEVE STB Bank"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Vérifier les premières pages pour les mots-clés STB
                for page_num in range(min(3, len(pdf.pages))):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        text_lower = text.lower()
                        # Mots-clés spécifiques pour STB
                        stb_keywords = [
                            'stb bank',
                            'société tunisienne de banque',
                            'الشركة التونسية للبنك',
                            'releve de compte',
                            'كشف حساب',
                            'duplicata',
                            'ancien solde au',
                            'solde à nouveau',
                            'الرصيد الجديد',
                            'releve',  # Ajout pour "RELEVEE STB"
                            'stb'      # Ajout pour "STB"
                        ]
                        found_keywords = sum(1 for keyword in stb_keywords if keyword in text_lower)
                        if found_keywords >= 2:  # Réduit à 2 mots-clés STB pour plus de flexibilité
                            print(f"DEBUG STB - PDF détecté comme STB avec {found_keywords} mots-clés")
                            return True
                
                # Fallback: vérifier le nom du fichier
                filename = os.path.basename(pdf_path).lower()
                if 'stb' in filename or 'releve' in filename:
                    print(f"DEBUG STB - PDF détecté comme STB depuis le nom de fichier: {filename}")
                    return True
                    
                return False
        except Exception as e:
            print(f"Erreur lors de la vérification du PDF STB: {e}")
            # En cas d'erreur, considérer comme STB si le nom contient "stb"
            filename = os.path.basename(pdf_path).lower()
            if 'stb' in filename:
                print(f"DEBUG STB - PDF considéré comme STB par défaut (nom: {filename})")
                return True
            return False
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF STB"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num in range(min(2, len(pdf.pages))):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        # Chercher des patterns d'année dans le texte
                        year_patterns = [
                            r'arrêté au\s*:\s*(\d{1,2}/\d{1,2}/(\d{4}))',  # "arrêté au : 31/07/2025"
                            r'(\d{4})',  # Toute année 4 chiffres
                        ]
                        
                        for pattern in year_patterns:
                            matches = re.findall(pattern, text, re.IGNORECASE)
                            if matches:
                                # Prendre la première année trouvée
                                year = matches[0][1] if isinstance(matches[0], tuple) else matches[0]
                                if 2020 <= int(year) <= 2030:  # Année raisonnable
                                    print(f"Année STB détectée: {year}")
                                    return year
                
                # Fallback: année actuelle
                current_year = datetime.now().year
                print(f"Année par défaut: {current_year}")
                return str(current_year)
                
        except Exception as e:
            print(f"Erreur détection année STB: {e}")
            return str(datetime.now().year)
    
    def extract_table_data(self, pdf_path):
        """Extrait les données RELEVE STB"""
        try:
            print(f"DEBUG STB - Début extraction pour: {os.path.basename(pdf_path)}")
            
            # D'abord, détecter l'année depuis le PDF
            year = self.detect_year_from_pdf(pdf_path)
            print(f"DEBUG STB - Année détectée: {year}")
            
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG STB - Traitement page {page_num + 1}/{len(pdf.pages)}")
                    
                    # Vérifier si le texte est vide (PDF scanné)
                    text = page.extract_text() or ''
                    if not text.strip() and page_num == 0:
                        print(f"DEBUG STB - PDF scanné détecté, tentative OCR...")
                        # Fallback OCR pour les PDFs scannés
                        ocr_text = self._extract_text_via_ocr(pdf_path)
                        if ocr_text.strip():
                            print(f"DEBUG STB - OCR réussi, traitement du texte...")
                            # Traiter le texte OCR comme un seul bloc
                            ocr_transactions = self._parse_ocr_text_stb(ocr_text, year)
                            if ocr_transactions:
                                print(f"DEBUG STB - {len(ocr_transactions)} transactions extraites via OCR")
                                return ocr_transactions
                    
                    # Extraire les tableaux de la page
                    tables = page.extract_tables()
                    print(f"DEBUG STB - {len(tables)} tableaux trouvés sur la page {page_num + 1}")
                    
                    for table_idx, table in enumerate(tables):
                        if table and len(table) > 1:  # Au moins un en-tête et des données
                            print(f"DEBUG STB - Tableau {table_idx + 1}: {len(table)} lignes")
                            
                            # Chercher l'en-tête du tableau
                            header_row = None
                            for i, row in enumerate(table):
                                if row and any(col for col in row if col):
                                    row_text = ' '.join([str(col) for col in row if col]).lower()
                                    print(f"DEBUG STB - Ligne {i}: {row_text}")
                                    # Mots-clés STB
                                    if any(keyword in row_text for keyword in ['date', 'libellé', 'libelle', 'débit', 'debit', 'crédit', 'credit', 'valeur']):
                                        header_row = i
                                        print(f"DEBUG STB - En-tête trouvé ligne {i}")
                                        break
                            
                            if header_row is not None:
                                print(f"DEBUG STB - Extraction à partir ligne {header_row + 1}")
                                # Parser les lignes de données
                                for row_idx, row in enumerate(table[header_row + 1:], header_row + 1):
                                    if row and len(row) >= 4:
                                        print(f"DEBUG STB - Ligne {row_idx}: {row}")
                                        transaction = self.parse_stb_transaction_row(row, year)
                                        if transaction:
                                            all_transactions.append(transaction)
                                            print(f"DEBUG STB - Transaction ajoutée: {transaction}")
                                        else:
                                            print(f"DEBUG STB - Transaction ignorée")
                            else:
                                print("DEBUG STB - Aucun en-tête trouvé dans le tableau")

                    # CORRECTION: Toujours parser le texte brut de la page pour une extraction complète
                    page_text = page.extract_text() or ""
                    if page_text:
                        print(f"DEBUG STB - Parsing du texte brut de la page {page_num + 1}...")
                        text_tx = self.parse_stb_text_block(page_text, year)
                        if text_tx:
                            print(f"DEBUG STB - {len(text_tx)} transactions depuis texte brut page {page_num + 1}")
                            all_transactions.extend(text_tx)
                        else:
                            print(f"DEBUG STB - Aucune transaction trouvée dans le texte brut page {page_num + 1}")
                
                print(f"DEBUG STB - Total transactions extraites: {len(all_transactions)}")
                return all_transactions
                
        except Exception as e:
            print(f"Erreur lors de l'extraction STB: {e}")
            import traceback
            traceback.print_exc()
            return []

    def parse_stb_text_block(self, text, year):
        """Parsing robuste du texte brut d'une page RELEVE STB"""
        transactions = []
        try:
            lines = [ln.rstrip() for ln in text.split("\n")]
            current = None
            
            # Pattern pour les montants STB (avec virgule comme séparateur décimal)
            amount_regex = re.compile(r"(?<!\d)(\d{1,3}(?:[ \u00A0]\d{3})*(?:,\d+)?)(?!\d)")
            # Pattern pour les dates STB (DD/MM/YYYY)
            date_regex = re.compile(r"^\s*(\d{1,2}/\d{1,2}/\d{4})\s*")
            
            def finalize_current():
                if not current:
                    return
                libelle = current.get('Libellé', '').strip()
                if not libelle or any(k in libelle.upper() for k in ['ANCIEN SOLDE AU', 'SOLDE À NOUVEAU', 'الرصيد الجديد', 'TOTAUX', 'SOLDE']):
                    return
                
                # CORRECTION: Vérifier si le libellé est une ligne de footer
                if self.is_footer_line(libelle):
                    print(f"DEBUG STB - Transaction de footer ignorée: {libelle}")
                    return
                
                # Vérifier que ce n'est pas un solde
                debit = current.get('Débit', '')
                credit = current.get('Crédit', '')
                
                if debit and self.is_amount(debit):
                    try:
                        debit_val = float(debit.replace(',', '.').replace(' ', ''))
                        if debit_val < 0 or debit_val > 1000000:  # Ignorer les soldes
                            return
                    except:
                        pass
                
                if credit and self.is_amount(credit):
                    try:
                        credit_val = float(credit.replace(',', '.').replace(' ', ''))
                        if credit_val < 0 or credit_val > 1000000:  # Ignorer les soldes
                            return
                    except:
                        pass
                
                # Choisir la date à utiliser
                date_to_use = current.get('DateOperation') or current.get('DateValeur')
                if not date_to_use:
                    return
                
                # CORRECTION: Nettoyer le libellé
                libelle_cleaned = self.clean_libelle(libelle)
                
                # Montants
                debit = self.format_amount(current.get('Débit', ''))
                credit = self.format_amount(current.get('Crédit', ''))
                transactions.append({
                    'Date': date_to_use,
                    'Libellé': libelle_cleaned,
                    'Débit': debit,
                    'Crédit': credit
                })
            
            def decide_debit_credit_by_keywords(libelle_upper):
                # CORRECTION: Mots-clés pour les CRÉDITS (entrées d'argent) - STB
                # Plus spécifiques pour éviter les faux positifs
                credit_kw = [
                    'VIREMENT RECU', 'VERSEMENT', 'ENCAISSEMENT', 'REMBOURSEMENT',
                    'CRÉDIT', 'CREDIT', 'CREDIT COMPTE', 'AVANCE'
                ]
                
                # CORRECTION: Mots-clés pour les DÉBITS (sorties d'argent) - STB
                # Plus complets et avec priorité sur les commissions
                debit_kw = [
                    'COMMISSION', 'FRAIS', 'AGIOS', 'PRÉLÈVEMENT', 'PRELEVEMENT',
                    'RETRAIT', 'CARTE', 'CHEQUE', 'VIREMENT EMIS', 'DÉBIT', 'DEBIT',
                    'EFFET', 'REJET', 'COTISATION', 'TVA', 'INTERET', 'PENALITE',
                    'RETARD', 'DEPASSEMENT', 'ADHESION', 'ACCEPT', 'REALISATION'
                ]
                
                # CORRECTION: Priorité aux débits pour éviter les erreurs de classification
                # Vérifier d'abord les débits (priorité)
                if any(k in libelle_upper for k in debit_kw):
                    return 'debit'
                # Puis les crédits
                if any(k in libelle_upper for k in credit_kw):
                    return 'credit'
                return ''
            
            for raw_line in lines:
                line = raw_line.strip()
                if not line:
                    continue
                uline = line.upper()
                
                # CORRECTION: Ignorer les lignes de footer (messages, soldes, contact)
                if self.is_footer_line(line):
                    print(f"DEBUG STB - Ligne de footer ignorée: {line}")
                    finalize_current()
                    current = None
                    continue
                
                # Ignorer les lignes de totaux/solde
                if any(kw in uline for kw in ['ANCIEN SOLDE AU', 'SOLDE À NOUVEAU', 'الرصيد الجديد', 'TOTAUX', 'SOLDE']) and not date_regex.match(line):
                    finalize_current()
                    current = None
                    continue
                
                m = date_regex.match(line)
                if m:
                    # Finaliser la transaction en cours
                    finalize_current()
                    current = {
                        'DateOperation': m.group(1),
                        'Libellé': ''
                    }
                    
                    # Retirer la date du début
                    rest = line[m.end():].strip()
                    
                    # Extraire les montants de la ligne
                    amounts = amount_regex.findall(rest)
                    
                    # Filtrer les montants plausibles
                    def plausible(a):
                        a_clean = a.replace(' ', '')
                        if (',' in a) or (' ' in a):
                            return True
                        try:
                            val = int(a_clean)
                            return val > 31
                        except Exception:
                            return False
                    amounts = [a for a in amounts if plausible(a)]
                    
                    # CORRECTION: Classification améliorée débit/crédit
                    if len(amounts) >= 2:
                        # Deux montants ou plus: classifier selon la nouvelle logique
                        # Utiliser la classification améliorée pour le premier montant
                        classification = self.classify_stb_transaction(rest, amounts[0])
                        if classification == 'debit':
                            current['Débit'] = amounts[0]
                            current['Crédit'] = amounts[1] if len(amounts) > 1 else ''
                        else:
                            current['Débit'] = amounts[1] if len(amounts) > 1 else ''
                            current['Crédit'] = amounts[0]
                    elif len(amounts) == 1:
                        # Un seul montant: classifier avec la nouvelle logique
                        amt = amounts[0]
                        classification = self.classify_stb_transaction(rest, amt)
                        if classification == 'debit':
                            current['Débit'] = amt
                            current['Crédit'] = ''
                        else:
                            current['Débit'] = ''
                            current['Crédit'] = amt
                    else:
                        current['Débit'] = ''
                        current['Crédit'] = ''
                    
                    # CORRECTION: Libellé = texte sans les montants finaux mais en préservant les numéros de référence
                    if amounts:
                        # Supprimer seulement les montants (pas les numéros de référence)
                        for amount in amounts:
                            # Vérifier que c'est bien un montant et pas un numéro de référence
                            if self.is_transaction_amount(amount):
                                rest = rest.replace(amount, '').strip()
                        label_text = rest.rstrip('-:')
                    else:
                        label_text = rest
                    
                    # CORRECTION: Nettoyer le libellé immédiatement
                    current['Libellé'] = self.clean_libelle(label_text.strip())
                else:
                    # Ligne de continuation: ajouter au libellé courant
                    if current:
                        if not any(kw in uline for kw in ['TOTAUX', 'SOLDE']):
                            sep = ' ' if current['Libellé'] else ''
                            current['Libellé'] = f"{current['Libellé']}{sep}{line}"
            
            # Finaliser la dernière
            finalize_current()
            
        except Exception as e:
            print(f"Erreur parse_stb_text_block: {e}")
        return transactions

    def parse_stb_transaction_row(self, row, year):
        """Parse une ligne de transaction STB"""
        try:
            # Nettoyer la ligne
            cleaned_row = [str(cell).strip() if cell else "" for cell in row]
            print(f"DEBUG STB PARSING - Ligne nettoyée: {cleaned_row}")
            
            # Vérifier si c'est une ligne d'en-tête ou de total
            row_text = ' '.join(cleaned_row).lower()
            if any(keyword in row_text for keyword in ['date', 'libellé', 'débit', 'crédit', 'solde', 'totaux', 'ancien solde']):
                print(f"DEBUG STB PARSING - Ligne d'en-tête ignorée: {row_text}")
                return None
            
            # CORRECTION: Vérifier si c'est une ligne de footer
            full_row_text = ' '.join(cleaned_row)
            if self.is_footer_line(full_row_text):
                print(f"DEBUG STB PARSING - Ligne de footer ignorée: {full_row_text}")
                return None
            
            # Structure STB: DATE, LIBELLE, VALEUR, DEBIT, CREDIT
            if len(cleaned_row) >= 5:
                date_operation = cleaned_row[0]  # DATE
                libelle = cleaned_row[1]         # LIBELLE
                date_valeur = cleaned_row[2]      # VALEUR
                debit = cleaned_row[3]           # DEBIT
                credit = cleaned_row[4]          # CREDIT
                
                print(f"DEBUG STB PARSING - Date op: '{date_operation}', Date val: '{date_valeur}', Libellé: '{libelle}'")
                print(f"DEBUG STB PARSING - Débit: '{debit}', Crédit: '{credit}'")
                
                # Utiliser DATE OPERATION si disponible
                date_to_use = ""
                if date_operation and self.is_date_stb(date_operation):
                    date_to_use = date_operation
                    print(f"DEBUG STB PARSING - Utilisation DATE OPERATION: {date_to_use}")
                elif date_valeur and self.is_date_stb(date_valeur):
                    date_to_use = date_valeur
                    print(f"DEBUG STB PARSING - Utilisation DATE VALEUR: {date_to_use}")
                
                # Vérifier que nous avons au moins une date et un libellé valide
                if date_to_use and libelle and len(libelle.strip()) > 2:
                    # CORRECTION: Re-classifier si nécessaire
                    final_debit = self.format_amount(debit)
                    final_credit = self.format_amount(credit)
                    
                    # Si les deux colonnes ont des montants, re-classifier
                    if final_debit and final_credit:
                        # Utiliser la classification améliorée
                        classification = self.classify_stb_transaction(libelle, final_debit)
                        if classification == 'debit':
                            final_credit = ''  # Garder seulement le débit
                        else:
                            final_debit = ''   # Garder seulement le crédit
                    
                    # Si une seule colonne a un montant, vérifier la classification
                    elif final_debit and not final_credit:
                        classification = self.classify_stb_transaction(libelle, final_debit)
                        if classification == 'credit':
                            final_credit = final_debit
                            final_debit = ''
                    elif final_credit and not final_debit:
                        classification = self.classify_stb_transaction(libelle, final_credit)
                        if classification == 'debit':
                            final_debit = final_credit
                            final_credit = ''
                    
                    # CORRECTION: Nettoyer le libellé
                    libelle_cleaned = self.clean_libelle(libelle.strip())
                    
                    result = {
                        'Date': self.format_date_stb(date_to_use),
                        'Libellé': libelle_cleaned,
                        'Débit': final_debit,
                        'Crédit': final_credit
                    }
                    print(f"DEBUG STB PARSING - Transaction créée: {result}")
                    return result
                else:
                    print(f"DEBUG STB PARSING - Transaction ignorée (date: '{date_to_use}', libellé: '{libelle}')")
            else:
                print(f"DEBUG STB PARSING - Pas assez de colonnes: {len(cleaned_row)}")
            
            return None
            
        except Exception as e:
            print(f"Erreur lors du parsing de la ligne STB: {e}")
            return None

    def is_transaction_amount(self, text):
        """CORRECTION: Détecte si un texte est un montant de transaction (pas un numéro de référence)"""
        if not text:
            return False
        
        # Nettoyer le texte
        clean_text = text.replace(' ', '').replace(',', '.')
        
        # Vérifier si c'est un nombre
        try:
            value = float(clean_text)
            
            # CORRECTION: Un montant de transaction est généralement > 0.01 et < 100000000
            # Les numéros de référence sont généralement des entiers longs
            if 0.01 <= value <= 100000000:
                # Vérifier si c'est un montant avec décimales (probablement un montant)
                if ',' in text or '.' in text:
                    return True
                # Vérifier si c'est un nombre court (probablement un montant)
                if value < 1000000:
                    return True
                # Sinon, c'est probablement un numéro de référence
                return False
            else:
                return False
        except:
            return False

    def clean_libelle(self, libelle):
        """CORRECTION: Nettoie le libellé en préservant les numéros de transaction importants"""
        if not libelle:
            return ""
        
        # CORRECTION: Supprimer seulement les dates au format DD/MM/YYYY (dates valeur)
        libelle = re.sub(r'\b\d{1,2}/\d{1,2}/\d{4}\b', '', libelle)
        
        # CORRECTION: Supprimer seulement les dates au format DD/MM/YY (dates courtes)
        libelle = re.sub(r'\b\d{1,2}/\d{1,2}/\d{2}\b', '', libelle)
        
        # CORRECTION: Préserver les numéros de transaction importants
        # - EFF N XXXXXXXX (numéros d'effets)
        # - Numéros de chèque
        # - Numéros de virement
        # - Autres numéros de référence bancaire
        
        # Supprimer les espaces multiples
        libelle = re.sub(r'\s+', ' ', libelle)
        
        # Supprimer les espaces en début et fin
        libelle = libelle.strip()
        
        # Supprimer les caractères de ponctuation en fin de libellé
        libelle = libelle.rstrip('.,:;')
        
        return libelle

    def is_footer_line(self, line):
        """CORRECTION: Détecte si une ligne est une ligne de footer (à ignorer)"""
        line_upper = line.upper()
        
        # Mots-clés de footer en français et arabe
        footer_keywords = [
            # Messages d'erreur
            'الرجاء إشعار فرعكم في حالة خطأ',
            'PRIÈRE AVISER VOTRE AGENCE EN CAS D\'ERREUR',
            'PRIERE AVISER VOTRE AGENCE EN CAS D\'ERREUR',
            'AVISER VOTRE AGENCE',
            'EN CAS D\'ERREUR',
            
            # Soldes
            'SOLDE À NOUVEAU',
            'SOLDE A NOUVEAU', 
            'الرصيد الجديد',
            'الرصيد الجديدا',
            'SOLDE AU',
            'SOLDE FINAL',
            'NOUVEAU SOLDE',
            
            # Informations de contact
            'SOCIÉTÉ TUNISIENNE DE BANQUE',
            'SOCIETE TUNISIENNE DE BANQUE',
            'الشركة التونسية للبنك',
            'STB BANK',
            'S.A. CAPITAL',
            'SIÈGE SOCIAL',
            'SIEGE SOCIAL',
            'RUE HÉDI NOUIRA',
            'RUE HEDI NOUIRA',
            'IDENTIFIANT UNIQUE',
            'IDENTIFIQUE UNIQUE',
            'TÉL:',
            'TEL:',
            'SITE WEB',
            'WWW.STB.COM.TN',
            
            # Totaux
            'TOTAUX',
            'TOTAL',
            'SOMME',
            
            # Lignes avec beaucoup de chiffres (probablement des soldes)
            '60 452,557',
            '60 452,557 عرف رشعاإعلرجا الجديدا لرصيدا 60 ف م أطخة لاح'
        ]
        
        # Vérifier les mots-clés de footer
        for keyword in footer_keywords:
            if keyword in line_upper:
                return True
        
        # Vérifier les patterns de footer
        # Lignes avec beaucoup de chiffres et de texte arabe
        if re.search(r'\d{2}\s+\d{3},\d{3}', line) and re.search(r'[\u0600-\u06FF]', line):
            return True
        
        # Lignes très longues avec beaucoup de texte arabe (informations de contact)
        if len(line) > 100 and re.search(r'[\u0600-\u06FF]', line):
            return True
            
        return False

    def classify_stb_transaction(self, libelle, amount):
        """CORRECTION: Classification améliorée pour les transactions STB"""
        libelle_upper = libelle.upper()
        
        # CORRECTION: Règles spécifiques pour les cas ambigus
        # 1. Toutes les commissions sont des débits
        if 'COMMISSION' in libelle_upper:
            return 'debit'
        
        # 2. Tous les frais sont des débits
        if any(kw in libelle_upper for kw in ['FRAIS', 'AGIOS', 'TVA', 'INTERET', 'PENALITE']):
            return 'debit'
        
        # 3. Tous les prélèvements sont des débits
        if any(kw in libelle_upper for kw in ['PRÉLÈVEMENT', 'PRELEVEMENT', 'RETRAIT']):
            return 'debit'
        
        # 4. Tous les virements émis sont des débits
        if 'VIREMENT EMIS' in libelle_upper:
            return 'debit'
        
        # 5. Tous les virements reçus sont des crédits (sauf si c'est une commission)
        if 'VIREMENT RECU' in libelle_upper and 'COMMISSION' not in libelle_upper:
            return 'credit'
        
        # 6. Tous les versements sont des crédits
        if any(kw in libelle_upper for kw in ['VERSEMENT', 'ENCAISSEMENT', 'REMBOURSEMENT']):
            return 'credit'
        
        # 7. Par défaut, analyser le montant
        try:
            amount_val = float(amount.replace(',', '.').replace(' ', ''))
            # Si le montant est très petit (< 100), c'est probablement un débit (frais/commissions)
            if amount_val < 100:
                return 'debit'
        except:
            pass
        
        return 'debit'  # Par défaut, débit

    def is_date_stb(self, text):
        """Vérifie si le texte est une date au format STB DD/MM/YYYY"""
        if not text:
            return False
        # Pattern pour date STB DD/MM/YYYY
        date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'
        return bool(re.match(date_pattern, text.strip()))
    
    def is_amount(self, text):
        """Vérifie si le texte est un montant"""
        if not text:
            return False
        # Pattern pour montant (avec virgule comme séparateur décimal)
        amount_pattern = r'^[\d\s.,-]+$'
        return bool(re.match(amount_pattern, text.strip())) and any(c.isdigit() for c in text)
    
    def format_date_stb(self, date_str):
        """Formate la date STB au format DD/MM/YYYY"""
        try:
            date_str = date_str.strip()
            # Si DD/MM/YYYY, garder tel quel
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
                return date_str
            return ""
        except:
            return ""
    
    def format_amount(self, amount_str):
        """Formate le montant comme BIAT (espace pour milliers, virgule pour décimales)"""
        if not amount_str or amount_str == "":
            return ""
        
        try:
            # Nettoyer le montant (même logique que BIAT)
            clean_amount = str(amount_str).replace(' ', '').replace('.', '')
            
            # Remplacer la virgule par un point pour la conversion
            if ',' in clean_amount:
                clean_amount = clean_amount.replace(',', '.')
            
            # Convertir en float puis formater (même format que BIAT)
            amount_float = float(clean_amount)
            # 3 décimales requises
            return f"{amount_float:,.3f}".replace(',', ' ').replace('.', ',')
            
        except (ValueError, TypeError):
            return str(amount_str)
    
    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel"""
        try:
            # Créer un nouveau workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "STB"  # Nom de la feuille
            
            # Définir les styles
            header_font = Font(name='Arial', size=12, bold=True, color='000000')
            header_fill = PatternFill(start_color='BFD4FF', end_color='BFD4FF', fill_type='solid')  # Bleu ciel transparent
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Écrire les en-têtes
            headers = ['Date', 'Libellé', 'Débit', 'Crédit']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Écrire les données
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Formatage spécial pour les montants (comme BIAT)
                    if col_idx in [3, 4]:  # Débit et Crédit
                        if value and str(value).strip():
                            # Garder le format texte comme BIAT
                            cell.value = str(value)
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx == 1:  # Date
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 2:  # Libellé
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 70  # Libellé
            ws.column_dimensions['C'].width = 16  # Débit
            ws.column_dimensions['D'].width = 16  # Crédit
            
            # Ajuster la hauteur des lignes
            for row in range(1, len(df) + 2):
                ws.row_dimensions[row].height = 20
            
            # Ajouter des bordures à toute la plage de données
            from openpyxl.utils import get_column_letter
            max_row = len(df) + 1
            max_col = len(headers)
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).border = border
            
            # Sauvegarder le fichier
            wb.save(excel_path)
            
        except Exception as e:
            print(f"Erreur lors du formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, engine='openpyxl')
    
    def convert_pdf_to_excel(self):
        """Fonction principale de conversion STB"""
        if not self.pdf_file_path.get():
            print("❌ Erreur: Veuillez sélectionner un fichier PDF")
            return
        
        if not self.excel_filename.get():
            print("❌ Erreur: Veuillez entrer un nom pour le fichier Excel")
            return

        self.result_text.delete(1.0, tk.END)
        
        try:
            pdf_path = self.pdf_file_path.get()
            print(f"DEBUG STB - Début conversion: {os.path.basename(pdf_path)}")
            
            # Vérifier que c'est un PDF STB Bank
            if not self.is_stb_releve_pdf(pdf_path):
                print("⚠️ Attention: Ce fichier ne semble pas être un RELEVE STB Bank. Tentative de conversion quand même...")
            
            # Extraire les données
            print(f"DEBUG STB - Extraction des données...")
            transactions = self.extract_table_data(pdf_path)
            
            if not transactions:
                print("❌ Erreur: Aucune transaction trouvée dans le PDF STB")
                self.result_text.insert(tk.END, "❌ Aucune transaction trouvée dans le PDF STB\n")
                self.result_text.insert(tk.END, "Vérifiez que le fichier est bien un relevé STB valide.\n")
                return
            
            print(f"DEBUG STB - {len(transactions)} transactions extraites")
            
            # Créer le DataFrame
            df = pd.DataFrame(transactions)
            print(f"DEBUG STB - DataFrame créé avec {len(df)} lignes et {len(df.columns)} colonnes")
            
            # Chemin de sortie dans le dossier Téléchargements
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = os.path.join(downloads_path, f"{self.excel_filename.get()}.xlsx")
            
            print(f"DEBUG STB - Sauvegarde en Excel: {excel_path}")
            
            # Sauvegarder en Excel avec formatage
            self.save_excel_with_formatting(df, excel_path)
            
            # Vérifier que le fichier a été créé
            if os.path.exists(excel_path):
                print(f"✅ DEBUG STB - Fichier Excel créé avec succès")
                
                # Afficher les résultats
                self.result_text.insert(tk.END, f"✅ Conversion STB réussie !\n")
                self.result_text.insert(tk.END, f"📁 Fichier créé: {excel_path}\n")
                self.result_text.insert(tk.END, f"📊 Nombre de transactions: {len(transactions)}\n\n")
                self.result_text.insert(tk.END, "Aperçu des données:\n")
                self.result_text.insert(tk.END, df.head(10).to_string(index=False))
                
                # Message de succès dans la console
                print(f"✅ Conversion STB terminée avec succès !")
                print(f"📁 Fichier créé: {excel_path}")
                print(f"📊 Nombre de transactions: {len(transactions)}")
                print(f"🎉 Votre fichier Excel est prêt à utiliser !")
            else:
                print(f"❌ DEBUG STB - Fichier Excel non créé")
                self.result_text.insert(tk.END, "❌ Erreur: Impossible de créer le fichier Excel\n")
            
        except Exception as e:
            error_msg = f"Erreur lors de la conversion STB: {str(e)}"
            self.result_text.insert(tk.END, error_msg)
            
            print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
        
    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            print(f"❌ Impossible de retourner à la page d'accueil: {e}")

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

    def _parse_ocr_text_stb(self, text: str, year: int):
        """Parse le texte OCR pour extraire les transactions STB"""
        transactions = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Chercher les patterns de date et montants STB
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            amount_matches = re.findall(r'[\d\s,\.]+', line)
            
            if date_match:
                date = date_match.group(1)
                
                # Extraire le libellé (texte entre la date et les montants)
                libelle = line[date_match.end():].strip()
                
                # Nettoyer le libellé des montants
                for amount in amount_matches:
                    if re.match(r'^\d', amount.strip()):
                        libelle = libelle.replace(amount, '').strip()
                
                # Déterminer débit/crédit basé sur les montants
                debit = credit = None
                amounts = [amt.strip() for amt in amount_matches if re.match(r'^\d', amt.strip())]
                
                if len(amounts) >= 2:
                    debit = self._format_amount(amounts[0])
                    credit = self._format_amount(amounts[1])
                elif len(amounts) == 1:
                    # Heuristique basée sur le libellé
                    if any(keyword in libelle.lower() for keyword in ['commission', 'frais', 'retrait', 'virement emis']):
                        debit = self._format_amount(amounts[0])
                    else:
                        credit = self._format_amount(amounts[0])
                
                if libelle:
                    # CORRECTION: Nettoyer le libellé
                    libelle_cleaned = self.clean_libelle(libelle)
                    
                    transactions.append({
                        'Date': date,
                        'Libellé': libelle_cleaned,
                        'Débit': debit,
                        'Crédit': credit
                    })
        
        return transactions

    def _format_amount(self, amount_str: str):
        """Formate un montant"""
        try:
            # Nettoyer le montant
            clean_amount = amount_str.replace(' ', '').replace(',', '.')
            return float(clean_amount)
        except:
            return None

def main():
    root = tk.Tk()
    app = STBReleveConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
