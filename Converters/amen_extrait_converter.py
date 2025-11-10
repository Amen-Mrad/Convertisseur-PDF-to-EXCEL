import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import unicodedata
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class AmenExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT AMEN BANK vers Excel")
        self.root.geometry("600x500")

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"EXTRAT_AMEN_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self.setup_ui()

    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur EXTRAT AMEN",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAT AMEN:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.browse_pdf, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                              command=self.convertir, 
                              font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF AMEN EXTRAT", filetypes=[["PDF", "*.pdf"]])
        if path:
            self.pdf_path.set(path)
            # Renouveler le nom Excel pour éviter toute confusion/écrasement
            base = os.path.splitext(os.path.basename(path))[0]
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.excel_name.set(f"EXTRAT_AMEN_{base}_{ts}")

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            print("⚠️ PDF manquant - Veuillez choisir un fichier PDF AMEN EXTRAT.")
            return
        self.progress['value'] = 10; self.root.update_idletasks()
        try:
            print(f"DEBUG AMEN EXTRAT - PDF sélectionné: {path}")
        except Exception:
            pass

        # Test de conversion des montants
        self.test_amount_conversion()
        
        rows = self.parse_pdf(path)
        if not rows:
            print("❌ Aucune transaction - Impossible d'extraire des transactions de l'extrait AMEN.")
            return

        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        
        # Trier les données par date en ordre décroissant (31 -> 01)
        df = self.sort_by_date(df)
        
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'EXTRAT_AMEN'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        self.progress['value'] = 100
        # Message de succès dans la console
        print(f"✅ Conversion EXTRAT terminée avec succès !")
        print(f"📁 Fichier enregistré: {out}")
        print(f"🎉 Votre fichier Excel est prêt à utiliser !")

    def test_amount_conversion(self):
        """Teste la conversion des montants pour vérifier le bon fonctionnement"""
        test_amounts = [
            "2,860",      # Petit montant
            "5 764,499",  # Grand montant avec espaces
            "285,950",    # Montant moyen
            "1 234 567,890",  # Très grand montant
            "25,000",     # Montant avec zéros
            "0,722"       # Petit décimal
        ]
        
        print("🧪 TEST DE CONVERSION DES MONTANTS")
        print("=" * 50)
        
        for test_amount in test_amounts:
            # Simuler la fonction to_float
            s = test_amount.replace('\u00a0', ' ').strip()
            neg = s.startswith('-')
            s = s.lstrip('-')
            
            has_comma = ',' in s
            has_spaces = ' ' in s
            has_dot = '.' in s
            
            if has_comma and has_spaces:
                s_clean = s.replace(' ', '').replace(',', '.')
            elif has_comma and not has_spaces:
                s_clean = s.replace(',', '.')
            elif has_dot and not has_comma:
                s_clean = s
            elif has_spaces and not has_comma:
                s_clean = s.replace(' ', '')
            else:
                s_clean = s
            
            try:
                val = float(s_clean)
                result = -val if neg else val
                print(f"✅ {test_amount:15} -> {result:12.3f}")
            except Exception as e:
                print(f"❌ {test_amount:15} -> ERREUR: {e}")
        
        print("=" * 50)

    def sort_by_date(self, df):
        """Trie le DataFrame par date en ordre décroissant (31 -> 01)"""
        try:
            # Convertir les dates du format DD/MM/YYYY vers datetime pour le tri
            df['date_parsed'] = pd.to_datetime(df['date'], format='%d/%m/%Y', errors='coerce')
            
            # Trier par date en ordre décroissant (plus récent en premier)
            df_sorted = df.sort_values('date_parsed', ascending=False)
            
            # Supprimer la colonne temporaire
            df_sorted = df_sorted.drop('date_parsed', axis=1)
            
            print(f"✅ Tri des dates effectué - {len(df_sorted)} lignes triées en ordre décroissant")
            return df_sorted
            
        except Exception as e:
            print(f"⚠️ Erreur lors du tri des dates: {e}")
            print("📋 Retour du DataFrame non trié")
            return df

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            print(f"❌ Impossible de retourner à la page d'accueil: {e}")

    def is_balance_or_total_line(self, line_text):
        """Détecte si une ligne est un solde, total ou footer (à bloquer)"""
        if not line_text:
            return False
        
        line_upper = line_text.upper()
        
        # BLOQUER: Lignes de soldes et totaux
        balance_keywords = [
            'SOLDE AU', 'SOLDE', 'TOTAUX', 'TOTAL', 'REPORT', 'RAPPORT',
            'SOLDE AU 31', 'SOLDE AU 30', 'SOLDE AU 29', 'SOLDE AU 28',
            'SOLDE AU 27', 'SOLDE AU 26', 'SOLDE AU 25', 'SOLDE AU 24',
            'SOLDE AU 23', 'SOLDE AU 22', 'SOLDE AU 21', 'SOLDE AU 20',
            'SOLDE AU 19', 'SOLDE AU 18', 'SOLDE AU 17', 'SOLDE AU 16',
            'SOLDE AU 15', 'SOLDE AU 14', 'SOLDE AU 13', 'SOLDE AU 12',
            'SOLDE AU 11', 'SOLDE AU 10', 'SOLDE AU 09', 'SOLDE AU 08',
            'SOLDE AU 07', 'SOLDE AU 06', 'SOLDE AU 05', 'SOLDE AU 04',
            'SOLDE AU 03', 'SOLDE AU 02', 'SOLDE AU 01'
        ]
        
        for keyword in balance_keywords:
            if keyword in line_upper:
                print(f"DEBUG AMEN EXTRAT - Ligne solde/total bloquée: {keyword}")
                return True
        
        # BLOQUER les footers de page AMEN BANK
        footer_keywords = [
            'S.A AU CAPITAL', 'R.C :', 'AVENUE MOHAMED V', 'TÉL :', 'FAX :', 'SWIFT :',
            'E-MAIL :', 'SITE WEB :', 'CENTRE DE RELATIONS CLIENTS :',
            'AMENBANK.COM.TN', 'AMENFIRSTBANK.COM.TN', 'CFCTINTTXXX',
            'DINARS', 'TUNIS', 'CLIENTS'
        ]
        
        for keyword in footer_keywords:
            if keyword in line_upper:
                print(f"DEBUG AMEN EXTRAT - Ligne footer bloquée: {keyword}")
                return True
        
        # BLOQUER les lignes avec des numéros de téléphone/fax
        if re.search(r'\b\d{2}\s\d{3}\s\d{3}\b', line_text):  # Format: 71 148 000
            print(f"DEBUG AMEN EXTRAT - Ligne numéro téléphone bloquée: {line_text}")
            return True
        
        # BLOQUER les lignes avec des montants très grands (capitaux, etc.)
        if re.search(r'\b\d{3}\.\d{3}\.\d{3}\b', line_text):  # Format: 174.600.000
            print(f"DEBUG AMEN EXTRAT - Ligne montant très grand bloquée: {line_text}")
            return True
        
        return False
    
    def is_large_amount(self, amount):
        """Détecte si un montant est probablement un solde (à bloquer)"""
        if not amount:
            return False
        
        try:
            # Convertir en float pour vérifier la valeur
            if isinstance(amount, str):
                # Nettoyer le montant
                clean_amount = amount.replace(' ', '').replace(',', '.')
                amount_value = float(clean_amount)
            else:
                amount_value = float(amount)
            
            # BLOQUER: Montants très grands (probablement des soldes)
            if amount_value > 1000000:  # Plus d'1 million
                print(f"DEBUG AMEN EXTRAT - Montant très grand bloqué (solde): {amount}")
                return True
            
            # BLOQUER: Montants négatifs (soldes débiteurs)
            if amount_value < 0:
                print(f"DEBUG AMEN EXTRAT - Montant négatif bloqué (solde): {amount}")
                return True
                
        except:
            pass
        
        return False

    def parse_pdf(self, pdf_path: str):
        # Amen: une colonne d'index précède la date (ex: "58 02/05/2025 ...")
        # Autoriser l'index optionnel avant la date
        date_line_re = re.compile(r"^\s*(?:\d+\s+)?(\d{2}/\d{2}/\d{4})")
        date_any_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        # Regex améliorée pour capturer TOUS les montants AMEN
        # Supporte: 5,950 | 5 764,499 | 285 950,000 | 1 234 567,890
        amount_re = re.compile(r"-?\d{1,3}(?:[ .]\d{3})*(?:[.,]\d{2,3})?|-?\d+[.,]\d{2,3}")

        def to_float(s: str):
            if not s:
                return None
            
            # Nettoyer le string
            s = s.replace('\u00a0', ' ').strip()
            neg = s.startswith('-')
            s = s.lstrip('-')
            
            print(f"DEBUG AMEN EXTRAT - Conversion montant: '{s}'")
            
            # Gestion robuste des formats AMEN
            # Formats supportés: 5,950 | 5 764,499 | 285 950,000 | 1 234 567,890
            
            # Étape 1: Identifier le format
            has_comma = ',' in s
            has_spaces = ' ' in s
            has_dot = '.' in s
            
            print(f"DEBUG AMEN EXTRAT - Format détecté: virgule={has_comma}, espaces={has_spaces}, point={has_dot}")
            
            if has_comma and has_spaces:
                # Format: 5 764,499 ou 285 950,000
                # Supprimer les espaces (milliers) et convertir virgule en point
                s_clean = s.replace(' ', '').replace(',', '.')
                print(f"DEBUG AMEN EXTRAT - Format avec espaces+virgule: '{s}' -> '{s_clean}'")
                
            elif has_comma and not has_spaces:
                # Format: 5,950
                # Convertir virgule en point
                s_clean = s.replace(',', '.')
                print(f"DEBUG AMEN EXTRAT - Format avec virgule: '{s}' -> '{s_clean}'")
                
            elif has_dot and not has_comma:
                # Format: 5764.499 (déjà au bon format)
                s_clean = s
                print(f"DEBUG AMEN EXTRAT - Format avec point: '{s}' -> '{s_clean}'")
                
            elif has_spaces and not has_comma:
                # Format: 5 764 (entier avec espaces)
                s_clean = s.replace(' ', '')
                print(f"DEBUG AMEN EXTRAT - Format entier avec espaces: '{s}' -> '{s_clean}'")
                
            else:
                # Format simple: 5764
                s_clean = s
                print(f"DEBUG AMEN EXTRAT - Format simple: '{s}' -> '{s_clean}'")
            
            try:
                val = float(s_clean)
                result = -val if neg else val
                print(f"DEBUG AMEN EXTRAT - Conversion réussie: '{s}' -> {result}")
                return result
            except Exception as e:
                print(f"DEBUG AMEN EXTRAT - Erreur conversion: '{s}' -> {e}")
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not text.strip():
                    # Fallback OCR pour les PDFs scannés
                    ocr_text = self._extract_text_via_ocr(pdf_path)
                    if ocr_text.strip():
                        text = ocr_text
                        break  # Utiliser le texte OCR pour toutes les pages
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                # Préparer les mots de la page pour récupérer les positions X
                try:
                    page_words = page.extract_words() or []
                except Exception:
                    page_words = []
                # Détecter l'emplacement des en-têtes DEBIT / CREDIT si présent
                debit_x_hdr = None
                credit_x_hdr = None
                for w in page_words:
                    wt = str(w.get('text', '')).strip().upper()
                    xmid = (float(w.get('x0', 0)) + float(w.get('x1', 0))) / 2.0
                    if wt in ('DÉBIT', 'DEBIT'):
                        debit_x_hdr = xmid
                    elif wt in ('CRÉDIT', 'CREDIT'):
                        credit_x_hdr = xmid
                i = 0
                while i < len(lines):
                    line = lines[i]
                    
                    # Bloquer les lignes de soldes, totaux et footers
                    if self.is_balance_or_total_line(line):
                        print(f"DEBUG AMEN EXTRAT - ❌ Ligne bloquée: {line[:50]}...")
                        i += 1
                        continue
                    
                    # Bloquer les lignes qui contiennent des informations de contact
                    if any(keyword in line.upper() for keyword in ['AMENBANK', 'TÉLÉPHONE', 'FAX', 'EMAIL', 'SITE WEB', 'SWIFT']):
                        print(f"DEBUG AMEN EXTRAT - ❌ Ligne contact bloquée: {line[:50]}...")
                        i += 1
                        continue
                    
                    m = date_line_re.match(line)
                    if not m:
                        i += 1
                        continue
                    op_date = m.group(1)

                    # Accumuler description jusqu'à prochaine date ou fin de page
                    j = i + 1
                    desc_lines = []
                    while j < len(lines) and not date_line_re.match(lines[j]):
                        desc_lines.append(lines[j])
                        j += 1
                    
                    # Si on est à la fin de la page, traiter la ligne actuelle même sans date suivante
                    if j >= len(lines):
                        print(f"DEBUG AMEN EXTRAT - Fin de page détectée, traitement de la dernière ligne")

                    combined = (line + ' ' + ' '.join(desc_lines)).strip()
                    # Trouver la 2ème date (date valeur) pour cibler la zone des montants
                    dates_iter = list(date_any_re.finditer(combined))
                    tail_start = dates_iter[1].end() if len(dates_iter) >= 2 else m.end()
                    tail = combined[tail_start:]

                    # Nettoyer la queue avant d'extraire les montants pour éviter les dates
                    tail_clean = tail
                    # Supprimer les dates du libellé qui peuvent interférer (format DD.MM.YYYY)
                    tail_clean = re.sub(r'\b\d{1,2}\.\d{1,2}\.\d{2,4}\b', '', tail_clean)
                    # Supprimer les dates partielles (format DD.MM)
                    tail_clean = re.sub(r'\b\d{1,2}\.\d{1,2}\b', '', tail_clean)
                    # Supprimer les années isolées
                    tail_clean = re.sub(r'\b(19|20)\d{2}\b', '', tail_clean)
                    tail_clean = re.sub(r'\s+', ' ', tail_clean).strip()
                    
                    print(f"DEBUG AMEN EXTRAT - Queue nettoyée: '{tail_clean}'")
                    
                    # Chercher des montants après la date valeur
                    raw_amounts = amount_re.findall(tail_clean)
                    # enlever le préfixe des 2 chiffres de l'année (ex: "25 1,978" ou "251,978")
                    try:
                        # Année de la date valeur si disponible, sinon année de la date d'opération
                        if len(dates_iter) >= 2:
                            y_src = dates_iter[1].group(0)
                        else:
                            y_src = op_date
                        yy = int(y_src[-4:]) % 100
                        pat_space = re.compile(rf'^{yy:02d}\s+(\d{{1,3}}(?:[ \.]\d{{3}})*[.,]\d{{2,3}})$')
                        pat_nospace = re.compile(rf'^{yy:02d}(\d{{1,3}}(?:[ \.]\d{{3}})*[.,]\d{{2,3}})$')
                        cleaned_amounts = []
                        for a in raw_amounts:
                            m1 = pat_space.match(a)
                            m2 = pat_nospace.match(a) if not m1 else None
                            if m1:
                                cleaned_amounts.append(m1.group(1))
                            elif m2:
                                cleaned_amounts.append(m2.group(1))
                            else:
                                cleaned_amounts.append(a)
                        raw_amounts = cleaned_amounts
                    except Exception:
                        pass
                    print(f"DEBUG AMEN EXTRAT - Queue analysée: '{tail}' -> Queue nettoyée: '{tail_clean}' -> Montants trouvés: {raw_amounts}")
                    
                    # Si pas de montants trouvés dans la queue, chercher dans toute la ligne
                    if not raw_amounts and j >= len(lines):
                        print(f"DEBUG AMEN EXTRAT - Aucun montant trouvé dans la queue, recherche dans toute la ligne")
                        print(f"DEBUG AMEN EXTRAT - Ligne complète: '{combined}'")
                        raw_amounts = amount_re.findall(combined)
                        print(f"DEBUG AMEN EXTRAT - Montants trouvés dans toute la ligne: {raw_amounts}")
                    
                    # Filtrer les faux positifs (références longues sans séparateurs ou motifs 05.2025)
                    filtered = []
                    for a in raw_amounts:
                        if re.search(r"\d{2}\.\d{4}", a):
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (date): {a}")
                            continue
                        if re.fullmatch(r"\d{7,}", a.replace(' ', '')):
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (référence longue): {a}")
                            continue
                        
                        # Filtrer les montants qui ressemblent à des numéros de téléphone
                        if re.match(r'^\d{2}\s\d{3}\s\d{3}$', a):  # Format: 71 148 000
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (téléphone): {a}")
                            continue
                        
                        # Filtrer les montants très grands (capitaux, etc.)
                        if re.match(r'^\d{3}\.\d{3}\.\d{3}$', a):  # Format: 174.600.000
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (capital): {a}")
                            continue
                        
                        # Filtrer les montants qui sont dans une ligne de footer
                        if any(keyword in combined.upper() for keyword in ['S.A AU CAPITAL', 'R.C :', 'AVENUE MOHAMED V', 'TÉL :', 'FAX :', 'SWIFT :', 'E-MAIL :', 'SITE WEB :']):
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (footer): {a}")
                            continue
                        
                        # Filtrer les dates du libellé qui sont confondues avec des montants
                        # Pattern: 15.05.202 -> 15,050 (date mal interprétée)
                        if re.match(r'^\d{1,2}\.\d{1,2}\.\d{1,3}$', a):  # Format: 15.05.202
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (date du libellé): {a}")
                            continue
                        
                        # Filtrer les montants qui ressemblent à des dates partielles
                        if re.match(r'^\d{1,2}\.\d{1,2}$', a):  # Format: 15.05
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (date partielle): {a}")
                            continue
                        
                        # Filtrer les montants qui sont clairement des dates
                        if re.match(r'^\d{1,2}\.\d{1,2}\.\d{2,4}$', a):  # Format: 15.05.2025
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (date complète): {a}")
                            continue
                        
                        filtered.append(a)

                    amounts_f = [to_float(x) for x in filtered if to_float(x) is not None]
                    
                    # Debug pour voir les montants extraits
                    if amounts_f:
                        print(f"DEBUG AMEN EXTRAT - Montants extraits: {filtered} -> {amounts_f}")
                        for i, (raw, converted) in enumerate(zip(filtered, amounts_f)):
                            print(f"DEBUG AMEN EXTRAT - Montant {i+1}: '{raw}' -> {converted}")
                        
                        # Si plusieurs montants, logique de sélection intelligente
                        if len(amounts_f) > 1:
                            print(f"DEBUG AMEN EXTRAT - Plusieurs montants trouvés: {amounts_f}")
                            amounts_f_sorted = sorted(amounts_f)
                            amounts_f = [amounts_f_sorted[0]]
                            print(f"DEBUG AMEN EXTRAT - Montant sélectionné: {amounts_f[0]}")
                    
                    # Positions X des montants extraits (quand on peut les retrouver dans les mots)
                    pos_list = []  # (xmid, value)
                    if filtered and page_words:
                        norm_filtered = [f.replace(' ', '') for f in filtered]
                        for w in page_words:
                            wt = str(w.get('text', ''))
                            wt_norm = wt.replace(' ', '')
                            if wt_norm in norm_filtered:
                                val = to_float(wt)
                                if val is not None:
                                    xmid = (float(w.get('x0', 0)) + float(w.get('x1', 0))) / 2.0
                                    pos_list.append((xmid, val))

                    # Construire libellé (entre date opération et valeurs)
                    lib = combined[m.end():tail_start].strip()
                    
                    # Supprimer TOUTES les dates du libellé pour éviter la confusion
                    # Supprimer les dates au format DD/MM/YYYY
                    lib = re.sub(r"\b\d{2}/\d{2}/\d{4}\b", "", lib)
                    # Supprimer les dates au format DD.MM.YYYY (dans les libellés)
                    lib = re.sub(r"\b\d{1,2}\.\d{1,2}\.\d{2,4}\b", "", lib)
                    # Supprimer les dates partielles DD.MM
                    lib = re.sub(r"\b\d{1,2}\.\d{1,2}\b", "", lib)
                    # Supprimer les années isolées
                    lib = re.sub(r"\b(19|20)\d{2}\b", "", lib)
                    
                    lib = re.sub(r"\s+", " ", lib).strip()
                    
                    print(f"DEBUG AMEN EXTRAT - Libellé nettoyé: '{lib}'")

                    debit = credit = None
                    # Filtrer les montants très grands (soldes) avant classification
                    filtered_amounts = []
                    for amount in amounts_f:
                        if not self.is_large_amount(amount):
                            filtered_amounts.append(amount)
                        else:
                            print(f"DEBUG AMEN EXTRAT - Montant filtré (solde): {amount}")
                    
                    # Utiliser les montants filtrés
                    amounts_f = filtered_amounts
                    
                    # 1) Si deux montants trouvés et positions disponibles, trier par X (gauche=Débit, droite=Crédit)
                    if len(pos_list) >= 2:
                        pos_sorted = sorted(pos_list, key=lambda z: z[0])
                        # Vérifier que les montants ne sont pas des soldes
                        if not self.is_large_amount(pos_sorted[0][1]):
                            debit = abs(pos_sorted[0][1])
                        if not self.is_large_amount(pos_sorted[-1][1]):
                            credit = abs(pos_sorted[-1][1])
                    elif len(amounts_f) >= 2:
                        # fallback si positions manquantes
                        if not self.is_large_amount(amounts_f[0]):
                            debit = abs(amounts_f[0])
                        if not self.is_large_amount(amounts_f[1]):
                            credit = abs(amounts_f[1])
                    elif len(amounts_f) == 1:
                        only = amounts_f[0]
                        # Vérifier que le montant n'est pas un solde
                        if not self.is_large_amount(only):
                            # 1bis) si une position et des en-têtes, départager par distance aux en-têtes
                            if len(pos_list) == 1 and (debit_x_hdr is not None and credit_x_hdr is not None):
                                xonly = pos_list[0][0]
                                if abs(xonly - debit_x_hdr) <= abs(xonly - credit_x_hdr):
                                    debit = abs(only)
                                else:
                                    credit = abs(only)
                            else:
                                # 2) Heuristiques par mots-clés enrichies (basées sur vos exemples)
                                upper_lib = lib.upper()
                                credit_kw = [
                                # Encaissements / dépôts / remises
                                'VERSEMENT', 'DEPOT', 'ENCAISSEMENT', 'REM COMMER', 'REM COM',
                                'REM COMMERCANT', 'REM COMMERÇANT', 'REM CHEQ', 'REMISE CHEQUE',
                                'CHEQUE DEPOSIT', 'CHQ DEPOSIT','PLV'
                                # Avoir TPE
                                'AV.TPE', 'AV TPE', 'AVTPE',
                                # Virements reçus
                                'VIREMENT RECU', 'VIR RECU', 'CREDIT'
                                ]
                                debit_kw = [
                                # Commissions / frais / taxes / intérêts
                                'COMMISSION', 'COMM SUR REMISE', 'FRAIS', 'AGIOS', 'TVA', 'INTERET',
                                # Retraits / paiements
                                'RETRAIT', 'PAIEMENT',
                                # Effets / cion
                                'CION', 'EFFET', 'CION/EFFET', 'CION EFFET', 'REG EFFET',
                                # Prélèvements / règlements
                                'PRELEVEMENT', 'REGLEMENT', 'PLV REGL', 'PLV REGLE',
                                # Virements émis
                                'VIREMENT EMIS', 'DEBIT',
                                # Ajouter les mots-clés manquants
                                'COTIS', 'COTIS.TPE', 'COTIS TPE', 'COTISATION'
                                ]
                                # Règles spécifiques (très fréquentes)
                                credit_strong = ['VERSEMENT', 'AV.TPE', 'AV TPE', 'AVTPE', 'CHEQUE DEPOSIT', 'CHQ DEPOSIT']
                                debit_strong = ['COMMISSION SUR REMISE', 'TVA SUR', 'PLV REGLE', 'PLV REGL', 'CION/EFFET', 'CION EFFET', 'COTIS', 'COTIS.TPE']

                                if any(k in upper_lib for k in debit_strong):
                                    debit = abs(only)
                                elif any(k in upper_lib for k in credit_strong):
                                    credit = abs(only)
                                else:
                                    is_credit_kw = any(k in upper_lib for k in credit_kw)
                                    is_debit_kw = any(k in upper_lib for k in debit_kw)
                                    if is_debit_kw and not is_credit_kw:
                                        debit = abs(only)
                                    elif is_credit_kw and not is_debit_kw:
                                        credit = abs(only)
                                    else:
                                        # Logique améliorée pour les petites transactions
                                        # Si le montant est très petit (< 1), c'est probablement un débit (frais/commissions)
                                        if only < 1.0:
                                            debit = abs(only)
                                            print(f"DEBUG AMEN EXTRAT - Petite transaction classée DÉBIT: {only}")
                                        # 3) Fallback: signe de la donnée extraite
                                        elif (filtered and filtered[0].strip().startswith('-')) or only < 0:
                                            debit = abs(only)
                                        else:
                                            credit = abs(only)

                    # Ignorer lignes de soldes et totaux
                    upper = lib.upper()
                    if re.search(r"(?i)solde|totaux|report", upper):
                        i = j
                        continue

                    # Éviter d'avoir des montants dans les deux colonnes
                    if debit and credit:
                        # Si les montants sont identiques, garder seulement le crédit (versements)
                        if abs(debit - credit) < 0.001:
                            if any(kw in lib.upper() for kw in ["VERSEMENT", "DEPOT", "ENCAISSEMENT", "REM COM", "AV.TPE", "VIREMENT RECU"]):
                                debit = None
                                print(f"DEBUG AMEN EXTRAT - Montant identique, gardé CRÉDIT: {credit}")
                            else:
                                credit = None
                                print(f"DEBUG AMEN EXTRAT - Montant identique, gardé DÉBIT: {debit}")
                        elif debit < 0.01 or credit < 0.01:
                            # Si l'un des montants est quasi-nul, le supprimer
                            if debit < 0.01:
                                debit = None
                                print(f"DEBUG AMEN EXTRAT - Montant quasi-nul supprimé du DÉBIT")
                            else:
                                credit = None
                                print(f"DEBUG AMEN EXTRAT - Montant quasi-nul supprimé du CRÉDIT")
                        else:
                            # Si deux montants différents, classifier selon le libellé
                            upper_lib = lib.upper()
                            credit_keywords = [
                                'VERSEMENT', 'DEPOT', 'ENCAISSEMENT', 'REM COM', 'AV.TPE', 
                                'VIREMENT RECU', 'VIR RECU', 'CREDIT', 'REMISE'
                            ]
                            debit_keywords = [
                                'COMMISSION', 'FRAIS', 'AGIOS', 'TVA', 'RETRAIT', 'PAIEMENT',
                                'CION', 'EFFET', 'COTIS'
                            ]
                            
                            if any(kw in upper_lib for kw in credit_keywords):
                                debit = None
                                print(f"DEBUG AMEN EXTRAT - Classification CRÉDIT selon libellé: {credit}")
                            elif any(kw in upper_lib for kw in debit_keywords):
                                credit = None
                                print(f"DEBUG AMEN EXTRAT - Classification DÉBIT selon libellé: {debit}")
                            else:
                                # Par défaut, garder le plus petit montant comme crédit
                                if debit < credit:
                                    debit = None
                                    print(f"DEBUG AMEN EXTRAT - Par défaut, gardé CRÉDIT (plus grand): {credit}")
                                else:
                                    credit = None
                                    print(f"DEBUG AMEN EXTRAT - Par défaut, gardé DÉBIT (plus grand): {debit}")
                    
                    # Préserver les montants originaux avant formatage
                    debit_original = debit
                    credit_original = credit
                    
                    # Formater les montants selon le standard BIAT
                    if debit:
                        debit_formatted = self._format_amount_biats_style(debit)
                        print(f"DEBUG AMEN EXTRAT - DÉBIT formaté: {debit} -> {debit_formatted}")
                    else:
                        debit_formatted = None
                        
                    if credit:
                        credit_formatted = self._format_amount_biats_style(credit)
                        print(f"DEBUG AMEN EXTRAT - CRÉDIT formaté: {credit} -> {credit_formatted}")
                    else:
                        credit_formatted = None
                    
                    results.append({
                        "date": op_date,
                        "libelle": lib,
                        "debit": debit_formatted,
                        "credit": credit_formatted
                    })

                    i = j
        return results

    def _format_amount_biats_style(self, amount_float):
        """Formate un montant selon le style AMEN: espace pour milliers, virgule pour décimales"""
        if amount_float is None:
            return None
        
        try:
            print(f"DEBUG AMEN EXTRAT - Formatage montant: {amount_float}")
            
            # Formater selon la taille du montant
            if amount_float >= 1000:
                # Pour les grands montants: espace pour milliers, virgule pour décimales
                # Exemple: 5764.499 -> 5 764,499
                # Exemple: 285950.000 -> 285 950,000
                formatted = f"{amount_float:,.3f}"
                formatted = formatted.replace(',', ' ').replace('.', ',')
                print(f"DEBUG AMEN EXTRAT - Grand montant formaté: {amount_float} -> {formatted}")
            else:
                # Pour les petits montants: seulement virgule pour décimales
                # Exemple: 2.860 -> 2,860
                formatted = f"{amount_float:.3f}"
                formatted = formatted.replace('.', ',')
                print(f"DEBUG AMEN EXTRAT - Petit montant formaté: {amount_float} -> {formatted}")
            
            return formatted
        except (ValueError, TypeError) as e:
            print(f"DEBUG AMEN EXTRAT - Erreur formatage: {amount_float} -> {e}")
            return str(amount_float) if amount_float is not None else None

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "J03"
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        # Les montants sont déjà formatés en format BIAT dans le DataFrame
        # Pas besoin de reformater dans Excel
        max_row = ws.max_row
        print(f"DEBUG AMEN EXTRAT - Formatage Excel: {max_row} lignes à traiter")
        # Bordures fines sur tout le tableau
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

def main():
    root = tk.Tk(); AmenExtraitConverter(root); root.mainloop()

if __name__ == '__main__':
    main()
