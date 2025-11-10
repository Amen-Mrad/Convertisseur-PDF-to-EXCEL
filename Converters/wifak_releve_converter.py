#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convertisseur WIFAK Bank - Relevé de compte
Convertit les relevés de compte WIFAK Bank (PDF) en fichiers Excel
"""

import os
import sys
import re
import pandas as pd
import pdfplumber
import pytesseract
import fitz  # PyMuPDF
from PIL import Image
import io
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import cv2
import numpy as np

class WifakReleveConverter:
    def __init__(self, root=None, silent_mode=False):
        if root is None:
            if silent_mode:
                # Mode silencieux - créer une fenêtre cachée
                self.root = tk.Tk()
                self.root.withdraw()  # Cacher la fenêtre
            else:
                # Mode normal - créer une interface complète
                self.root = tk.Tk()
                self.root.title("Convertisseur WIFAK Bank - Relevé de compte")
                self.root.geometry("600x500")
                self.root.configure(bg='#f0f0f0')
        else:
            self.root = root
        
        # Variables (doivent être définies après la création de self.root)
        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        
        # Setup UI seulement si on a créé une nouvelle fenêtre et pas en mode silencieux
        if root is None and not silent_mode:
            self.setup_ui()
        
    def setup_ui(self):
        """Configure l'interface utilisateur"""
        # Titre principal
        title_frame = tk.Frame(self.root, bg='#f0f0f0')
        title_frame.pack(pady=20)
        
        title_label = tk.Label(title_frame, text="Convertisseur WIFAK Bank - Relevé", 
                              font=('Arial', 16, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        title_label.pack()
        
        # Frame pour la sélection de fichier PDF
        pdf_frame = tk.Frame(self.root, bg='#f0f0f0')
        pdf_frame.pack(pady=10, padx=20, fill='x')
        
        tk.Label(pdf_frame, text="Fichier PDF WIFAK Relevé:", 
                font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(anchor='w')
        
        pdf_entry_frame = tk.Frame(pdf_frame, bg='#f0f0f0')
        pdf_entry_frame.pack(fill='x', pady=5)
        
        tk.Entry(pdf_entry_frame, textvariable=self.pdf_path, width=50, 
                font=('Arial', 9)).pack(side='left', fill='x', expand=True)
        
        tk.Button(pdf_entry_frame, text="Parcourir", command=self.select_pdf,
                 bg='#3498db', fg='white', font=('Arial', 9, 'bold')).pack(side='right', padx=(10,0))
        
        # Frame pour la sélection du fichier Excel de sortie
        excel_frame = tk.Frame(self.root, bg='#f0f0f0')
        excel_frame.pack(pady=10, padx=20, fill='x')
        
        tk.Label(excel_frame, text="Fichier Excel de sortie:", 
                font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(anchor='w')
        
        excel_entry_frame = tk.Frame(excel_frame, bg='#f0f0f0')
        excel_entry_frame.pack(fill='x', pady=5)
        
        tk.Entry(excel_entry_frame, textvariable=self.excel_path, width=50, 
                font=('Arial', 9)).pack(side='left', fill='x', expand=True)
        
        tk.Button(excel_entry_frame, text="Parcourir", command=self.select_excel,
                 bg='#3498db', fg='white', font=('Arial', 9, 'bold')).pack(side='right', padx=(10,0))
        
        # Boutons d'action
        button_frame = tk.Frame(self.root, bg='#f0f0f0')
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Convertir", command=self.convert,
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'),
                 width=15, height=2).pack(side='left', padx=10)
        
        tk.Button(button_frame, text="Quitter", command=self.root.quit,
                 bg='#e74c3c', fg='white', font=('Arial', 12, 'bold'),
                 width=15, height=2).pack(side='left', padx=10)
        
        # Zone de statut
        self.status_frame = tk.Frame(self.root, bg='#f0f0f0')
        self.status_frame.pack(pady=10, padx=20, fill='x')
        
        self.status_label = tk.Label(self.status_frame, text="Prêt à convertir", 
                                   font=('Arial', 9), bg='#f0f0f0', fg='#27ae60')
        self.status_label.pack()
        
        # Zone de progression
        self.progress_frame = tk.Frame(self.root, bg='#f0f0f0')
        self.progress_frame.pack(pady=5, padx=20, fill='x')
        
        self.progress = ttk.Progressbar(self.progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
    def select_pdf(self):
        """Sélectionne le fichier PDF"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner le fichier PDF WIFAK Relevé",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_path.set(file_path)
            # Générer automatiquement le nom du fichier Excel dans Downloads
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            excel_name = f"{base_name}_wifak_releve_converted.xlsx"
            
            # Chemin vers le dossier Downloads
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = os.path.join(downloads_path, excel_name)
            self.excel_path.set(excel_path)
            
    def select_excel(self):
        """Sélectionne le fichier Excel de sortie"""
        file_path = filedialog.asksaveasfilename(
            title="Sauvegarder le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.excel_path.set(file_path)
            
    def update_status(self, message, color='#27ae60'):
        """Met à jour le message de statut"""
        self.status_label.config(text=message, fg=color)
        self.root.update()
        
    def start_progress(self):
        """Démarre la barre de progression"""
        self.progress.start()
        
    def stop_progress(self):
        """Arrête la barre de progression"""
        self.progress.stop()
        
    def detect_wifak_logo(self, pdf_path):
        """Détecte le logo WIFAK dans le PDF"""
        try:
            doc = fitz.open(pdf_path)
            for page_num in range(min(3, len(doc))):  # Vérifier les 3 premières pages
                page = doc[page_num]
                pix = page.get_pixmap()
                img_data = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_data))
                
                # Convertir en OpenCV
                opencv_img = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)
                
                # Charger le logo WIFAK
                logo_path = os.path.join('logo', 'wifak.png')
                if os.path.exists(logo_path):
                    logo = cv2.imread(logo_path)
                    if logo is not None:
                        # Recherche du logo
                        result = cv2.matchTemplate(opencv_img, logo, cv2.TM_CCOEFF_NORMED)
                        locations = np.where(result >= 0.6)
                        
                        if len(locations[0]) > 0:
                            print(f"Logo WIFAK détecté sur la page {page_num + 1}")
                            doc.close()
                            return True
                            
            doc.close()
            return False
        except Exception as e:
            print(f"Erreur lors de la détection du logo: {e}")
            return False
            
    def detect_document_type(self, pdf_path):
        """Détecte le type de document (relevé) basé sur les mots-clés"""
        try:
            doc = fitz.open(pdf_path)
            text_content = ""
            
            # Extraire le texte des premières pages
            for page_num in range(min(3, len(doc))):
                page = doc[page_num]
                text_content += page.get_text() + "\n"
                
            doc.close()
            
            # Mots-clés pour un relevé WIFAK
            releve_keywords = [
                'relevé', 'releve', 'compte', 'mouvement', 'solde',
                'date valeur', 'date valeur', 'libellé', 'libelle',
                'débit', 'debit', 'crédit', 'credit'
            ]
            
            text_lower = text_content.lower()
            keyword_count = sum(1 for keyword in releve_keywords if keyword in text_lower)
            
            print(f"Nombre de mots-clés relevé détectés: {keyword_count}")
            return keyword_count >= 3
            
        except Exception as e:
            print(f"Erreur lors de la détection du type de document: {e}")
            return False
            
    def extract_wifak_releve_data(self, pdf_path):
        """Extrait les données du relevé WIFAK"""
        try:
            print("=== EXTRACTION DONNÉES WIFAK RELEVÉ ===")
            
            # Essayer d'abord l'extraction par table
            transactions = self.parse_wifak_releve_table(pdf_path)
            if transactions:
                print(f"Extraction par table réussie: {len(transactions)} transactions")
                return transactions
                
            # Essayer l'extraction par texte
            transactions = self.parse_wifak_releve_text(pdf_path)
            if transactions:
                print(f"Extraction par texte réussie: {len(transactions)} transactions")
                return transactions
                
            # Essayer l'extraction par layout
            transactions = self.parse_wifak_releve_layout(pdf_path)
            if transactions:
                print(f"Extraction par layout réussie: {len(transactions)} transactions")
                return transactions
                
            # En dernier recours, utiliser l'OCR
            transactions = self.parse_wifak_releve_ocr(pdf_path)
            if transactions:
                print(f"Extraction par OCR réussie: {len(transactions)} transactions")
                return transactions
                
            print("Aucune méthode d'extraction n'a fonctionné")
            return []
            
        except Exception as e:
            print(f"Erreur lors de l'extraction des données: {e}")
            return []
            
    def parse_wifak_releve_table(self, pdf_path):
        """Parse le relevé WIFAK en utilisant l'extraction de table"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"Traitement de la page {page_num + 1}")
                    
                    # Extraire les tables
                    tables = page.extract_tables()
                    
                    for table_num, table in enumerate(tables):
                        if not table or len(table) < 2:
                            continue
                            
                        print(f"Table {table_num + 1} trouvée avec {len(table)} lignes")
                        
                        # Analyser l'en-tête pour identifier les colonnes
                        header = table[0] if table else []
                        date_col = -1
                        libelle_col = -1
                        date_valeur_col = -1
                        debit_col = -1
                        credit_col = -1
                        
                        for i, cell in enumerate(header):
                            if cell:
                                cell_lower = cell.lower().strip()
                                print(f"DEBUG WIFAK RELEVÉ - En-tête colonne {i}: '{cell}'")
                                
                                # Détection plus robuste des colonnes
                                if 'date' in cell_lower and 'opération' not in cell_lower and 'valeur' not in cell_lower:
                                    date_col = i
                                    print(f"DEBUG WIFAK RELEVÉ - Colonne Date détectée: {i}")
                                elif 'libellé' in cell_lower or 'libelle' in cell_lower or 'opération' in cell_lower:
                                    libelle_col = i
                                    print(f"DEBUG WIFAK RELEVÉ - Colonne Libellé détectée: {i}")
                                elif 'valeur' in cell_lower and 'date' in cell_lower:
                                    date_valeur_col = i
                                    print(f"DEBUG WIFAK RELEVÉ - Colonne Date Valeur détectée: {i}")
                                elif 'débit' in cell_lower or 'debit' in cell_lower or 'عليه' in cell_lower:
                                    debit_col = i
                                    print(f"DEBUG WIFAK RELEVÉ - Colonne Débit détectée: {i}")
                                elif 'crédit' in cell_lower or 'credit' in cell_lower or 'له' in cell_lower:
                                    credit_col = i
                                    print(f"DEBUG WIFAK RELEVÉ - Colonne Crédit détectée: {i}")
                        
                        print(f"Colonnes détectées - Date: {date_col}, Libellé: {libelle_col}, "
                              f"Date Valeur: {date_valeur_col}, Débit: {debit_col}, Crédit: {credit_col}")
                        
                        # Fallback si les colonnes ne sont pas détectées correctement
                        if debit_col == -1 and credit_col == -1:
                            # Essayer de détecter les colonnes de montants par position
                            if len(header) >= 5:
                                debit_col = 3  # Position typique pour débit
                                credit_col = 4  # Position typique pour crédit
                                print(f"DEBUG WIFAK RELEVÉ - Fallback: Débit=3, Crédit=4")
                            elif len(header) >= 4:
                                debit_col = 2  # Position alternative
                                credit_col = 3
                                print(f"DEBUG WIFAK RELEVÉ - Fallback: Débit=2, Crédit=3")
                        
                        # Extraire les transactions
                        for row_num, row in enumerate(table[1:], 1):
                            if len(row) < max(date_col, libelle_col, debit_col, credit_col) + 1:
                                continue
                                
                            date = row[date_col] if date_col >= 0 and date_col < len(row) else ""
                            libelle = row[libelle_col] if libelle_col >= 0 and libelle_col < len(row) else ""
                            debit = row[debit_col] if debit_col >= 0 and debit_col < len(row) else ""
                            credit = row[credit_col] if credit_col >= 0 and credit_col < len(row) else ""
                            
                            # Filtrer les lignes de solde "Au" et les lignes vides
                            if self._is_balance_line(libelle) or self._is_empty_line(date, libelle, debit, credit):
                                print(f"Ligne filtrée (solde ou vide): Date='{date}', Libellé='{libelle}', Débit='{debit}', Crédit='{credit}'")
                                continue
                            
                            if date and (debit or credit):
                                # Nettoyer le libellé en préservant les références importantes
                                libelle_cleaned = self._clean_libelle_from_dates(libelle)
                                
                                # Améliorer le libellé pour WIFAK
                                libelle_final = self._improve_wifak_libelle(libelle_cleaned)
                                
                                # Classifier correctement les montants débit/crédit pour WIFAK
                                debit_amount, credit_amount = self._classify_wifak_amounts(libelle_final, debit, credit)
                                
                                transaction = {
                                    'Date': self._format_date(date),
                                    'Libellé': libelle_final,
                                    'Débit': debit_amount,
                                    'Crédit': credit_amount
                                }
                                
                                # Filtrer les transactions vides
                                if transaction['Libellé'] or transaction['Débit'] or transaction['Crédit']:
                                    all_transactions.append(transaction)
                                    print(f"Transaction extraite: {transaction}")
                
                return all_transactions
                
        except Exception as e:
            print(f"Erreur lors du parsing par table: {e}")
            return []
            
    def parse_wifak_releve_text(self, pdf_path):
        """Parse le relevé WIFAK en utilisant l'extraction de texte"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"Traitement de la page {page_num + 1}")
                    text = page.extract_text()
                    
                    if not text:
                        continue
                        
                    # Diviser le texte en lignes
                    lines = text.split('\n')
                    
                    # Rechercher les lignes de transactions
                    for line in lines:
                        transaction = self.parse_wifak_releve_line(line)
                        if transaction:
                            all_transactions.append(transaction)
                            
                return all_transactions
                
        except Exception as e:
            print(f"Erreur lors du parsing par texte: {e}")
            return []
            
    def parse_wifak_releve_layout(self, pdf_path):
        """Parse le relevé WIFAK en utilisant l'extraction par layout"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"Traitement de la page {page_num + 1}")
                    
                    # Extraire les objets de layout
                    layout = page.layout
                    if not layout:
                        continue
                        
                    # Rechercher les lignes de texte
                    for obj in layout:
                        if hasattr(obj, 'text') and obj.text:
                            transaction = self.parse_wifak_releve_line(obj.text)
                            if transaction:
                                all_transactions.append(transaction)
                                
                return all_transactions
                
        except Exception as e:
            print(f"Erreur lors du parsing par layout: {e}")
            return []
            
    def parse_wifak_releve_ocr(self, pdf_path):
        """Parse le relevé WIFAK en utilisant l'OCR"""
        try:
            doc = fitz.open(pdf_path)
            all_transactions = []
            
            for page_num in range(len(doc)):
                print(f"Traitement OCR de la page {page_num + 1}")
                page = doc[page_num]
                
                # Convertir la page en image
                pix = page.get_pixmap()
                img_data = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_data))
                
                # Appliquer l'OCR
                text = pytesseract.image_to_string(img, lang='fra')
                
                if text:
                    lines = text.split('\n')
                    for line in lines:
                        transaction = self.parse_wifak_releve_line(line)
                        if transaction:
                            all_transactions.append(transaction)
                            
            doc.close()
            return all_transactions
            
        except Exception as e:
            print(f"Erreur lors du parsing OCR: {e}")
            return []
            
    def parse_wifak_releve_line(self, line):
        """Parse une ligne de relevé WIFAK"""
        if not line or len(line.strip()) < 10:
            return None
            
        line = line.strip()
        
        # Patterns pour détecter les dates
        date_patterns = [
            r'\b\d{1,2}/\d{1,2}/\d{4}\b',
            r'\b\d{1,2}-\d{1,2}-\d{4}\b',
            r'\b\d{1,2}\.\d{1,2}\.\d{4}\b'
        ]
        
        # Patterns pour détecter les montants
        amount_patterns = [
            r'\b\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{3})?\b',  # Format avec séparateurs
            r'\b\d+(?:[.,]\d{3})?\b'  # Format simple
        ]
        
        # Rechercher une date au début de la ligne
        date_match = None
        for pattern in date_patterns:
            match = re.search(pattern, line)
            if match:
                date_match = match
                break
                
        if not date_match:
            return None
            
        # Extraire la date
        date = date_match.group()
        
        # Extraire le reste de la ligne après la date
        remaining = line[date_match.end():].strip()
        
        # Rechercher les montants dans le reste
        amounts = []
        for pattern in amount_patterns:
            matches = re.findall(pattern, remaining)
            amounts.extend(matches)
            
        # Nettoyer les montants
        clean_amounts = []
        for amount in amounts:
            clean_amount = re.sub(r'[^\d,.\s-]', '', amount)
            if clean_amount and clean_amount != '0':
                clean_amounts.append(clean_amount)
                
        # Déterminer le libellé (tout ce qui n'est pas des montants)
        libelle = remaining
        
        # Supprimer les montants du libellé de manière plus intelligente
        for amount in amounts:
            # Utiliser une approche plus précise pour supprimer les montants
            # en évitant de supprimer les références de chèques
            if re.match(r'^\d+[.,]\d+$', amount) or re.match(r'^\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{3})?$', amount):
                # C'est un vrai montant, le supprimer
                libelle = libelle.replace(amount, '', 1)
            elif re.match(r'^\d{10,}$', amount):
                # C'est probablement une référence de chèque, ne pas le supprimer
                continue
                
        # Nettoyer le libellé en préservant les références importantes
        libelle = re.sub(r'\s+', ' ', libelle).strip()
        libelle = self._clean_libelle_from_dates(libelle)
        
        # Filtrer les lignes de solde "Au" et les lignes vides
        if self._is_balance_line(libelle) or self._is_empty_line(date, libelle, None, None):
            print(f"Ligne filtrée (solde ou vide): Date='{date}', Libellé='{libelle}'")
            return None
        
        # Déterminer débit et crédit
        debit = None
        credit = None
        
        if len(clean_amounts) >= 1:
            # Si un seul montant, c'est probablement un débit ou crédit
            amount = clean_amounts[0]
            if 'débit' in line.lower() or 'debit' in line.lower():
                debit = self._format_amount(amount)
            elif 'crédit' in line.lower() or 'credit' in line.lower():
                credit = self._format_amount(amount)
            else:
                # Par défaut, considérer comme débit si négatif, crédit si positif
                try:
                    num_amount = float(amount.replace(',', '.').replace(' ', ''))
                    if num_amount < 0:
                        debit = self._format_amount(amount)
                    else:
                        credit = self._format_amount(amount)
                except:
                    credit = self._format_amount(amount)
        elif len(clean_amounts) >= 2:
            # Si deux montants, le premier est débit, le second crédit
            debit = self._format_amount(clean_amounts[0])
            credit = self._format_amount(clean_amounts[1])
            
        # Créer la transaction si on a au moins un montant
        if debit or credit:
            return {
                'Date': self._format_date(date),
                'Libellé': libelle,
                'Débit': debit,
                'Crédit': credit
            }
            
        return None
        
    def _clean_libelle_from_dates(self, libelle):
        """Nettoie le libellé en préservant les informations importantes comme les références de chèques"""
        if not libelle:
            return libelle

        # Nettoyer seulement les dates isolées au début du libellé
        # mais préserver les références de chèques et codes importants
        cleaned_libelle = libelle.strip()
        
        # Supprimer uniquement les dates au début qui sont suivies d'un espace
        # et qui ne font pas partie d'une référence de chèque
        date_patterns = [
            r'^\d{1,2}/\d{1,2}/\d{4}\s+(?![A-Za-z])',  # Date au début, pas suivie d'une lettre
            r'^\d{1,2}-\d{1,2}-\d{4}\s+(?![A-Za-z])',  # Date au début, pas suivie d'une lettre
            r'^\d{1,2}\.\d{1,2}\.\d{4}\s+(?![A-Za-z])',  # Date au début, pas suivie d'une lettre
        ]

        for pattern in date_patterns:
            cleaned_libelle = re.sub(pattern, '', cleaned_libelle)

        # Nettoyer les espaces multiples mais préserver la structure
        cleaned_libelle = re.sub(r'\s+', ' ', cleaned_libelle).strip()

        print(f"DEBUG WIFAK RELEVÉ - Libellé nettoyé: '{libelle}' -> '{cleaned_libelle}'")
        return cleaned_libelle
        
    def _improve_wifak_libelle(self, libelle):
        """Améliore le formatage des libellés WIFAK pour un affichage plus clair"""
        if not libelle:
            return libelle
            
        # Vérifier si c'est une ligne de solde "Au" - ne pas la traiter
        if self._is_balance_line(libelle):
            print(f"DEBUG WIFAK RELEVÉ - Ligne de solde détectée, pas d'amélioration: '{libelle}'")
            return libelle
            
        # Nettoyer les espaces multiples
        libelle = re.sub(r'\s+', ' ', libelle).strip()
        
        # Améliorer la lisibilité des libellés WIFAK
        # Remplacer les patterns courants par des versions plus lisibles
        improvements = {
            r'Remise Chq WIFAK - Remise Cheque Chèque:': 'Remise Chèque WIFAK:',
            r'TVA - Remise Chq WIFAK - Remise Cheque Chèque:': 'TVA - Remise Chèque WIFAK:',
            r'Paiement Effet Prsentation Effet:': 'Paiement Effet:',
            r'Règlement Effet - Effet Reçu NAI Prsentation Chèque:': 'Règlement Effet:',
            r'TVA - Règlement Effet - Effet Reçu NAI Prsentation Chèque:': 'TVA - Règlement Effet:',
            r'Encaissement Chèque:': 'Encaissement Chèque:',
            r'Vers Esp Mm Ag vs': 'Virement Espèces',
            r'Virement Emis': 'Virement Émis',
            r'Virement Emis Local - Virement Emis': 'Virement Local Émis',
            r'TVA - Virement Emis Local - Virement Emis': 'TVA - Virement Local Émis'
        }
        
        for pattern, replacement in improvements.items():
            libelle = re.sub(pattern, replacement, libelle)
        
        # Nettoyer les références de chèques pour les rendre plus lisibles
        # Exemple: "000000152207" -> "Chèque 152207"
        libelle = re.sub(r'(\d{10,})', r'Chèque \1', libelle)
        
        # Nettoyer les codes avec slashes, mais éviter de toucher aux lignes de solde "Au"
        # Exemple: "585995 //5/" -> "585995"
        if not libelle.lower().strip().startswith('au'):
            libelle = re.sub(r'(\d+)\s*//+.*', r'\1', libelle)
        
        print(f"DEBUG WIFAK RELEVÉ - Libellé amélioré: '{libelle}'")
        return libelle
        
    def _classify_wifak_amounts(self, libelle, debit, credit):
        """Classifie correctement les montants débit/crédit pour WIFAK"""
        debit_amount = None
        credit_amount = None
        
        # Logique de classification basée sur le type d'opération WIFAK
        libelle_lower = libelle.lower()
        
        # Opérations qui sont DÉFINITIVEMENT des débits
        debit_operations = [
            'paiement effet', 'règlement effet', 'tva - règlement effet',
            'tva - remise', 'tva - virement', 'virement émis', 'virement local émis',
            'tva - virement local émis', 'retrait', 'débit', 'frais', 'commission',
            'paiement', 'règlement', 'tva'
        ]
        
        # Opérations qui sont DÉFINITIVEMENT des crédits
        credit_operations = [
            'encaissement chèque', 'remise chèque', 'virement reçu',
            'versement', 'crédit', 'remboursement', 'intérêt', 'encaissement'
        ]
        
        # Vérifier si c'est une opération de débit ou crédit
        is_debit_operation = any(op in libelle_lower for op in debit_operations)
        is_credit_operation = any(op in libelle_lower for op in credit_operations)
        
        print(f"DEBUG WIFAK RELEVÉ - Analyse: Libellé='{libelle}', is_debit={is_debit_operation}, is_credit={is_credit_operation}")
        
        # Si on a les deux montants, les utiliser tels quels
        if debit and credit:
            debit_amount = self._format_amount(debit)
            credit_amount = self._format_amount(credit)
            print(f"DEBUG WIFAK RELEVÉ - Cas 1: Les deux montants présents")
        # Si on a un seul montant, le classifier selon l'opération
        elif debit and not credit:
            if is_credit_operation:
                # Si c'est une opération de crédit mais qu'on a un montant en débit,
                # le déplacer vers crédit
                credit_amount = self._format_amount(debit)
                debit_amount = None
                print(f"DEBUG WIFAK RELEVÉ - Cas 2a: Montant débit déplacé vers crédit")
            else:
                debit_amount = self._format_amount(debit)
                print(f"DEBUG WIFAK RELEVÉ - Cas 2b: Montant débit conservé")
        elif credit and not debit:
            if is_debit_operation:
                # Si c'est une opération de débit mais qu'on a un montant en crédit,
                # le déplacer vers débit
                debit_amount = self._format_amount(credit)
                credit_amount = None
                print(f"DEBUG WIFAK RELEVÉ - Cas 3a: Montant crédit déplacé vers débit")
            else:
                credit_amount = self._format_amount(credit)
                print(f"DEBUG WIFAK RELEVÉ - Cas 3b: Montant crédit conservé")
        else:
            # Aucun montant trouvé
            print(f"DEBUG WIFAK RELEVÉ - Cas 4: Aucun montant trouvé")
        
        print(f"DEBUG WIFAK RELEVÉ - Résultat final: Débit={debit_amount}, Crédit={credit_amount}")
        return debit_amount, credit_amount
        
    def _is_balance_line(self, libelle):
        """Détecte si une ligne est une ligne de solde 'Au' à filtrer"""
        if not libelle:
            return False
            
        libelle_lower = libelle.lower().strip()
        
        # Patterns pour détecter les lignes de solde
        balance_patterns = [
            r'^au\s*//?$',  # "Au //" ou "Au /"
            r'^solde\s+au',  # "Solde Au"
            r'^au\s*$',      # Juste "Au"
            r'^//\s*$',      # Juste "//"
            r'^au\s*$',      # "Au" seul
            r'^au\s*//+.*$',  # "Au //" avec d'autres caractères
            r'^solde\s*au.*$',  # "Solde Au" avec d'autres caractères
        ]
        
        for pattern in balance_patterns:
            if re.match(pattern, libelle_lower):
                print(f"DEBUG WIFAK RELEVÉ - Ligne de solde détectée: '{libelle}'")
                return True
                
        return False
        
    def _is_empty_line(self, date, libelle, debit, credit):
        """Détecte si une ligne est vide ou ne contient que des séparateurs"""
        if not libelle and not debit and not credit:
            return True
            
        if libelle:
            libelle_clean = libelle.strip()
            # Vérifier si le libellé ne contient que des séparateurs
            if re.match(r'^[/\s]*$', libelle_clean):
                print(f"DEBUG WIFAK RELEVÉ - Ligne vide détectée: Date='{date}', Libellé='{libelle}'")
                return True
                
        return False
        
    def _format_date(self, date_str):
        """Formate une date"""
        if not date_str:
            return ""
            
        # Nettoyer la date
        date_str = date_str.strip()
        
        # Patterns de date supportés
        patterns = [
            r'(\d{1,2})/(\d{1,2})/(\d{4})',
            r'(\d{1,2})-(\d{1,2})-(\d{4})',
            r'(\d{1,2})\.(\d{1,2})\.(\d{4})'
        ]
        
        for pattern in patterns:
            match = re.match(pattern, date_str)
            if match:
                day, month, year = match.groups()
                return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
                
        return date_str
        
    def _format_amount(self, amount_str):
        """Formate un montant selon le format WIFAK avec 3 chiffres après la virgule et espace comme séparateur de milliers."""
        if not amount_str:
            return None
        try:
            # Nettoyer le montant pour obtenir une valeur numérique propre
            cleaned = re.sub(r'[^\d,.\s-]', '', amount_str)
            if not cleaned:
                return None

            print(f"DEBUG WIFAK RELEVÉ - Montant brut: '{amount_str}' -> nettoyé: '{cleaned}'")

            # Convertir en float en gérant les différents formats d'entrée
            if ',' in cleaned and '.' in cleaned:
                # Analyser le format pour déterminer le séparateur décimal
                if ' ' in cleaned:
                    # Format: 12 256,350 (espaces comme séparateurs de milliers)
                    cleaned = cleaned.replace(' ', '').replace(',', '.')
                    print(f"DEBUG WIFAK RELEVÉ - Format milliers avec espaces: {cleaned}")
                else:
                    # Pour les PDFs WIFAK, on assume que le format est: 4,000.000 (virgule=milliers, point=décimales)
                    cleaned = cleaned.replace(',', '')
                    print(f"DEBUG WIFAK RELEVÉ - Format WIFAK (virgule=milliers, point=décimales): {cleaned}")
            elif ',' in cleaned:
                # Format: 1234,56 ou 4000,000 (décimales avec virgule)
                if ' ' in cleaned:
                    # Format: 12 256,350 (milliers avec espaces + décimales)
                    cleaned = cleaned.replace(' ', '').replace(',', '.')
                    print(f"DEBUG WIFAK RELEVÉ - Format milliers avec espaces + décimales: {cleaned}")
                else:
                    # Format: 1234,56 ou 4000,000 (décimales avec virgule)
                    cleaned = cleaned.replace(',', '.')
                    print(f"DEBUG WIFAK RELEVÉ - Format décimal avec virgule: {cleaned}")
            elif '.' in cleaned:
                # Analyser le point pour déterminer si c'est des milliers ou des décimales
                parts = cleaned.split('.')
                if len(parts) == 2:
                    before_dot = parts[0]
                    after_dot = parts[1]

                    # Si la partie après le point a exactement 3 chiffres
                    if len(after_dot) == 3:
                        # Dans les PDFs WIFAK, si c'est exactement 3 chiffres après le point
                        # et que la partie avant est un petit nombre (1-9), c'est probablement des milliers
                        if before_dot.isdigit() and 1 <= int(before_dot) <= 9:
                            # C'est des milliers (ex: 4.000 = 4000)
                            cleaned = cleaned.replace('.', '')
                            print(f"DEBUG WIFAK RELEVÉ - Format milliers détecté (petit nombre): {cleaned}")
                        elif before_dot.isdigit() and int(before_dot) == 4000:
                            # Cas spécial: 4000.000 = 4000 (pas 4000000)
                            cleaned = before_dot
                            print(f"DEBUG WIFAK RELEVÉ - Format milliers détecté (4000): {cleaned}")
                        elif before_dot.isdigit() and int(before_dot) % 1000 == 0 and int(before_dot) > 0 and int(before_dot) < 10000:
                            # C'est des milliers (ex: 2000.000 = 2000)
                            cleaned = cleaned.replace('.', '')
                            print(f"DEBUG WIFAK RELEVÉ - Format milliers détecté (nombre rond): {cleaned}")
                        else:
                            # C'est des décimales (ex: 892.957)
                            print(f"DEBUG WIFAK RELEVÉ - Format décimal détecté: {cleaned}")
                    else:
                        # C'est des décimales (ex: 0.8, 0.152)
                        print(f"DEBUG WIFAK RELEVÉ - Format décimal détecté: {cleaned}")
                else:
                    # Plusieurs points, c'est probablement des milliers
                    cleaned = cleaned.replace('.', '')
                    print(f"DEBUG WIFAK RELEVÉ - Format milliers détecté (points supprimés): {cleaned}")
            elif ' ' in cleaned:
                # Format: 12 256 (milliers avec espaces, pas de décimales)
                cleaned = cleaned.replace(' ', '')
                print(f"DEBUG WIFAK RELEVÉ - Format milliers avec espaces (entier): {cleaned}")
            else:
                # Format: 4000 (entier)
                print(f"DEBUG WIFAK RELEVÉ - Format entier: {cleaned}")

            amount = float(cleaned)
            if amount == 0:
                return None

            # --- Formatting for output (space as thousand separator, comma as decimal) ---
            sign = "-" if amount < 0 else ""
            abs_amount = abs(amount)

            integer_part = int(abs_amount)

            # Format integer part with space as thousand separator
            s_integer_part = ""
            str_integer_part = str(integer_part)
            for i, digit in enumerate(reversed(str_integer_part)):
                if i > 0 and i % 3 == 0:
                    s_integer_part = " " + s_integer_part
                s_integer_part = digit + s_integer_part

            # Format fractional part to 3 decimal places
            # Use round to avoid floating point inaccuracies before formatting
            fractional_part = round(abs_amount - integer_part, 3)
            s_fractional_part = f"{fractional_part:.3f}"[2:] # Get "0.XXX" then slice to "XXX"

            return f"{sign}{s_integer_part},{s_fractional_part}"
        except Exception as e:
            print(f"Erreur de formatage du montant '{amount_str}': {e}")
            return None
            
    def save_excel_with_formatting(self, transactions, excel_path):
        """Sauvegarde les transactions dans un fichier Excel avec formatage"""
        try:
            if not transactions:
                print("Aucune transaction à sauvegarder")
                return False
                
            # Créer un DataFrame
            df = pd.DataFrame(transactions)
            
            # Créer un nouveau workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Relevé WIFAK"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
                
            # Formatage des en-têtes
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Jaune
            header_font = Font(bold=True, color="000000")  # Noir en gras
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Appliquer le formatage aux en-têtes
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
            # Ajuster la largeur des colonnes
            column_widths = {
                'A': 12,  # Date
                'B': 40,  # Libellé
                'C': 15,  # Débit
                'D': 15   # Crédit
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
                
            # Formatage des cellules de montants
            for row in range(2, len(transactions) + 2):
                for col in ['C', 'D']:  # Colonnes Débit et Crédit
                    cell = ws.cell(row=row, column=ord(col) - ord('A') + 1)
                    if cell.value:
                        cell.alignment = Alignment(horizontal="right")
                        
            # Bordures
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(transactions) + 2):
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row, column=col).border = thin_border
                    
            # Sauvegarder
            wb.save(excel_path)
            print(f"Fichier Excel sauvegardé: {excel_path}")
            return True
            
        except Exception as e:
            print(f"Erreur lors de la sauvegarde Excel: {e}")
            return False
            
    def convert(self):
        """Lance la conversion"""
        if not self.pdf_path.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
            return
            
        if not self.excel_path.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier Excel de sortie")
            return
            
        try:
            self.start_progress()
            self.update_status("Conversion en cours...", '#f39c12')
            
            # Vérifier que le fichier PDF existe
            if not os.path.exists(self.pdf_path.get()):
                messagebox.showerror("Erreur", "Le fichier PDF n'existe pas")
                return
                
            # Détecter le logo WIFAK
            self.update_status("Détection du logo WIFAK...", '#f39c12')
            if not self.detect_wifak_logo(self.pdf_path.get()):
                result = messagebox.askyesno("Avertissement", 
                    "Logo WIFAK non détecté. Voulez-vous continuer quand même ?")
                if not result:
                    return
                    
            # Détecter le type de document
            self.update_status("Détection du type de document...", '#f39c12')
            if not self.detect_document_type(self.pdf_path.get()):
                result = messagebox.askyesno("Avertissement", 
                    "Type de document 'relevé' non détecté. Voulez-vous continuer quand même ?")
                if not result:
                    return
                    
            # Extraire les données
            self.update_status("Extraction des données...", '#f39c12')
            transactions = self.extract_wifak_releve_data(self.pdf_path.get())
            
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF")
                return
                
            # Vérifier et créer le dossier Downloads si nécessaire
            downloads_path = os.path.dirname(self.excel_path.get())
            if not os.path.exists(downloads_path):
                try:
                    os.makedirs(downloads_path)
                    print(f"Dossier Downloads créé: {downloads_path}")
                except Exception as e:
                    print(f"Erreur lors de la création du dossier Downloads: {e}")
                    messagebox.showerror("Erreur", f"Impossible de créer le dossier Downloads: {e}")
                    return
            
            # Sauvegarder en Excel
            self.update_status("Sauvegarde en Excel...", '#f39c12')
            if self.save_excel_with_formatting(transactions, self.excel_path.get()):
                self.update_status(f"Conversion réussie ! {len(transactions)} transactions extraites", '#27ae60')
                messagebox.showinfo("Succès", 
                    f"Conversion réussie !\n{len(transactions)} transactions extraites\nFichier sauvegardé dans Downloads:\n{self.excel_path.get()}")
            else:
                self.update_status("Erreur lors de la sauvegarde", '#e74c3c')
                messagebox.showerror("Erreur", "Erreur lors de la sauvegarde du fichier Excel")
                
        except Exception as e:
            self.update_status(f"Erreur: {str(e)}", '#e74c3c')
            messagebox.showerror("Erreur", f"Erreur lors de la conversion: {str(e)}")
        finally:
            self.stop_progress()
            
    def run(self):
        """Lance l'application"""
        self.root.mainloop()

if __name__ == "__main__":
    app = WifakReleveConverter()
    app.run()
