import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
from PIL import Image, ImageTk
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime as dt
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class STBExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT STB Bank PDF vers Excel")
        self.root.geometry("600x500")

        # Variables
        self.pdf_file_path = tk.StringVar()
        self.excel_filename = tk.StringVar()
        self.excel_filename.set("extrait_stb_" + datetime.now().strftime("%d%m%Y_%H%M"))
        
        self.setup_ui()
        
    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(self.root, text="Convertisseur EXTRAT STB BANK",
                              font=("Arial", 16, "bold"), fg="#1E3A8A")  # Couleur bleue STB
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(self.root, text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame(self.root)
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section sélection fichier PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAT STB Bank:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_file_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.select_pdf_file, 
                              font=("Segoe UI", 10, "bold"), bg="#1E3A8A", fg="white")
        browse_btn.pack(pady=5)
        
        # Section nom fichier Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_filename, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section conversion
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        # Boutons
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                               command=self.convert_pdf_to_excel, 
                               font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Zone de résultats
        self.result_text = tk.Text(main_frame, height=8, width=70, font=("Arial", 8))
        self.result_text.pack(fill='both', expand=True, pady=10)
        
    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF EXTRAT STB Bank",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_file_path.set(file_path)
            
    def is_stb_extrait_pdf(self, pdf_path):
        """Détection automatique STB Extrait - basée sur la sélection utilisateur"""
        # CORRECTION: Détection automatique car pas de logo/mots-clés dans le PDF
        # L'utilisateur a sélectionné STB + Extrait, donc c'est automatiquement STB Extrait
        return True
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF STB Extrait"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num in range(min(2, len(pdf.pages))):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        # Chercher des patterns d'année dans le texte
                        year_patterns = [
                            r'(\d{4})',  # Toute année 4 chiffres
                        ]
                        
                        for pattern in year_patterns:
                            matches = re.findall(pattern, text, re.IGNORECASE)
                            if matches:
                                # Prendre la première année trouvée
                                year = matches[0]
                                if 2020 <= int(year) <= 2030:  # Année raisonnable
                                    print(f"Année détectée: {year}")
                                    return year
                
                # Fallback: année actuelle
                current_year = datetime.now().year
                print(f"Année par défaut: {current_year}")
                return str(current_year)
                
        except Exception as e:
            print(f"Erreur détection année: {e}")
            return str(datetime.now().year)
    
    def extract_table_data(self, pdf_path):
        """Extrait les données STB Extrait.
        1) Essaye via extract_tables
        2) Fallback robuste via parsing du texte brut spécifique STB Extrait
        """
        try:
            # D'abord, détecter l'année depuis le PDF
            year = self.detect_year_from_pdf(pdf_path)
            
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG EXTRAT - Page {page_num + 1}")
                    
                    # Vérifier si le texte est vide (PDF scanné)
                    text = page.extract_text() or ''
                    if not text.strip() and page_num == 0:
                        # Fallback OCR pour les PDFs scannés
                        ocr_text = self._extract_text_via_ocr(pdf_path)
                        if ocr_text.strip():
                            # Traiter le texte OCR comme un seul bloc
                            return self._parse_ocr_text_extrait(ocr_text, year)
                    
                    # Extraire les tableaux de la page
                    tables = page.extract_tables()
                    print(f"DEBUG EXTRAT - {len(tables)} tableaux trouvés")
                    
                    for table_idx, table in enumerate(tables):
                        if table and len(table) > 1:  # Au moins un en-tête et des données
                            print(f"DEBUG EXTRAT - Tableau {table_idx + 1}: {len(table)} lignes")
                            
                            # Chercher l'en-tête STB Extrait
                            header_row = None
                            for i, row in enumerate(table):
                                if row and any(col for col in row if col):
                                    row_text = ' '.join([str(col) for col in row if col]).lower()
                                    print(f"DEBUG EXTRAT - Ligne {i}: {row_text}")
                                    # Mots-clés pour STB Extrait
                                    if any(keyword in row_text for keyword in ['date opération', 'date valeur', 'intitulé', 'débit', 'crédit']):
                                        header_row = i
                                        print(f"DEBUG EXTRAT - En-tête trouvé ligne {i}")
                                        break
                            
                            if header_row is not None:
                                print(f"DEBUG EXTRAT - Extraction à partir ligne {header_row + 1}")
                                # Parser les lignes de données
                                for row_idx, row in enumerate(table[header_row + 1:], header_row + 1):
                                    if row and len(row) >= 4:
                                        print(f"DEBUG EXTRAT - Ligne {row_idx}: {row}")
                                        print(f"DEBUG EXTRAT - Ligne complète analysée: {[str(cell) for cell in row]}")
                                        transaction = self.parse_stb_extrait_transaction_row(row, year)
                                        if transaction:
                                            all_transactions.append(transaction)
                                            print(f"DEBUG EXTRAT - Transaction ajoutée: {transaction}")
                                        else:
                                            print(f"DEBUG EXTRAT - Transaction ignorée")
                            else:
                                print("DEBUG EXTRAT - Aucun en-tête trouvé")

                    # CORRECTION: Parser le texte brut SEULEMENT si peu de transactions trouvées dans les tableaux
                    page_text = page.extract_text() or ""
                    if page_text and len(all_transactions) < 5:  # Seulement si moins de 5 transactions trouvées
                        text_tx = self.parse_stb_extrait_text_block(page_text, year)
                        if text_tx:
                            print(f"DEBUG EXTRAT - {len(text_tx)} transactions depuis texte brut page {page_num + 1} (fallback)")
                            all_transactions.extend(text_tx)
                    else:
                        print(f"DEBUG EXTRAT - Texte brut ignoré car {len(all_transactions)} transactions déjà trouvées dans les tableaux")
                
                # CORRECTION: Trier toutes les transactions par date décroissante
                all_transactions = self.sort_transactions_by_date_desc(all_transactions)
                
                # CORRECTION: Analyser et corriger les transactions problématiques
                all_transactions = self.analyze_and_fix_transactions(all_transactions)
                
                return all_transactions
                
        except Exception as e:
            print(f"Erreur lors de l'extraction EXTRAT: {e}")
            return []

    def parse_stb_extrait_text_block(self, text, year):
        """Parsing robuste du texte brut d'une page STB Extrait.
        - Détecte les lignes qui commencent par une date: DD/MM/YYYY
        - Identifie les montants en fin de ligne (débit et/ou crédit)
        - Agrège les lignes suivantes sans date comme continuation du libellé
        - Ignore les lignes de totaux/solde
        """
        transactions = []
        try:
            lines = [ln.rstrip() for ln in text.split("\n")]
            current = None
            # Nombre au format montant STB (ex: 2 246.000 | 10 000.000 | 1 530.304)
            amount_regex = re.compile(r"(?<!\d)(\d{1,3}(?:[ \u00A0]\d{3})*(?:\.\d+)?)(?!\d)")
            date_regex = re.compile(r"^(\d{1,2}/\d{1,2}/\d{4})")
            
            def finalize_current():
                if not current:
                    print(f"DEBUG FINALIZE - Aucune transaction en cours")
                    return
                libelle = current.get('Libellé', '').strip()
                print(f"DEBUG FINALIZE - Transaction en cours: {current}")
                
                if not libelle or any(k in libelle.upper() for k in ['TOTAUX', 'SOLDE AU', 'SOLDE', 'SOLDE FINAL', 'SOLDE PRÉCÉDENT']):
                    print(f"DEBUG FINALIZE - Transaction ignorée (libellé vide ou solde): '{libelle}'")
                    return
                
                # CORRECTION: Utiliser SEULEMENT la Date Opération, ignorer la Date Valeur
                date_to_use = current.get('DateOperation')
                if not date_to_use:
                    print(f"DEBUG FINALIZE - Transaction ignorée (pas de date opération)")
                    return
                
                # CORRECTION: Nettoyer le libellé des dates valeur qui pourraient s'y être glissées
                libelle_cleaned = self.clean_libelle_from_dates(libelle)
                
                # Montants
                debit = self.format_amount_stb(current.get('Débit', ''))
                credit = self.format_amount_stb(current.get('Crédit', ''))
                
                transaction = {
                    'Date': date_to_use,
                    'Libellé': libelle_cleaned,
                    'Débit': debit,
                    'Crédit': credit
                }
                
                print(f"DEBUG FINALIZE - Transaction créée: {transaction}")
                transactions.append(transaction)
            
            def decide_debit_credit_by_keywords(libelle_upper):
                # CORRECTION: Mots-clés pour les CRÉDITS (entrées d'argent) - STB Extrait
                credit_kw = [
                    'VIREMENT RECU', 'VERSEMENT', 'ENCAISSEMENT', 'REMBOURSEMENT', 
                    'CRÉDIT', 'CREDIT', 'CREDIT COMPTE', 'AVANCE', 'EFF N',
                    'ENCAISSEMENT EFF'
                ]
                
                # CORRECTION: Mots-clés pour les DÉBITS (sorties d'argent) - STB Extrait
                # CORRECTION CRITIQUE: COMMISSION et TVA sont TOUJOURS des débits
                debit_kw = [
                    'COMMISSION', 'FRAIS', 'AGIOS', 'PRÉLÈVEMENT', 'PRELEVEMENT',
                    'RETRAIT', 'CARTE', 'CHEQUE', 'VIREMENT EMIS', 'DÉBIT', 'DEBIT',
                    'EFFET', 'REJET', 'COTISATION', 'TVA', 'INTERET', 'PENALITE',
                    'RETARD', 'DEPASSEMENT', 'ADHESION', 'ACCEPT', 'REALISATION',
                    'REG CHEQUE', 'COM PREAVIS', 'TVA sur', 'COM ACCEPT',
                    'TVA sur COMMISSION', 'COMMISSION VIREMENT', 'COMMISSION DE REJET',
                    'COMMISSION REJET', 'COM ETUDE', 'COM ACCEPT', 'COM PREAVIS'
                ]
                
                # CORRECTION: Vérifier d'abord les débits (priorité pour éviter les erreurs de commission)
                if any(k in libelle_upper for k in debit_kw):
                    return 'debit'
                # Puis les crédits
                if any(k in libelle_upper for k in credit_kw):
                    return 'credit'
                return ''
            
            for raw_line in lines:
                line = raw_line.strip()
                if not line:
                    continue
                uline = line.upper()
                
                # DEBUG: Log de chaque ligne analysée
                print(f"DEBUG LIGNE - Analysant: '{line}'")
                
                # Ignorer les lignes de totaux/solde
                if any(kw in uline for kw in ['TOTAUX', 'SOLDE AU', 'SOLDE', 'SOLDE FINAL', 'SOLDE PRÉCÉDENT']) and not date_regex.match(line):
                    print(f"DEBUG LIGNE - Ligne de solde ignorée: '{line}'")
                    finalize_current()
                    current = None
                    continue
                
                m = date_regex.match(line)
                if m:
                    print(f"DEBUG LIGNE - NOUVELLE TRANSACTION détectée avec date: {m.group(1)}")
                    # Finaliser la transaction en cours
                    finalize_current()
                    current = {
                        'DateOperation': m.group(1),
                        'Libellé': ''
                    }
                    # Retirer la date du début
                    rest = line[m.end():].strip()
                    
                    # Extraire les montants de manière plus précise (uniquement tout à droite)
                    tail_window = 120
                    tail_segment = rest[-tail_window:]
                    amount_matches = list(amount_regex.finditer(tail_segment))
                    
                    # Garder uniquement les montants alignés à droite (zone Débit/Crédit)
                    filtered_matches = []
                    for m in amount_matches:
                        start = m.start()
                        prefix = tail_segment[:start]
                        # Compter les espaces consécutifs juste avant
                        space_run = 0
                        for ch in reversed(prefix):
                            if ch == ' ' or ch == '\u00A0':
                                space_run += 1
                            else:
                                break
                        # Position relative dans la fenêtre
                        rel = start / float(len(tail_segment)) if len(tail_segment) > 0 else 1.0
                        # Conserver uniquement si suffisamment à droite et avec séparation par espaces
                        if space_run >= 1 and rel >= 0.65:
                            filtered_matches.append(m)
                    
                    amounts = [m.group(1) for m in filtered_matches]
                    
                    # Filtrer les montants peu plausibles
                    def plausible(a):
                        a_clean = a.replace(' ', '')
                        # Accepter si contient point (décimales) ou séparateurs d'espace (milliers)
                        if ('.' in a) or (' ' in a):
                            return True
                        # Refuser les entiers simples, surtout <= 31 (typiquement des jours de date)
                        try:
                            val = int(a_clean)
                            return val > 31
                        except Exception:
                            return False
                    amounts = [a for a in amounts if plausible(a)]
                    
                    # Classification débit/crédit
                    # CORRECTION: Classification débit/crédit
                    print(f"DEBUG MONTANTS - Montants trouvés: {amounts}")
                    print(f"DEBUG MONTANTS - Ligne complète: '{line}'")
                    
                    if len(amounts) >= 2:
                        # Deux montants ou plus: classifier selon les mots-clés
                        hint = decide_debit_credit_by_keywords(uline)
                        print(f"DEBUG MONTANTS - Classification: {hint}")
                        if hint == 'debit':
                            current['Débit'] = amounts[0]
                            current['Crédit'] = amounts[1] if len(amounts) > 1 else ''
                        elif hint == 'credit':
                            current['Débit'] = amounts[1] if len(amounts) > 1 else ''
                            current['Crédit'] = amounts[0]
                        else:
                            # Analyser la position des montants dans le texte
                            current['Débit'] = amounts[0]
                            current['Crédit'] = amounts[1] if len(amounts) > 1 else ''
                    elif len(amounts) == 1:
                        # Un seul montant: classifier par mots-clés
                        amt = amounts[0]
                        hint = decide_debit_credit_by_keywords(uline)
                        print(f"DEBUG MONTANTS - Un seul montant: {amt}, Classification: {hint}")
                        if hint == 'debit':
                            current['Débit'] = amt
                            current['Crédit'] = ''
                        elif hint == 'credit':
                            current['Débit'] = ''
                            current['Crédit'] = amt
                        else:
                            # Par défaut, analyser le contexte
                            if amounts[0] in line[-20:]:  # Montant dans les 20 derniers caractères
                                current['Débit'] = amt
                                current['Crédit'] = ''
                            else:
                                current['Débit'] = ''
                                current['Crédit'] = amt
                    else:
                        print(f"DEBUG MONTANTS - Aucun montant trouvé")
                        current['Débit'] = ''
                        current['Crédit'] = ''
                    
                    # Libellé = texte sans les montants finaux
                    if amounts:
                        # Supprimer les montants détectés depuis la fin
                        for i in range(min(2, len(amounts))):
                            last_amount = amounts[-(i+1)]
                            # Rechercher à partir de la fin uniquement
                            cut_pos = rest.rfind(last_amount, max(0, len(rest) - tail_window))
                            if cut_pos != -1:
                                rest = rest[:cut_pos].rstrip()
                        label_text = rest.rstrip('-:')
                    else:
                        label_text = rest
                    current['Libellé'] = label_text.strip()
                else:
                    # Ligne de continuation: ajouter au libellé courant
                    if current:
                        if not any(kw in uline for kw in ['TOTAUX', 'SOLDE']):
                            sep = ' ' if current['Libellé'] else ''
                            current['Libellé'] = f"{current['Libellé']}{sep}{line}"
            
            # Finaliser la dernière
            finalize_current()
            
            # Éliminer les doublons de transactions
            transactions = self.remove_duplicate_transactions(transactions)
            
            # CORRECTION: Trier les transactions par date décroissante (comme le PDF)
            transactions = self.sort_transactions_by_date_desc(transactions)
            
        except Exception as e:
            print(f"Erreur parse_stb_extrait_text_block: {e}")
        return transactions
    
    def remove_duplicate_transactions(self, transactions):
        """CORRECTION: Élimine les transactions dupliquées basées sur la date, le libellé et le montant"""
        try:
            unique_transactions = []
            seen_transactions = set()
            
            for transaction in transactions:
                # CORRECTION: Créer une clé unique plus robuste
                libelle_clean = transaction.get('Libellé', '').strip()
                # Nettoyer le libellé pour la comparaison (supprimer les espaces multiples et les détails supplémentaires)
                libelle_clean = re.sub(r'\s+', ' ', libelle_clean)
                
                # CORRECTION: Supprimer les détails supplémentaires qui créent des faux doublons
                # Exemple: "VIREMENT RECU SOCIETE ALMERIA 30039074808 SOCIETE ALMERIA" 
                # devient "VIREMENT RECU SOCIETE ALMERIA"
                libelle_base = self.extract_base_libelle(libelle_clean)
                
                key = (
                    transaction.get('Date', ''),
                    libelle_base,
                    transaction.get('Débit', ''),
                    transaction.get('Crédit', '')
                )
                
                if key not in seen_transactions:
                    seen_transactions.add(key)
                    unique_transactions.append(transaction)
                    print(f"DEBUG - Transaction unique ajoutée: {transaction}")
                else:
                    print(f"DEBUG - Transaction dupliquée ignorée: {transaction}")
            
            print(f"DEBUG - {len(transactions)} transactions originales, {len(unique_transactions)} uniques")
            return unique_transactions
            
        except Exception as e:
            print(f"Erreur remove_duplicate_transactions: {e}")
            return transactions
    
    def extract_base_libelle(self, libelle):
        """CORRECTION: Extrait le libellé de base sans les détails supplémentaires qui créent des doublons"""
        try:
            if not libelle:
                return ""
            
            # CORRECTION: Supprimer les numéros de référence longs (ex: 30039074808)
            libelle = re.sub(r'\s+\d{10,}\s+', ' ', libelle)
            
            # CORRECTION: Supprimer les répétitions de noms de société
            # Exemple: "SOCIETE ALMERIA SOCIETE ALMERIA" -> "SOCIETE ALMERIA"
            libelle = re.sub(r'\b(\w+)\s+\1\b', r'\1', libelle)
            
            # CORRECTION: Supprimer les détails de bénéficiaire en fin de libellé
            # Exemple: "VIREMENT RECU SOCIETE ALMERIA 30039074808 SOCIETE ALMERIA" 
            # -> "VIREMENT RECU SOCIETE ALMERIA"
            patterns_to_remove = [
                r'\s+\d+\s+[A-Z\s]+MARKIS\s+FOOD$',  # "464 STE MEDIYAS DISTRIBUTION MARKIS FOOD"
                r'\s+\d+\s+[A-Z\s]+$',  # "480 COMPTE DE PASSAGE STEG MARKIS FOOD"
                r'\s+\d{10,}\s+[A-Z\s]+$',  # "30039074808 SOCIETE ALMERIA"
            ]
            
            for pattern in patterns_to_remove:
                libelle = re.sub(pattern, '', libelle)
            
            # Nettoyer les espaces multiples
            libelle = re.sub(r'\s+', ' ', libelle).strip()
            
            print(f"DEBUG - Libellé nettoyé: '{libelle}'")
            return libelle
            
        except Exception as e:
            print(f"Erreur extract_base_libelle: {e}")
            return libelle
    
    def sort_transactions_by_date_desc(self, transactions):
        """CORRECTION: Trie les transactions par date décroissante (comme le PDF)"""
        try:
            def parse_date(date_str):
                """Parse une date DD/MM/YYYY en objet datetime pour le tri"""
                try:
                    if not date_str:
                        return dt.min
                    # Format DD/MM/YYYY
                    return dt.strptime(date_str, '%d/%m/%Y')
                except:
                    return dt.min
            
            # Trier par date décroissante (plus récent en premier)
            sorted_transactions = sorted(transactions, 
                                       key=lambda x: parse_date(x.get('Date', '')), 
                                       reverse=True)
            
            print(f"DEBUG - Transactions triées par date décroissante")
            return sorted_transactions
            
        except Exception as e:
            print(f"Erreur sort_transactions_by_date_desc: {e}")
            return transactions
    
    def analyze_and_fix_transactions(self, transactions):
        """CORRECTION: Analyse et corrige les transactions problématiques ligne par ligne"""
        try:
            print(f"DEBUG ANALYSE - Analyse de {len(transactions)} transactions")
            
            fixed_transactions = []
            problematic_transactions = []
            
            for i, transaction in enumerate(transactions):
                print(f"DEBUG ANALYSE - Transaction {i+1}: {transaction}")
                
                # Vérifier les problèmes courants
                issues = []
                
                # 1. Vérifier si le libellé contient des dates valeur
                libelle = transaction.get('Libellé', '')
                if re.match(r'^\d{1,2}/\d{1,2}/\d{4}', libelle):
                    issues.append("Date valeur dans le libellé")
                    # Corriger en supprimant la date
                    transaction['Libellé'] = re.sub(r'^\d{1,2}/\d{1,2}/\d{4}\s+', '', libelle).strip()
                
                # 2. Vérifier si le libellé contient des informations parasites
                if 'Page' in libelle:
                    issues.append("Information de page dans le libellé")
                    transaction['Libellé'] = re.sub(r'\s+Page\s+\d+/\d+', '', libelle).strip()
                
                # 3. Vérifier la classification débit/crédit
                debit = transaction.get('Débit', '')
                credit = transaction.get('Crédit', '')
                libelle_upper = libelle.upper()
                
                # Si c'est une commission mais classée en crédit, corriger
                if any(kw in libelle_upper for kw in ['COMMISSION', 'TVA', 'FRAIS']) and credit and not debit:
                    issues.append("Commission classée en crédit au lieu de débit")
                    transaction['Débit'] = credit
                    transaction['Crédit'] = ''
                
                # 4. Vérifier si les deux colonnes ont des montants (impossible)
                if debit and credit:
                    issues.append("Montant dans les deux colonnes (impossible)")
                    # Garder seulement le débit par défaut
                    transaction['Crédit'] = ''
                
                # 5. CORRECTION: Nettoyer le libellé des détails supplémentaires qui créent des doublons
                libelle_original = transaction.get('Libellé', '')
                libelle_cleaned = self.extract_base_libelle(libelle_original)
                if libelle_cleaned != libelle_original:
                    issues.append("Libellé avec détails supplémentaires nettoyé")
                    transaction['Libellé'] = libelle_cleaned
                
                if issues:
                    problematic_transactions.append((i+1, issues, transaction))
                    print(f"DEBUG ANALYSE - Problèmes détectés: {issues}")
                
                fixed_transactions.append(transaction)
            
            if problematic_transactions:
                print(f"DEBUG ANALYSE - {len(problematic_transactions)} transactions problématiques détectées et corrigées")
                for line_num, issues, transaction in problematic_transactions:
                    print(f"DEBUG ANALYSE - Ligne {line_num}: {issues} -> {transaction}")
            else:
                print("DEBUG ANALYSE - Aucune transaction problématique détectée")
            
            return fixed_transactions
            
        except Exception as e:
            print(f"Erreur analyze_and_fix_transactions: {e}")
            return transactions
    
    def clean_libelle_from_dates(self, libelle):
        """CORRECTION: Nettoie le libellé des dates valeur et informations parasites"""
        try:
            if not libelle:
                return ""
            
            # CORRECTION: Supprimer les dates au début du libellé (dates valeur)
            # Pattern pour date DD/MM/YYYY au début
            date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}\s+'
            cleaned = re.sub(date_pattern, '', libelle)
            
            # CORRECTION: Supprimer les informations parasites (Page X/Y, etc.)
            page_pattern = r'\s+Page\s+\d+/\d+'
            cleaned = re.sub(page_pattern, '', cleaned)
            
            # CORRECTION: Supprimer les doublons de bénéficiaires
            # Pattern pour détecter les bénéficiaires répétés
            beneficiaire_pattern = r'\s+([A-Z\s]+MARKIS\s+FOOD)\s+\1'
            cleaned = re.sub(beneficiaire_pattern, r' \1', cleaned)
            
            # Supprimer les espaces multiples
            cleaned = re.sub(r'\s+', ' ', cleaned)
            
            # Supprimer les espaces en début et fin
            cleaned = cleaned.strip()
            
            print(f"DEBUG - Libellé nettoyé: '{libelle}' -> '{cleaned}'")
            return cleaned
            
        except Exception as e:
            print(f"Erreur clean_libelle_from_dates: {e}")
            return libelle
    
    def parse_stb_extrait_transaction_row(self, row, year):
        """Parse une ligne de transaction STB Extrait selon la structure: DATE OPÉRATION, DATE VALEUR, INTITULÉ, RÉFÉRENCE, BÉNÉFICIAIRE, N° DÉCLARATION, DÉBIT, CRÉDIT"""
        try:
            # Nettoyer la ligne
            cleaned_row = [str(cell).strip() if cell else "" for cell in row]
            print(f"DEBUG PARSING - Ligne nettoyée: {cleaned_row}")
            
            # Vérifier si c'est une ligne d'en-tête ou de total
            row_text = ' '.join(cleaned_row).lower()
            if any(keyword in row_text for keyword in ['date opération', 'date valeur', 'intitulé', 'débit', 'crédit', 'totaux']):
                print(f"DEBUG PARSING - Ligne d'en-tête ignorée: {row_text}")
                return None
            
            # Structure STB Extrait: DATE OPÉRATION, DATE VALEUR, INTITULÉ, RÉFÉRENCE, BÉNÉFICIAIRE, N° DÉCLARATION, DÉBIT, CRÉDIT
            if len(cleaned_row) >= 8:
                date_operation = cleaned_row[0]  # DATE OPÉRATION
                date_valeur = cleaned_row[1]      # DATE VALEUR
                intitule = cleaned_row[2]        # INTITULÉ
                reference = cleaned_row[3]       # RÉFÉRENCE
                beneficiaire = cleaned_row[4]    # BÉNÉFICIAIRE
                declaration = cleaned_row[5]      # N° DÉCLARATION
                debit = cleaned_row[6]           # DÉBIT
                credit = cleaned_row[7]           # CRÉDIT
                
                print(f"DEBUG PARSING - Date op: '{date_operation}', Date val: '{date_valeur}', Intitulé: '{intitule}'")
                print(f"DEBUG PARSING - Débit: '{debit}', Crédit: '{credit}'")
                
                # CORRECTION: Utiliser SEULEMENT la DATE OPÉRATION, ignorer la DATE VALEUR
                date_to_use = ""
                if date_operation and self.is_date_stb_extrait(date_operation):
                    date_to_use = date_operation
                    print(f"DEBUG PARSING - Utilisation DATE OPÉRATION: {date_to_use}")
                # CORRECTION: Ne plus utiliser la DATE VALEUR pour éviter les fausses transactions
                
                # Vérifier que nous avons au moins une date et un intitulé valide
                if date_to_use and intitule and len(intitule.strip()) > 2:
                    # CORRECTION: Nettoyer le libellé des dates valeur
                    libelle_cleaned = self.clean_libelle_from_dates(intitule.strip())
                    result = {
                        'Date': self.format_date_stb_extrait(date_to_use),
                        'Libellé': libelle_cleaned,
                        'Débit': self.format_amount_stb(debit),
                        'Crédit': self.format_amount_stb(credit)
                    }
                    print(f"DEBUG PARSING - Transaction créée: {result}")
                    return result
                else:
                    print(f"DEBUG PARSING - Transaction ignorée (date: '{date_to_use}', intitulé: '{intitule}')")
            else:
                print(f"DEBUG PARSING - Pas assez de colonnes: {len(cleaned_row)}")
            
            return None
            
        except Exception as e:
            print(f"Erreur lors du parsing de la ligne STB Extrait: {e}")
            return None
    
    def is_date_stb_extrait(self, text):
        """Vérifie si le texte est une date au format DD/MM/YYYY"""
        if not text:
            return False
        # Pattern pour date DD/MM/YYYY
        date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'
        return bool(re.match(date_pattern, text.strip()))
    
    def format_date_stb_extrait(self, date_str):
        """Formate la date STB Extrait au format DD/MM/YYYY"""
        try:
            date_str = date_str.strip()
            
            # Si DD/MM/YYYY, garder tel quel
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
                return date_str
            
            return ""
        except:
            return ""
    
    def format_amount_stb(self, amount_str):
        """CORRECTION: Formate le montant STB Extrait avec point comme séparateur décimal"""
        try:
            if not amount_str or amount_str.strip() == "":
                return ""
            
            # Nettoyer le montant
            amount_str = amount_str.strip()
            
            # CORRECTION: Format STB Extrait: Point comme séparateur décimal, espace comme séparateur de milliers
            # Exemple: "2 246.000" ou "10 000.000"
            
            # CORRECTION: Convertir les virgules en points pour les décimales
            if ',' in amount_str and '.' not in amount_str:
                # Si c'est un format avec virgule comme décimal (ex: "1,500")
                amount_str = amount_str.replace(',', '.')
            elif ',' in amount_str and '.' in amount_str:
                # Si c'est un format mixte (ex: "1,015,114" -> "1 015.114")
                # Remplacer la dernière virgule par un point (décimal)
                parts = amount_str.split(',')
                if len(parts) > 1:
                    # Garder la dernière partie comme décimales
                    decimal_part = parts[-1]
                    # Rejoindre le reste avec des espaces
                    integer_part = ' '.join(parts[:-1])
                    amount_str = f"{integer_part}.{decimal_part}"
            
            # Vérifier si c'est un montant valide
            if self.is_amount_stb(amount_str):
                return amount_str
            return ""
        except:
            return ""
    
    def is_amount_stb(self, text):
        """Vérifie si le texte est un montant STB Extrait"""
        if not text:
            return False
        # Pattern pour montant STB (avec point comme séparateur décimal)
        amount_pattern = r'^[\d\s.]+$'
        return bool(re.match(amount_pattern, text.strip())) and any(c.isdigit() for c in text)
    
    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel STB Extrait"""
        try:
            # Créer un nouveau workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "STB Extrait"  # Nom de la feuille
            
            # Définir les styles
            header_font = Font(name='Arial', size=12, bold=True, color='000000')
            header_fill = PatternFill(start_color='BFD4FF', end_color='BFD4FF', fill_type='solid')  # Bleu ciel transparent
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Écrire les en-têtes
            headers = ['Date', 'Libellé', 'Débit', 'Crédit']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Écrire les données
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Formatage spécial pour les montants (comme BIAT)
                    if col_idx in [3, 4]:  # Débit et Crédit
                        if value and str(value).strip():
                            # Garder le format texte comme BIAT
                            cell.value = str(value)
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx == 1:  # Date
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 2:  # Libellé
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 70  # Libellé
            ws.column_dimensions['C'].width = 16  # Débit
            ws.column_dimensions['D'].width = 16  # Crédit
            
            # Ajuster la hauteur des lignes
            for row in range(1, len(df) + 2):
                ws.row_dimensions[row].height = 20
            
            # Ajouter des bordures à toute la plage de données
            from openpyxl.utils import get_column_letter
            max_row = len(df) + 1
            max_col = len(headers)
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).border = border
            
            # Sauvegarder le fichier
            wb.save(excel_path)
            
        except Exception as e:
            print(f"Erreur lors du formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, engine='openpyxl')
    
    def convert_pdf_to_excel(self):
        """Fonction principale de conversion STB Extrait"""
        if not self.pdf_file_path.get():
            print("❌ Erreur: Veuillez sélectionner un fichier PDF")
            return
        
        if not self.excel_filename.get():
            print("❌ Erreur: Veuillez entrer un nom pour le fichier Excel")
            return

        self.result_text.delete(1.0, tk.END)
        
        try:
            # Détection automatique STB Extrait (basée sur la sélection utilisateur)
            if not self.is_stb_extrait_pdf(self.pdf_file_path.get()):
                print("⚠️ Attention: Ce fichier ne semble pas être un EXTRAT STB Bank. La conversion peut ne pas fonctionner correctement.")
            
            # Extraire les données
            transactions = self.extract_table_data(self.pdf_file_path.get())
            
            if not transactions:
                print("❌ Erreur: Aucune transaction trouvée dans le PDF STB Extrait")
                return
            
            # Créer le DataFrame
            df = pd.DataFrame(transactions)
            
            # Chemin de sortie dans le dossier Téléchargements
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = os.path.join(downloads_path, f"{self.excel_filename.get()}.xlsx")
            
            # Sauvegarder en Excel avec formatage
            self.save_excel_with_formatting(df, excel_path)
            
            # Afficher les résultats
            self.result_text.insert(tk.END, f"Conversion STB Extrait réussie !\n")
            self.result_text.insert(tk.END, f"Fichier créé: {excel_path}\n")
            self.result_text.insert(tk.END, f"Nombre de transactions: {len(transactions)}\n\n")
            self.result_text.insert(tk.END, "Aperçu des données:\n")
            self.result_text.insert(tk.END, df.head(10).to_string(index=False))
            
            # Message de succès dans la console
            print(f"✅ Conversion STB Extrait terminée avec succès !")
            print(f"📁 Fichier créé: {excel_path}")
            print(f"📊 Nombre de transactions: {len(transactions)}")
            print(f"🎉 Votre fichier Excel est prêt à utiliser !")
            
        except Exception as e:
            error_msg = f"Erreur lors de la conversion STB Extrait: {str(e)}"
            self.result_text.insert(tk.END, error_msg)
            
            print(f"❌ {error_msg}")
        

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            print(f"❌ Impossible de retourner à la page d'accueil: {e}")

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

    def _parse_ocr_text_extrait(self, text: str, year: int):
        """Parse le texte OCR pour extraire les transactions STB Extrait"""
        transactions = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Chercher les patterns de date et montants
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            amount_matches = re.findall(r'[\d\s.]+', line)
            
            if date_match:
                date = date_match.group(1)
                
                # Extraire le libellé (texte entre la date et les montants)
                libelle = line[date_match.end():].strip()
                
                # Nettoyer le libellé des montants
                for amount in amount_matches:
                    if re.match(r'^\d', amount.strip()):
                        libelle = libelle.replace(amount, '').strip()
                
                # Déterminer débit/crédit basé sur les montants
                debit = credit = None
                amounts = [amt.strip() for amt in amount_matches if re.match(r'^\d', amt.strip())]
                
                if len(amounts) >= 2:
                    debit = self._format_amount_stb(amounts[0])
                    credit = self._format_amount_stb(amounts[1])
                elif len(amounts) == 1:
                    # Heuristique basée sur le libellé
                    if any(keyword in libelle.lower() for keyword in ['retrait', 'paiement', 'frais', 'commission']):
                        debit = self._format_amount_stb(amounts[0])
                    else:
                        credit = self._format_amount_stb(amounts[0])
                
                if libelle:
                    transactions.append({
                        'date': date,
                        'libelle': libelle,
                        'debit': debit,
                        'credit': credit
                    })
        
        return transactions

    def _format_amount_stb(self, amount_str: str):
        """Formate un montant STB Extrait"""
        try:
            # Nettoyer le montant
            clean_amount = amount_str.replace(' ', '').replace(',', '.')
            return float(clean_amount)
        except:
            return None

def main():
    root = tk.Tk()
    app = STBExtraitConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
