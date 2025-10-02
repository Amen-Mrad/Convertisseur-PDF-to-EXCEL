import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
from PIL import Image, ImageTk
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class AttijariBankConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur Attijari Bank PDF vers Excel")
        self.root.geometry("600x500")

        # Variables
        self.pdf_file_path = tk.StringVar()
        self.excel_filename = tk.StringVar()
        self.excel_filename.set("extrait_attijari_" + datetime.now().strftime("%d%m%Y_%H%M"))
        
        self.setup_ui()
        
    def setup_ui(self):
        # Titre principal
        title_label = tk.Label(self.root, text="Convertisseur EXTRAT ATTIJARI",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(self.root, text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame(self.root)
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section sélection fichier PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAT Attijari Bank:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_file_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.select_pdf_file, bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section nom fichier Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_filename, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section conversion
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        # Boutons
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                               command=self.convert_pdf_to_excel, bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Zone de résultats
        self.result_text = tk.Text(main_frame, height=8, width=70, font=("Arial", 8))
        self.result_text.pack(fill='both', expand=True, pady=10)
        
    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF Attijari Bank",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_file_path.set(file_path)
            
    def is_attijari_bank_pdf(self, pdf_path):
        """Vérifie si le PDF est un extrait Attijari Bank - Version améliorée"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Vérifier les premières pages pour les mots-clés
                for page_num in range(min(3, len(pdf.pages))):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        text_lower = text.lower()
                        # Mots-clés pour identifier Attijari Bank (plus flexibles)
                        keywords = [
                            'attijari', 'extrait de compte', 'releve de compte',
                            'التجاري بنك', 'كشف حساب', 'attijari bank',
                            'date', 'libelle', 'debit', 'credit'
                        ]
                        found_keywords = sum(1 for keyword in keywords if keyword in text_lower)
                        # Si on trouve au moins 2 mots-clés, c'est probablement Attijari
                        if found_keywords >= 2:
                            print(f"DEBUG - PDF Attijari détecté avec {found_keywords} mots-clés")
                            return True
                return False
        except Exception as e:
            print(f"Erreur lors de la vérification du PDF: {e}")
            return False
    
    def extract_table_data(self, pdf_path):
        """Extrait les données du tableau des transactions - Version améliorée avec gestion multi-lignes"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG - Traitement page {page_num + 1}")
                    
                    # Extraire le texte brut de la page
                    page_text = page.extract_text() or ""
                    if not page_text.strip():
                        print(f"DEBUG - Page {page_num + 1} vide, passage à la suivante")
                        continue
                    
                    # CORRECTION: Parser le texte pour extraire les transactions avec gestion multi-lignes
                    try:
                        transactions = self.parse_attijari_text_improved(page_text)
                        if transactions:
                            print(f"DEBUG - {len(transactions)} transactions extraites de la page {page_num + 1}")
                            all_transactions.extend(transactions)
                        else:
                            # Fallback - analyser le texte brut si aucune transaction trouvée
                            print(f"DEBUG - Aucune transaction trouvée, analyse fallback de la page {page_num + 1}")
                            fallback_transactions = self.parse_fallback_text_attijari(page_text)
                            if fallback_transactions:
                                print(f"DEBUG - {len(fallback_transactions)} transactions trouvées en fallback")
                                all_transactions.extend(fallback_transactions)
                    except Exception as e:
                        print(f"DEBUG - Erreur parse_attijari_text_improved: {e}")
                        # Fallback simple
                        print(f"DEBUG - Utilisation du fallback simple")
                        fallback_transactions = self.parse_simple_fallback(page_text)
                        if fallback_transactions:
                            print(f"DEBUG - {len(fallback_transactions)} transactions trouvées en fallback simple")
                            all_transactions.extend(fallback_transactions)
                
                # Nettoyer et dédupliquer
                all_transactions = self.clean_and_deduplicate_transactions(all_transactions)
                print(f"DEBUG - Total final: {len(all_transactions)} transactions")
                
                return all_transactions
                
        except Exception as e:
            print(f"Erreur lors de l'extraction: {e}")
            return []
    
    def parse_attijari_text_improved(self, text):
        """Parse le texte Attijari EXTRAT pour extraire Date, Libellé, Débit, Crédit avec gestion multi-lignes"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG - {len(lines)} lignes à analyser")
            
            # DEBUG: Afficher les premières lignes pour diagnostic
            print("DEBUG - Premières 10 lignes du texte:")
            for i, line in enumerate(lines[:10]):
                print(f"  Ligne {i+1}: {line}")
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                if not line or len(line) < 5:  # Réduire la longueur minimale
                    i += 1
                    continue
                
                print(f"DEBUG LIGNE {i + 1}: {line}")
                
                # CORRECTION: Gérer les libellés multi-lignes
                # Chercher une ligne qui commence par une date
                if self.is_transaction_start_line_attijari(line):
                    print(f"DEBUG - Ligne de transaction détectée: {line}")
                    # Collecter toutes les lignes suivantes qui ne commencent pas par une date
                    full_transaction_text = line
                    j = i + 1
                    
                    while j < len(lines):
                        next_line = lines[j].strip()
                        if not next_line:
                            j += 1
                            continue
                        
                        # Si la ligne suivante commence par une date, c'est une nouvelle transaction
                        if self.is_transaction_start_line_attijari(next_line):
                            break
                        
                        # Ajouter la ligne au libellé
                        full_transaction_text += " " + next_line
                        j += 1
                    
                    print(f"DEBUG - Texte complet de transaction: {full_transaction_text}")
                    
                    # Extraire la transaction du texte complet
                    transaction_match = self.extract_transaction_from_line_attijari(full_transaction_text)
                    
                    if transaction_match:
                        transactions.append(transaction_match)
                        print(f"DEBUG - Transaction extraite: {transaction_match}")
                    else:
                        print(f"DEBUG - Aucune transaction extraite de: {full_transaction_text}")
                    
                    i = j  # Passer à la ligne suivante
                else:
                    print(f"DEBUG - Ligne ignorée (pas de date): {line}")
                    i += 1
            
            print(f"DEBUG - Total transactions trouvées: {len(transactions)}")
            return transactions
            
        except Exception as e:
            print(f"Erreur parse_attijari_text_improved: {e}")
            return []
    
    def is_transaction_start_line_attijari(self, line):
        """Vérifie si une ligne commence une nouvelle transaction Attijari"""
        try:
            # Pattern pour détecter les dates au début de ligne
            date_patterns = [
                r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{4})',  # DD/MM/YYYY ou DD MM YYYY
                r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{2})',   # DD/MM/YY ou DD MM YY
                r'^(\d{1,2}\s+\d{1,2})',               # DD MM (sans année)
                r'^(\d{1,2}\s+\d{1,2}\s+\d{4})',      # DD MM YYYY
                r'^(\d{1,2}/\d{1,2}/\d{4})',          # DD/MM/YYYY
                r'^(\d{1,2}/\d{1,2}/\d{2})',          # DD/MM/YY
            ]
            
            for pattern in date_patterns:
                if re.match(pattern, line):
                    print(f"DEBUG - Date détectée avec pattern {pattern}: {line}")
                    return True
            return False
        except Exception as e:
            print(f"DEBUG - Erreur is_transaction_start_line_attijari: {e}")
            return False
    
    def extract_transaction_from_line_attijari(self, line):
        """Extrait une transaction d'une ligne de texte Attijari EXTRAT"""
        try:
            print(f"DEBUG - Analyse de la ligne: {line}")
            
            # Pattern plus flexible pour détecter les dates
            date_patterns = [
                r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{4})',  # DD/MM/YYYY ou DD MM YYYY
                r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{2})',   # DD/MM/YY ou DD MM YY
                r'^(\d{1,2}\s+\d{1,2})',               # DD MM (sans année)
                r'^(\d{1,2}\s+\d{1,2}\s+\d{4})',      # DD MM YYYY
                r'^(\d{1,2}/\d{1,2}/\d{4})',          # DD/MM/YYYY
                r'^(\d{1,2}/\d{1,2}/\d{2})',          # DD/MM/YY
            ]
            
            date_match = None
            for pattern in date_patterns:
                date_match = re.match(pattern, line)
                if date_match:
                    print(f"DEBUG - Date trouvée avec pattern {pattern}: {date_match.group(1)}")
                    break
            
            if not date_match:
                print(f"DEBUG - Aucune date trouvée dans: {line}")
                return None
            
            date_str = date_match.group(1).replace(' ', '/')
            print(f"DEBUG - Date formatée: {date_str}")
            
            # Extraire le reste de la ligne après la date
            rest_line = line[date_match.end():].strip()
            print(f"DEBUG - Reste de ligne: {rest_line}")
            
            # Chercher les montants avec patterns flexibles
            amount_patterns = [
                r'([\d\s,\.]+)',           # Format général
                r'(\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?)',  # Format avec séparateurs
                r'(\d+(?:[.,]\d+)?)',      # Format simple
            ]
            
            amount_matches = []
            for pattern in amount_patterns:
                matches = re.findall(pattern, rest_line)
                amount_matches.extend(matches)
            
            print(f"DEBUG - Montants trouvés: {amount_matches}")
            
            if not amount_matches:
                print(f"DEBUG - Aucun montant trouvé dans: {rest_line}")
                return None
            
            # Filtrer les montants plausibles
            plausible_amounts = []
            for amount in amount_matches:
                amount_clean = amount.replace(' ', '').replace(',', '.')
                try:
                    value = float(amount_clean)
                    if 0.01 < value < 1000000:  # Ignorer les soldes et très petits montants
                        plausible_amounts.append(amount.strip())
                        print(f"DEBUG - Montant plausible: {amount} (valeur: {value})")
                except:
                    print(f"DEBUG - Montant non valide: {amount}")
                    pass
            
            if not plausible_amounts:
                print(f"DEBUG - Aucun montant plausible trouvé")
                return None
            
            # Prendre le dernier montant trouvé
            amount_str = plausible_amounts[-1]
            print(f"DEBUG - Montant sélectionné: {amount_str}")
            
            # Extraire le libellé en supprimant TOUS les montants
            libelle = rest_line
            for amount in plausible_amounts:
                libelle = libelle.replace(amount, '').strip()
            
            print(f"DEBUG - Libellé avant nettoyage: {libelle}")
            
            # Nettoyer le libellé
            libelle = self.clean_libelle_from_dates_attijari(libelle)
            libelle = re.sub(r'\s+', ' ', libelle)
            
            print(f"DEBUG - Libellé après nettoyage: {libelle}")
            
            # Classifier le montant
            debit, credit = self.classify_amount_attijari(libelle, amount_str)
            
            print(f"DEBUG - Classification: Débit={debit}, Crédit={credit}")
            
            # Créer la transaction
            transaction = {
                'Date': date_str,
                'Libellé': libelle,
                'Débit': debit,
                'Crédit': credit
            }
            
            print(f"DEBUG - Transaction créée: {transaction}")
            return transaction
            
        except Exception as e:
            print(f"Erreur extract_transaction_from_line_attijari: {e}")
            return None
    
    def classify_amount_attijari(self, libelle, amount_str):
        """Classifie un montant comme débit ou crédit basé sur le libellé Attijari"""
        try:
            libelle_upper = libelle.upper()
            
            # Mots-clés pour les DÉBITS (sorties d'argent)
            debit_keywords = [
                'COM & TVA', 'REJET', 'REGLEMENT CHEQUE', 'ABONNEMENT', 'AGIOS',
                'COMMISSION', 'COM DEPASSEMENT', 'FRAIS', 'RETRAIT', 'CARTE',
                'PRÉLÈVEMENT', 'PRELEVEMENT', 'COTISATION', 'TVA', 'PRLV',
                'DÉBIT', 'DEBIT', 'VIR EMIS', 'VIREMENT EMIS', 'FRAIS PHOTO',
                'FRAIS HUISSIER', 'REGL HUISS', 'PACK BUSINESS', 'SOLUTIONS',
                'DONT TVA', 'MEME BQ', 'DEPASSEMENT PONCTUEL', 'REGLEMENT PRELEVEMENT',
                'COMMISSION PRELEVEMENT', 'ABONNEMENT OFFRE'
            ]
            
            # Mots-clés pour les CRÉDITS (entrées d'argent)
            credit_keywords = [
                'VIR RECU', 'VIREMENT RECU', 'ENCAISSEMENT', 'VERSEMENT', 
                'REMBOURSEMENT', 'CRÉDIT', 'CREDIT', 'VIR. RECU',
                'VIREMENT REÇU', 'RECU', 'REÇU', 'VIR RECU TN',
                'VIR RECU TN MEME', 'VERSEMENT ESPECE', 'ENCAISSEMENT CHEQUE',
                'REMBOURSEMENT FRAIS', 'CREDIT COMPTE', 'AVANCE', 'FACT'
            ]
            
            # Vérifier d'abord les débits
            if any(keyword in libelle_upper for keyword in debit_keywords):
                print(f"DEBUG - Classification DÉBIT: {libelle}")
                return self.format_amount_attijari(amount_str), ""
            
            # Puis les crédits
            if any(keyword in libelle_upper for keyword in credit_keywords):
                print(f"DEBUG - Classification CRÉDIT: {libelle}")
                return "", self.format_amount_attijari(amount_str)
            
            # Par défaut, analyser le contexte
            if any(word in libelle_upper for word in ['PAIEMENT', 'FRAIS', 'COM', 'AGIOS']):
                print(f"DEBUG - Classification DÉBIT (contexte): {libelle}")
                return self.format_amount_attijari(amount_str), ""
            else:
                print(f"DEBUG - Classification CRÉDIT (défaut): {libelle}")
                return "", self.format_amount_attijari(amount_str)
                
        except Exception as e:
            print(f"Erreur classify_amount_attijari: {e}")
            return "", ""
    
    def clean_libelle_from_dates_attijari(self, libelle):
        """Supprime les dates valeur du début du libellé et les montants SOLDE de la fin - Version Attijari"""
        try:
            if not libelle:
                return libelle
            
            # Supprimer les dates au début
            date_pattern = r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{4})\s*'
            libelle_cleaned = re.sub(date_pattern, '', libelle).strip()
            
            # CORRECTION: Supprimer TOUS les montants de la fin (plus agressif)
            amount_patterns = [
                r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Montants simples
                r'\s*\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?-\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Format montant-balance
                r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?-\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Format négatif-balance
                r'\s*\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Montants positifs simples
            ]
            
            for pattern in amount_patterns:
                libelle_cleaned = re.sub(pattern, '', libelle_cleaned).strip()
            
            # CORRECTION: Supprimer aussi les fragments de montants restants
            while re.search(r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$', libelle_cleaned):
                libelle_cleaned = re.sub(r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$', '', libelle_cleaned).strip()
            
            print(f"DEBUG - Libellé avant nettoyage: '{libelle}'")
            print(f"DEBUG - Libellé après nettoyage: '{libelle_cleaned}'")
            
            return libelle_cleaned
            
        except Exception as e:
            print(f"Erreur clean_libelle_from_dates_attijari: {e}")
            return libelle
    
    def format_amount_attijari(self, amount_str):
        """Formate le montant avec virgule comme séparateur décimal - Version Attijari"""
        try:
            if not amount_str or amount_str.strip() == "":
                return ""
            
            # Nettoyer le montant
            amount_str = amount_str.strip()
            cleaned = amount_str.replace(" ", "")
            
            # Si le point est utilisé comme séparateur décimal, le remplacer par virgule
            if "." in cleaned and "," not in cleaned:
                parts = cleaned.split(".")
                if len(parts) == 2 and len(parts[1]) <= 3:  # Probablement décimal
                    cleaned = cleaned.replace(".", ",")
            elif "," in cleaned and "." in cleaned:
                # Format "1,234.56" -> "1234,56"
                cleaned = cleaned.replace(",", "").replace(".", ",")
            
            # Vérifier que c'est un montant valide
            try:
                test_value = float(cleaned.replace(",", "."))
                if test_value > 0:
                    return cleaned
            except:
                pass
            
            return ""
        except:
            return ""
    
    def parse_simple_fallback(self, text):
        """Fonction de fallback simple pour analyser le texte brut Attijari"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG SIMPLE FALLBACK - {len(lines)} lignes à analyser")
            
            # Afficher les premières lignes pour diagnostic
            print("DEBUG SIMPLE FALLBACK - Premières 20 lignes:")
            for i, line in enumerate(lines[:20]):
                print(f"  Ligne {i+1}: {line}")
            
            for line_num, line in enumerate(lines):
                line = line.strip()
                if not line or len(line) < 5:
                    continue
                
                print(f"DEBUG SIMPLE FALLBACK - Ligne {line_num + 1}: {line}")
                
                # Chercher des patterns de transaction simples
                # Pattern: Date + Description + Montant
                transaction_patterns = [
                    r'(\d{1,2}[/\s]\d{1,2}[/\s]?\d{0,4})\s+(.+?)\s+([\d\s,\.]+)$',  # Date + description + montant
                    r'(\d{1,2}[/\s]\d{1,2})\s+(.+?)\s+([\d\s,\.]+)$',              # Date + description + montant (sans année)
                    r'(\d{1,2}\s+\d{1,2})\s+(.+?)\s+([\d\s,\.]+)$',               # Date + description + montant (format DD MM)
                ]
                
                for pattern in transaction_patterns:
                    match = re.search(pattern, line)
                    if match:
                        print(f"DEBUG SIMPLE FALLBACK - Pattern match: {pattern}")
                        date_part = match.group(1)
                        description = match.group(2)
                        amount = match.group(3)
                        
                        print(f"DEBUG SIMPLE FALLBACK - Date: {date_part}, Description: {description}, Montant: {amount}")
                        
                        # Nettoyer la date
                        if len(date_part.split('/')) == 2 or len(date_part.split(' ')) == 2:
                            date_str = f"{date_part}/{datetime.now().year}" if '/' in date_part else f"{date_part.replace(' ', '/')}/{datetime.now().year}"
                        else:
                            date_str = date_part.replace(' ', '/')
                        
                        # Nettoyer la description
                        description = re.sub(r'\s+', ' ', description).strip()
                        
                        # Classifier le montant
                        debit, credit = self.classify_amount_attijari(description, amount)
                        
                        if debit or credit:
                            transaction = {
                                'Date': date_str,
                                'Libellé': description,
                                'Débit': debit,
                                'Crédit': credit
                            }
                            transactions.append(transaction)
                            print(f"DEBUG SIMPLE FALLBACK - Transaction créée: {transaction}")
                        break
            
            print(f"DEBUG SIMPLE FALLBACK - Total transactions: {len(transactions)}")
            return transactions
            
        except Exception as e:
            print(f"Erreur parse_simple_fallback: {e}")
            return []
    
    def parse_fallback_text_attijari(self, text):
        """Fonction de fallback pour analyser le texte brut Attijari si aucune transaction n'est trouvée"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG FALLBACK - {len(lines)} lignes à analyser")
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                if not line or len(line) < 10:
                    i += 1
                    continue
                
                # CORRECTION: Gérer les libellés multi-lignes dans le fallback aussi
                if self.is_transaction_start_line_attijari(line):
                    # Collecter toutes les lignes suivantes qui ne commencent pas par une date
                    full_transaction_text = line
                    j = i + 1
                    
                    while j < len(lines):
                        next_line = lines[j].strip()
                        if not next_line:
                            j += 1
                            continue
                        
                        # Si la ligne suivante commence par une date, c'est une nouvelle transaction
                        if self.is_transaction_start_line_attijari(next_line):
                            break
                        
                        # Ajouter la ligne au libellé
                        full_transaction_text += " " + next_line
                        j += 1
                    
                    print(f"DEBUG FALLBACK - Texte complet: {full_transaction_text}")
                    
                    # Chercher des patterns de transaction plus larges
                    transaction_patterns = [
                        r'(\d{1,2}[/\s]\d{1,2}[/\s]?\d{0,4})\s+(.+?)\s+([\d\s,\.]+)$',  # Date + description + montant
                        r'(\d{1,2}[/\s]\d{1,2})\s+(.+?)\s+([\d\s,\.]+)$',              # Date + description + montant (sans année)
                    ]
                    
                    for pattern in transaction_patterns:
                        match = re.search(pattern, full_transaction_text)
                        if match:
                            date_part = match.group(1)
                            description = match.group(2)
                            amount = match.group(3)
                            
                            # Nettoyer la date
                            if len(date_part.split('/')) == 2 or len(date_part.split(' ')) == 2:
                                date_str = f"{date_part}/{datetime.now().year}" if '/' in date_part else f"{date_part.replace(' ', '/')}/{datetime.now().year}"
                            else:
                                date_str = date_part.replace(' ', '/')
                            
                            # Nettoyer la description
                            description = self.clean_libelle_from_dates_attijari(description)
                            description = re.sub(r'\s+', ' ', description).strip()
                            
                            # Classifier le montant
                            debit, credit = self.classify_amount_attijari(description, amount)
                            
                            if debit or credit:
                                transaction = {
                                    'Date': date_str,
                                    'Libellé': description,
                                    'Débit': debit,
                                    'Crédit': credit
                                }
                                transactions.append(transaction)
                                print(f"DEBUG FALLBACK - Transaction créée: {transaction}")
                            break
                    
                    i = j  # Passer à la ligne suivante
                else:
                    i += 1
            
            return transactions
            
        except Exception as e:
            print(f"Erreur parse_fallback_text_attijari: {e}")
            return []
    
    def parse_text_transactions(self, text):
        """Parse les transactions depuis le texte brut - Format Attijari RELEVE amélioré"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG TEXT - {len(lines)} lignes à analyser")
            
            # Détecter l'année depuis le texte
            year = self.detect_year_from_text(text)
            print(f"DEBUG TEXT - Année détectée: {year}")
            
            current_transaction = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Ignorer les lignes de totaux/solde
                if any(keyword in line.upper() for keyword in ['TOTAUX', 'SOLDE AU', 'SOLDE']):
                    if current_transaction:
                        transactions.append(current_transaction)
                        current_transaction = None
                    continue
                
                # Chercher une ligne qui commence par une date au format DD MM
                date_match = re.match(r'^(\d{1,2})\s+(\d{1,2})\s+(.+)', line)
                if date_match:
                    # Finaliser la transaction précédente
                    if current_transaction:
                        transactions.append(current_transaction)
                    
                    day = date_match.group(1)
                    month = date_match.group(2)
                    rest = date_match.group(3)
                    
                    # Chercher la date valeur dans le reste (format DD MM YYYY)
                    date_val_match = re.search(r'(\d{1,2}\s+\d{1,2}\s+\d{4})', rest)
                    date_valeur = ""
                    if date_val_match:
                        date_valeur = date_val_match.group(1)
                        # Retirer la date valeur du reste
                        rest = rest.replace(date_valeur, '').strip()
                    
                    # PARSER LES MONTANTS SELON LE FORMAT ATTIJARI - VERSION AMÉLIORÉE
                    # Le format est: LIBELLE [MONTANT_DEBIT] [MONTANT_CREDIT]
                    # Chercher les montants à la fin de la ligne - PATTERN PLUS PRÉCIS
                    
                    # Pattern pour montants avec virgule décimale (format Attijari)
                    amount_pattern = r'(\d{1,3}(?:\s+\d{3})*(?:,\d{1,3})?)'
                    amounts = re.findall(amount_pattern, rest)
                    
                    # Si aucun montant trouvé, essayer un pattern plus large
                    if not amounts:
                        # Pattern plus large pour capturer tous les montants
                        amount_pattern_large = r'(\d+(?:[\s,]\d+)*(?:,\d+)?)'
                        amounts = re.findall(amount_pattern_large, rest)
                        print(f"DEBUG TEXT - Pattern large: {amounts}")
                    
                    print(f"DEBUG TEXT - Montants trouvés: {amounts}")
                    
                    # Extraire le libellé (tout sauf les montants)
                    libelle = rest
                    
                    # Filtrer les montants valides (ignorer les dates et petits nombres)
                    valid_amounts = []
                    for amount in amounts:
                        # Ignorer les petits nombres qui sont probablement des dates
                        try:
                            amount_value = float(amount.replace(' ', '').replace(',', '.'))
                            # Garder seulement les montants > 1 (ignorer les dates comme "30", "09", etc.)
                            # Et qui ont une virgule (décimales) ou sont > 50
                            if amount_value > 1 and (',' in amount or amount_value > 50):
                                valid_amounts.append(amount)
                        except:
                            # Si pas un nombre, garder seulement si assez long
                            if len(amount) > 2:  # Ignorer les très courts
                                valid_amounts.append(amount)
                    
                    print(f"DEBUG TEXT - Montants valides: {valid_amounts}")
                    print(f"DEBUG TEXT - Libellé AVANT retrait: '{libelle}'")
                    
                    # Retirer les montants valides du libellé - VERSION ULTRA AMÉLIORÉE
                    for amount in valid_amounts:
                        # Retirer le montant du libellé avec plusieurs méthodes
                        original_libelle = libelle
                        
                        # Méthode 1: Retrait direct
                        libelle = libelle.replace(amount, '').strip()
                        
                        # Méthode 2: Si pas de changement, essayer avec espaces
                        if libelle == original_libelle:
                            libelle = libelle.replace(f" {amount}", '').strip()
                            libelle = libelle.replace(f"{amount} ", '').strip()
                        
                        # Méthode 3: Si toujours pas de changement, essayer avec regex
                        if libelle == original_libelle:
                            # Échapper les caractères spéciaux pour regex
                            escaped_amount = re.escape(amount)
                            libelle = re.sub(escaped_amount, '', libelle).strip()
                        
                        # Nettoyer les espaces multiples après chaque retrait
                        libelle = re.sub(r'\s+', ' ', libelle).strip()
                    
                    # Nettoyer le libellé final - supprimer TOUS les caractères indésirables
                    libelle = re.sub(r'\s+', ' ', libelle).strip()
                    libelle = re.sub(r'[,\s]+$', '', libelle)  # Supprimer les virgules et espaces en fin
                    libelle = re.sub(r'^\s*[,.\s]+', '', libelle)  # Supprimer les virgules et points en début
                    libelle = re.sub(r'\s+$', '', libelle)  # Supprimer les espaces en fin
                    libelle = re.sub(r'^\s+', '', libelle)  # Supprimer les espaces en début
                    libelle = re.sub(r'[,\s]*$', '', libelle)  # Supprimer les virgules et espaces en fin
                    libelle = re.sub(r'^\s*[,.\s]*', '', libelle)  # Supprimer les virgules et points en début
                    
                    print(f"DEBUG TEXT - Libellé APRÈS retrait: '{libelle}'")
                    print(f"DEBUG TEXT - Montants extraits: {valid_amounts}")
                    
                    # Utiliser les montants valides pour la classification
                    amounts = valid_amounts
                    
                    # LOGIQUE DE CLASSIFICATION DES MONTANTS AMÉLIORÉE
                    debit = ""
                    credit = ""
                    
                    if len(amounts) == 2:
                        # 2 montants: classifier selon le libellé et la valeur
                        libelle_upper = libelle.upper()
                        
                        # Mots-clés pour DÉBIT - VERSION ÉTENDUE
                        debit_keywords = [
                            'AGIOS', 'FRAIS', 'COMMISSION', 'ABONNEMENT', 
                            'COM DEPASSEMENT', 'REGLEMENT', 'PRELEVEMENT', 'TVA',
                            'PRODUCTIONS', 'COTISATION', 'PRLV', 'FRAIS PHOTO',
                            'FRAIS HUISSIER', 'COMMISSION VIR', 'ABONNEMENT PACK',
                            'COM ET TVA', 'COMMISSION PRELEVEMENT', 'RETRAIT',
                            'DONT TVA', 'COM & TVA', 'FRAIS HUISSIER',
                            'COMMISSION PRELEVEMENT', 'ABONNEMENT OFFRE'
                        ]
                        
                        # Mots-clés pour CRÉDIT - VERSION ÉTENDUE
                        credit_keywords = [
                            'VIR RECU', 'ENCAISSEMENT', 'VERSEMENT', 'REMBOURSEMENT',
                            'VIREMENT RECU', 'RECEPTION', 'VIR RECU TN', 'VERSEMENT ESPECE',
                            'ENCAISSEMENT CHEQUE', 'REMBOURSEMENT FRAIS', 'VIR RECU TN AUT',
                            'C.F.A', 'ATFP'
                        ]
                        
                        # Analyser le libellé pour déterminer la classification
                        if any(kw in libelle_upper for kw in debit_keywords):
                            # Transaction de débit - premier montant = débit, second = crédit
                            debit = self.format_amount(amounts[0])
                            credit = self.format_amount(amounts[1])
                        elif any(kw in libelle_upper for kw in credit_keywords):
                            # Transaction de crédit - premier montant = crédit, second = débit
                            credit = self.format_amount(amounts[0])
                            debit = self.format_amount(amounts[1])
                        else:
                            # Par défaut: analyser les valeurs
                            try:
                                val1 = float(amounts[0].replace(' ', '').replace(',', '.'))
                                val2 = float(amounts[1].replace(' ', '').replace(',', '.'))
                                # Le plus petit montant est souvent un débit (frais), le plus grand un crédit
                                if val1 < val2:
                                    debit = self.format_amount(amounts[0])
                                    credit = self.format_amount(amounts[1])
                                else:
                                    debit = self.format_amount(amounts[1])
                                    credit = self.format_amount(amounts[0])
                            except:
                                # En cas d'erreur, premier = débit, second = crédit
                                debit = self.format_amount(amounts[0])
                                credit = self.format_amount(amounts[1])
                    elif len(amounts) == 1:
                        # Un seul montant: déterminer selon le libellé
                        libelle_upper = libelle.upper()
                        amount = amounts[0]
                        
                        # Mots-clés pour DÉBIT - VERSION ÉTENDUE
                        debit_keywords = [
                            'AGIOS', 'FRAIS', 'COMMISSION', 'ABONNEMENT', 
                            'COM DEPASSEMENT', 'REGLEMENT', 'PRELEVEMENT', 'TVA',
                            'PRODUCTIONS', 'COTISATION', 'PRLV', 'FRAIS PHOTO',
                            'FRAIS HUISSIER', 'COMMISSION VIR', 'ABONNEMENT PACK',
                            'COM ET TVA', 'COMMISSION PRELEVEMENT', 'RETRAIT',
                            'DONT TVA', 'COM & TVA', 'FRAIS HUISSIER',
                            'COMMISSION PRELEVEMENT', 'ABONNEMENT OFFRE'
                        ]
                        
                        # Mots-clés pour CRÉDIT - VERSION ÉTENDUE
                        credit_keywords = [
                            'VIR RECU', 'ENCAISSEMENT', 'VERSEMENT', 'REMBOURSEMENT',
                            'VIREMENT RECU', 'RECEPTION', 'VIR RECU TN', 'VERSEMENT ESPECE',
                            'ENCAISSEMENT CHEQUE', 'REMBOURSEMENT FRAIS', 'VIR RECU TN AUT',
                            'C.F.A', 'ATFP'
                        ]
                        
                        if any(kw in libelle_upper for kw in debit_keywords):
                            debit = self.format_amount(amount)
                        elif any(kw in libelle_upper for kw in credit_keywords):
                            credit = self.format_amount(amount)
                        else:
                            # Par défaut, analyser le montant et le libellé
                            try:
                                amount_value = float(amount.replace(' ', '').replace(',', '.'))
                                # Si c'est un petit montant (< 100), probablement débit (frais)
                                if amount_value < 100:
                                    debit = self.format_amount(amount)
                                else:
                                    # Analyser le libellé pour des indices supplémentaires
                                    if any(word in libelle_upper for word in ['FRAIS', 'COMMISSION', 'AGIOS', 'TVA']):
                                        debit = self.format_amount(amount)
                                    else:
                                        credit = self.format_amount(amount)
                            except:
                                # En cas d'erreur, analyser le libellé
                                if any(word in libelle_upper for word in ['FRAIS', 'COMMISSION', 'AGIOS', 'TVA']):
                                    debit = self.format_amount(amount)
                                else:
                                    credit = self.format_amount(amount)
                    
                    # Créer la transaction
                    current_transaction = {
                        'Date': f"{day.zfill(2)}/{month.zfill(2)}/{year}",
                        'Libellé': libelle,
                        'Débit': debit,
                        'Crédit': credit
                    }
                    print(f"DEBUG TEXT - Nouvelle transaction: {current_transaction}")
                
                elif current_transaction and line and not re.match(r'^\d', line):
                    # Ligne de continuation du libellé
                    current_transaction['Libellé'] += f" {line}"
                    print(f"DEBUG TEXT - Libellé étendu: {current_transaction['Libellé']}")
            
            # Finaliser la dernière transaction
            if current_transaction:
                transactions.append(current_transaction)
            
            print(f"DEBUG TEXT - {len(transactions)} transactions extraites")
            return transactions
            
        except Exception as e:
            print(f"Erreur parsing texte: {e}")
            return []
    
    def detect_year_from_text(self, text):
        """Détecte l'année depuis le texte"""
        try:
            # Chercher des patterns d'année
            year_patterns = [
                r'au\s*:\s*\d{1,2}/\d{1,2}/(\d{4})',  # "Au : 31/07/2025"
                r'(\d{4})',  # Toute année 4 chiffres
            ]
            
            for pattern in year_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    year = matches[0]
                    if 2020 <= int(year) <= 2030:  # Année raisonnable
                        return year
            
            # Fallback: année actuelle
            return str(datetime.now().year)
            
        except Exception:
            return str(datetime.now().year)

    def parse_transaction_row(self, row):
        """Parse une ligne de transaction et extrait les 4 colonnes nécessaires - Version améliorée"""
        try:
            # Nettoyer la ligne
            cleaned_row = [str(cell).strip() if cell else "" for cell in row]
            print(f"DEBUG ROW - Ligne: {cleaned_row}")
            
            # Trouver les colonnes par position et contenu
            date = ""
            libelle = ""
            debit = ""
            credit = ""
            
            # La première colonne contient généralement la date
            for i, cell in enumerate(cleaned_row):
                if self.is_date(cell):
                    date = cell
                    print(f"DEBUG ROW - Date trouvée: {date}")
                    break
            
            # Chercher les montants (débit et crédit) en ignorant la colonne "Solde"
            amounts_found = []
            for i, cell in enumerate(cleaned_row):
                if self.is_amount(cell) and not self.is_balance_column(cell, i, cleaned_row):
                    amounts_found.append((i, cell))
                    print(f"DEBUG ROW - Montant trouvé: {cell} à la position {i}")
            
            # CORRECTION: Séparer débit et crédit basé sur la position ET le libellé
            if len(amounts_found) >= 2:
                # Généralement: Date, Valeur, Libellé, Débit, Crédit, Solde
                # Analyser le libellé pour mieux classifier
                libelle_upper = libelle.upper()
                
                # Mots-clés pour DÉBIT (sorties d'argent) - AMÉLIORÉ
                debit_keywords = [
                    'RETRAIT', 'FRAIS', 'COMMISSION', 'AGIOS', 'ABONNEMENT', 
                    'COM DEPASSEMENT', 'REGLEMENT', 'PRELEVEMENT', 'TVA',
                    'PRODUCTIONS', 'COTISATION', 'PRLV', 'FRAIS PHOTO',
                    'FRAIS HUISSIER', 'COMMISSION VIR', 'ABONNEMENT PACK',
                    'COM ET TVA', 'COMMISSION PRELEVEMENT', 'VIR EMIS', 'VIREMENT EMIS'
                ]
                
                # Mots-clés pour CRÉDIT (entrées d'argent) - AMÉLIORÉ
                credit_keywords = [
                    'VIR RECU', 'ENCAISSEMENT', 'VERSEMENT', 'REMBOURSEMENT',
                    'VIREMENT RECU', 'RECEPTION', 'VIR RECU TN', 'VERSEMENT ESPECE',
                    'ENCAISSEMENT CHEQUE', 'REMBOURSEMENT FRAIS', 'VIR. RECU',
                    'VIREMENT REÇU', 'RECU', 'REÇU', 'VIR RECU TN MEME'
                ]
                
                # CORRECTION: Classification améliorée
                if any(kw in libelle_upper for kw in debit_keywords):
                    # Transaction de débit - premier montant = débit
                    debit = amounts_found[0][1]
                    credit = amounts_found[1][1] if len(amounts_found) > 1 else ""
                    print(f"DEBUG ROW - Classification DÉBIT: {libelle} -> {debit}")
                elif any(kw in libelle_upper for kw in credit_keywords):
                    # Transaction de crédit - premier montant = crédit
                    credit = amounts_found[0][1]
                    debit = amounts_found[1][1] if len(amounts_found) > 1 else ""
                    print(f"DEBUG ROW - Classification CRÉDIT: {libelle} -> {credit}")
                else:
                    # CORRECTION: Par défaut, analyser la position des montants
                    # Le premier montant est généralement le débit, le second le crédit
                    debit = amounts_found[0][1]
                    credit = amounts_found[1][1] if len(amounts_found) > 1 else ""
                    print(f"DEBUG ROW - Classification par défaut: Débit={debit}, Crédit={credit}")
            elif len(amounts_found) == 1:
                # CORRECTION: Une seule transaction - classifier selon le libellé
                pos, amount = amounts_found[0]
                libelle_upper = libelle.upper()
                
                # Mots-clés pour DÉBIT (sorties d'argent) - AMÉLIORÉ
                debit_keywords = [
                    'RETRAIT', 'FRAIS', 'COMMISSION', 'AGIOS', 'ABONNEMENT', 
                    'COM DEPASSEMENT', 'REGLEMENT', 'PRELEVEMENT', 'TVA',
                    'PRODUCTIONS', 'COTISATION', 'PRLV', 'FRAIS PHOTO',
                    'FRAIS HUISSIER', 'COMMISSION VIR', 'ABONNEMENT PACK',
                    'COM ET TVA', 'COMMISSION PRELEVEMENT', 'VIR EMIS', 'VIREMENT EMIS'
                ]
                
                # Mots-clés pour CRÉDIT (entrées d'argent) - AMÉLIORÉ
                credit_keywords = [
                    'VIR RECU', 'ENCAISSEMENT', 'VERSEMENT', 'REMBOURSEMENT',
                    'VIREMENT RECU', 'RECEPTION', 'VIR RECU TN', 'VERSEMENT ESPECE',
                    'ENCAISSEMENT CHEQUE', 'REMBOURSEMENT FRAIS', 'VIR. RECU',
                    'VIREMENT REÇU', 'RECU', 'REÇU', 'VIR RECU TN MEME'
                ]
                
                # CORRECTION: Classification améliorée pour un seul montant
                if any(kw in libelle_upper for kw in debit_keywords):
                    debit = amount
                    print(f"DEBUG ROW - Classification DÉBIT (1 montant): {libelle} -> {debit}")
                elif any(kw in libelle_upper for kw in credit_keywords):
                    credit = amount
                    print(f"DEBUG ROW - Classification CRÉDIT (1 montant): {libelle} -> {credit}")
                else:
                    # CORRECTION: Par défaut, analyser le contexte
                    # Si le montant est à la fin de la ligne, c'est probablement un débit
                    if pos <= 3:  # Plus proche de la position débit
                        debit = amount
                        print(f"DEBUG ROW - Classification DÉBIT (position): {libelle} -> {debit}")
                    else:  # Plus proche de la position crédit
                        credit = amount
                        print(f"DEBUG ROW - Classification CRÉDIT (position): {libelle} -> {credit}")
            
            # CORRECTION CRITIQUE: Extraire le libellé AVANT la classification des montants
            for i, cell in enumerate(cleaned_row):
                if (len(cell) > len(libelle) and 
                    not self.is_date(cell) and 
                    not self.is_amount(cell) and 
                    not self.is_balance_column(cell, i, cleaned_row) and
                    cell.strip()):
                    libelle = cell
            
            # CORRECTION CRITIQUE: Nettoyer le libellé AVANT de classifier les montants
            libelle_original = libelle
            libelle = self.clean_libelle(libelle)
            
            # BLOQUER: Si le libellé est vide après nettoyage, ignorer la transaction
            if not libelle:
                print("DEBUG ROW - Libellé vide après nettoyage, transaction ignorée")
                return None
            
            # CORRECTION CRITIQUE: Re-extraire les montants depuis le libellé original si nécessaire
            if not amounts_found and libelle_original:
                # Chercher les montants dans le libellé original
                amount_pattern = r'(\d{1,3}(?:\s+\d{3})*(?:,\d{1,3})?)'
                amounts_in_libelle = re.findall(amount_pattern, libelle_original)
                
                for amount in amounts_in_libelle:
                    try:
                        amount_value = float(amount.replace(' ', '').replace(',', '.'))
                        # Si c'est un montant valide (> 0.1), l'ajouter
                        if amount_value > 0.1:
                            amounts_found.append((len(cleaned_row), amount))  # Position fictive
                            print(f"DEBUG ROW - Montant trouvé dans libellé: {amount}")
                    except:
                        pass
                
                # Re-classifier avec les montants trouvés dans le libellé
                if amounts_found:
                    # Re-extraire débit et crédit
                    debit = ""
                    credit = ""
                    libelle_upper = libelle.upper()
                    
                    # Mots-clés pour DÉBIT (sorties d'argent)
                    debit_keywords = [
                        'RETRAIT', 'FRAIS', 'COMMISSION', 'AGIOS', 'ABONNEMENT', 
                        'COM DEPASSEMENT', 'REGLEMENT', 'PRELEVEMENT', 'TVA',
                        'PRODUCTIONS', 'COTISATION', 'PRLV', 'FRAIS PHOTO',
                        'FRAIS HUISSIER', 'COMMISSION VIR', 'ABONNEMENT PACK',
                        'COM ET TVA', 'COMMISSION PRELEVEMENT', 'VIR EMIS', 'VIREMENT EMIS'
                    ]
                    
                    # Mots-clés pour CRÉDIT (entrées d'argent)
                    credit_keywords = [
                        'VIR RECU', 'ENCAISSEMENT', 'VERSEMENT', 'REMBOURSEMENT',
                        'VIREMENT RECU', 'RECEPTION', 'VIR RECU TN', 'VERSEMENT ESPECE',
                        'ENCAISSEMENT CHEQUE', 'REMBOURSEMENT FRAIS', 'VIR. RECU',
                        'VIREMENT REÇU', 'RECU', 'REÇU', 'VIR RECU TN MEME'
                    ]
                    
                    # Classification basée sur le libellé nettoyé
                    if any(kw in libelle_upper for kw in debit_keywords):
                        debit = amounts_found[0][1]
                        print(f"DEBUG ROW - Montant DÉBIT extrait du libellé: {debit}")
                    elif any(kw in libelle_upper for kw in credit_keywords):
                        credit = amounts_found[0][1]
                        print(f"DEBUG ROW - Montant CRÉDIT extrait du libellé: {credit}")
                    else:
                        # Par défaut, considérer comme débit
                        debit = amounts_found[0][1]
                        print(f"DEBUG ROW - Montant par défaut DÉBIT extrait du libellé: {debit}")
            
            print(f"DEBUG ROW - Résultat: Date='{date}', Libellé='{libelle}', Débit='{debit}', Crédit='{credit}'")
            
            # Vérifier que nous avons au moins une date et un libellé
            if date and libelle:
                result = {
                    'Date': self.format_date(date),
                    'Libellé': libelle,
                    'Débit': self.format_amount(debit),
                    'Crédit': self.format_amount(credit)
                }
                print(f"DEBUG ROW - Transaction créée: {result}")
                return result
            
            print("DEBUG ROW - Transaction ignorée (date ou libellé manquant)")
            return None
            
        except Exception as e:
            print(f"Erreur lors du parsing de la ligne: {e}")
            return None
    
    def is_balance_column(self, cell, position, row):
        """CORRECTION: Vérifie si la colonne est probablement la colonne Solde"""
        if not cell or not self.is_amount(cell):
            return False
        
        try:
            # Nettoyer le montant pour vérifier sa valeur
            clean_amount = cell.replace(' ', '').replace(',', '.')
            amount_value = float(clean_amount)
            
            # BLOQUER: Montants négatifs (soldes débiteurs)
            if amount_value < 0:
                return True
            
            # BLOQUER: Montants très grands (soldes créditeurs)
            if amount_value > 1000000:
                return True
                
        except:
            pass
        
        # Si c'est la dernière colonne avec un montant, c'est probablement le solde
        if position == len(row) - 1:
            return True
        
        # Si c'est la 6ème colonne (position 5), c'est généralement le solde
        if position == 5:
            return True
            
        return False
    
    def clean_libelle(self, libelle):
        """CORRECTION: Nettoie et valide le libellé d'une transaction"""
        if not libelle:
            return ""
        
        libelle_clean = libelle.strip()
        libelle_clean = libelle_clean.strip('-,.:/')
        libelle_upper = libelle_clean.upper()
        
        # BLOQUER: Lignes non-transactionnelles
        if (libelle_upper.startswith('DATE') or 
            libelle_upper.startswith('PÉRIODE') or 
            libelle_upper.startswith('PERIODE') or
            libelle_upper.startswith('SOLDE') or 
            libelle_upper.startswith('TOTAL')):
            print(f"DEBUG - Libellé commençant par mot interdit ignoré: '{libelle_clean}'")
            return ""
        
        # BLOQUER: Libellés trop courts ou non valides
        if (len(libelle_clean) < 3 or
            libelle_clean.isdigit() or
            libelle_clean in ['-', '.', ',', ':', '/', 'TND', 'TOTAL', 'SOLDE', 'Période', 'Date'] or
            (any(char.isdigit() for char in libelle_clean) and len([c for c in libelle_clean if c.isalpha()]) < 2)):
            return ""
        
        # CORRECTION: Nettoyer TOUS les montants qui traînent dans le libellé
        amount_pattern = r'(\d{1,3}(?:\s+\d{3})*(?:,\d{1,3})?)'
        amounts_in_libelle = re.findall(amount_pattern, libelle_clean)
        
        for amount in amounts_in_libelle:
            try:
                amount_value = float(amount.replace(' ', '').replace(',', '.'))
                # CORRECTION: Retirer TOUS les montants valides (> 0.1) du libellé
                if amount_value > 0.1:
                    libelle_clean = libelle_clean.replace(amount, '').strip()
                    print(f"DEBUG - Montant retiré du libellé: {amount}")
            except:
                pass
        
        # Nettoyer le libellé final
        libelle_clean = re.sub(r'\s+', ' ', libelle_clean).strip()
        libelle_clean = re.sub(r'[,\s]+$', '', libelle_clean)
        libelle_clean = re.sub(r'^\s*[,.\s]+', '', libelle_clean)
        
        return libelle_clean
    
    def is_transaction_row(self, row):
        """CORRECTION: Vérifie si une ligne est une transaction valide (pas un solde/total)"""
        if not row or len(row) < 3:
            return False
        
        # Joindre toutes les cellules pour analyser le contenu
        row_text = ' '.join([str(cell) for cell in row if cell]).upper()
        
        # BLOQUER: Lignes non-transactionnelles
        non_transaction_keywords = [
            'SOLDE AU', 'SOLDE', 'TOTAL', 'TOTAUX', 'DATE', 'PÉRIODE', 'PERIODE',
            'SOLDE FINAL', 'SOLDE PRÉCÉDENT', 'SOLDE PRECEDENT'
        ]
        
        # Vérifier si la ligne commence par un mot-clé non-transactionnel
        for keyword in non_transaction_keywords:
            if row_text.startswith(keyword):
                print(f"DEBUG - Ligne non-transactionnelle détectée: {keyword}")
                return False
        
        # Vérifier si la ligne contient uniquement des montants (probablement un solde)
        amount_count = 0
        for cell in row:
            if cell and self.is_amount(str(cell)):
                amount_count += 1
        
        # Si la ligne contient principalement des montants, c'est probablement un solde
        if amount_count >= len([cell for cell in row if cell]) * 0.7:  # 70% de montants
            print(f"DEBUG - Ligne probablement un solde (trop de montants): {row}")
            return False
        
        return True
    
    def extract_alternative_transactions(self, text):
        """CORRECTION: Extraction alternative depuis le texte brut pour capturer les montants dans les libellés"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG ALTERNATIVE - {len(lines)} lignes à analyser")
            
            for line in lines:
                line = line.strip()
                if not line or len(line) < 10:
                    continue
                
                print(f"DEBUG ALTERNATIVE - Ligne: {line}")
                
                # Chercher les patterns de date et montants
                date_pattern = r'(\d{1,2}/\d{1,2}/\d{4})'
                amount_pattern = r'(\d{1,3}(?:\s+\d{3})*(?:,\d{1,3})?)'
                
                date_match = re.search(date_pattern, line)
                amount_matches = re.findall(amount_pattern, line)
                
                if date_match and amount_matches:
                    date = date_match.group(1)
                    
                    # Extraire le libellé (tout sauf la date et les montants)
                    libelle = line
                    for amount in amount_matches:
                        libelle = libelle.replace(amount, '').strip()
                    libelle = libelle.replace(date, '').strip()
                    
                    # Nettoyer le libellé
                    libelle = self.clean_libelle(libelle)
                    
                    if not libelle:
                        continue
                    
                    # Classifier les montants
                    debit = ""
                    credit = ""
                    libelle_upper = libelle.upper()
                    
                    # Mots-clés pour DÉBIT
                    debit_keywords = [
                        'RETRAIT', 'FRAIS', 'COMMISSION', 'AGIOS', 'ABONNEMENT', 
                        'COM DEPASSEMENT', 'REGLEMENT', 'PRELEVEMENT', 'TVA',
                        'PRODUCTIONS', 'COTISATION', 'PRLV', 'FRAIS PHOTO',
                        'FRAIS HUISSIER', 'COMMISSION VIR', 'ABONNEMENT PACK',
                        'COM ET TVA', 'COMMISSION PRELEVEMENT', 'VIR EMIS', 'VIREMENT EMIS'
                    ]
                    
                    # Mots-clés pour CRÉDIT
                    credit_keywords = [
                        'VIR RECU', 'ENCAISSEMENT', 'VERSEMENT', 'REMBOURSEMENT',
                        'VIREMENT RECU', 'RECEPTION', 'VIR RECU TN', 'VERSEMENT ESPECE',
                        'ENCAISSEMENT CHEQUE', 'REMBOURSEMENT FRAIS', 'VIR. RECU',
                        'VIREMENT REÇU', 'RECU', 'REÇU', 'VIR RECU TN MEME'
                    ]
                    
                    # Filtrer les montants valides
                    valid_amounts = []
                    for amount in amount_matches:
                        try:
                            amount_value = float(amount.replace(' ', '').replace(',', '.'))
                            if 0.1 < amount_value < 1000000:  # Éviter les soldes
                                valid_amounts.append(amount)
                        except:
                            pass
                    
                    if valid_amounts:
                        if any(kw in libelle_upper for kw in debit_keywords):
                            debit = valid_amounts[0]
                        elif any(kw in libelle_upper for kw in credit_keywords):
                            credit = valid_amounts[0]
                        else:
                            # Par défaut, débit
                            debit = valid_amounts[0]
                        
                        transaction = {
                            'Date': date,
                            'Libellé': libelle,
                            'Débit': self.format_amount(debit),
                            'Crédit': self.format_amount(credit)
                        }
                        
                        transactions.append(transaction)
                        print(f"DEBUG ALTERNATIVE - Transaction créée: {transaction}")
        
        except Exception as e:
            print(f"Erreur extraction alternative: {e}")
        
        return transactions
    
    def is_date(self, text):
        """Vérifie si le texte est une date au format DD/MM/YYYY"""
        if not text:
            return False
        # Pattern pour date DD/MM/YYYY
        date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'
        return bool(re.match(date_pattern, text.strip()))
    
    def is_amount(self, text):
        """Vérifie si le texte est un montant - Version améliorée pour Attijari"""
        if not text:
            return False
        
        # Pattern pour montant Attijari (espaces pour milliers, virgule pour décimales)
        # Exemples: "1 122,685", "0,595", "23,800"
        amount_pattern = r'^\d{1,3}(?:\s+\d{3})*(?:,\d+)?$'
        return bool(re.match(amount_pattern, text.strip()))
    
    def format_date(self, date_str):
        """Formate la date au format DD/MM/YYYY"""
        try:
            # Nettoyer la date
            date_str = date_str.strip()
            if self.is_date(date_str):
                return date_str
            return ""
        except:
            return ""
    
    def format_amount(self, amount_str):
        """Formate le montant - Version améliorée pour Attijari"""
        try:
            if not amount_str or amount_str.strip() == "":
                return ""
            
            # Nettoyer le montant
            amount_str = amount_str.strip()
            
            # Pour Attijari : espaces pour milliers, virgule pour décimales
            # Exemple: "1 122,685" -> "1 122,685"
            # Ne pas toucher aux espaces, juste nettoyer
            cleaned = amount_str
            
            # Vérifier si c'est un montant valide
            if self.is_amount(cleaned):
                return cleaned
            return ""
        except:
            return ""
    
    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel"""
        try:
            # Créer un nouveau workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "J03"  # Nom de la feuille
            
            # Définir les styles
            header_font = Font(name='Arial', size=12, bold=True, color='000000')
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Jaune
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Écrire les en-têtes
            headers = ['Date', 'Libellé', 'Débit', 'Crédit']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Écrire les données
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Alignement spécial pour les colonnes
                    if col_idx == 1:  # Date
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 2:  # Libellé
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    elif col_idx in [3, 4]:  # Débit et Crédit
                        cell.alignment = Alignment(horizontal='right')
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 50  # Libellé
            ws.column_dimensions['C'].width = 15  # Débit
            ws.column_dimensions['D'].width = 15  # Crédit
            
            # Ajuster la hauteur des lignes
            for row in range(1, len(df) + 2):
                ws.row_dimensions[row].height = 20
            
            # Ajouter des bordures à toute la plage de données
            from openpyxl.utils import get_column_letter
            max_row = len(df) + 1
            max_col = len(headers)
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).border = border
            
            # Sauvegarder le fichier
            wb.save(excel_path)
            
        except Exception as e:
            print(f"Erreur lors du formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, engine='openpyxl')
    
    def convert_pdf_to_excel(self):
        """Fonction principale de conversion"""
        if not self.pdf_file_path.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
            return
        
        if not self.excel_filename.get():
            messagebox.showerror("Erreur", "Veuillez entrer un nom pour le fichier Excel")
            return
        
        self.result_text.delete(1.0, tk.END)
        
        try:
            # Vérifier que c'est un PDF Attijari Bank
            if not self.is_attijari_bank_pdf(self.pdf_file_path.get()):
                messagebox.showwarning("Attention"
                    "Ce fichier ne semble pas être un extrait Attijari Bank.\n"
                    "La conversion peut ne pas fonctionner correctement.")
            
            # Extraire les données
            transactions = self.extract_table_data(self.pdf_file_path.get())
            
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF")
                return
            
            # Créer le DataFrame
            df = pd.DataFrame(transactions)
            
            # Chemin de sortie dans le dossier Téléchargements
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = os.path.join(downloads_path, f"{self.excel_filename.get()}.xlsx")
            
            # Sauvegarder en Excel avec formatage
            self.save_excel_with_formatting(df, excel_path)
            
            # Afficher les résultats
            self.result_text.insert(tk.END, f"Conversion réussie !\n")
            self.result_text.insert(tk.END, f"Fichier créé: {excel_path}\n")
            self.result_text.insert(tk.END, f"Nombre de transactions: {len(transactions)}\n\n")
            self.result_text.insert(tk.END, "Aperçu des données:\n")
            self.result_text.insert(tk.END, df.head(10).to_string(index=False))

            # Message de succès plus positif
            success_msg = f"✅ Conversion terminée avec succès !\n\n"
            success_msg += f"📁 Fichier créé: {excel_path}\n"
            success_msg += f"📊 Nombre de transactions: {len(transactions)}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            messagebox.showinfo("✅ Conversion réussie", success_msg)
            
        except Exception as e:
            error_msg = f"Erreur lors de la conversion: {str(e)}"
            self.result_text.insert(tk.END, error_msg)
            messagebox.showerror("Erreur", error_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

def main():
    root = tk.Tk()
    app = AttijariBankConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()