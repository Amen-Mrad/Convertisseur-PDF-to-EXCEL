import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import re
from datetime import datetime
import subprocess
import sys
from pdf_detector import PDFBankDetector


class UIBReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVÉ UIB vers Excel")
        self.root.geometry("620x440")
        self.root.configure(bg="#f0f0f0")

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"RELEVE_UIB_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self._build_ui()

    def _build_ui(self):
        tk.Label(self.root, text="Convertisseur RELEVÉ UIB", font=("Arial", 18, "bold"), bg="#f0f0f0", fg="#2c3e50").pack(pady=(20, 4))
        tk.Label(self.root, text="Conversion PDF vers Excel", font=("Arial", 12), bg="#f0f0f0", fg="#666").pack()

        frm = tk.Frame(self.root, bg="#f0f0f0"); frm.pack(pady=10, padx=40, fill='x')
        tk.Label(frm, text="Fichier PDF RELEVÉ UIB:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(frm, width=50, textvariable=self.pdf_path).grid(row=1, column=0, padx=(0, 10))
        tk.Button(frm, text="Parcourir", command=self.browse_pdf, bg="#3498db").grid(row=1, column=1)

        tk.Label(frm, text="Nom du fichier Excel:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', pady=(15, 5))
        tk.Entry(frm, width=50, textvariable=self.excel_name).grid(row=3, column=0, padx=(0, 10))

        buttons = tk.Frame(self.root, bg="#f0f0f0"); buttons.pack(pady=10)
        tk.Button(buttons, text="Convertir en Excel", command=self.convertir, bg="#16A34A", font=("Arial", 13, "bold")).pack(side='left', padx=5)
        tk.Button(buttons, text="Retour page d'accueil", command=self.retour_accueil, bg="#95a5a6", font=("Arial", 11, "bold")).pack(side='left', padx=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF UIB RELEVÉ", filetypes=[["PDF", "*.pdf"]])
        if not path:
            return
        try:
            det = PDFBankDetector().detect_document_type(path)
            if det.get('type') == 'releve_uib':
                self.pdf_path.set(path)
            else:
                summary = PDFBankDetector().get_detection_summary(det)
                if messagebox.askyesno("Type détecté", f"{summary}\n\nContinuer quand même avec UIB ?"):
                    self.pdf_path.set(path)
        except Exception:
            self.pdf_path.set(path)

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF UIB RELEVÉ.")
            return
        rows = self.parse_pdf(path)
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire les opérations du relevé UIB.")
            return
        # Nettoyage: supprimer lignes de solde/total/nouveau/UIBphone et lignes sans libellé
        cleaned = []
        for r in rows:
            lib = (r.get('libelle') or '').strip()
            if not lib:
                continue
            # Enlever lignes de synthèse qui COMMENCENT par ces mots
            # (Totaux des mouvements, Solde, Nouveau solde ...)
            if re.match(r"^\s*(solde|totaux?)\b", lib, flags=re.IGNORECASE):
                continue
            if re.match(r"^\s*nouveau\b", lib, flags=re.IGNORECASE):
                continue
            if re.match(r"^\s*uibphone\b", lib, flags=re.IGNORECASE):
                continue
            # Enlever uniquement l'entête de période si le libellé est exactement du/au (sans autres mots)
            if re.match(r"^\s*du\s+\d{2}/\d{2}/\d{2,4}\s+au\s+\d{2}/\d{2}/\d{2,4}\s*$", lib, flags=re.IGNORECASE):
                continue
            # Supprimer toute ligne qui CONTIENT 'totaux' ou 'libellé'
            if re.search(r"totaux|libell[é|e]", lib, flags=re.IGNORECASE):
                continue
            cleaned.append(r)
        # Exporter avec Date et sans Valeur
        df = pd.DataFrame(cleaned, columns=["date", "libelle", "debit", "credit"]).rename(columns={
            'date': 'Date',
            'libelle': 'Libellé',
            'debit': 'Débit',
            'credit': 'Crédit'
        })
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'RELEVE_UIB'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        messagebox.showinfo("✅ Conversion réussie", f"Fichier enregistré:\n{out}")

    def retour_accueil(self):
        try:
            self.root.destroy()
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def parse_pdf(self, pdf_path: str):
        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words() or []
                    if not words:
                        continue
                    # Trouver en-têtes: Date | Libellé de l'opération | Débit | Crédit | Valeur
                    header_words = [w for w in words if any(k in w['text'].lower() for k in ['date', "libellé", 'débit', 'debit', 'crédit', 'credit', 'valeur'])]
                    if not header_words:
                        continue
                    header_y = min(w['top'] for w in header_words)
                    # X des en-têtes principaux
                    def find_x(lbl):
                        xs = [w['x0'] for w in words if lbl in w['text'].lower() and abs(w['top'] - header_y) < 25]
                        return min(xs) if xs else None
                    x_date = find_x('date')
                    x_lib = find_x("libellé") or find_x('libelle')
                    x_deb = find_x('débit') or find_x('debit')
                    x_cre = find_x('crédit') or find_x('credit')
                    x_val = find_x('valeur')
                    ordered = [x for x in [x_date, x_lib, x_deb, x_cre, x_val] if x is not None]
                    if len(ordered) < 4:
                        continue
                    page_width = page.width
                    edges = [0]
                    for i in range(len(ordered) - 1):
                        edges.append((ordered[i] + ordered[i+1]) / 2)
                    edges.append(page_width)

                    def bounds(x):
                        for i in range(len(edges)-1):
                            if ordered[i] == x:
                                return (edges[i], edges[i+1])
                        return (x, x+60)
                    dmax = bounds(x_date)[1]
                    lmin, lmax = bounds(x_lib)
                    dbmin, dbmax = bounds(x_deb)
                    crmin, crmax = bounds(x_cre)
                    vmin, vmax = bounds(x_val) if x_val is not None else (page_width*0.85, page_width)
                    # Zone verticale
                    # Étendre la zone lisible jusqu'au bas de la page pour éviter de perdre la dernière ligne
                    footer_y = page.height * 0.985
                    lines = {}
                    for w in words:
                        yc = round((w['top'] + w['bottom'])/2)
                        # Réduire la zone ignorée au-dessus de l'en-tête pour ne pas perdre la 1ère ligne
                        if yc <= header_y + 5 or yc >= footer_y:
                            continue
                        lines.setdefault(yc, []).append(w)
                    # Nombre robuste: accepte espaces de milliers, virgules/points
                    amount_re = re.compile(r"^-?\d[\d\s\.,]*$")
                    for yk in sorted(lines.keys()):
                        ws = sorted(lines[yk], key=lambda m: m['x0'])
                        # ignorer explicitement les pieds de page s'ils apparaissent
                        if any(any(k in w['text'].lower() for k in ['identifiant unique', 'www.uib.com.tn', 'swift', 'avenue habib', 'société anonyme']) for w in ws):
                            continue
                        d = ''
                        lib_parts = []
                        deb = ''
                        cre = ''
                        val = ''
                        for w in ws:
                            t = w['text']
                            x0, x1 = w['x0'], w['x1']
                            if x1 <= dmax and re.match(r"^\d{2}/\d{2}/\d{4}$", t):
                                d = t
                            elif lmin <= x0 and x1 <= lmax:
                                # éviter de ramasser les montants dans libellé
                                if not amount_re.match(t):
                                    lib_parts.append(t)
                            elif dbmin <= x0 and x1 <= dbmax and amount_re.match(t):
                                deb = t
                            elif crmin <= x0 and x1 <= crmax and amount_re.match(t):
                                cre = t
                            elif vmin <= x0 and x1 <= vmax and re.match(r"^\d{2}/\d{2}/\d{4}$", t):
                                val = t
                        # on ne filtre plus ici pour garder les lignes de continuation du libellé
                        # Nettoyer libellé: supprimer tirets, tokens arabes et mots de pied de page
                        footer_stoppers = ['identifiant', 'unique', 'swift', 'uibkntt', 'www.uib.com.tn', 'avenue', 'habib', 'bourgiba', 'société', 'anonyme', 'capital', 'tunis', 'suite', 'releve', 'de', 'compte', 'groupe', 'societe', 'generale', 'page']
                        def is_arabic(s: str) -> bool:
                            return any(
                                ('\u0600' <= ch <= '\u06FF') or
                                ('\u0750' <= ch <= '\u077F') or
                                ('\u08A0' <= ch <= '\u08FF') or
                                ('\uFB50' <= ch <= '\uFDFF') or
                                ('\uFE70' <= ch <= '\uFEFF')
                                for ch in s
                            )
                        cleaned_tokens = []
                        for tok in lib_parts:
                            low = tok.lower()
                            if tok in ['-', ':']:
                                continue
                            if is_arabic(tok):
                                continue
                            if any(word in low for word in footer_stoppers):
                                continue
                            cleaned_tokens.append(tok)
                        lib = ' '.join(cleaned_tokens).strip()
                        # Si des fragments arabes restent collés, supprimer via regex des blocs arabes
                        lib = re.sub(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]+", "", lib).strip()
                        def fmt(a: str):
                            if not a:
                                return None
                            v = a.replace('\u00A0',' ').replace(' ','').replace('.','').replace(',', '.')
                            try:
                                f = float(v)
                            except Exception:
                                return None
                            # Format milliers espace et décimales virgule
                            return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
                        rows.append({
                            'date': d,
                            'libelle': lib,
                            'debit': fmt(deb),
                            'credit': fmt(cre),
                            'valeur': val
                        })
        except Exception:
            return []
        # Consolidation: fusionner lignes de continuation pour ne pas perdre des libellés multi-lignes
        consolidated = []
        current = None
        last_date = ''
        for r in rows:
            d = (r.get('date') or '').strip()
            lib = (r.get('libelle') or '').strip()
            deb = r.get('debit')
            cre = r.get('credit')
            if d:
                # nouvelle opération; valider la précédente
                if current:
                    consolidated.append(current)
                current = {'date': d, 'libelle': lib, 'debit': deb, 'credit': cre}
                last_date = d
                continue
            # pas de date -> continuation
            if not current:
                current = {'date': last_date, 'libelle': lib, 'debit': deb, 'credit': cre}
                continue
            # concaténer libellé si présent
            if lib:
                current['libelle'] = (current.get('libelle', '') + ' ' + lib).strip()
            # compléter montants si manquants
            if deb and not current.get('debit'):
                current['debit'] = deb
            if cre and not current.get('credit'):
                current['credit'] = cre
        if current:
            consolidated.append(current)
        # Propager dates manquantes par sécurité
        last_d = ''
        for r in consolidated:
            if r.get('date'):
                last_d = r['date']
            else:
                r['date'] = last_d
        return consolidated

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "Relevé UIB"
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        # Quatre colonnes: Date | Libellé | Débit | Crédit
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        max_row = ws.max_row
        # Trouver indices colonnes par libellé pour robustesse
        headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        col_debit = headers.get('Débit', 3)
        col_credit = headers.get('Crédit', 4)
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_debit).number_format = '# ##0,000'
            ws.cell(row=r, column=col_credit).number_format = '# ##0,000'
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)


def main():
    root = tk.Tk()
    UIBReleveConverter(root)
    root.mainloop()


if __name__ == '__main__':
    main()


