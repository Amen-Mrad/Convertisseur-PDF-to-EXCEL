import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import subprocess
import sys
import pdfplumber
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class BNAReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVÉ BNA vers Excel")
        self.root.geometry("600x500")
        
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"RELEVE_BNA_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        self._ui()

    def _ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur RELEVÉ BNA",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVÉ BNA:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.browse, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                              command=self.convertir, 
                              font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.back, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def browse(self):
        f = filedialog.askopenfilename(title="Choisir un PDF BNA RELEVÉ", filetypes=[["PDF", "*.pdf"]])
        if f: self.pdf_path.set(f)

    def back(self):
        try:
            self.root.destroy(); subprocess.Popen([sys.executable, "lancer_convertisseur.py"])  # noqa
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF BNA relevé.")
            return
        
        self.progress['value'] = 10
        self.root.update_idletasks()
        rows, ocr_used = self._parse_bna(path)
        self._ocr_used = ocr_used  # Stocker le flag OCR pour le message de succès
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire des transactions BNA.")
            return
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'RELEVE_BNA'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)

        # Message de succès plus positif
        success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n"
        success_msg += f"📊 Nombre de transactions: {len(rows)}\n\n"
        
        # Ajouter une note si l'OCR a été utilisé
        if hasattr(self, '_ocr_used') and self._ocr_used:
            success_msg += f"🔍 PDF scanné traité avec OCR\n\n"
        
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        messagebox.showinfo("✅ Conversion réussie", success_msg)

    # --- Parsing ---
    def _parse_bna(self, pdf_path: str):
        rows = []
        month = None; year = None
        in_table = False
        ocr_used = False  # Flag pour indiquer si l'OCR a été utilisé
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ''
                    if not text.strip():
                        # Fallback OCR pour les PDFs scannés
                        print("🔍 PDF scanné détecté - Utilisation de l'OCR...")
                        self.status_label.config(text="🔍 Analyse OCR en cours...")
                        self.root.update()
                        
                        ocr_text = self._extract_text_via_ocr(pdf_path)
                        if ocr_text.strip():
                            text = ocr_text
                            ocr_used = True
                            print(f"✅ OCR terminé - {len(ocr_text)} caractères extraits")
                            break  # Utiliser le texte OCR pour toutes les pages
                        else:
                            print("⚠️ OCR échoué - Aucun texte extrait")
                            continue
                    if not month or not year:
                        # Chercher l'en-tête "DU MOIS DE : MAI 2025" (mois en lettres)
                        m = re.search(r"DU\s+MOIS\s+DE\s*:\s*([A-ZÉÈÂÎÛÀÙÔÇ]+)\s+(\d{4})", text, re.IGNORECASE)
                        if m:
                            month = self._mois_to_num(m.group(1))
                            year = int(m.group(2))
                    
                    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                    print(f"DEBUG: {len(lines)} lignes à analyser")
                    
                    for i, ln in enumerate(lines):
                        print(f"DEBUG LIGNE {i+1}: {ln[:80]}...")
                        
                        # Détecter le début du tableau de transactions (plus flexible)
                        if (re.search(r"JOUR.*LIBELLE.*VALEUR.*DEBIT.*CREDIT", ln, re.IGNORECASE) or
                            re.search(r"JOUR.*LIBELLE.*DEBIT.*CREDIT", ln, re.IGNORECASE) or
                            re.search(r"DATE.*LIBELLE.*MONTANT", ln, re.IGNORECASE) or
                            re.search(r"JOUR.*OPERATION.*MONTANT", ln, re.IGNORECASE)):
                            in_table = True
                            print(f"DEBUG: Début du tableau détecté ligne {i+1}")
                            continue
                        
                        # Détecter la fin du tableau (lignes de résumé)
                        if in_table and (re.search(r"SOLDE\s+FINAL|TOTAL|RÉSUMÉ|RESUME|SOLDE\s+PRÉCÉDENT", ln, re.IGNORECASE)):
                            print(f"DEBUG: Fin du tableau détectée ligne {i+1}")
                            break
                        
                        # Si on n'est pas dans le tableau, vérifier si c'est une ligne de transaction potentielle
                        if not in_table:
                            # Vérifier si la ligne ressemble à une transaction (jour + montant)
                            has_day = re.match(r"^(\d{1,2})\b", ln)
                            has_amount = re.search(r"\d{1,3}(?:[\.,]\d{3})*(?:,\d{3}|\.\d{3})", ln)
                            if has_day and has_amount:
                                print(f"DEBUG: Transaction potentielle détectée ligne {i+1} (hors tableau)")
                                in_table = True  # Commencer le traitement
                            else:
                                continue
                        
                        # Ligne de transaction: jour | libellé | valeur | debit | credit
                        # On repère d'abord un jour (2 chiffres) au début de ligne
                        mday = re.match(r"^(\d{1,2})\b", ln)
                        if not mday:
                            continue
                        
                        day = int(mday.group(1))
                        
                        # Vérifier que c'est une vraie transaction (doit avoir des montants)
                        # Patterns plus flexibles pour les montants BNA
                        amts = re.findall(r"\d{1,3}(?:[\.,]\d{3})*(?:,\d{3}|\.\d{3})", ln)
                        # Essayer aussi des patterns plus simples
                        if not amts:
                            amts = re.findall(r"\d+[.,]\d{2,3}", ln)
                        # Essayer des montants sans séparateurs de milliers
                        if not amts:
                            amts = re.findall(r"\d+[.,]\d+", ln)
                        
                        print(f"DEBUG: Montants trouvés: {amts}")
                        if not amts:
                            print(f"DEBUG: Aucun montant trouvé dans: {ln[:50]}...")
                            continue
                        
                        debit = credit = None
                        
                        # Classification améliorée basée sur la structure BNA
                        # Format BNA: JOUR | LIBELLE | VALEUR | DEBIT | CREDIT
                        # Les montants sont dans les 2 dernières colonnes
                        
                        # Classification basée sur la position dans le tableau BNA
                        # Format BNA: JOUR | LIBELLE | VALEUR | DEBIT | CREDIT
                        
                        if len(amts) >= 2:
                            # Deux montants détectés -> utiliser la position dans le tableau
                            # Le montant dans l'avant-dernière position est le débit
                            # Le montant dans la dernière position est le crédit
                            debit = self._fmt_amount(amts[-2])
                            credit = self._fmt_amount(amts[-1])
                        elif len(amts) == 1:
                            # Un seul montant -> classification par mots-clés ET position
                            single = self._fmt_amount(amts[-1])
                            upper = ln.upper()
                            
                            # Mots-clés pour les CRÉDITS (priorité absolue)
                            credit_keywords = [
                                'DÉBLOCAGE CRÉD', 'DEBLOCAGE CRED', 'DÉRÉSERVATION', 'DERESERVATION',
                                'ENCAISSEMENT', 'VIREMENT REÇU', 'VIR REÇU', 'CRÉDIT', 'CREDIT',
                                'REMBOURSEMENT', 'RETOUR', 'ANNULATION', 'CANCELLATION'
                            ]
                            
                            # Mots-clés pour les DÉBITS
                            debit_keywords = [
                                'RETRAIT', 'ACHAT', 'PAIEMENT', 'PRÉLÈV', 'PRELEV', 'COMMISSION', 'COMM',
                                'TVA/COMM', 'CHÈQUE', 'CHEQUE', 'VIR PONCTUEL', 'VIR EMIS', 'DÉCISION',
                                'RÉSERVATION', 'RESERVATION', 'REJET', 'CERTIF', 'GAB', 'ATM'
                            ]
                            
                            # Vérifier d'abord les crédits (priorité absolue)
                            if any(k in upper for k in credit_keywords):
                                credit = single
                            # Puis les débits
                            elif any(k in upper for k in debit_keywords):
                                debit = single
                            # Par défaut, considérer comme débit
                            else:
                                debit = single
                        
                        # Libellé = texte entre le jour et les montants
                        # Retirer jour en tête
                        lib = ln[mday.end():].strip()
                        if amts:
                            last_amt = amts[-1]
                            pos = lib.rfind(last_amt)
                            if pos != -1:
                                lib = lib[:pos].strip()
                        
                        # Supprimer la date de valeur (format: DD MM YYYY) du libellé
                        # La date de valeur est généralement à la fin du libellé
                        lib = re.sub(r'\s+\d{2}\s+\d{2}\s+\d{4}\s*$', '', lib)
                        
                        # Nettoyage libellé
                        lib = re.sub(r"\s+", " ", lib)
                        
                        # Filtrer les libellés qui ne sont pas des transactions (adresses, etc.)
                        if not lib or len(lib) < 3:
                            continue
                        
                        # Ignorer les lignes qui ressemblent à des adresses
                        if any(word in lib.upper() for word in ['RUE', 'AVENUE', 'BOULEVARD', 'PLACE', 'MONTPLAISIR', 'TUNIS', 'SOCIETE']):
                            continue
                        
                        # Construire date complète
                        if not (month and year):
                            # fallback: essayer valeur dans la ligne (colonne valeur: JJ MM AAAA)
                            mv = re.search(r"(\d{2})\s*(\d{2})\s*(\d{4})", ln)
                            if mv:
                                day, month, year = int(mv.group(1)), int(mv.group(2)), int(mv.group(3))
                        
                        date_str = self._build_date(day, month, year)
                        rows.append({
                            'date': date_str,
                            'libelle': lib,
                            'debit': debit,
                            'credit': credit
                        })
        except Exception as e:
            print(f"DEBUG BNA PARSE ERROR: {e}")
        
        # Si aucune transaction trouvée, essayer une approche plus permissive
        if not rows:
            print("DEBUG: Aucune transaction trouvée avec le parser standard, tentative avec approche permissive...")
            rows = self._parse_bna_permissive(pdf_path)
            if rows:
                print(f"DEBUG: {len(rows)} transactions trouvées avec l'approche permissive")
        
        return rows, ocr_used

    def _mois_to_num(self, mois_str: str) -> int:
        mois = mois_str.strip().upper()
        mapping = {
            'JANVIER': 1, 'FÉVRIER': 2, 'FEVRIER': 2, 'MARS': 3, 'AVRIL': 4, 'MAI': 5, 'JUIN': 6,
            'JUILLET': 7, 'AOÛT': 8, 'AOUT': 8, 'SEPTEMBRE': 9, 'OCTOBRE': 10, 'NOVEMBRE': 11, 'DÉCEMBRE': 12, 'DECEMBRE': 12
        }
        return mapping.get(mois, 1)

    def _build_date(self, day: int, month: int, year: int) -> str:
        try:
            dt = datetime(year, int(month or 1), int(day or 1))
            return dt.strftime('%d/%m/%Y')
        except Exception:
            return ''

    def _fmt_amount(self, s: str):
        if not s:
            return None
        # Normaliser: enlever séparateurs de milliers . ou , et garder 3 décimales avec virgule
        v = s.replace('\u00A0', ' ').replace(' ', '')
        # Remplacer dernier séparateur décimal en point
        # Cas "1,220.517" ou "3,550.770"
        v = v.replace(',', '')
        try:
            f = float(v)
            return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
        except Exception:
            # Autre forme: 3.550,770
            v2 = s.replace('.', '').replace(',', '.')
            try:
                f = float(v2)
                return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
            except Exception:
                return None

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "J03"
        # entêtes
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for c in ws[1]:
            c.fill = yellow; c.font = bold; c.alignment = center
        # tailles colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        # formats
        for r in range(2, ws.max_row + 1):
            ws[f'C{r}'].number_format = '# ##0,000'
            ws[f'D{r}'].number_format = '# ##0,000'
        # bordures
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR optimisé pour les relevés bancaires BNA scannés.
        Utilise plusieurs techniques de pré-traitement pour améliorer la qualité.
        """
        if not _OCR_AVAILABLE:
            print("⚠️ OCR non disponible - Installez pytesseract, opencv-python et PyMuPDF")
            return ""
        
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            
            for page_num, page in enumerate(doc, 1):
                print(f"🔍 Traitement OCR page {page_num}/{len(doc)}...")
                
                # Rendu à haute résolution pour un OCR plus fiable
                mat = fitz.Matrix(3, 3)  # Augmenter la résolution
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                
                # Convertir en numpy array
                arr = np.array(img)
                
                # Pré-traitement avancé pour les documents bancaires
                # 1. Conversion en niveaux de gris
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                
                # 2. Réduction du bruit
                denoised = cv2.medianBlur(gray, 3)
                
                # 3. Amélioration du contraste
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                enhanced = clahe.apply(denoised)
                
                # 4. Seuillage adaptatif pour les documents bancaires
                _, binary = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                
                # 5. Morphologie pour nettoyer les caractères
                kernel = np.ones((2,2), np.uint8)
                cleaned = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
                
                # 6. Configuration OCR optimisée pour les relevés bancaires
                config = '--psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyzÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞßàáâãäåæçèéêëìíîïðñòóôõöøùúûüýþÿ.,/-:() '
                
                # OCR en français avec configuration optimisée
                txt = pytesseract.image_to_string(cleaned, lang='fra', config=config)
                
                if txt.strip():
                    # Nettoyage du texte OCR
                    cleaned_text = self._clean_ocr_text(txt)
                    text_parts.append(cleaned_text)
                    print(f"✅ Page {page_num}: {len(cleaned_text)} caractères extraits")
                else:
                    print(f"⚠️ Page {page_num}: Aucun texte détecté")
            
            result = "\n".join(text_parts)
            print(f"📊 OCR terminé - Total: {len(result)} caractères sur {len(doc)} pages")
            return result
            
        except Exception as e:
            print(f"❌ Erreur OCR: {str(e)}")
            return ""
    
    def _clean_ocr_text(self, text: str) -> str:
        """Nettoie le texte OCR pour améliorer la reconnaissance des données bancaires."""
        if not text:
            return ""
        
        # Nettoyage spécifique aux relevés bancaires
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # Supprimer les lignes trop courtes (probablement du bruit)
            if len(line.strip()) < 3:
                continue
            
            # Corriger les erreurs OCR communes dans les relevés bancaires
            line = line.replace('|', ' ')  # Remplacer les | par des espaces
            line = line.replace('O', '0')  # Corriger les O par des 0 dans les montants
            line = re.sub(r'\s+', ' ', line)  # Normaliser les espaces
            
            # Garder seulement les lignes qui semblent contenir des données bancaires
            if (re.search(r'\d', line) or  # Contient des chiffres
                any(word in line.upper() for word in ['JOUR', 'LIBELLE', 'DEBIT', 'CREDIT', 'SOLDE', 'TOTAL']) or
                re.search(r'\d{1,2}/\d{1,2}/\d{4}', line) or  # Contient des dates
                re.search(r'\d+[.,]\d+', line)):  # Contient des montants
                cleaned_lines.append(line.strip())
        
        return '\n'.join(cleaned_lines)
    
    def _parse_bna_permissive(self, pdf_path: str):
        """Parser permissif pour les relevés BNA qui ne suivent pas le format standard"""
        rows = []
        month = None
        year = None
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ''
                    if not text.strip():
                        # Essayer l'OCR si pas de texte
                        ocr_text = self._extract_text_via_ocr(pdf_path)
                        if ocr_text.strip():
                            text = ocr_text
                    
                    # Extraire mois/année si pas encore fait
                    if not month or not year:
                        m = re.search(r"DU\s+MOIS\s+DE\s*:\s*([A-ZÉÈÂÎÛÀÙÔÇ]+)\s+(\d{4})", text, re.IGNORECASE)
                        if m:
                            month = self._mois_to_num(m.group(1))
                            year = int(m.group(2))
                    
                    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                    print(f"DEBUG PERMISSIF: {len(lines)} lignes à analyser")
                    
                    for i, ln in enumerate(lines):
                        # Chercher des lignes qui ressemblent à des transactions
                        # Pattern: jour + texte + montant(s)
                        day_match = re.match(r"^(\d{1,2})\b", ln)
                        if not day_match:
                            continue
                        
                        day = int(day_match.group(1))
                        
                        # Chercher des montants dans la ligne
                        amounts = re.findall(r"\d+[.,]\d{2,3}", ln)
                        if not amounts:
                            # Essayer des patterns plus simples
                            amounts = re.findall(r"\d+[.,]\d+", ln)
                        
                        if not amounts:
                            continue
                        
                        print(f"DEBUG PERMISSIF: Transaction potentielle ligne {i+1}: {ln[:60]}...")
                        print(f"DEBUG PERMISSIF: Montants: {amounts}")
                        
                        # Extraire le libellé (tout sauf jour et montants)
                        libelle = ln[day_match.end():].strip()
                        for amt in amounts:
                            pos = libelle.find(amt)
                            if pos != -1:
                                libelle = libelle[:pos].strip()
                                break
                        
                        # Classification simple des montants
                        debit = credit = None
                        if len(amounts) >= 2:
                            # Deux montants: premier = débit, second = crédit
                            debit = self._fmt_amount(amounts[0])
                            credit = self._fmt_amount(amounts[1])
                        elif len(amounts) == 1:
                            # Un seul montant: classifier par mots-clés
                            amount = self._fmt_amount(amounts[0])
                            libelle_upper = libelle.upper()
                            
                            # Mots-clés pour crédit
                            credit_keywords = ['ENCAISSEMENT', 'VIREMENT REÇU', 'CRÉDIT', 'REMBOURSEMENT']
                            # Mots-clés pour débit  
                            debit_keywords = ['RETRAIT', 'PAIEMENT', 'COMMISSION', 'FRAIS', 'DÉBIT']
                            
                            if any(k in libelle_upper for k in credit_keywords):
                                credit = amount
                            elif any(k in libelle_upper for k in debit_keywords):
                                debit = amount
                            else:
                                # Par défaut, considérer comme débit
                                debit = amount
                        
                        # Construire la date
                        date_str = self._build_date(day, month, year)
                        
                        if libelle.strip() and (debit or credit):
                            print(f"DEBUG PERMISSIF: Transaction ajoutée - Date: {date_str}, Libellé: {libelle[:30]}..., Débit: {debit}, Crédit: {credit}")
                            rows.append({
                                'date': date_str,
                                'libelle': libelle,
                                'debit': debit,
                                'credit': credit
                            })
        
        except Exception as e:
            print(f"DEBUG PERMISSIF ERROR: {e}")
        
        return rows

def main():
    root = tk.Tk()
    BNAReleveConverter(root)
    root.mainloop()

if __name__ == '__main__':
    main()

