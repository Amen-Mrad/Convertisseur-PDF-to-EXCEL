import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
from PIL import Image, ImageTk
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime as dt
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class AttijariReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVE Attijari Bank PDF vers Excel")
        self.root.geometry("600x500")

        # Variables
        self.pdf_file_path = tk.StringVar()
        self.excel_filename = tk.StringVar()
        self.excel_filename.set("releve_attijari_" + datetime.now().strftime("%d%m%Y_%H%M"))
        
        self.setup_ui()
        
    def setup_ui(self):
        # Titre principal moderne
        title_label = tk.Label(self.root, text="Convertisseur RELEVE ATTIJARI",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(self.root, text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame(self.root)
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section sélection fichier PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVE Attijari Bank:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_file_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.select_pdf_file, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section nom fichier Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_filename, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section conversion
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        # Boutons
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                               command=self.convert_pdf_to_excel, font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Zone de résultats
        self.result_text = tk.Text(main_frame, height=8, width=70, font=("Arial", 8))
        self.result_text.pack(fill='both', expand=True, pady=10)
        
    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF RELEVE Attijari Bank",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_file_path.set(file_path)
            
    def is_attijari_releve_pdf(self, pdf_path):
        """Vérifie si le PDF est un RELEVE Attijari Bank"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Vérifier les premières pages pour les mots-clés RELEVE
                for page_num in range(min(3, len(pdf.pages))):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        text_lower = text.lower()
                        # Mots-clés spécifiques pour RELEVE
                        releve_keywords = [
                            'releve de compte',
                            'كشف حساب',
                            'attijari',
                            'date operation',
                            'date valeur',
                            'debit (tnd)',
                            'credit (tnd)'
                        ]
                        found_keywords = sum(1 for keyword in releve_keywords if keyword in text_lower)
                        if found_keywords >= 2:  # Au moins 2 mots-clés RELEVE
                            return True
                return False
        except Exception as e:
            print(f"Erreur lors de la vérification du PDF RELEVE: {e}")
            return False
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF RELEVE"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num in range(min(2, len(pdf.pages))):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        # Chercher des patterns d'année dans le texte
                        year_patterns = [
                            r'au\s*:\s*\d{1,2}/\d{1,2}/(\d{4})',  # "Au : 31/07/2024"
                            r'(\d{4})',  # Toute année 4 chiffres
                        ]
                        
                        for pattern in year_patterns:
                            matches = re.findall(pattern, text, re.IGNORECASE)
                            if matches:
                                # Prendre la première année trouvée
                                year = matches[0]
                                if 2020 <= int(year) <= 2030:  # Année raisonnable
                                    print(f"Année détectée: {year}")
                                    return year
                
                # Fallback: année actuelle
                current_year = datetime.now().year
                print(f"Année par défaut: {current_year}")
                return str(current_year)
                
        except Exception as e:
            print(f"Erreur détection année: {e}")
            return str(datetime.now().year)
    
    def extract_table_data(self, pdf_path):
        """Nouvelle logique d'extraction pour Attijari RELEVE - Simple et efficace"""
        try:
            # Détecter l'année depuis le PDF
            year = self.detect_year_from_pdf(pdf_path)
            print(f"DEBUG - Année détectée: {year}")
            
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG - Traitement page {page_num + 1}")
                    
                    # Extraire le texte brut de la page
                    page_text = page.extract_text() or ""
                    if not page_text.strip():
                        print(f"DEBUG - Page {page_num + 1} vide, passage à la suivante")
                        continue
                    
                    # Parser le texte pour extraire les transactions
                    transactions = self.parse_attijari_releve_text(page_text, year)
                    if transactions:
                        print(f"DEBUG - {len(transactions)} transactions extraites de la page {page_num + 1}")
                        all_transactions.extend(transactions)
                                else:
                        # Fallback - analyser le texte brut si aucune transaction trouvée
                        print(f"DEBUG - Aucune transaction trouvée, analyse fallback de la page {page_num + 1}")
                        fallback_transactions = self.parse_fallback_text(page_text, year)
                        if fallback_transactions:
                            print(f"DEBUG - {len(fallback_transactions)} transactions trouvées en fallback")
                            all_transactions.extend(fallback_transactions)
                
                # Nettoyer et dédupliquer
                all_transactions = self.clean_and_deduplicate_transactions(all_transactions)
                print(f"DEBUG - Total final: {len(all_transactions)} transactions")
                
                return all_transactions
                
        except Exception as e:
            print(f"Erreur extraction RELEVE: {e}")
            return []

    def parse_attijari_releve_text(self, text, year):
        """Parse le texte Attijari RELEVE pour extraire Date, Libellé, Débit, Crédit"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG - {len(lines)} lignes à analyser")
            
            for line_num, line in enumerate(lines):
                line = line.strip()
                if not line or len(line) < 10:
                    continue
                
                print(f"DEBUG LIGNE {line_num + 1}: {line}")
                
                # Pattern pour détecter une ligne de transaction Attijari RELEVE
                transaction_match = self.extract_transaction_from_line(line, year)
                
                if transaction_match:
                    transactions.append(transaction_match)
                    print(f"DEBUG - Transaction extraite: {transaction_match}")
                else:
                    print(f"DEBUG - Ligne ignorée: {line}")
            
            return transactions
            
        except Exception as e:
            print(f"Erreur parse_attijari_releve_text: {e}")
            return []
    
    def extract_transaction_from_line(self, line, year):
        """Extrait une transaction d'une ligne de texte Attijari RELEVE"""
        try:
            # Pattern plus flexible pour détecter les dates
            date_patterns = [
                r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{4})',  # DD/MM/YYYY ou DD MM YYYY
                r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{2})',   # DD/MM/YY ou DD MM YY
                r'^(\d{1,2}\s+\d{1,2})'               # DD MM (sans année)
            ]
            
            date_match = None
            for pattern in date_patterns:
                date_match = re.match(pattern, line)
                if date_match:
                                break
            
            if not date_match:
                return None
            
            date_str = date_match.group(1).replace(' ', '/')
            print(f"DEBUG - Date trouvée: {date_str}")
            
            # Extraire le reste de la ligne après la date
            rest_line = line[date_match.end():].strip()
            print(f"DEBUG - Reste de ligne: {rest_line}")
            
            # Chercher les montants avec patterns flexibles
            amount_patterns = [
                r'([\d\s,\.]+)',           # Format général
                r'(\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?)',  # Format avec séparateurs
                r'(\d+(?:[.,]\d+)?)',      # Format simple
            ]
            
            amount_matches = []
            for pattern in amount_patterns:
                matches = re.findall(pattern, rest_line)
                amount_matches.extend(matches)
            
            if not amount_matches:
                return None
            
            # Filtrer les montants plausibles
            plausible_amounts = []
            for amount in amount_matches:
                amount_clean = amount.replace(' ', '').replace(',', '.')
                try:
                    value = float(amount_clean)
                    if 1 < value < 1000000:  # Ignorer les soldes
                        plausible_amounts.append(amount.strip())
                    except:
                        pass
                
            if not plausible_amounts:
                return None
            
            # Prendre le dernier montant trouvé
            amount_str = plausible_amounts[-1]
            print(f"DEBUG - Montant trouvé: {amount_str}")
            
            # Extraire le libellé en supprimant TOUS les montants
            libelle = rest_line
            for amount in plausible_amounts:
                libelle = libelle.replace(amount, '').strip()
            
            # Nettoyer le libellé
            libelle = self.clean_libelle_from_dates(libelle)
            libelle = re.sub(r'\s+', ' ', libelle)
            
            # Classifier le montant
            debit, credit = self.classify_amount(libelle, amount_str)
            
            # Créer la transaction
            transaction = {
                'Date': date_str,
                'Libellé': libelle,
                'Débit': debit,
                'Crédit': credit
            }
            
            return transaction
            
        except Exception as e:
            print(f"Erreur extract_transaction_from_line: {e}")
            return None
    
    def classify_amount(self, libelle, amount_str):
        """Classifie un montant comme débit ou crédit basé sur le libellé"""
        try:
            libelle_upper = libelle.upper()
            
            # Mots-clés pour les DÉBITS (sorties d'argent)
            debit_keywords = [
                'COM & TVA', 'REJET', 'REGLEMENT CHEQUE', 'ABONNEMENT', 'AGIOS',
                'COMMISSION', 'COM DEPASSEMENT', 'FRAIS', 'RETRAIT', 'CARTE',
                'PRÉLÈVEMENT', 'PRELEVEMENT', 'COTISATION', 'TVA', 'PRLV',
                'DÉBIT', 'DEBIT', 'VIR EMIS', 'VIREMENT EMIS', 'FRAIS PHOTO',
                'FRAIS HUISSIER', 'REGL HUISS', 'PACK BUSINESS', 'SOLUTIONS',
                'DONT TVA', 'MEME BQ', 'DEPASSEMENT PONCTUEL'
            ]
            
            # Mots-clés pour les CRÉDITS (entrées d'argent)
            credit_keywords = [
                'VIR RECU', 'VIREMENT RECU', 'ENCAISSEMENT', 'VERSEMENT', 
                'REMBOURSEMENT', 'CRÉDIT', 'CREDIT', 'VIR. RECU',
                'VIREMENT REÇU', 'RECU', 'REÇU', 'VIR RECU TN',
                'VIR RECU TN MEME', 'VERSEMENT ESPECE', 'ENCAISSEMENT CHEQUE',
                'REMBOURSEMENT FRAIS', 'CREDIT COMPTE', 'AVANCE', 'FACT'
            ]
            
            # Vérifier d'abord les débits
            if any(keyword in libelle_upper for keyword in debit_keywords):
                print(f"DEBUG - Classification DÉBIT: {libelle}")
                return self.format_amount(amount_str), ""
            
            # Puis les crédits
            if any(keyword in libelle_upper for keyword in credit_keywords):
                print(f"DEBUG - Classification CRÉDIT: {libelle}")
                return "", self.format_amount(amount_str)
            
            # Par défaut, analyser le contexte
            if any(word in libelle_upper for word in ['PAIEMENT', 'FRAIS', 'COM', 'AGIOS']):
                print(f"DEBUG - Classification DÉBIT (contexte): {libelle}")
                return self.format_amount(amount_str), ""
                    else:
                print(f"DEBUG - Classification CRÉDIT (défaut): {libelle}")
                return "", self.format_amount(amount_str)
                
        except Exception as e:
            print(f"Erreur classify_amount: {e}")
            return "", ""
    
    def clean_libelle_from_dates(self, libelle):
        """Supprime les dates valeur du début du libellé et les montants SOLDE de la fin"""
        try:
            if not libelle:
                return libelle
            
            # Supprimer les dates au début
            date_pattern = r'^(\d{1,2}[/\s]\d{1,2}[/\s]\d{4})\s*'
            libelle_cleaned = re.sub(date_pattern, '', libelle).strip()
            
            # CORRECTION: Supprimer TOUS les montants de la fin (plus agressif)
            # Pattern pour tous les montants à la fin (positifs et négatifs)
            amount_patterns = [
                r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Montants simples
                r'\s*\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?-\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Format montant-balance
                r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?-\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Format négatif-balance
                r'\s*\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$',  # Montants positifs simples
            ]
            
            for pattern in amount_patterns:
                libelle_cleaned = re.sub(pattern, '', libelle_cleaned).strip()
            
            # CORRECTION: Supprimer aussi les fragments de montants restants
            # Exemple: "-2,713.339" -> "-2,713" -> "-2" -> ""
            while re.search(r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$', libelle_cleaned):
                libelle_cleaned = re.sub(r'\s*-?\d{1,3}(?:[,\s]\d{3})*(?:[.,]\d{1,3})?\s*$', '', libelle_cleaned).strip()
            
            print(f"DEBUG - Libellé avant nettoyage: '{libelle}'")
            print(f"DEBUG - Libellé après nettoyage: '{libelle_cleaned}'")
            
            return libelle_cleaned
            
        except Exception as e:
            print(f"Erreur clean_libelle_from_dates: {e}")
            return libelle
    
    def parse_fallback_text(self, text, year):
        """Fonction de fallback pour analyser le texte brut si aucune transaction n'est trouvée"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG FALLBACK - {len(lines)} lignes à analyser")
            
            for line_num, line in enumerate(lines):
                line = line.strip()
                if not line or len(line) < 10:
                    continue
                
                # Chercher des patterns de transaction plus larges
                transaction_patterns = [
                    r'(\d{1,2}[/\s]\d{1,2}[/\s]?\d{0,4})\s+(.+?)\s+([\d\s,\.]+)$',  # Date + description + montant
                    r'(\d{1,2}[/\s]\d{1,2})\s+(.+?)\s+([\d\s,\.]+)$',              # Date + description + montant (sans année)
                ]
                
                for pattern in transaction_patterns:
                    match = re.search(pattern, line)
                if match:
                        date_part = match.group(1)
                        description = match.group(2)
                        amount = match.group(3)
                        
                        # Nettoyer la date
                        if len(date_part.split('/')) == 2 or len(date_part.split(' ')) == 2:
                            date_str = f"{date_part}/{year}" if '/' in date_part else f"{date_part.replace(' ', '/')}/{year}"
                        else:
                            date_str = date_part.replace(' ', '/')
                        
                        # Nettoyer la description
                        description = self.clean_libelle_from_dates(description)
                        description = re.sub(r'\s+', ' ', description).strip()
                        
                        # Classifier le montant
                        debit, credit = self.classify_amount(description, amount)
                        
                        if debit or credit:
                    transaction = {
                                'Date': date_str,
                                'Libellé': description,
                                'Débit': debit,
                                'Crédit': credit
                            }
                    transactions.append(transaction)
                            print(f"DEBUG FALLBACK - Transaction créée: {transaction}")
                        break
        
        return transactions
    
        except Exception as e:
            print(f"Erreur parse_fallback_text: {e}")
            return []
    
    def clean_and_deduplicate_transactions(self, transactions):
        """Nettoie et déduplique les transactions"""
        try:
            unique_transactions = []
            seen_transactions = set()
            
            for transaction in transactions:
                # Créer une clé unique
                key = (
                    transaction.get('Date', ''),
                    transaction.get('Libellé', '').strip(),
                    transaction.get('Débit', ''),
                    transaction.get('Crédit', '')
                )
                
                if key not in seen_transactions:
                    seen_transactions.add(key)
                    unique_transactions.append(transaction)
                else:
                    print(f"DEBUG - Transaction dupliquée ignorée: {transaction}")
            
            print(f"DEBUG - {len(transactions)} transactions originales, {len(unique_transactions)} uniques")
            return unique_transactions
            
        except Exception as e:
            print(f"Erreur clean_and_deduplicate_transactions: {e}")
            return transactions
    
    def format_amount(self, amount_str):
        """Formate le montant avec virgule comme séparateur décimal"""
        try:
            if not amount_str or amount_str.strip() == "":
                return ""
            
            # Nettoyer le montant
            amount_str = amount_str.strip()
            cleaned = amount_str.replace(" ", "")
            
            # Si le point est utilisé comme séparateur décimal, le remplacer par virgule
            if "." in cleaned and "," not in cleaned:
                parts = cleaned.split(".")
                if len(parts) == 2 and len(parts[1]) <= 3:  # Probablement décimal
                    cleaned = cleaned.replace(".", ",")
            elif "," in cleaned and "." in cleaned:
                # Format "1,234.56" -> "1234,56"
                cleaned = cleaned.replace(",", "").replace(".", ",")
            
            # Vérifier que c'est un montant valide
            try:
                test_value = float(cleaned.replace(",", "."))
                if test_value > 0:
                return cleaned
            except:
                pass
            
            return ""
        except:
            return ""
    
    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel"""
        try:
            # Créer un nouveau workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "J03"  # Nom de la feuille
            
            # Définir les styles
            header_font = Font(name='Arial', size=12, bold=True, color='000000')
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Jaune
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Écrire les en-têtes
            headers = ['Date', 'Libellé', 'Débit', 'Crédit']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Écrire les données
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Formatage spécial pour les montants
                    if col_idx in [3, 4]:  # Débit et Crédit
                        if value and str(value).strip():
                            try:
                                if ',' in str(value):
                                    excel_value = str(value).replace(',', '.')
                                    cell.value = float(excel_value)
                                    cell.number_format = '#,##0.000'
                                else:
                                    cell.value = float(str(value))
                                    cell.number_format = '#,##0.000'
                            except:
                                cell.value = value
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx == 1:  # Date
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 2:  # Libellé
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 70  # Libellé
            ws.column_dimensions['C'].width = 16  # Débit
            ws.column_dimensions['D'].width = 16  # Crédit
            
            # Ajuster la hauteur des lignes
            for row in range(1, len(df) + 2):
                ws.row_dimensions[row].height = 20
            
            # Ajouter des bordures à toute la plage de données
            from openpyxl.utils import get_column_letter
            max_row = len(df) + 1
            max_col = len(headers)
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).border = border
            
            # Sauvegarder le fichier
            wb.save(excel_path)
            
        except Exception as e:
            print(f"Erreur lors du formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, engine='openpyxl')
    
    def convert_pdf_to_excel(self):
        """Fonction principale de conversion RELEVE"""
        if not self.pdf_file_path.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
            return
        
        if not self.excel_filename.get():
            messagebox.showerror("Erreur", "Veuillez entrer un nom pour le fichier Excel")
            return

        self.result_text.delete(1.0, tk.END)
        
        try:
            # Vérifier que c'est un PDF RELEVE Attijari Bank
            if not self.is_attijari_releve_pdf(self.pdf_file_path.get()):
                messagebox.showwarning("Attention",
                    "Ce fichier ne semble pas être un RELEVE Attijari Bank.\n"
                    "La conversion peut ne pas fonctionner correctement.")
            
            # Extraire les données
            transactions = self.extract_table_data(self.pdf_file_path.get())
            
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF RELEVE")
                return
            
            # Créer le DataFrame
            df = pd.DataFrame(transactions)
            
            # Chemin de sortie dans le dossier Téléchargements
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = os.path.join(downloads_path, f"{self.excel_filename.get()}.xlsx")
            
            # Sauvegarder en Excel avec formatage
            self.save_excel_with_formatting(df, excel_path)
            
            # Afficher les résultats
            self.result_text.insert(tk.END, f"Conversion RELEVE réussie !\n")
            self.result_text.insert(tk.END, f"Fichier créé: {excel_path}\n")
            self.result_text.insert(tk.END, f"Nombre de transactions: {len(transactions)}\n\n")
            self.result_text.insert(tk.END, "Aperçu des données:\n")
            self.result_text.insert(tk.END, df.head(10).to_string(index=False))
            
            # Message de succès plus positif
            success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
            success_msg += f"📁 Fichier créé: {excel_path}\n"
            success_msg += f"📊 Nombre de transactions: {len(transactions)}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            messagebox.showinfo("✅ Conversion réussie", success_msg)
            
        except Exception as e:
            error_msg = f"Erreur lors de la conversion RELEVE: {str(e)}"
            self.result_text.insert(tk.END, error_msg)
            
            messagebox.showerror("Erreur", error_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

def main():
    root = tk.Tk()
    app = AttijariReleveConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
