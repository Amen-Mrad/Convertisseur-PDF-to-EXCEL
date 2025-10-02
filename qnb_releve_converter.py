import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import unicodedata
import fitz  # PyMuPDF
import cv2
import numpy as np
from PIL import Image
import subprocess
import sys
try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False
import io
import threading

class QNBReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVÉ QNB vers Excel")
        self.root.geometry("600x420")
        self.root.configure(bg='#f0f0f0')

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"RELEVE_QNB_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self.setup_ui()

    def setup_ui(self):
        title = tk.Label(self.root, text="Convertisseur RELEVÉ QNB", font=("Arial", 18, "bold"), bg="#f0f0f0", fg="#2c3e50")
        title.pack(pady=(20,4))
        tk.Label(self.root, text="Conversion PDF vers Excel", font=("Arial", 12), bg="#f0f0f0", fg="#666").pack()

        frm = tk.Frame(self.root, bg="#f0f0f0")
        frm.pack(pady=10, padx=40, fill='x')

        tk.Label(frm, text="Fichier PDF RELEVÉ QNB:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(frm, width=48, textvariable=self.pdf_path).grid(row=1, column=0, padx=(0,10))
        tk.Button(frm, text="Parcourir", command=self.browse_pdf, bg="#3498db").grid(row=1, column=1)

        tk.Label(frm, text="Nom du fichier Excel:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', pady=(15,5))
        tk.Entry(frm, width=48, textvariable=self.excel_name).grid(row=3, column=0, padx=(0,10))

        # Boutons
        buttons_frame = tk.Frame(self.root, bg="#f0f0f0")
        buttons_frame.pack(pady=10)
        
        self.convert_button = tk.Button(buttons_frame, text="Convertir en Excel", command=self.convertir, bg="#820422", font=("Arial", 13, "bold"))
        self.convert_button.pack(side='left', padx=5)
        
        self.retour_button = tk.Button(buttons_frame, text="Retour page d'accueil", command=self.retour_accueil, bg="#95a5a6", font=("Arial", 11, "bold"))
        self.retour_button.pack(side='left', padx=5)

        tk.Label(self.root, text="Étape 1: Détection du fichier (logo + mots-clés)", bg="#f0f0f0", fg="#7f8c8d", font=("Arial", 9)).pack(side='bottom', pady=10)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF QNB", filetypes=[["PDF", "*.pdf"]])
        if path:
            self.pdf_path.set(path)
            base = os.path.splitext(os.path.basename(path))[0]
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.excel_name.set(f"RELEVE_QNB_{base}_{ts}")

    # --- Détection ---
    def detect_qnb(self, pdf_path: str) -> bool:
        """Détection robuste: mots-clés accent-insensibles + présence d'images du bandeau."""
        def norm(s: str) -> str:
            s = unicodedata.normalize('NFKD', s or '')
            s = ''.join(ch for ch in s if not unicodedata.combining(ch))
            return s.lower()

        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Concaténer le texte des 3 premières pages
                pages = pdf.pages[:3] if len(pdf.pages) >= 3 else pdf.pages
                text_joined = '\n'.join([(p.extract_text() or '') for p in pages])
                T = norm(text_joined)

                # Mots-clés QNB (sans accents) - plus larges
                keywords_any = [
                    'qnb', 'releve bancaire', 'releve bancaire', 'current account'
                    'debit', 'credit', 'solde', 'tunisian dinar', 'resume du compte'
                    'type de compte', 'account (iban)', 'inward clearing', 'chq deposit'
                    'cheque returned', 'charges/fees', 'local transfer', 'pos'
                ]
                # Au moins 2 indices au total (plus tolérant)
                found = sum(1 for k in keywords_any if k in T)

                # Indices forts: "qnb" + (releve bancaire ou resume du compte)
                strong = ('qnb' in T) and (('releve bancaire' in T) or ('resume du compte' in T))

                if found >= 2 or strong:
                    return True

                # Si le texte est pauvre, tenter: image(s) présentes + au moins 1 mot-clé
                has_images = any(len(p.images) > 0 for p in pages)
                if has_images and found >= 1:
                    return True

        except Exception:
            return False
        return False

    def is_scanned_pdf(self, pdf_path: str) -> bool:
        """Détecte si le PDF est scanné (image-based)"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Vérifier les 3 premières pages
                for page in pdf.pages[:3]:
                    text = page.extract_text()
                    if text and len(text.strip()) > 50:  # Si on trouve du texte substantiel
                        return False
                return True
        except Exception:
            return True

    def extract_text_with_ocr(self, pdf_path: str) -> str:
        """Extrait le texte d'un PDF scanné avec OCR"""
        if not TESSERACT_AVAILABLE:
            print("Tesseract OCR non disponible. Installation requise.")
            return ""
        try:
            doc = fitz.open(pdf_path)
            full_text = ""
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                # Convertir la page en image
                mat = fitz.Matrix(2.0, 2.0)  # Zoom 2x pour meilleure qualité
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                # Convertir en image PIL
                img = Image.open(io.BytesIO(img_data))
                img_array = np.array(img)
                
                # Convertir en niveaux de gris
                if len(img_array.shape) == 3:
                    gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
                else:
                    gray = img_array
                
                # Améliorer l'image pour OCR
                gray = cv2.medianBlur(gray, 3)
                gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
                
                # OCR avec Tesseract
                text = pytesseract.image_to_string(gray, lang='fra+eng')
                full_text += text + "\n"
            
            doc.close()
            return full_text
        except Exception as e:
            print(f"Erreur OCR: {e}")
            return ""

def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF QNB.")
            return

        # Désactiver le bouton pendant le traitement
        self.convert_button.config(state='disabled')
        self.progress['value'] = 10
        self.root.update_idletasks()
        
        try:
            print(f"DEBUG QNB - PDF sélectionné: {path}")
        except Exception:
            pass

        # Lancer la conversion dans un thread séparé
        thread = threading.Thread(target=self._convertir_thread, args=(path))
        thread.daemon = True
        thread.start()

    def _convertir_thread(self, path):
        """Conversion dans un thread séparé pour éviter le blocage de l'UI"""
        try:
            # Détection avant conversion (tolérante). Même si échec, on tente la conversion.
            detected = self.detect_qnb(path)
            self.root.after(0, lambda: self.progress.config(value=20))

            transactions = self.parse_qnb_pdf(path)
            self.root.after(0, lambda: self.progress.config(value=80))

            if not transactions:
                msg = "Aucune transaction trouvée dans le PDF."
                if not detected:
                    msg += " (Fichier non reconnu comme QNB)"
                self.root.after(0, lambda: messagebox.showerror("Aucune transaction", msg))
                self.root.after(0, lambda: self.convert_button.config(state='normal'))
                return

            # Sauvegarde Excel (date, libelle, debit, credit)
            df = pd.DataFrame(transactions, columns=["date", "libelle", "debit", "credit"])
            downloads = os.path.join(os.path.expanduser("~"), "Downloads")
            out_name = f"{self.excel_name.get().strip() or 'RELEVE_QNB'}.xlsx"
            out_path = os.path.join(downloads, out_name)
            df.to_excel(out_path, index=False)
            self._format_excel(out_path)
            
            self.root.after(0, lambda: self.progress.config(value=100))
            # Message de succès plus positif
            success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
            success_msg += f"📁 Fichier enregistré: {out_path}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            self.root.after(0, lambda: messagebox.showinfo("✅ Conversion réussie", success_msg))
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Erreur", f"Erreur: {e}"))
        finally:
            # Réactiver le bouton
            self.root.after(0, lambda: self.convert_button.config(state='normal'))

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    # --- Parsing QNB ---
    def parse_qnb_pdf(self, pdf_path: str):
        """Parse QNB statement into a list of dicts with keys: date, libelle, debit, credit"""
        results = []
        date_re = re.compile(r"^(\d{2}/\d{2}/\d{4,6})")
        # Pattern pour QNB: montants avec point pour milliers et virgule pour décimales
        amount_re = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{3}")

        def to_float(s: str):
            if not s:
                return None
            # Nettoyer le string
            s = s.strip()
            negative = s.startswith('-')
            s = s.lstrip('-')
            
            # Gérer les formats QNB variés
            if ',' in s and '.' in s:
                # Format: 6.000,000 -> retirer les points (milliers) puis remplacer virgule
                s = s.replace('.', '').replace(',', '.')
            elif ',' in s:
                # Format: 5,700 ou 6688, 198 -> remplacer virgule
                s = s.replace(',', '.')
            elif ' ' in s:
                # Format: 22 596 712 -> remplacer espaces par points
                s = s.replace(' ', '.')
            
            try:
                val = float(s)
                return -val if negative else val
            except Exception:
                return None

        # Vérifier si le PDF est scanné
        is_scanned = self.is_scanned_pdf(pdf_path)
        print(f"DEBUG QNB - PDF scanné: {is_scanned}")
        
        if is_scanned:
            # Utiliser OCR pour extraire tout le texte
            print("DEBUG QNB - Utilisation OCR...")
            full_text = self.extract_text_with_ocr(pdf_path)
            lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]
            print(f"DEBUG QNB - OCR: {len(lines)} lignes extraites")
            if lines:
                print(f"DEBUG QNB - Premières lignes OCR: {lines[:5]}")
        else:
            # Utiliser pdfplumber normal
            with pdfplumber.open(pdf_path) as pdf:
                all_lines = []
                for page_num, page in enumerate(pdf.pages):
                    text = page.extract_text() or ''
                    page_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                    all_lines.extend(page_lines)
                    print(f"DEBUG QNB - Page {page_num + 1}: {len(page_lines)} lignes")
                lines = all_lines
                print(f"DEBUG QNB - Total: {len(lines)} lignes")
                if lines:
                    print(f"DEBUG QNB - Premières lignes: {lines[:5]}")

        # Parser les lignes
        i = 0
        while i < len(lines):
            line = lines[i]
            m = date_re.match(line)
            if not m:
                i += 1
                continue

            # Corriger années mal OCRisées à 6 chiffres
            raw_date = m.group(1)
            if len(raw_date.split('/')[-1]) == 6:
                y = raw_date.split('/')[-1][-4:]
                op_date = f"{raw_date[:6]}{y}"
            else:
                op_date = raw_date

            # Accumuler toutes les lignes de cette transaction jusqu'à la prochaine date
            j = i + 1
            transaction_lines = [line]
            while j < len(lines) and not date_re.match(lines[j]):
                transaction_lines.append(lines[j])
                j += 1

            # Chercher le montant dans toutes les lignes de la transaction
            amount = None
            for tx_line in transaction_lines:
                # Pattern plus large pour capturer tous les formats de montants QNB
                # Format: -6.000,000 ou 5.700,000 ou 22596.712 ou 12501 156 ou 6688, 198
                amounts = re.findall(r'-?\d{1,3}(?:[\.\s]\d{3})*(?:[\.\s]\d{3})?', tx_line)
                if amounts:
                    # Prendre le premier montant trouvé qui ressemble à un vrai montant
                    for amt_str in amounts:
                        amt_val = to_float(amt_str)
                        if amt_val is not None and amt_val != 0 and abs(amt_val) > 0.01:
                            amount = amt_val
                            break
                    if amount is not None:
                        break

            # Construire libellé complet (préserver les retours à la ligne comme dans le PDF)
            libelle_parts = []
            for tx_line in transaction_lines:
                # Enlever la date du début de la première ligne
                if tx_line == transaction_lines[0]:
                    libelle_part = tx_line[m.end():].strip()
                else:
                    libelle_part = tx_line
                if libelle_part:
                    # Préserver intégralement le libellé (ne pas supprimer les montants)
                    libelle_part = re.sub(r'\s+', ' ', libelle_part.strip())
                    if libelle_part and len(libelle_part) > 2:  # Ne garder que si il reste du texte
                        libelle_parts.append(libelle_part)
            
            # Utiliser des sauts de ligne pour refléter les multiples lignes du PDF
            libelle_full = '\n'.join(libelle_parts).strip()

            # Classification débit/crédit
            debit = None
            credit = None
            if amount is not None:
                # Classification basée sur les mots-clés QNB spécifiques
                upper = libelle_full.lower()
                
                # Mots-clés pour les crédits QNB
                credit_keywords = [
                    'chq deposit-clearing', 'deposit', 'clearing', 'transfer'
                    'versement', 'encaissement', 'remboursement', 'recu', 'credit'
                    'cash', 'pos', 'virement recu'
                ]
                
                # Mots-clés pour les débits QNB
                debit_keywords = [
                    'inward clearing', 'charges/fees', 'tax on charges', 'cheque returned'
                    'debit', 'paiement', 'withdrawal', 'retrait', 'chgs-cheque returned'
                    'chgs-coll', 'frais', 'commission'
                ]
                
                # Vérifier les crédits en premier
                is_credit = any(kw in upper for kw in credit_keywords)
                is_debit = any(kw in upper for kw in debit_keywords)
                
                if is_credit and not is_debit:
                    # Transaction de crédit
                    credit = abs(amount)  # Toujours positif en crédit
                elif is_debit and not is_credit:
                    # Transaction de débit
                    debit = abs(amount)  # Toujours positif en débit
                else:
                    # Par défaut, utiliser le signe du montant
                    if amount < 0:
                        debit = abs(amount)
                    else:
                        credit = amount

            # Filtrer lignes évidentes de solde initial/total
            upper_lib = libelle_full.upper()
            if re.search(r"(?i)solde\s+initial|resume du compte", upper_lib):
                i = j
                continue

            # Éviter d'avoir le même montant en débit et crédit
            if debit and credit and abs(debit - credit) < 0.001:
                # Si les montants sont identiques, garder seulement le crédit (versements)
                if any(kw in libelle_full.upper() for kw in ["VERSEMENT", "DEPOT", "ENCAISSEMENT", "REM COM", "AV.TPE", "VIREMENT RECU"]):
                    debit = None
                else:
                    credit = None
            elif debit and credit and (debit < 0.01 or credit < 0.01):
                # Si l'un des montants est quasi-nul, le supprimer
                if debit < 0.01:
                    debit = None
                else:
                    credit = None
            
            print(f"DEBUG QNB - Transaction: {op_date} | {libelle_full[:50]}... | D:{debit} C:{credit}")
            results.append({
                'date': op_date
                'libelle': libelle_full
                'debit': debit
                'credit': credit
            })

            i = j

        # Nettoyage simple: retirer lignes sans montant du tout si elles sont vides
        cleaned = []
        for r in results:
            if (r['debit'] is None and r['credit'] is None) and len(r['libelle']) < 3:
                continue
            cleaned.append(r)
        return cleaned

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "J03"
        # Header style
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        # Number format 3 decimals
        max_row = ws.max_row
        pattern = '#\u00A0##0,000'
        for r in range(2, max_row + 1):
            ws[f'C{r}'].number_format = pattern
            ws[f'D{r}'].number_format = pattern
        # Borders
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)

def main():
    root = tk.Tk()
    app = QNBReleveConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()

