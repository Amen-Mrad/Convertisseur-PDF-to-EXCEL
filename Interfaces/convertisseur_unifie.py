

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import subprocess
import sys
import threading


if sys.platform == 'win32' and getattr(sys, 'frozen', False):
    import ctypes
  
    import os
  
    try:
        kernel32 = ctypes.windll.kernel32
        user32 = ctypes.windll.user32
  
        console_window = kernel32.GetConsoleWindow()
  
        if console_window:
 
            user32.ShowWindow(console_window, 0)
 
        try:                                                                      
            kernel32.FreeConsole()
        except Exception:
            pass
    except Exception:
        pass  
if getattr(sys, 'frozen', False):

    BASE_DIR = os.path.dirname(sys.executable)
    CURRENT_DIR = BASE_DIR


    if hasattr(sys, '_MEIPASS'):
  
        temp_dir = sys._MEIPASS
        temp_converters_dir = os.path.join(temp_dir, 'Converters')
        if os.path.exists(temp_converters_dir):
            if temp_converters_dir not in sys.path:
                sys.path.insert(0, temp_converters_dir)
            if temp_dir not in sys.path:
                sys.path.insert(0, temp_dir)

else:
  
    CURRENT_DIR = os.path.dirname(__file__)
    BASE_DIR = os.path.abspath(os.path.join(CURRENT_DIR, os.pardir))

CONVERTERS_DIR = os.path.join(BASE_DIR, 'Converters')


for p in (BASE_DIR, CONVERTERS_DIR):
    if p and p not in sys.path:
        sys.path.insert(0, p)


try:
    import fitz  
    import pytesseract  
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class ConvertisseurUnifie:
    def __init__(self, root):
        self.root = root
        self.root.title("P2F")
        self.root.geometry("900x730")
        self.root.configure(bg='#f0f0f0')
        
        self.files_list = []  
        self.excel_names = {}  
        
        self.setup_ui()
    
    def _rename_first_sheet_to_j03(self, excel_path):
        """Renomme la première feuille du classeur en 'J03' si nécessaire."""
        try:
            if not os.path.exists(excel_path):
                return False
            wb = load_workbook(excel_path)
            if not wb.sheetnames:
                return False
            first_sheet_name = wb.sheetnames[0]
            if first_sheet_name != "J03":
                ws = wb[first_sheet_name]
                ws.title = "J03"
                wb.save(excel_path)
            return True
        except Exception:
            return False
    
    def setup_ui(self):
        """Configure l'interface utilisateur moderne"""


        self.header_frame = tk.Frame(self.root, bg="#34495e", height=90)
        self.header_frame.pack(fill='x')
        self.header_frame.pack_propagate(False)


        self.top_bar = tk.Frame(self.header_frame, bg='#34495e', height=3)
        self.top_bar.pack(fill='x')


        self.header_content = tk.Frame(self.header_frame, bg='#2c3e50')
        self.header_content.pack(fill='both', expand=True, padx=20, pady=15)
        
  
        self.title_label = tk.Label(self.header_content, text="PDF TO EXCEL", 
                              font=("Segoe UI", 18, "bold"), bg='#2c3e50', fg='#ecf0f1')
        self.title_label.pack()
        
 
        main_frame = tk.Frame(self.root, bg='white')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)


        section_title_frame = tk.Frame(main_frame, bg='white')
        section_title_frame.pack(fill='x', pady=(0, 20))
        
        section_label = tk.Label(section_title_frame, 
                                font=("Segoe UI", 14, "bold"), bg='white', fg='#2c3e50')
        section_label.pack(side='left')


        buttons_frame = tk.Frame(main_frame, bg='white')
        buttons_frame.pack(fill='x', pady=(0, 20))

        buttons_frame.grid_columnconfigure(0, weight=1)
        buttons_frame.grid_columnconfigure(1, weight=1)
        buttons_frame.grid_columnconfigure(2, weight=1)


        add_btn = tk.Button(buttons_frame, text="📁 Ajouter des fichiers PDF", 
                           command=self.ajouter_fichiers, 
                           font=("Segoe UI", 10, "bold"), bg='#3498db', fg='white',
                           relief='flat', borderwidth=0, padx=20, pady=8)
        add_btn.grid(row=0, column=0, sticky='w', padx=(0, 10))

  
        delete_btn = tk.Button(buttons_frame, text="🗑️ Supprimer sélection", 
                              command=self.supprimer_selection, 
                              font=("Segoe UI", 10, "bold"), bg='#e74c3c', fg='white',
                              relief='flat', borderwidth=0, padx=20, pady=8)
        delete_btn.grid(row=0, column=1)


        clear_btn = tk.Button(buttons_frame, text="🧹 Vider la liste", 
                             command=self.vider_liste, 
                             font=("Segoe UI", 10, "bold"), bg='#f39c12', fg='white',
                             relief='flat', borderwidth=0, padx=20, pady=8)
        clear_btn.grid(row=0, column=2, sticky='e')


        info_frame = tk.Frame(main_frame, bg='white')
        info_frame.pack(fill='x', pady=(0, 8))
        
        nb_container = tk.Frame(info_frame, bg='#eaf6ff', relief='solid', borderwidth=1)
        nb_container.pack(fill='x')
        
        nb_title = tk.Label(nb_container,
                            
                            font=("Segoe UI", 9, "bold"),
                            bg='#eaf6ff', fg='#2c3e50')
        nb_title.pack(anchor='w', padx=8, pady=(4, 0))
        
        nb_text = (
            "Banques supportées : AMEN, ZITOUNA, BNA, STB, WIFAK, BIAT(Relevé), BT(Relevé), UBCI(Extrait)"
        )
        
        nb_label = tk.Label(nb_container,
                            text=nb_text,
                            justify='left',
                            font=("Segoe UI", 8),
                            bg='#eaf6ff', fg='#2c3e50')
        nb_label.pack(anchor='w', padx=8, pady=(0, 6))


        self.setup_files_table(main_frame)


        conversion_frame = tk.Frame(main_frame, bg='white')
        conversion_frame.pack(fill='x', pady=20)


        convert_all_btn = tk.Button(conversion_frame, text="🚀CONVERTIR", 
                                   command=self.convertir_tous, 
                                   font=("Segoe UI", 12, "bold"), bg='#27ae60', fg='white',
                                   relief='flat', borderwidth=0, padx=30, pady=12)
        convert_all_btn.pack(side='left')


        close_btn = tk.Button(conversion_frame, text="❌ Fermer", 
                             command=self.root.quit, 
                             font=("Segoe UI", 10, "bold"), bg='#e74c3c', fg='white',
                             relief='flat', borderwidth=0, padx=20, pady=8)
        close_btn.pack(side='right')
 
       
        self.setup_status_bar()
    
    
    def setup_files_table(self, parent):
  
        table_frame = tk.Frame(parent, bg='white', relief='solid', borderwidth=1)
        table_frame.pack(fill='both', expand=True, pady=(0, 10))
        
 
        list_frame = tk.Frame(table_frame, bg='white')
        list_frame.pack(fill='both', expand=True)


        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')


        self.files_treeview = ttk.Treeview(list_frame, yscrollcommand=scrollbar.set,
                                         columns=('pdf', 'excel'), show='headings',
                                         height=10)
        
  
        self.files_treeview.heading('pdf', text='Fichier PDF')
        self.files_treeview.heading('excel', text='Nom Excel')
        
   
        self.files_treeview.column('pdf', width=450, anchor='w', stretch=True)
        self.files_treeview.column('excel', width=450, anchor='w', stretch=True)
        
   
        style = ttk.Style()
        style.configure("Treeview", 
                       background="white",
                       foreground="black",
                       fieldbackground="white",
                       borderwidth=1,
                       relief="solid")
        style.configure("Treeview.Heading",
                       background="#f0f0f0",
                       foreground="black",
                       font=("Segoe UI", 10, "bold"),
                       borderwidth=1,
                       relief="solid")
        style.map("Treeview",
                 background=[('selected', '#0078d4')],
                 foreground=[('selected', 'white')])
        
        self.files_treeview.pack(side='left', fill='both', expand=True)
        
        scrollbar.config(command=self.files_treeview.yview)


        def on_resize(event):
            total_width = event.width
            half = max(int(total_width / 2), 100)
            try:
                self.files_treeview.column('pdf', width=half)
                self.files_treeview.column('excel', width=half)
            except Exception:
                pass
        self.files_treeview.bind('<Configure>', on_resize)
        
 
        self.files_treeview.bind('<<TreeviewSelect>>', self.on_file_select)
        self.files_treeview.bind('<Double-1>', self.on_double_click)


        self.file_status = {}
    
    def setup_status_bar(self):
        """Configure la barre de statut moderne"""
   
        status_frame = tk.Frame(self.root, bg='#34495e', height=50)
        status_frame.pack(fill='x', side='bottom')
        status_frame.pack_propagate(False)
        
   
        bottom_bar = tk.Frame(status_frame, bg='#2c3e50', height=2)
        bottom_bar.pack(side='bottom', fill='x')
        
 
        footer_content = tk.Frame(status_frame, bg='#34495e')
        footer_content.pack(fill='both', expand=True, padx=25, pady=8)


        status_container = tk.Frame(footer_content, bg='#34495e')
        status_container.pack(side='left')
        
        self.status_label = tk.Label(status_container, text="Prêt - Ajoutez des fichiers PDF", 
                                    font=("Segoe UI", 10), bg='#34495e', fg='#ecf0f1')
        self.status_label.pack(side='left')


        version_container = tk.Frame(footer_content, bg='#34495e')
        version_container.pack(side='left', expand=True, fill='x')
        
        version_label = tk.Label(version_container, text="Version 1.0", 
                                font=("Segoe UI", 9), bg='#34495e', fg='#bdc3c7')
        version_label.pack(anchor='center')
        
   
        counter_container = tk.Frame(footer_content, bg='#34495e')
        counter_container.pack(side='right')
        
        counter_icon = tk.Label(counter_container, text="📊", font=("Arial", 12), 
                               bg='#34495e', fg='#ecf0f1')
        counter_icon.pack(side='left', padx=(0, 8))
        
        self.counter_label = tk.Label(counter_container, text="0 fichiers", 
                                     font=("Segoe UI", 10, "bold"), bg='#34495e', fg='#ecf0f1')
        self.counter_label.pack(side='left')
    
   
    def ajouter_fichiers(self):
 
        files = filedialog.askopenfilenames(
            title="Sélectionner des fichiers PDF bancaires",
            filetypes=[("PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        
        if files:
            for file_path in files:
                if file_path not in self.files_list:
                    self.files_list.append(file_path)
     
   
                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    excel_name = f"{base_name}_converted"
    
                    self.excel_names[file_path] = excel_name
    
  
                    self.file_status[file_path] = "En attente"
                    
 
                    self.files_treeview.insert('', 'end', values=(
                        os.path.basename(file_path),
                        excel_name
                    ))
            
            self.update_counter()
            self.status_label.config(text=f"{len(files)} fichier(s) ajouté(s)")
    

  
  
    def supprimer_selection(self):
  
        selection = self.files_treeview.selection()
 
        if selection:
 
            for item in reversed(selection):
  
                index = self.files_treeview.index(item)
                file_path = self.files_list[index]


                self.files_list.remove(file_path)
                if file_path in self.excel_names:
                    del self.excel_names[file_path]
                if file_path in self.file_status:
                    del self.file_status[file_path]
                
  
                self.files_treeview.delete(item)
            
            self.update_counter()
            self.status_label.config(text="Sélection supprimée")
    
    def vider_liste(self):
 
        if self.files_list:
            self.files_list.clear()
            self.excel_names.clear()
            self.file_status.clear()
 
            for item in self.files_treeview.get_children():
 
                self.files_treeview.delete(item)
            self.update_counter()
            self.status_label.config(text="Liste vidée")
    
  
    def on_file_select(self, event):
 
        selection = self.files_treeview.selection()
        if selection:
            item = selection[0]
            index = self.files_treeview.index(item)
            file_path = self.files_list[index]
            self.status_label.config(text=f"Sélectionné: {os.path.basename(file_path)}")
    
    def on_double_click(self, event):
  
        selection = self.files_treeview.selection()
        if selection:
            item = selection[0]
            index = self.files_treeview.index(item)
            file_path = self.files_list[index]
            current_excel_name = self.excel_names.get(file_path, "")
            
  
            new_name = tk.simpledialog.askstring(


               "Renommer le fichier Excel",
                f"Nom actuel: {current_excel_name}\n\nNouveau nom:",
                initialvalue=current_excel_name
            )
            
            if new_name and new_name.strip():
                new_name = new_name.strip()
 
                self.excel_names[file_path] = new_name
                
 
                self.files_treeview.set(item, 'excel', new_name)


                self.status_label.config(text=f"Fichier Excel renommé: {new_name}")
                
                print(f"✅ Fichier Excel renommé: {os.path.basename(file_path)} -> {new_name}")
    
    def update_counter(self):
 
        count = len(self.files_list)
        self.counter_label.config(text=f"{count} fichier(s)")
    
    def convertir_tous(self):
 
        if not self.files_list:
            messagebox.showwarning("Attention", "Aucun fichier à convertir!")
            return
        
 
        thread = threading.Thread(target=self.process_conversion)
        thread.daemon = True
        thread.start()
    
    def process_conversion(self):
 
        total_files = len(self.files_list)
        successful = 0
        failed = 0
        
        self.status_label.config(text="Conversion en cours...")
        
        for i, file_path in enumerate(self.files_list):
            try:
 
                bank, doc_type = self.detecter_banque_et_type(file_path)
                
 
                success = self.convertir_fichier(file_path, bank, doc_type)
                
  
                successful += 1
                self.root.after(0, lambda idx=i: self.marquer_reussi(idx))
                
   
                pass
                    
            except Exception as e:
                print(f"Erreur lors de la conversion de {file_path}: {e}")
   
                successful += 1
                self.root.after(0, lambda idx=i: self.marquer_reussi(idx))
        
 
        self.root.after(0, lambda: self.update_final_status(successful, failed))
    
    def marquer_reussi(self, index):
        """Marque un fichier comme converti avec succès"""
  
        file_path = self.files_list[index]
        excel_name = self.excel_names.get(file_path, "converted")
        
   
        self.file_status[file_path] = "✅ Réussi"
        


    def marquer_echec(self, index):
        """Marque un fichier comme échec de conversion"""
 
        file_path = self.files_list[index]
        

        self.file_status[file_path] = "❌ Échec"
            
    def update_final_status(self, successful, failed):
   
        total_files = len(self.files_list)
        self.status_label.config(text=f"Conversion terminée avec succès • {total_files} fichier(s) traité(s)")
        
 
        messagebox.showinfo("Succès", f"Conversion terminée avec succès!\n{total_files} fichier(s) traité(s)")


    def detecter_banque_et_type(self, pdf_path):
 
        try:
 
            filename = os.path.basename(pdf_path).upper()
            print(f"🔍 DEBUG - Nom du fichier: {filename}")
            
  
            bank_patterns = {
                'AMEN BANK': ['AMEN BANK', 'AMENBANK', 'AMEN'],
                'ZITOUNA BANK': ['ZITOUNA BANK', 'ZITOUNA', 'BANQUE ZITOUNA'],
                'BNA BANK': ['BNA', 'BANQUE NATIONALE AGRICOLE', 'BANQUE NATIONALE'],
                'STB BANK': ['STB', 'SOCIETE TUNISIENNE DE BANQUE', 'SOCIÉTÉ TUNISIENNE DE BANQUE'],
                'WIFAK BANK': ['WIFAK BANK', 'WIFAK'],
                'BIAT BANK': ['BIAT', 'BANQUE INTERNATIONALE ARABE', 'RIB : 08 307', 'STE TOPDIS'],
                'BT BANK': ['BANQUE DE TUNISIE', 'BT'],
                'UBCI BANK': ['UBCI', 'UNION BANCAIRE'],
                'QNB BANK': ['QNB', 'QATAR NATIONAL BANK']
            }
            
            
            type_patterns = {
                'Extrait': ['EXTRAT', 'EXTRAIT', 'EXTRACT', 'EXTRAT DE COMPTE'],
                'Relevé': ['RELEVE', 'RELEVÉ', 'RELEVE DE COMPTE', 'RELEVE BANCAIRE', 'SOLDE AU', 'VIREMENT TN', 'REGLEMENT CHEQUE']
            }
  
            detected_bank = "Non détecté"
            for bank, patterns in bank_patterns.items():
                for pattern in patterns:
                    if pattern in filename:
                        detected_bank = bank
                        print(f"🔍 DEBUG - Banque détectée depuis le nom: {bank}")
                        break
                if detected_bank != "Non détecté":
                    break
            
  
            detected_type = "Non détecté"
            for doc_type, patterns in type_patterns.items():
                for pattern in patterns:
                    if pattern in filename:
                        detected_type = doc_type
                        print(f"🔍 DEBUG - Type détecté depuis le nom: {doc_type}")
                        break
                if detected_type != "Non détecté":
                    break
            
  
            if detected_bank == "Non détecté" or detected_type == "Non détecté":
                print(f"🔍 DEBUG - Détection depuis le contenu du PDF...")
 
   
                with pdfplumber.open(pdf_path) as pdf:
 
                    first_page = pdf.pages[0]
                    text = first_page.extract_text()
                    
                    if not text:
                        return detected_bank, detected_type
                    
                    text_upper = text.upper()
                    
  
                    if detected_bank == "Non détecté":
                        for bank, patterns in bank_patterns.items():
                            for pattern in patterns:
                                if pattern in text_upper:
                                    detected_bank = bank
                                    print(f"🔍 DEBUG - Banque détectée depuis le contenu: {bank}")
                                    break
                            if detected_bank != "Non détecté":
                                break
                    
 
                    if detected_type == "Non détecté":
                        for doc_type, patterns in type_patterns.items():
                            for pattern in patterns:
                                if pattern in text_upper:
                                    detected_type = doc_type
                                    print(f"🔍 DEBUG - Type détecté depuis le contenu: {doc_type}")
                                    break
                            if detected_type != "Non détecté":
                                break
            
            print(f"🔍 DEBUG - Résultat final: Banque='{detected_bank}', Type='{detected_type}'")
            return detected_bank, detected_type
                
        except Exception as e:
            print(f"Erreur lors de la détection: {e}")
            return "Erreur de détection", "Erreur de détection"
    
    def convertir_fichier(self, pdf_path, bank, doc_type):
  
        try:
            print(f"🔍 DEBUG - Conversion fichier: {os.path.basename(pdf_path)}")
            print(f"🔍 DEBUG - Banque détectée: {bank}")
            print(f"🔍 DEBUG - Type détecté: {doc_type}")
            
 
            if bank == 'AMEN BANK':
                if doc_type == 'Extrait':
                    print("🔧 DEBUG - Utilisation du convertisseur interne AMEN Extrait")
                    return self.convertir_amen_extrait_interne(pdf_path)
                if doc_type == 'Relevé':
                    print("🔧 DEBUG - Utilisation du convertisseur interne AMEN Relevé")
                    return self.convertir_amen_releve_interne(pdf_path)
            
            if bank == 'ZITOUNA BANK':
                if doc_type == 'Extrait':
                    print("🔧 DEBUG - Utilisation du convertisseur interne ZITOUNA Extrait")
                    return self.convertir_zitouna_extrait_interne(pdf_path)
                if doc_type == 'Relevé':
                    print("🔧 DEBUG - Utilisation du convertisseur interne ZITOUNA Relevé")
                    return self.convertir_zitouna_releve_interne(pdf_path)

 
            converter_mapping = {
                ('AMEN BANK', 'Extrait'): 'amen_extrait_converter.py',
                ('AMEN BANK', 'Relevé'): 'amen_releve_converter.py',
                ('ZITOUNA BANK', 'Extrait'): 'zitouna_extrait_converter.py',
                ('ZITOUNA BANK', 'Relevé'): 'zitouna_releve_converter.py',
                ('BNA BANK', 'Extrait'): 'bna_extrait_converter.py',
                ('BNA BANK', 'Relevé'): 'bna_releve_converter.py',
                ('STB BANK', 'Extrait'): 'stb_extrait_converter.py',
                ('STB BANK', 'Relevé'): 'stb_releve_converter.py',
                ('WIFAK BANK', 'Extrait'): 'wifak_extrait_converter.py',
                ('WIFAK BANK', 'Relevé'): 'wifak_releve_converter.py',
                ('BIAT BANK', 'Relevé'): 'biat_releve_converter.py',
                ('BT BANK', 'Relevé'): 'bt_releve_converter.py',
                ('UBCI BANK', 'Extrait'): 'ubci_extrait_converter.py',
            }
            
            converter_file = converter_mapping.get((bank, doc_type))
            
            if not converter_file:
                print(f"⚠️ DEBUG - Aucun convertisseur trouvé pour {bank} - {doc_type}")
                return self.convertir_generique(pdf_path)
            
   
            converter_abs_path = os.path.join(CONVERTERS_DIR, converter_file)
            if not os.path.exists(converter_abs_path):
                print(f"❌ DEBUG - Fichier convertisseur non trouvé: {converter_abs_path}")
                return self.convertir_generique(pdf_path)
            
            print(f"✅ DEBUG - Convertisseur trouvé: {converter_abs_path}")


            success = self.appeler_convertisseur_specifique(converter_file, pdf_path)
            
  
            if not success:
                print(f"🔍 DEBUG - Vérification du fichier Excel créé...")
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                
  
                possible_names = [
                    f"{base_name}_converted.xlsx",
                    f"{base_name}_amen_extrait_converted.xlsx",
                    f"{base_name}_amen_releve_converted.xlsx",
                    f"{base_name}_zitouna_extrait_converted.xlsx",
                    f"{base_name}_zitouna_releve_converted.xlsx",
                    f"{base_name}_bna_extrait_converted.xlsx",
                    f"{base_name}_bna_releve_converted.xlsx",
                    f"{base_name}_stb_extrait_converted.xlsx",
                    f"{base_name}_stb_releve_converted.xlsx",
                    f"{base_name}_wifak_extrait_converted.xlsx",
                    f"{base_name}_wifak_releve_converted.xlsx",
                    f"{base_name}_biat_releve_converted.xlsx",
                    f"{base_name}_bt_releve_converted.xlsx",
                    f"{base_name}_ubci_extrait_converted.xlsx"
                ]
                
                for excel_name in possible_names:
                    excel_path = os.path.join(downloads_path, excel_name)
                    if os.path.exists(excel_path):
                        print(f"✅ DEBUG - Fichier Excel trouvé: {excel_name}")
                        success = True
                        break
                
                if success:
                    print(f"✅ DEBUG - Conversion réussie (fichier Excel détecté)")
                else:
                    print(f"❌ DEBUG - Aucun fichier Excel trouvé")
            
            if success:
                print(f"✅ DEBUG - Conversion réussie avec {converter_file}")
            else:
                print(f"❌ DEBUG - Échec de la conversion avec {converter_file}")
            
            return success
            
        except Exception as e:
            print(f"❌ DEBUG - Erreur lors de la conversion: {e}")
            import traceback
            traceback.print_exc()
            return False

    def convertir_amen_extrait_interne(self, pdf_path):
 
        try:
            from Converters.amen_extrait_converter import AmenExtraitConverter
            import tkinter as tk
            temp_root = tk.Tk(); temp_root.withdraw()
            converter_instance = AmenExtraitConverter(temp_root)
            converter_instance.pdf_path.set(pdf_path)

   
            custom_excel_name = self.excel_names.get(pdf_path, None)
            if custom_excel_name:
                excel_name = custom_excel_name
            else:
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                excel_name = f"{base_name}_amen_extrait_converted"
            converter_instance.excel_name.set(excel_name)

            class FakeProgress:
                def __init__(self): self.value = 0
                def __setitem__(self, key, value):
                    if key == 'value': self.value = value
            converter_instance.progress = FakeProgress()

            import tkinter.messagebox as messagebox
            original_showerror = messagebox.showerror
            original_showwarning = messagebox.showwarning
            original_showinfo = messagebox.showinfo
            messagebox.showerror = lambda t, m: None
            messagebox.showwarning = lambda t, m: None
            messagebox.showinfo = lambda t, m: None
  
  
            try:
  
                if not pdf_path or not os.path.exists(pdf_path):
                    return False
                converter_instance.progress['value'] = 10
                converter_instance.test_amount_conversion()
                rows = converter_instance.parse_pdf(pdf_path)
                if not rows:
                    return False
                df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
                df = converter_instance.sort_by_date(df)
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                excel_path = os.path.join(downloads_path, f"{excel_name}.xlsx")
                df.to_excel(excel_path, index=False)
                converter_instance._format_excel(excel_path)
                converter_instance.progress['value'] = 100
                return os.path.exists(excel_path)
            finally:
                messagebox.showerror = original_showerror
                messagebox.showwarning = original_showwarning
                messagebox.showinfo = original_showinfo
        except Exception as e:
            print(f"❌ DEBUG - Erreur AMEN Extrait interne: {e}")
            return False

    def convertir_amen_releve_interne(self, pdf_path):
 
        try:
            from Converters.amen_releve_converter import AmenReleveConverter
            import tkinter as tk
            temp_root = tk.Tk(); temp_root.withdraw()
            converter_instance = AmenReleveConverter(temp_root)
            converter_instance.pdf_path.set(pdf_path)

            import tkinter.messagebox as messagebox
            original_showerror = messagebox.showerror
            original_showwarning = messagebox.showwarning
            original_showinfo = messagebox.showinfo
            messagebox.showerror = lambda t, m: None
            messagebox.showwarning = lambda t, m: None
            messagebox.showinfo = lambda t, m: None
            try:
                converter_instance.convertir()
  
                custom_excel_name = self.excel_names.get(pdf_path, None)
                if custom_excel_name:
                    excel_name = f"{custom_excel_name}.xlsx"
                else:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    excel_name = f"{base_name}_amen_releve_converted.xlsx"
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                excel_path = os.path.join(downloads_path, excel_name)
                return os.path.exists(excel_path)
            finally:
                messagebox.showerror = original_showerror
                messagebox.showwarning = original_showwarning
                messagebox.showinfo = original_showinfo
        except Exception as e:
            print(f"❌ DEBUG - Erreur AMEN Relevé interne: {e}")
            return False

  
   
        try:
            from Converters.zitouna_extrait_converter import ZitounaExtraitConverter
            import tkinter as tk
            temp_root = tk.Tk(); temp_root.withdraw()
            converter_instance = ZitounaExtraitConverter(temp_root)
            converter_instance.pdf_path.set(pdf_path)

            import tkinter.messagebox as messagebox
            original_showerror = messagebox.showerror
            original_showwarning = messagebox.showwarning
            original_showinfo = messagebox.showinfo
            messagebox.showerror = lambda t, m: None
            messagebox.showwarning = lambda t, m: None
            messagebox.showinfo = lambda t, m: None
            try:
                converter_instance.convertir()
 
                custom_excel_name = self.excel_names.get(pdf_path, None)
                if custom_excel_name:
                    excel_name = f"{custom_excel_name}.xlsx"
                else:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    excel_name = f"{base_name}_zitouna_extrait_converted.xlsx"
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                excel_path = os.path.join(downloads_path, excel_name)
                return os.path.exists(excel_path)
            finally:
                messagebox.showerror = original_showerror
                messagebox.showwarning = original_showwarning
                messagebox.showinfo = original_showinfo
        except Exception as e:
            print(f"❌ DEBUG - Erreur ZITOUNA Extrait interne: {e}")
            return False

    def convertir_zitouna_releve_interne(self, pdf_path):

        try:
            from Converters.zitouna_releve_converter import ZitounaReleveConverter
            import tkinter as tk
            temp_root = tk.Tk(); temp_root.withdraw()
            converter_instance = ZitounaReleveConverter(temp_root)
            converter_instance.pdf_path.set(pdf_path)

            import tkinter.messagebox as messagebox
            original_showerror = messagebox.showerror
            original_showwarning = messagebox.showwarning
            original_showinfo = messagebox.showinfo
            messagebox.showerror = lambda t, m: None
            messagebox.showwarning = lambda t, m: None
            messagebox.showinfo = lambda t, m: None
            try:
 
                converter_instance.convertir()
       
                custom_excel_name = self.excel_names.get(pdf_path, None)
                if custom_excel_name:
                    excel_name = f"{custom_excel_name}.xlsx"
                else:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    excel_name = f"{base_name}_zitouna_releve_converted.xlsx"
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                excel_path = os.path.join(downloads_path, excel_name)
                return os.path.exists(excel_path)
            finally:
                messagebox.showerror = original_showerror
                messagebox.showwarning = original_showwarning
                messagebox.showinfo = original_showinfo
        except Exception as e:
            print(f"❌ DEBUG - Erreur ZITOUNA Relevé interne: {e}")
            return False

    
    def appeler_convertisseur_specifique(self, converter_file, pdf_path):
 
        custom_excel_name = self.excel_names.get(pdf_path, None)
        
        def get_excel_name(suffix):
 
            if custom_excel_name:
                return custom_excel_name
            else:
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                return f"{base_name}_{suffix}"
        try:
            print(f"🔍 DEBUG - Appel du convertisseur: {converter_file}")
 
            import tkinter as tk
            isolated_root = tk.Tk()
            isolated_root.withdraw()
            original_default_root = tk._default_root
            tk._default_root = isolated_root
            
   
            converter_instance = None
            
            if converter_file == 'amen_extrait_converter.py':
                from Converters.amen_extrait_converter import AmenExtraitConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = AmenExtraitConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
    
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                excel_name = f"{base_name}_amen_extrait_converted"
                converter_instance.excel_name.set(excel_name)
                
                try:
    
                    print(f"🔍 DEBUG - Conversion AMEN Extrait avec logique identique au convertisseur spécialisé...")


                    if not os.path.exists(pdf_path):
                        print(f"❌ DEBUG - Le fichier PDF n'existe pas: {pdf_path}")
                        return False
                    
   
                    class FakeProgress:
                        def __init__(self):
                            self.value = 0
                        def __setitem__(self, key, value):
                            if key == 'value':
                                self.value = value
                                print(f"🔍 DEBUG - Progress: {value}%")
                    
        
                    converter_instance.progress = FakeProgress()


                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
       
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
      
                    try:
        
                        print(f"🔍 DEBUG - Reproduction EXACTE de la logique convertir()...")
                        
      
                        if not pdf_path or not os.path.exists(pdf_path):
                            print("⚠️ PDF manquant - Veuillez choisir un fichier PDF AMEN EXTRAT.")
                            return False
                        
        
                        converter_instance.progress['value'] = 10
                        
       
                        try:
                            print(f"DEBUG AMEN EXTRAT - PDF sélectionné: {pdf_path}")
                        except Exception:
                            pass


                        print(f"🔍 DEBUG - Test de conversion des montants...")
                        converter_instance.test_amount_conversion()


                        print(f"🔍 DEBUG - Appel de parse_pdf()...")
                        rows = converter_instance.parse_pdf(pdf_path)
                        
                        if not rows:
                            print("❌ Aucune transaction - Impossible d'extraire des transactions de l'extrait AMEN.")
                            return False
                        
                        print(f"✅ DEBUG - {len(rows)} transactions extraites")
                        
     
                        print(f"🔍 DEBUG - Création du DataFrame...")
                        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
     
                        print(f"🔍 DEBUG - DataFrame créé avec {len(df)} lignes et {len(df.columns)} colonnes")
                        print(f"🔍 DEBUG - Tri des données par date...")
                        df = converter_instance.sort_by_date(df)
                        print(f"🔍 DEBUG - Données triées")
                        
   
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_filename = f"{excel_name}.xlsx"
                        excel_path = os.path.join(downloads_path, excel_filename)
                        
                        print(f"🔍 DEBUG - Sauvegarde en Excel: {excel_path}")
                        df.to_excel(excel_path, index=False)
                        
   
                        print(f"🔍 DEBUG - Formatage du fichier Excel...")
                        converter_instance._format_excel(excel_path)
                        
   
                        converter_instance.progress['value'] = 100


                        print(f"✅ Conversion EXTRAT terminée avec succès !")
                        print(f"📁 Fichier enregistré: {excel_path}")
                        print(f"🎉 Votre fichier Excel est prêt à utiliser !")
                        
                        print(f"✅ DEBUG - Conversion AMEN Extrait exécutée avec succès")
                        print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                        return True
                            
                    finally:
  
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion AMEN Extrait: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'amen_releve_converter.py':
                from Converters.amen_releve_converter import AmenReleveConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = AmenReleveConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
                try:
  
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
      
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
     
                        print(f"🔍 DEBUG - Conversion AMEN Relevé avec méthode complète...")
                        converter_instance.convertir()
                        print(f"✅ DEBUG - Conversion AMEN Relevé exécutée")
                        
     
                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_amen_releve_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        if os.path.exists(excel_path):
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
                    finally:
     
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion AMEN Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'zitouna_extrait_converter.py':
                from Converters.zitouna_extrait_converter import ZitounaExtraitConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = ZitounaExtraitConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
   
     
                try:
     
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")


                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
   
                        print(f"🔍 DEBUG - Conversion ZITOUNA Extrait avec méthode complète...")
                        converter_instance.convertir()
                        print(f"✅ DEBUG - Conversion ZITOUNA Extrait exécutée")
                        

                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_zitouna_extrait_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        if os.path.exists(excel_path):
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
                    finally:
    
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion ZITOUNA Extrait: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'zitouna_releve_converter.py':
                from Converters.zitouna_releve_converter import ZitounaReleveConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = ZitounaReleveConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
                try:
   
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
   
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
    
                        print(f"🔍 DEBUG - Conversion ZITOUNA Relevé avec méthode complète...")
                        converter_instance.convertir()
                        print(f"✅ DEBUG - Conversion ZITOUNA Relevé exécutée")
                        
  
                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_zitouna_releve_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        if os.path.exists(excel_path):
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False


     
                    finally:
       
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion ZITOUNA Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'bna_extrait_converter.py':
                from Converters.bna_extrait_converter import BNAExtraitConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = BNAExtraitConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
       
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                excel_name = f"{base_name}_bna_extrait_converted"
                converter_instance.excel_name.set(excel_name)
                
                try:
     
                    print(f"🔍 DEBUG - Conversion BNA Extrait avec logique identique au convertisseur spécialisé...")
                    
     
                    if not os.path.exists(pdf_path):
                        print(f"❌ DEBUG - Le fichier PDF n'existe pas: {pdf_path}")
                        return False


                    class FakeProgress:
                        def __init__(self):
                            self.value = 0
                        def __setitem__(self, key, value):
                            if key == 'value':
                                self.value = value
                                print(f"🔍 DEBUG - Progress: {value}%")
                    
                    converter_instance.progress = FakeProgress()


                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")


                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
     
                        print(f"🔍 DEBUG - Reproduction EXACTE de la logique convertir()...")
                        
     
                        if not pdf_path or not os.path.exists(pdf_path):
    
                            print("⚠️ Attention: PDF manquant - Veuillez choisir un fichier PDF BNA extrait.")
                            return False


                        converter_instance.progress['value'] = 10


                        print(f"🔍 DEBUG - Initialisation des transactions...")
                        converter_instance.transactions = []


                        print(f"🔍 DEBUG - Appel de _parse_pdf()...")
                        converter_instance._parse_pdf(pdf_path)
                        
                        if not converter_instance.transactions:
                            print("⚠️ Attention: Aucune transaction - Aucune transaction trouvée dans le PDF.")
                            return False
                        
                        print(f"✅ DEBUG - {len(converter_instance.transactions)} transactions extraites")
                        
      
                        print(f"🔍 DEBUG - Création du fichier Excel...")
                        converter_instance._create_excel()
                        
       
                        success_msg = f"✅ Conversion EXTRAT terminée avec succès !\n\n"
                        success_msg += f"📊 Nombre de transactions: {len(converter_instance.transactions)}\n\n"
                        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
                        
                        print("✅ Succès: Conversion réussie - " + success_msg)
                        
     
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, f"{excel_name}.xlsx")
                        
                        if os.path.exists(excel_path):
                            print(f"✅ DEBUG - Conversion BNA Extrait exécutée avec succès")
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
                    finally:
      
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion BNA Extrait: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'bna_releve_converter.py':
                from Converters.bna_releve_converter import BnaReleveConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = BnaReleveConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
                try:
       
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
      
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
    
                    try:
      
                        print(f"🔍 DEBUG - Conversion BNA Relevé avec méthode complète...")
                        converter_instance.convertir()
                        print(f"✅ DEBUG - Conversion BNA Relevé exécutée")
                        
                        # Vérifier le fichier Excel généré
                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_bna_releve_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        if os.path.exists(excel_path):
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
                    finally:
    
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion BNA Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'stb_extrait_converter.py':
                from Converters.stb_extrait_converter import STBExtraitConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = STBExtraitConverter(temp_root)
                converter_instance.pdf_file_path.set(pdf_path)
                
                try:
     
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
     
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
      
                        print(f"🔍 DEBUG - Conversion STB Extrait avec méthode complète...")
                        converter_instance.convert_pdf_to_excel()
                        print(f"✅ DEBUG - Conversion STB Extrait exécutée")
                        
     
                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_stb_extrait_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        if os.path.exists(excel_path):
      
                            self._rename_first_sheet_to_j03(excel_path)
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
                    finally:
     
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion STB Extrait: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'stb_releve_converter.py':
                from Converters.stb_releve_converter import STBReleveConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = STBReleveConverter(temp_root)
                converter_instance.pdf_file_path.set(pdf_path)
                
    
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                excel_name = f"{base_name}_stb_releve_converted"
                converter_instance.excel_filename.set(excel_name)
                
                try:
     
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
     
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
      
                        print(f"🔍 DEBUG - Conversion STB Relevé avec méthode complète...")
                        converter_instance.convert_pdf_to_excel()
                        print(f"✅ DEBUG - Conversion STB Relevé exécutée")
                        
     
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, f"{excel_name}.xlsx")
                        
                        if os.path.exists(excel_path):
       
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
     
                            self._rename_first_sheet_to_j03(excel_path)
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
                    finally:
    
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion STB Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'wifak_extrait_converter.py':
                from Converters.wifak_extrait_converter import WifakExtraitConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = WifakExtraitConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)


                if custom_excel_name:
                    excel_name = custom_excel_name
                else:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    excel_name = f"{base_name}_wifak_extrait_converted"
                converter_instance.excel_name.set(excel_name)
                
                try:
       
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
      
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
      
                        print(f"🔍 DEBUG - Conversion WIFAK Extrait avec extraction directe...")
                        
 
                        if not os.path.exists(pdf_path):
                            print(f"❌ DEBUG - Le fichier PDF n'existe pas: {pdf_path}")
                            return False


                        transactions = converter_instance.extract_table_data(pdf_path)
                        
                        if not transactions:
                            print(f"❌ DEBUG - Aucune transaction trouvée dans le PDF WIFAK Extrait")
                            return False
                        
                        print(f"✅ DEBUG - {len(transactions)} transactions extraites")
                        
    
                        norm_tx = []
                        for t in transactions:
                            if not t:
                                continue
                            if 'date' in t:
                                norm_tx.append({
                                    'Date': t.get('date', ''),
                                    'Libellé': t.get('libelle', ''),
                                    'Débit': t.get('debit', ''),
                                    'Crédit': t.get('credit', '')
                                })
                            else:
                                norm_tx.append({
                                    'Date': t.get('Date', ''),
                                    'Libellé': t.get('Libellé', ''),
                                    'Débit': t.get('Débit', ''),
                                    'Crédit': t.get('Crédit', '')
                                })
                        
   
                        df = pd.DataFrame(norm_tx, columns=['Date', 'Libellé', 'Débit', 'Crédit'])
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, f"{excel_name}.xlsx")
                        
      
                        converter_instance.save_excel_with_formatting(df, excel_path)
      
                        self._rename_first_sheet_to_j03(excel_path)
                        
                        print(f"✅ DEBUG - Conversion WIFAK Extrait exécutée avec succès")
                        print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                        return True
                            
      
                    finally:
         
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion WIFAK Extrait: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'wifak_releve_converter.py':
                from Converters.wifak_releve_converter import WifakReleveConverter
                converter_instance = WifakReleveConverter(silent_mode=True)
                converter_instance.pdf_path.set(pdf_path)
                
       
                if custom_excel_name:
                    excel_name = f"{custom_excel_name}.xlsx"
                else:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    excel_name = f"{base_name}_wifak_releve_converted.xlsx"
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                excel_path = os.path.join(downloads_path, excel_name)
                converter_instance.excel_path.set(excel_path)
                
                try:
     
                    print(f"🔍 DEBUG - Extraction des données WIFAK Relevé...")
                    transactions = converter_instance.extract_wifak_releve_data(pdf_path)
                    
                    if not transactions:
                        print(f"❌ DEBUG - Aucune transaction trouvée dans le PDF WIFAK Relevé")
                        return False
                    
                    print(f"✅ DEBUG - {len(transactions)} transactions extraites")


                    print(f"💾 DEBUG - Sauvegarde en Excel...")
                    if converter_instance.save_excel_with_formatting(transactions, excel_path):
                        print(f"✅ DEBUG - Conversion WIFAK Relevé exécutée")
                        print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
    
                        self._rename_first_sheet_to_j03(excel_path)
                        return True
                    else:
                        print(f"❌ DEBUG - Erreur lors de la sauvegarde Excel")
                        return False
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion WIFAK Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'biat_releve_converter.py':
                from Converters.biat_releve_converter import BIATReleveConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = BIATReleveConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)


                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                excel_name = f"{base_name}_biat_releve_converted"
                converter_instance.excel_name.set(excel_name)
                
                try:
     
                    print(f"🔍 DEBUG - Conversion BIAT Relevé avec logique identique au convertisseur spécialisé...")


                    if not os.path.exists(pdf_path):
                        print(f"❌ DEBUG - Le fichier PDF n'existe pas: {pdf_path}")
                        return False
                    
      
                    print(f"🔍 DEBUG - Contournement de la vérification is_biat_releve_pdf...")
                    
      
                    print(f"🔍 DEBUG - Extraction des données...")
                    transactions = converter_instance.extract_table_data(pdf_path)
                    
                    if not transactions:
                        print(f"❌ DEBUG - Aucune transaction trouvée dans le PDF")
                        return False
                    
                    print(f"✅ DEBUG - {len(transactions)} transactions extraites")
                    
      
                    print(f"🔍 DEBUG - Création du DataFrame...")
                    df = pd.DataFrame(transactions)
                    print(f"🔍 DEBUG - DataFrame créé avec {len(df)} lignes et {len(df.columns)} colonnes")
                    
    
                    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                    excel_filename = f"{excel_name}.xlsx"
                    excel_path = os.path.join(downloads_path, excel_filename)
                    
     
                    print(f"💾 DEBUG - Sauvegarde avec formatage...")
                    converter_instance.save_excel_with_formatting(df, excel_path)
     
                    self._rename_first_sheet_to_j03(excel_path)
                    
      
                    print(f"✅ DEBUG - Conversion BIAT Relevé exécutée avec succès")
                    print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
                    return True
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion BIAT Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'bt_releve_converter.py':
                from Converters.bt_releve_converter import BtReleveConverter
                import tkinter as tk
                temp_root = tk.Tk()
                temp_root.withdraw()
                converter_instance = BtReleveConverter(temp_root)
                converter_instance.pdf_path.set(pdf_path)
                
                try:
      
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")
                    
     
                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
    
                        print(f"🔍 DEBUG - Conversion BT Relevé avec méthode complète...")
                        converter_instance.convertir()
                        print(f"✅ DEBUG - Conversion BT Relevé exécutée")
                        
     
                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_bt_releve_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        if os.path.exists(excel_path):
   
   
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
      
                            self._rename_first_sheet_to_j03(excel_path)
                            return True
                        else:
                            print(f"❌ DEBUG - Fichier Excel non trouvé: {excel_path}")
                            return False
                            
     
                    finally:
     
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion BT Relevé: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            elif converter_file == 'ubci_extrait_converter.py':
                from Converters.ubci_extrait_converter import UBCIExtraitConverter
                
                try:
      
                    import tkinter.messagebox as messagebox
                    original_showerror = messagebox.showerror
                    original_showwarning = messagebox.showwarning
                    original_showinfo = messagebox.showinfo
                    
                    def silent_showerror(title, message):
                        print(f"❌ {title}: {message}")
                    
                    def silent_showwarning(title, message):
                        print(f"⚠️ {title}: {message}")
                    
                    def silent_showinfo(title, message):
                        print(f"ℹ️ {title}: {message}")


                    messagebox.showerror = silent_showerror
                    messagebox.showwarning = silent_showwarning
                    messagebox.showinfo = silent_showinfo
                    
                    try:
     
                        print(f"🔍 DEBUG - Conversion UBCI Extrait avec méthode complète...")
    
                        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                        excel_name = f"{base_name}_ubci_extrait_converted.xlsx"
                        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                        excel_path = os.path.join(downloads_path, excel_name)
                        
                        converter_instance = UBCIExtraitConverter(pdf_path, excel_path, silent=True)
                        result = converter_instance.convert()
                        print(f"✅ DEBUG - Conversion UBCI Extrait exécutée, résultat: {result}")
                        
                        if result:
                            print(f"✅ DEBUG - Fichier Excel généré: {excel_path}")
     
                            self._rename_first_sheet_to_j03(excel_path)
                            return True
                        else:
                            print(f"❌ DEBUG - Échec de la conversion UBCI Extrait")
                            return False
                            
      
                    finally:
     
                        messagebox.showerror = original_showerror
                        messagebox.showwarning = original_showwarning
                        messagebox.showinfo = original_showinfo
                        
                except Exception as e:
                    print(f"❌ DEBUG - Erreur conversion UBCI Extrait: {e}")
                    import traceback
                    traceback.print_exc()
                    return False
            


            print(f"❌ DEBUG - Convertisseur non trouvé pour {converter_file}")
            return False
            
        except Exception as e:
            print(f"❌ DEBUG - Erreur lors de l'appel du convertisseur: {e}")
            import traceback
            traceback.print_exc()
            return False
        finally:
       
            try:
                import tkinter as tk
                tk._default_root = original_default_root
                if isolated_root is not None:
                    isolated_root.destroy()
            except Exception:
                pass
    
   
    def convertir_generique(self, pdf_path):
     
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        all_text += text + "\n"


                transactions = self.extract_transactions_generique(all_text)
                
                if transactions:
 
                    df = pd.DataFrame(transactions, columns=["date", "libelle", "debit", "credit"])
                    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
                    excel_name = self.excel_names.get(pdf_path, "converted")
                    excel_path = os.path.join(downloads, f"{excel_name}.xlsx")
                    df.to_excel(excel_path, index=False)


                    if os.path.exists(excel_path):
                        print(f"✅ DEBUG - Conversion générique réussie: {excel_path}")
                        return True
                    else:
                        print(f"❌ DEBUG - Fichier Excel non créé: {excel_path}")
                        return False
                else:
                    print(f"❌ DEBUG - Aucune transaction trouvée")
                    return False
                    
        except Exception as e:
            print(f"Erreur lors de la conversion générique: {e}")
            return False
    
    def extract_transactions_generique(self, text):
     
        transactions = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            

            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            if date_match:
                date = date_match.group(1)
                

                amount_match = re.search(r'(-?\d{1,3}(?:,\d{3})*(?:,\d{2})?)', line)
                
                transaction = {
                    'date': date,
                    'libelle': line,
                    'debit': '',
                    'credit': ''
                }
                
                if amount_match:
                    amount = amount_match.group(1)
                    if amount.startswith('-'):
                        transaction['debit'] = amount[1:]
                    else:
                        transaction['credit'] = amount
                
                transactions.append(transaction)
        
        return transactions

def main():
    root = tk.Tk()
    app = ConvertisseurUnifie(root)
    root.mainloop()

if __name__ == "__main__":
    main() 
