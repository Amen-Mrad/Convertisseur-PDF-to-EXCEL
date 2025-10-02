import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import unicodedata
import subprocess
import sys
from pdf_detector import PDFBankDetector

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class BtReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVE BT vers Excel")
        # Harmoniser avec l'interface BIAT
        self.root.geometry("600x500")
        self.root.configure(bg='#f0f0f0')

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"RELEVE_BT_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self.setup_ui()

    def setup_ui(self):
        # Titres (alignés avec BIAT)
        tk.Label(self.root, text="Convertisseur RELEVE BT", font=("Arial", 16, "bold"), bg="#f0f0f0", fg="#2c3e50").pack(pady=(20,8))
        tk.Label(self.root, text="Conversion PDF vers Excel", font=("Arial", 10), bg="#f0f0f0", fg="#666").pack(pady=(0,20))

        # Frame principal
        main_frame = tk.Frame(self.root, bg="#f0f0f0")
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)

        # Section PDF
        pdf_frame = tk.Frame(main_frame, bg="#f0f0f0")
        pdf_frame.pack(fill='x', pady=15)
        tk.Label(pdf_frame, text="Fichier PDF RELEVE BT:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 10, "bold")).pack(anchor='w', pady=(0,5))
        pdf_entry = tk.Entry(pdf_frame, width=60, textvariable=self.pdf_path, font=("Arial", 9), state='readonly')
        pdf_entry.pack(pady=5, fill='x')
        tk.Button(pdf_frame, text="Parcourir", command=self.browse_pdf, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white").pack(pady=5)

        # Section Excel
        excel_frame = tk.Frame(main_frame, bg="#f0f0f0")
        excel_frame.pack(fill='x', pady=15)
        tk.Label(excel_frame, text="Nom du fichier Excel:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 10, "bold")).pack(anchor='w', pady=(0,5))
        tk.Entry(excel_frame, width=60, textvariable=self.excel_name, font=("Arial", 9)).pack(pady=5, fill='x')

        # Boutons
        buttons_frame = tk.Frame(main_frame, bg="#f0f0f0")
        buttons_frame.pack(pady=40, fill='x')
        tk.Button(buttons_frame, text="Convertir en Excel", command=self.convertir, font=("Segoe UI", 12, "bold"), bg="green", fg="white").pack(side='left', padx=10)
        tk.Button(buttons_frame, text="Retour page d'accueil", command=self.retour_accueil, font=("Segoe UI", 12, "bold"), bg="red", fg="white").pack(side='right', padx=10)

        # Barre de progression + statut (comme BIAT)
        progress_frame = tk.Frame(main_frame, bg="#f0f0f0")
        progress_frame.pack(fill='x', pady=10)
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        self.status_label = tk.Label(progress_frame, text="Prêt", font=("Arial", 9), fg="green", bg="#f0f0f0")
        self.status_label.pack(pady=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF BT RELEVE", filetypes=[["PDF", "*.pdf"]])
        if path:
            # Détection automatique du type de document
            try:
                detector = PDFBankDetector()
                detection_result = detector.detect_document_type(path)
                
                # Vérifier si c'est un relevé BT
                if detection_result['type'] == 'releve_bt':
                    self.pdf_path.set(path)
                    base = os.path.splitext(os.path.basename(path))[0]
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    self.excel_name.set(f"RELEVE_BT_{base}_{ts}")
                    
                    # Message de confirmation discret
                    print(f"✅ Document détecté: RELEVE BT (Confiance: {detection_result['confidence']:.1%})")
                    
                else:
                    # Afficher le type détecté et demander confirmation
                    summary = detector.get_detection_summary(detection_result)
                    response = messagebox.askyesno(
                        "⚠️ Type de document détecté",
                        f"Document détecté: {summary}\n\n"
                        f"Ce convertisseur est conçu pour les relevés BT.\n"
                        f"Voulez-vous continuer quand même ?"
                    )
                    
                    if response:
                        self.pdf_path.set(path)
                        base = os.path.splitext(os.path.basename(path))[0]
                        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                        self.excel_name.set(f"RELEVE_BT_{base}_{ts}")
                    else:
                        return  # Ne pas sélectionner le fichier
                        
            except Exception as e:
                # En cas d'erreur de détection, continuer quand même
                print(f"⚠️ Erreur de détection: {str(e)} - Continuation avec le fichier sélectionné")
                self.pdf_path.set(path)
                base = os.path.splitext(os.path.basename(path))[0]
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                self.excel_name.set(f"RELEVE_BT_{base}_{ts}")

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF BT RELEVE.")
            return
        
        try:
            print(f"DEBUG BT RELEVE - PDF sélectionné: {path}")
        except Exception:
            pass
        
        rows = self.parse_pdf(path)
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire des transactions du relevé BT.")
            return
        
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'RELEVE_BT'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        
        # Message de succès plus positif
        success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n\n"
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        messagebox.showinfo("✅ Conversion réussie", success_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            self.root.destroy()
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def parse_pdf(self, pdf_path: str):
        """Parser pour les relevés BT
        Format BT: JJ | LIBELLE | VALEUR | DEBIT | CREDIT
        Dates: DD.MM.YY
        Montants: Point pour milliers, virgule pour décimales
        """
        # Expressions régulières pour capturer les données
        date_re = re.compile(r"(\d{1,2}\.\d{1,2}\.\d{2})")
        
        # Pattern pour les montants BT : point pour milliers, virgule pour décimales
        # Exemples: 744,750, 1.300,000, 2.000,000
        amount_re = re.compile(r'\d+(?:\.\d{3})*(?:,\d{2,3})?')
        
        # Patterns pour détecter les éléments à exclure des montants
        cheque_pattern = re.compile(r'\b\d{7,8}\b')  # Numéros de 7-8 chiffres
        virement_pattern = re.compile(r'BV\d{8}')     # Numéros BV
        date_pattern = re.compile(r'\d{1,2}\.\d{1,2}\.\d{2,4}')  # Dates DD.MM.YY ou DD.MM.YYYY
        year_pattern = re.compile(r'\b(19|20)\d{2}\b')  # Années 1900-2099
        day_pattern = re.compile(r'^\d{1,2}$')  # Jours seuls (1-31)
        reference_pattern = re.compile(r'\b\d{6,8}\b')  # Numéros de référence 6-8 chiffres
        
        def clean_amount(amount_str: str) -> str:
            """Nettoie et formate un montant BT"""
            if not amount_str:
                return None
            
            # Nettoyer le montant
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]
            
            # Filtrer les montants à zéro
            if cleaned in ['0', '0,000', '0.000', '000', '000,000', '000.000']:
                return None
            
            try:
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 1.234,567 (point = milliers, virgule = décimales)
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 1234,567 (virgule = décimales)
                    cleaned = cleaned.replace(',', '.')
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est des décimales ou des milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales
                        pass
                    else:
                        # Probablement des milliers
                        cleaned = cleaned.replace('.', '') + '.000'
                else:
                    # Nombre entier sans séparateurs - ajouter .000
                    cleaned = cleaned + '.000'
                
                # Convertir en float pour validation
                amount_float = float(cleaned)
                
                # Formater TOUJOURS avec 3 décimales (format BT standard)
                # Format: 1.234,567 (point pour milliers, virgule pour décimales)
                formatted = f"{abs(amount_float):,.3f}".replace(',', '.').replace('.', ',', 1)
                
                # Appliquer le signe négatif si nécessaire
                if is_negative:
                    formatted = f"-{formatted}"
                
                return formatted
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not text.strip():
                    # Fallback OCR pour les PDFs scannés
                    ocr_text = self._extract_text_via_ocr(pdf_path)
                    if ocr_text.strip():
                        text = ocr_text
                        break  # Utiliser le texte OCR pour toutes les pages
                
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                print(f"DEBUG: Page avec {len(lines)} lignes")
                
                # Afficher toutes les lignes pour debug
                for i, line in enumerate(lines):
                    print(f"DEBUG Ligne {i+1}: {line}")
                
                # 1. Extraire le mois et l'année depuis l'en-tête
                month_year = self._extract_month_year_from_header(lines)
                print(f"DEBUG: Mois/Année extrait: {month_year}")
                
                # 2. Parser les transactions avec le format BT
                for i, line in enumerate(lines):
                    print(f"DEBUG: Analyse ligne {i+1}: {line}")
                    
                    # Chercher un jour (JJ) au début de la ligne
                    day_match = re.match(r'^(\d{1,2})\s+', line)
                    
                    if day_match:
                        try:
                            day = int(day_match.group(1))
                            
                            # Construire la date complète
                            if month_year:
                                date_operation = f"{day:02d}/{month_year['month']:02d}/{month_year['year']}"
                                print(f"DEBUG: Date complète construite: {date_operation}")
                            else:
                                # Fallback: utiliser la date de la colonne VALEUR si disponible
                                dates = date_re.findall(line)
                                if dates:
                                    date_operation = self._normalize_date_str(dates[0])
                                    print(f"DEBUG: Date de fallback: {date_operation}")
                                else:
                                    continue
                            
                            # Parser les colonnes BT de manière plus précise
                            # Format: JJ | LIBELLE | VALEUR | DEBIT | CREDIT
                            
                            # Trouver les positions des colonnes en cherchant les patterns
                            parts = line.split()
                            libelle_parts = []
                            debit_amount = None
                            credit_amount = None
                            
                            # État du parsing
                            in_libelle = True
                            found_date = False
                            
                            for j, part in enumerate(parts[1:], 1):  # Ignorer le jour
                                print(f"DEBUG: Analyse partie {j}: '{part}'")
                                
                                # Vérifier si c'est une date (colonne VALEUR)
                                if date_re.match(part):
                                    print(f"DEBUG: Date trouvée (colonne VALEUR): {part}")
                                    found_date = True
                                    in_libelle = False
                                    continue
                                
                                # Vérifier si c'est un montant (colonne DEBIT ou CREDIT)
                                elif amount_re.match(part):
                                    # Cas spéciaux: certains nombres appartiennent au libellé (références, chèques, BV...)
                                    alnum_ref_pattern = re.compile(r'^[A-Z]*\d{4,12}[A-Z]*$', re.IGNORECASE)
                                    is_reference_like = (
                                        bool(cheque_pattern.match(part)) or
                                        bool(virement_pattern.match(part)) or
                                        bool(reference_pattern.match(part)) or
                                        bool(alnum_ref_pattern.match(part))
                                    )

                                    if is_reference_like and not found_date:
                                        # Conserver dans le libellé
                                        libelle_parts.append(part)
                                        print(f"DEBUG: Numéro de référence conservé dans libellé: {part}")
                                        continue

                                    # Exclure uniquement les éléments qui ne doivent pas faire partie du libellé
                                    is_excluded_non_libelle = (
                                        date_pattern.match(part) or
                                        year_pattern.match(part) or
                                        day_pattern.match(part)
                                    )

                                    if not is_excluded_non_libelle:
                                        # Nettoyer et valider le montant
                                        try:
                                            clean_num = part.replace('.', '').replace(',', '.')
                                            value = float(clean_num)
                                            if 0.001 <= value <= 999999999:
                                                # C'est un montant valide
                                                cleaned_amount = clean_amount(part)
                                                if cleaned_amount:
                                                    if not debit_amount:
                                                        debit_amount = cleaned_amount
                                                        print(f"DEBUG: Montant débit trouvé: {part} -> {cleaned_amount}")
                                                    elif not credit_amount:
                                                        credit_amount = cleaned_amount
                                                        print(f"DEBUG: Montant crédit trouvé: {part} -> {cleaned_amount}")
                                                    else:
                                                        print(f"DEBUG: Trop de montants, ignoré: {part}")
                                                else:
                                                    print(f"DEBUG: Montant non nettoyé: {part}")
                                            else:
                                                print(f"DEBUG: Montant exclu (taille): {part}")
                                        except:
                                            print(f"DEBUG: Montant exclu (format): {part}")

                                    # Une fois qu'on a rencontré un montant (réel), on n'est plus dans le libellé
                                    in_libelle = False
                                
                                # Si on est encore dans le libellé et ce n'est ni une date ni un montant
                                elif in_libelle and not found_date:
                                    # Pour le libellé, on garde TOUT sauf les dates et les jours seuls
                                    # Les numéros de chèque/virement sont importants pour le libellé
                                    is_excluded_from_libelle = (
                                        date_pattern.match(part) or  # Dates complètes seulement
                                        year_pattern.match(part) or  # Années seules
                                        (day_pattern.match(part) and len(part) <= 2)  # Jours seuls (1-31)
                                    )
                                    
                                    if not is_excluded_from_libelle:
                                        libelle_parts.append(part)
                                        print(f"DEBUG: Partie libellé ajoutée: {part}")
                                    else:
                                        print(f"DEBUG: Élément exclu du libellé: {part}")
                                
                                # Si on a trouvé une date, on ignore le reste du libellé
                                elif found_date and not amount_re.match(part):
                                    # Ignorer les parties après la date qui ne sont pas des montants
                                    print(f"DEBUG: Partie ignorée après date: {part}")
                            
                            libelle = ' '.join(libelle_parts)
                            print(f"DEBUG: Libellé final: {libelle}")
                            print(f"DEBUG: Débit trouvé: {debit_amount}, Crédit trouvé: {credit_amount}")
                            
                            # Classification des montants
                            debit = credit = None
                            
                            # Si on a deux montants, le premier est généralement débit, le second crédit
                            if debit_amount and credit_amount:
                                debit = debit_amount
                                credit = credit_amount
                            elif debit_amount and not credit_amount:
                                # Un seul montant - déterminer s'il s'agit d'un débit ou crédit
                                if any(keyword in libelle.upper() for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT', 'CREDIT', 'SOLDE', 'VIREMENT RE?U']):
                                    credit = debit_amount
                                else:
                                    debit = debit_amount
                            elif credit_amount and not debit_amount:
                                # Un seul montant crédit
                                credit = credit_amount
                            
                            print(f"DEBUG: Classification finale - Débit: {debit}, Crédit: {credit}")
                            
                            # Ajouter la transaction si elle est valide
                            if libelle.strip() and (debit or credit):
                                print(f"DEBUG: ✅ Transaction valide ajoutée - Date: {date_operation}, Libellé: {libelle[:30]}..., Débit: {debit}, Crédit: {credit}")
                                results.append({
                                    "date": date_operation,
                                    "libelle": libelle,
                                    "debit": debit,
                                    "credit": credit
                                })
                            else:
                                print(f"DEBUG: ❌ Transaction invalide - Libellé vide ou pas de montant")
                        except Exception as e:
                            print(f"DEBUG: Erreur lors du parsing de la ligne: {e}")
                            continue
        
        print(f"DEBUG: Total de {len(results)} transactions trouvées")
        return results

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "Relevé BT"
        
        # Styles
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        
        # En-têtes
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        
        # Largeurs des colonnes
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 50  # Libellé
        ws.column_dimensions['C'].width = 16  # Débit
        ws.column_dimensions['D'].width = 16  # Crédit
        
        # Format des nombres
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            ws[f'C{r}'].number_format = '# ##0,000'  # Débit
            ws[f'D{r}'].number_format = '# ##0,000'  # Crédit
        
        # Bordures
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        
        wb.save(path)

    def _extract_month_year_from_header(self, lines):
        """Extrait le mois et l'année depuis l'en-tête du document BT"""
        month_names = {
            'JANVIER': 1, 'FEVRIER': 2, 'MARS': 3, 'AVRIL': 4, 'MAI': 5, 'JUIN': 6,
            'JUILLET': 7, 'AOUT': 8, 'SEPTEMBRE': 9, 'OCTOBRE': 10, 'NOVEMBRE': 11, 'DECEMBRE': 12
        }
        
        for line in lines:
            print(f"DEBUG: Recherche mois/année dans: {line}")
            # Chercher des patterns comme "MAI 2025", "JANVIER 2025", etc.
            for month_name, month_num in month_names.items():
                pattern = rf'{month_name}\s+(\d{{4}})'
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    year = int(match.group(1))
                    print(f"DEBUG: Trouvé {month_name} {year}")
                    return {'month': month_num, 'year': year}
        
        print("DEBUG: Aucun mois/année trouvé dans l'en-tête")
        return None

    def _normalize_date_str(self, raw: str) -> str:
        """Extrait et normalise une date au format DD/MM/YYYY depuis une chaîne quelconque.
        Gère les formes: 2.05.25, 5.05.25, etc.
        """
        if not raw:
            return ''
        try:
            txt = str(raw).strip()
            # Extraire le premier motif de date
            m = re.search(r"(\d{1,2}\.\d{1,2}\.\d{2})", txt)
            if not m:
                return txt  # retourner tel quel si pas de motif
            d = m.group(1)
            # Format DD.MM.YY -> DD/MM/YYYY
            parts = d.split('.')
            day = int(parts[0])
            month = int(parts[1])
            year = int(parts[2])
            # Convertir l'année à 4 chiffres
            if year < 50:  # Années 00-49 = 2000-2049
                year += 2000
            else:  # Années 50-99 = 1950-1999
                year += 1900
            
            dt = datetime(year, month, day)
            return dt.strftime('%d/%m/%Y')
        except Exception:
            return ''

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

def main():
    root = tk.Tk()
    BtReleveConverter(root)
    root.mainloop()

if __name__ == '__main__':
    main()
