import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import re
from datetime import datetime
import subprocess
import sys
from pdf_detector import PDFBankDetector


class UBCIExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT UBCI vers Excel")
        self.root.geometry("600x420")
        self.root.configure(bg="#f0f0f0")

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"EXTRAT_UBCI_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self._build_ui()

    def _build_ui(self):
        tk.Label(self.root, text="Convertisseur EXTRAT UBCI", font=("Arial", 18, "bold"), bg="#f0f0f0", fg="#2c3e50").pack(pady=(20, 4))
        tk.Label(self.root, text="Conversion PDF vers Excel", font=("Arial", 12), bg="#f0f0f0", fg="#666").pack()

        frm = tk.Frame(self.root, bg="#f0f0f0"); frm.pack(pady=10, padx=40, fill='x')
        tk.Label(frm, text="Fichier PDF EXTRAT UBCI:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(frm, width=48, textvariable=self.pdf_path).grid(row=1, column=0, padx=(0, 10))
        tk.Button(frm, text="Parcourir", command=self.browse_pdf, bg="#3498db").grid(row=1, column=1)

        tk.Label(frm, text="Nom du fichier Excel:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', pady=(15, 5))
        tk.Entry(frm, width=48, textvariable=self.excel_name).grid(row=3, column=0, padx=(0, 10))

        buttons = tk.Frame(self.root, bg="#f0f0f0"); buttons.pack(pady=10)
        tk.Button(buttons, text="Convertir en Excel", command=self.convertir, bg="#16A34A", font=("Arial", 13, "bold")).pack(side='left', padx=5)
        tk.Button(buttons, text="Retour page d'accueil", command=self.retour_accueil, bg="#95a5a6", font=("Arial", 11, "bold")).pack(side='left', padx=5)

        self.progress = tk.ttk.Progressbar(self.root, orient='horizontal', length=500, mode='determinate') if hasattr(tk, 'ttk') else None
        if self.progress:
            self.progress.pack(pady=8)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF UBCI EXTRAT", filetypes=[["PDF", "*.pdf"]])
        if not path:
            return
        try:
            detector = PDFBankDetector()
            det = detector.detect_document_type(path)
            if det.get('type') == 'extrait_ubci':
                self.pdf_path.set(path)
                base = os.path.splitext(os.path.basename(path))[0]
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                self.excel_name.set(f"EXTRAT_UBCI_{base}_{ts}")
            else:
                summary = detector.get_detection_summary(det)
                if messagebox.askyesno("Type détecté", f"{summary}\n\nContinuer avec ce fichier dans le convertisseur UBCI ?"):
                    self.pdf_path.set(path)
                else:
                    return
        except Exception:
            self.pdf_path.set(path)

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF UBCI EXTRAT.")
            return
        if self.progress:
            self.progress['value'] = 10; self.root.update_idletasks()
        rows = self.parse_pdf(path)
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire des opérations de l'extrait UBCI.")
            return
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'EXTRAT_UBCI'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        if self.progress:
            self.progress['value'] = 100
        messagebox.showinfo("✅ Conversion réussie", f"Fichier enregistré:\n{out}")

    def retour_accueil(self):
        try:
            self.root.destroy()
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def parse_pdf(self, pdf_path: str):
        # Extraction par layout: colonnes Date opération | Natures des opérations | Débit | Crédit
        rows = self._parse_by_layout(pdf_path)
        if rows:
            return rows
        # Fallback textuel simple
        return self._parse_textual(pdf_path)

    def _parse_by_layout(self, pdf_path: str):
        def fmt_amount(s: str) -> str:
            if not s:
                return None
            v = s.replace('\u00A0', ' ').replace(' ', '').replace('.', '').replace(',', '.')
            try:
                f = float(v)
            except Exception:
                return None
            return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)

        header_stop = [
            'solde', 'total', 'page ', 'ubci - société', 'swift', 'identifiant unique', 'www.ubci.tn'
        ]
        results = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words() or []
                    if not words:
                        continue
                    # 1) Localiser la rangée d'en-tête pour connaître la zone du tableau et les limites de colonnes
                    header_candidates = [w for w in words if any(h in w['text'].lower() for h in ['date opération', 'date', 'natures', 'débit', 'debit', 'crédit', 'credit', 'valeur', 'ref'])]
                    if not header_candidates:
                        continue
                    header_y = min(w['top'] for w in header_candidates)  # ligne d'en-tête

                    # Trouver x des libellés d'en-têtes principaux
                    def find_x(label_sub):
                        xs = [w['x0'] for w in words if label_sub in w['text'].lower() and abs(w['top'] - header_y) < 30]
                        return min(xs) if xs else None

                    x_date = find_x('date')
                    x_natures = find_x('natures')
                    x_debit = find_x('débit') or find_x('debit')
                    x_credit = find_x('crédit') or find_x('credit')
                    x_date_val = find_x('valeur')
                    x_ref = find_x('ref') or find_x('banque')

                    # Si certains en-têtes manquent, estimer à partir de la largeur de page
                    page_width = page.width
                    candidates = [x for x in [x_date, x_natures, x_debit, x_credit, x_date_val, x_ref] if x is not None]
                    if not candidates:
                        continue
                    xs_sorted = sorted(set(candidates))
                    # Ordonner colonnes attendues si possibles
                    ordered = []
                    for x in [x_date, x_natures, x_debit, x_credit, x_date_val, x_ref]:
                        if x is not None:
                            ordered.append(x)
                    # Calculer les bords comme milieux entre en-têtes successifs
                    edges = [0]
                    for i in range(len(ordered) - 1):
                        edges.append((ordered[i] + ordered[i + 1]) / 2)
                    edges.append(page_width)

                    def col_bounds(x_left):
                        # Retourner (min,max) pour une colonne, basée sur l'ordre dans edges
                        if x_left is None:
                            return (0, 0)
                        for i in range(len(edges) - 1):
                            if abs(edges[i] - x_left) < 2 or edges[i] <= x_left < edges[i + 1]:
                                return (edges[i], edges[i + 1])
                        return (x_left, x_left + 50)

                    date_max = col_bounds(x_date)[1]
                    lib_min, lib_max = col_bounds(x_natures)
                    deb_min, deb_max = col_bounds(x_debit)
                    cre_min, cre_max = col_bounds(x_credit)
                    # Limiter la zone tableau verticalement: au-dessus du pied de page
                    footer_y = page.height * 0.88

                    # 2) Grouper par lignes et extraire seulement si la ligne est sous l'en-tête et au-dessus du pied
                    lines = {}
                    for w in words:
                        y_center = round((w['top'] + w['bottom']) / 2)
                        if y_center <= header_y + 10 or y_center >= footer_y:
                            continue
                        lines.setdefault(y_center, []).append(w)

                    amount_pat = re.compile(r"^-?\d[\d\s\.,]*$")
                    for yk in sorted(lines.keys()):
                        ws = sorted(lines[yk], key=lambda m: m['x0'])
                        # Ignorer lignes qui contiennent des mots de pied de page
                        if any(any(h in (w.get('text','').lower()) for h in header_stop) for w in ws):
                            continue
                        date_text = ''
                        lib_parts = []
                        debit_text = ''
                        credit_text = ''
                        for w in ws:
                            t = w.get('text', '')
                            x0, x1 = w['x0'], w['x1']
                            if x1 <= date_max and re.match(r"^\d{2}/\d{2}/\d{4}$", t):
                                date_text = t
                            elif lib_min <= x0 and x1 <= lib_max:
                                lib_parts.append(t)
                            elif deb_min <= x0 and x1 <= deb_max and amount_pat.match(t):
                                debit_text = t
                            elif cre_min <= x0 and x1 <= cre_max and amount_pat.match(t):
                                credit_text = t
                        # N'inclure que les vraies lignes du tableau: besoin d'une date ou d'un montant
                        if not (date_text or debit_text or credit_text):
                            continue
                        # Nettoyer le libellé: enlever dates et montants parasites
                        libelle_tokens = []
                        for tok in lib_parts:
                            if re.match(r"^\d{2}/\d{2}/\d{4}$", tok):
                                continue
                            if amount_pat.match(tok):
                                continue
                            libelle_tokens.append(tok)
                        libelle = ' '.join(libelle_tokens).strip()
                        results.append({
                            'date': date_text,
                            'libelle': libelle,
                            'debit': fmt_amount(debit_text),
                            'credit': fmt_amount(credit_text)
                        })
        except Exception:
            return []

        # Nettoyage dates manquantes en propageant la dernière date vue
        last_date = ''
        cleaned = []
        for r in results:
            d = r.get('date') or ''
            if re.match(r"^\d{2}/\d{2}/\d{4}$", d):
                last_date = d
            else:
                r['date'] = last_date
            # Vider chaînes vides en None pour montants
            if not r.get('debit'):
                r['debit'] = None
            if not r.get('credit'):
                r['credit'] = None
            # Filtrer les faux montants qui sont des références: séquences très longues (> 11 chiffres) sans virgule
            def is_reference(val: str) -> bool:
                if not val:
                    return False
                raw = val.replace(' ', '').replace(',', '').replace('.', '')
                return raw.isdigit() and len(raw) >= 11
            if r['debit'] and is_reference(r['debit']):
                r['debit'] = None
            if r['credit'] and is_reference(r['credit']):
                r['credit'] = None
            cleaned.append(r)
        # Garder seulement lignes avec au moins un montant et un libellé/date
        final = []
        for r in cleaned:
            if (r.get('debit') or r.get('credit')) and (r.get('libelle') or r.get('date')):
                final.append(r)
        return final

    def _parse_textual(self, pdf_path: str):
        date_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        amount_re = re.compile(r"-?\d+(?:[ \.]\d{3})*(?:[,\.]\d+)?")
        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ''
                    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                    i = 0
                    while i < len(lines):
                        line = lines[i]
                        if not date_re.match(line):
                            i += 1
                            continue
                        # Accumuler libellé jusqu'à rencontrer des montants (débit/crédit)
                        j = i
                        merged = line
                        while j + 1 < len(lines) and not date_re.match(lines[j + 1]):
                            merged += ' ' + lines[j + 1]
                            j += 1
                        amts = amount_re.findall(merged)
                        debit = credit = None
                        if len(amts) >= 2:
                            debit, credit = amts[-2], amts[-1]
                        elif len(amts) == 1:
                            # Heuristique: versement => crédit sinon débit
                            if any(k in merged.upper() for k in ['VERSEMENT', 'REMISE', 'ENCAISSEMENT']):
                                credit = amts[0]
                            else:
                                debit = amts[0]
                        # Libellé = entre date et premier montant
                        first_amt_pos = min((merged.find(a) for a in amts), default=-1)
                        libelle = merged[len(line):first_amt_pos].strip() if first_amt_pos > 0 else merged[len(line):].strip()
                        def clean_amt(a: str):
                            if not a:
                                return None
                            v = a.replace(' ', '').replace('.', '').replace(',', '.')
                            try:
                                f = float(v)
                            except Exception:
                                return None
                            return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
                        rows.append({
                            'date': date_re.match(line).group(1),
                            'libelle': libelle,
                            'debit': clean_amt(debit),
                            'credit': clean_amt(credit)
                        })
                        i = j + 1
        except Exception:
            return []
        return rows

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "Extrait UBCI"

        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            ws[f'C{r}'].number_format = '# ##0,000'
            ws[f'D{r}'].number_format = '# ##0,000'
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)


def main():
    root = tk.Tk()
    UBCIExtraitConverter(root)
    root.mainloop()


if __name__ == '__main__':
    main()


