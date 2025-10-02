import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import unicodedata
import subprocess
import sys
from pdf_detector import PDFBankDetector

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class AmenExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT AMEN BANK vers Excel")
        self.root.geometry("600x420")
        self.root.configure(bg='#f0f0f0')

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"EXTRAT_AMEN_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self.setup_ui()

    def setup_ui(self):
        tk.Label(self.root, text="Convertisseur EXTRAT AMEN", font=("Arial", 18, "bold"), bg="#f0f0f0", fg="#2c3e50").pack(pady=(20,4))
        tk.Label(self.root, text="Conversion PDF vers Excel", font=("Arial", 12), bg="#f0f0f0", fg="#666").pack()
        frm = tk.Frame(self.root, bg="#f0f0f0"); frm.pack(pady=10, padx=40, fill='x')
        tk.Label(frm, text="Fichier PDF EXTRAT AMEN:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(frm, width=48, textvariable=self.pdf_path).grid(row=1, column=0, padx=(0,10))
        tk.Button(frm, text="Parcourir", command=self.browse_pdf, bg="#3498db").grid(row=1, column=1)
        tk.Label(frm, text="Nom du fichier Excel:", bg="#f0f0f0", fg="#2c3e50", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', pady=(15,5))
        tk.Entry(frm, width=48, textvariable=self.excel_name).grid(row=3, column=0, padx=(0,10))

        # Boutons
        buttons_frame = tk.Frame(self.root, bg="#f0f0f0")
        buttons_frame.pack(pady=10)
        
        tk.Button(buttons_frame, text="Convertir en Excel", command=self.convertir, bg="#16A34A", font=("Arial", 13, "bold")).pack(side='left', padx=5)
        tk.Button(buttons_frame, text="Retour page d'accueil", command=self.retour_accueil, bg="#95a5a6", font=("Arial", 11, "bold")).pack(side='left', padx=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF AMEN EXTRAT", filetypes=[["PDF", "*.pdf"]])
        if path:
            # Détection automatique du type de document
            try:
                detector = PDFBankDetector()
                detection_result = detector.detect_document_type(path)
                
                # Vérifier si c'est un extrait Amen Bank
                if detection_result['type'] == 'extrait_amen':
                    self.pdf_path.set(path)
                    base = os.path.splitext(os.path.basename(path))[0]
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    self.excel_name.set(f"EXTRAT_AMEN_{base}_{ts}")
                    
                    # Message de confirmation discret
                    print(f"✅ Document détecté: EXTRAT AMEN BANK (Confiance: {detection_result['confidence']:.1%})")
                    
                else:
                    # Afficher le type détecté et demander confirmation
                    summary = detector.get_detection_summary(detection_result)
                    response = messagebox.askyesno(
                        "⚠️ Type de document détecté"
                        f"Document détecté: {summary}\n\n"
                        f"Ce convertisseur est conçu pour les extraits AMEN BANK.\n"
                        f"Voulez-vous continuer quand même ?"
                    )
                    
                    if response:
                        self.pdf_path.set(path)
                        base = os.path.splitext(os.path.basename(path))[0]
                        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                        self.excel_name.set(f"EXTRAT_AMEN_{base}_{ts}")
                    else:
                        return  # Ne pas sélectionner le fichier
                        
            except Exception as e:
                # En cas d'erreur de détection, continuer quand même
                print(f"⚠️ Erreur de détection: {str(e)} - Continuation avec le fichier sélectionné")
                self.pdf_path.set(path)
                base = os.path.splitext(os.path.basename(path))[0]
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                self.excel_name.set(f"EXTRAT_AMEN_{base}_{ts}")

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF AMEN EXTRAT.")
            return
        self.progress['value'] = 10; self.root.update_idletasks()
        try:
            print(f"DEBUG AMEN EXTRAT - PDF sélectionné: {path}")
        except Exception:
            pass
        rows = self.parse_pdf(path)
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire des transactions de l'extrait AMEN.")
            return
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'EXTRAT_AMEN'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        self.progress['value'] = 100
        # Message de succès plus positif
        success_msg = f"✅ Conversion EXTRAT terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n\n"
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        messagebox.showinfo("✅ Conversion réussie", success_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            self.root.destroy()
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def parse_pdf(self, pdf_path: str):
        """Parser pour les extraits AMEN BANK
        Stratégie:
        1) Layout (x/y) pour lire les colonnes Date | Libellé | Débit | Crédit
        2) Fallback: parsing textuel existant
        """
        # D'abord tenter via mise en page
        rows_layout = self._parse_layout_amen(pdf_path)
        if rows_layout:
            print(f"DEBUG AMEN EXTRAT - Layout a extrait {len(rows_layout)} lignes")
            return rows_layout
        # Expressions régulières pour capturer les données
        # Dates possibles: 31/07/2025, 31.07.2025, 31-07-2025, 31072025
        date_re = re.compile(r"(\d{2}[./-]\d{2}[./-]\d{4}|\d{8})")
        amount_re = re.compile(r"-?\d+(?:[ .]\d{3})*(?:[.,]\d+)?|-?\d+[.,]\d+")
        
        def clean_amount(amount_str: str) -> str:
            """Nettoie et formate un montant AMEN EXTRAT"""
            if not amount_str:
                return None
            
            # Nettoyer le montant
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            
            if not cleaned or cleaned == '-':
                return None
            
            # Détecter si c'est un montant négatif
            is_negative = cleaned.startswith('-')
            if is_negative:
                cleaned = cleaned[1:]
            
            # Filtrer les montants à zéro
            if cleaned in ['0', '0,000', '0.000', '000', '000,000', '000.000']:
                return None
            
            try:
                # Gérer les formats avec points et virgules
                if ',' in cleaned and '.' in cleaned:
                    # Format: 1 234,567 (espace = milliers, virgule = décimales)
                    cleaned = cleaned.replace(' ', '').replace(',', '.')
                elif ',' in cleaned and '.' not in cleaned:
                    # Format: 1234,567 (virgule = décimales)
                    cleaned = cleaned.replace(',', '.')
                elif ' ' in cleaned and ',' not in cleaned:
                    # Format: 1 234 (espace = milliers, pas de décimales)
                    cleaned = cleaned.replace(' ', '') + '.000'
                elif '.' in cleaned and ',' not in cleaned:
                    # Vérifier si c'est des décimales ou des milliers
                    parts = cleaned.split('.')
                    if len(parts) == 2 and len(parts[1]) <= 3:
                        # Probablement des décimales
                        pass
                    else:
                        # Probablement des milliers
                        cleaned = cleaned.replace('.', '') + '.000'
                
                # Convertir en float pour validation
                amount_float = float(cleaned)
                
                # Appliquer le signe négatif si nécessaire
                if is_negative:
                    amount_float = -amount_float
                
                # Formater selon le standard AMEN (virgule pour décimales, espace pour milliers)
                if amount_float == int(amount_float):
                    # Nombre entier
                    return f"{int(abs(amount_float)):,}".replace(',', ' ')
                else:
                    # Nombre avec décimales
                    formatted = f"{abs(amount_float):,.3f}".replace(',', ' ').replace('.', ',')
                    return f"-{formatted}" if is_negative else formatted
                    
            except (ValueError, TypeError):
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not text.strip():
                    # Fallback OCR pour les PDFs scannés
                    ocr_text = self._extract_text_via_ocr(pdf_path)
                    if ocr_text.strip():
                        text = ocr_text
                        break  # Utiliser le texte OCR pour toutes les pages
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                
                # Chercher les lignes de transaction
                print(f"DEBUG: Page avec {len(lines)} lignes")
                for i, line in enumerate(lines):
                    # Vérifier si la ligne contient une date
                    dates = date_re.findall(line)
                    amounts = amount_re.findall(line)
                    
                    print(f"DEBUG: Ligne {i+1}: {line[:50]}... - Dates: {len(dates)}, Montants: {len(amounts)}")
                    
                    if len(dates) >= 1:  # Au moins 1 date (plus flexible) (date opération et date valeur)
                        # Extraire les montants
                        
                        if len(amounts) >= 1:
                            # Nettoyer la ligne
                            parts = line.split()
                            
                            # Extraire les dates
                            date_operation = self._normalize_date_str(dates[0] if len(dates) > 0 else "")
                            date_valeur = self._normalize_date_str(dates[1] if len(dates) > 1 else date_operation)  # Utiliser la même date si pas de date valeur
                            
                            # Extraire le libellé (entre les dates et les montants)
                            libelle_parts = []
                            amount_started = False
                            
                            for part in parts:
                                if amount_re.match(part):
                                    amount_started = True
                                elif not amount_started and part not in dates:
                                    libelle_parts.append(part)
                            
                            libelle = ' '.join(libelle_parts)
                            
                            # Extraire le numéro de pièce (généralement après le libellé)
                            numero_piece = ""
                            if libelle_parts:
                                # Chercher un pattern de numéro de pièce
                                for part in libelle_parts:
                                    if re.match(r'^[A-Z0-9]+$', part) and len(part) > 3:
                                        numero_piece = part
                                        break
                            
                            # Traiter les montants
                            debit = credit = None
                            for amount in amounts:
                                cleaned_amount = clean_amount(amount)
                                if cleaned_amount:
                                    # Déterminer si c'est un débit ou crédit
                                    # Dans les extraits AMEN, les montants sont généralement positifs
                                    # On détermine le type par la position ou le contexte
                                    if not debit:
                                        debit = cleaned_amount
                                    elif not credit:
                                        credit = cleaned_amount
                            
                            # Si on a un seul montant, déterminer s'il s'agit d'un débit ou crédit
                            if debit and not credit:
                                # Par défaut, considérer comme crédit si c'est un versement
                                if any(keyword in libelle.upper() for keyword in ['VERSEMENT', 'ENCAISSEMENT', 'REMISE', 'DEPOT']):
                                    credit = debit
                                    debit = None
                                else:
                                    # Sinon, considérer comme débit
                                    pass
                            
                            # Ajouter la transaction si elle est valide
                            if libelle.strip() and (debit or credit):
                                print(f"DEBUG: Transaction trouvée - Date: {date_operation}, Libellé: {libelle[:30]}..., Débit: {debit}, Crédit: {credit}")
                                results.append({
                                    "date": date_operation,  # Date d'opération du PDF
                                    "libelle": libelle,      # Opération du PDF
                                    "debit": debit,          # Débit du PDF
                                    "credit": credit         # Crédit du PDF
                                })
        
        print(f"DEBUG: Total de {len(results)} transactions trouvées")
        return results

    def _parse_layout_amen(self, pdf_path: str):
        """Lecture par coordonnées pour éviter les pertes de lignes et toute classification.
        On mappe strictement les colonnes Débit et Crédit selon leur position horizontale.
        """
        def fmt_amount(s: str) -> str:
            if not s:
                return ''
            v = s.replace('\u00A0', ' ').replace(' ', '').replace('.', '').replace(',', '.')
            try:
                f = float(v)
            except Exception:
                return ''
            return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)

        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(use_text_flow=True)
                    if not words:
                        continue
                    # Détecter colonnes via en-têtes si présents
                    header_words = [w for w in words if w.get('text') and any(h in w['text'].upper() for h in ['DATE', 'LIBELL', 'NUMERO', 'NUMÉRO', 'VALEUR', 'DEBIT', 'CRÉDIT', 'CREDIT'])]
                    page_width = page.width
                    # Bornes par défaut adaptées au format Amen (A4 portrait)
                    edges = [0, 120, 520, 640, 740, page_width]
                    if header_words:
                        xs = sorted(w['x0'] for w in header_words)
                        # Construire edges grossiers
                        xs = [min(xs)] + xs + [max(w['x1'] for w in words)]
                        xs_sorted = []
                        for x in xs:
                            if not xs_sorted or x - xs_sorted[-1] > 25:
                                xs_sorted.append(x)
                        if len(xs_sorted) >= 5:
                            # Date | Libellé | Numéro+DateValeur (ignorés) | Débit | Crédit
                            # On ne garde que 6 edges
                            base = xs_sorted[:5]
                            edges = [base[0], base[1], base[2], base[3], base[4], page_width]

                    # Grouper les mots par ligne y
                    line_map = {}
                    for w in words:
                        y = int(round((w['top'] + w['bottom']) / 2))
                        line_map.setdefault(y, []).append(w)
                    for y, ws in sorted(line_map.items()):
                        ws.sort(key=lambda t: t['x0'])
                        text_line = ' '.join(w['text'] for w in ws)
                        up = text_line.upper()
                        if any(k in up for k in ['DATE OP', 'DATE OPÉ', 'LIBELLÉ', 'LIBELLE', 'NUMERO', 'NUMÉRO', 'DATE VALEUR', 'DEBIT', 'CRÉDIT', 'CREDIT', 'TOTAL', 'SOLDE', 'PAGE ']):
                            continue
                        buckets = [[] for _ in range(5)]
                        for w in ws:
                            x = w['x0']
                            for i in range(5):
                                if edges[i] <= x < edges[i+1]:
                                    buckets[i].append(w['text'])
                                    break
                        date = ' '.join(buckets[0]).strip()
                        # Construire un libellé enrichi: colonne libellé + morceaux utiles de la colonne suivante
                        def clean_token(t: str) -> str:
                            # Supprimer uniquement les vraies dates, garder le reste (même les numéros longs)
                            if re.match(r"^\d{2}[./-]\d{2}[./-]\d{4}$", t):
                                return ""
                            return t

                        lib_main = ' '.join(buckets[1]).strip()
                        lib_extra = ' '.join(clean_token(tok.strip()) for tok in buckets[2]).strip()
                        libelle = ' '.join([lib_main, lib_extra]).strip()

                        debit_raw = ' '.join(buckets[3]).strip()
                        credit_raw = ' '.join(buckets[4]).strip()
                        if not (date or libelle or debit_raw or credit_raw):
                            continue
                        # Normaliser la date (peut contenir des suffixes comme "VMA").
                        # Si ce n'est pas une date valide, vider la cellule pour n'afficher que les vraies dates.
                        date = self._normalize_date_str(date)
                        if not re.match(r'^\d{2}/\d{2}/\d{4}$', str(date)):
                            date = ''
                        debit = fmt_amount(debit_raw)
                        credit = fmt_amount(credit_raw)
                        # Ne pas classifier: on laisse vides si aucune valeur
                        # Si libellé vide mais il y a montant, garder la ligne quand même
                        if date or libelle or debit or credit:
                            rows.append({'date': date, 'libelle': libelle or '', 'debit': debit or None, 'credit': credit or None})
        except Exception as e:
            print(f"DEBUG AMEN LAYOUT ERROR: {e}")
        # Consolidation: fusionner les lignes continues (libellé sur plusieurs lignes
        # ou montants sur la ligne suivante) pour garantir 1 ligne = 1 opération.
        consolidated = []
        current = None
        last_valid_date = ''
        for r in rows:
            d = (r.get('date') or '').strip()
            lib = (r.get('libelle') or '').strip()
            deb = r.get('debit')
            cre = r.get('credit')
            if d:
                if current:
                    consolidated.append(current)
                current = {'date': d, 'libelle': lib, 'debit': deb, 'credit': cre}
                last_valid_date = d
            else:
                # Pas de date sur la ligne
                if deb or cre:
                    # Ligne avec montants: peut être une nouvelle opération dont la date est sur la ligne précédente
                    if current and (current.get('debit') or current.get('credit')):
                        # L'opération actuelle est complète, démarrer une nouvelle avec la dernière date connue
                        consolidated.append(current)
                        current = {'date': last_valid_date, 'libelle': lib, 'debit': deb, 'credit': cre}
                    else:
                        # Ajouter ces montants/libellé à l'opération en cours (ou créer si absente)
                        if not current:
                            current = {'date': last_valid_date, 'libelle': lib, 'debit': deb, 'credit': cre}
                        else:
                            if lib:
                                current['libelle'] = (current['libelle'] + ' ' + lib).strip()
                            if deb and not current.get('debit'):
                                current['debit'] = deb
                            if cre and not current.get('credit'):
                                current['credit'] = cre
                else:
                    # Pas de date ni montants: continuation de libellé
                    if current and lib:
                        current['libelle'] = (current['libelle'] + ' ' + lib).strip()
        if current:
            consolidated.append(current)

        # Pass 2: fiabiliser date/amounts et fusionner les fragments restants
        final_rows = []
        last_date = ''
        pending = None
        for r in consolidated:
            # Compléter date manquante
            if not (r.get('date') or '').strip():
                r['date'] = last_date
            else:
                last_date = r['date']
            # Démarrer/continuer l'agrégation
            if not pending:
                pending = dict(r)
                continue
            same_date = (r.get('date') == pending.get('date'))
            r_has_amount = bool(r.get('debit') or r.get('credit'))
            pending_has_amount = bool(pending.get('debit') or pending.get('credit'))
            # Cas de continuation: même date et l'une des deux lignes manque de montants
            if same_date and (not r_has_amount or not pending_has_amount):
                # Fusionner libellé
                if r.get('libelle'):
                    pending['libelle'] = (pending.get('libelle', '') + ' ' + r['libelle']).strip()
                # Compléter montants si absents
                if not pending.get('debit') and r.get('debit'):
                    pending['debit'] = r['debit']
                if not pending.get('credit') and r.get('credit'):
                    pending['credit'] = r['credit']
                continue
            # Sinon, valider la ligne en cours et passer à la suivante
            final_rows.append(pending)
            pending = dict(r)
        if pending:
            final_rows.append(pending)

        # Remplir les dates manquantes par propagation vers l'avant
        last_seen_date = ''
        for r in final_rows:
            if r.get('date'):
                last_seen_date = r['date']
            else:
                r['date'] = last_seen_date
        # Nettoyer valeurs vides -> None pour débits/crédits
        for r in final_rows:
            if not r.get('debit'):
                r['debit'] = None
            if not r.get('credit'):
                r['credit'] = None
        return final_rows

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "Extrait AMEN"
        
        # Styles
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        
        # En-têtes
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        
        # Largeurs des colonnes
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 50  # Libellé
        ws.column_dimensions['C'].width = 16  # Débit
        ws.column_dimensions['D'].width = 16  # Crédit
        
        # Format des nombres
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            ws[f'C{r}'].number_format = '# ##0,000'  # Débit
            ws[f'D{r}'].number_format = '# ##0,000'  # Crédit
        
        # Bordures
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        
        wb.save(path)

    def _normalize_date_str(self, raw: str) -> str:
        """Extrait et normalise une date au format DD/MM/YYYY depuis une chaîne quelconque.
        Gère les formes: 31/07/2025, 31.07.2025, 31-07-2025, 31072025, et ignore les suffixes (p.ex. 'VMA').
        """
        if not raw:
            return ''
        try:
            txt = str(raw).strip()
            # Extraire le premier motif de date
            m = re.search(r"(\d{2}[./-]\d{2}[./-]\d{4}|\d{8})", txt)
            if not m:
                return txt  # retourner tel quel si pas de motif
            d = m.group(1)
            if len(d) == 8 and d.isdigit():
                # Supposé DDMMYYYY
                day = int(d[0:2])
                month = int(d[2:4])
                year = int(d[4:8])
                if year < 1900 or year > 2100:
                    return ''
                dt = datetime(year, month, day)
            else:
                d = d.replace('.', '/').replace('-', '/')
                dt = datetime.strptime(d, '%d/%m/%Y')
                if dt.year < 1900 or dt.year > 2100:
                    return ''
            return dt.strftime('%d/%m/%Y')
        except Exception:
            return ''

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

def main():
    root = tk.Tk()
    AmenExtraitConverter(root)
    root.mainloop()

if __name__ == '__main__':
    main()
