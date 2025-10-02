import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import subprocess
import sys
import pdfplumber
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class BNAExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAIT BNA vers Excel")
        self.root.geometry("600x500")
        
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"EXTRAIT_BNA_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        self._ui()

    def _ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur EXTRAIT BNA",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAIT BNA:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.browse, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                              command=self.convertir, 
                              font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.back, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def browse(self):
        f = filedialog.askopenfilename(title="Choisir un PDF BNA EXTRAIT", filetypes=[["PDF", "*.pdf"]])
        if f: self.pdf_path.set(f)

    def back(self):
        try:
            self.root.destroy(); subprocess.Popen([sys.executable, "lancer_convertisseur.py"])  # noqa
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF BNA extrait.")
            return
        
        self.progress['value'] = 10
        self.root.update_idletasks()
        try:
            self.transactions = []
            self._parse_pdf(path)
            if not self.transactions:
                messagebox.showwarning("Aucune transaction", "Aucune transaction trouvée dans le PDF.")
                return
            self._create_excel()
            # Message de succès plus positif
            success_msg = f"✅ Conversion EXTRAT terminée avec succès !\n\n"
            success_msg += f"📊 Nombre de transactions: {len(self.transactions)}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            messagebox.showinfo("✅ Conversion réussie", success_msg)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la conversion : {str(e)}")
        finally:
            pass
    
    def _parse_pdf(self, path):
        with pdfplumber.open(path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        if not text.strip():
            raise ValueError("Impossible d'extraire le texte du PDF")
        
        # Extraire les informations de base
        day, month, year = None, None, None
        
        # Chercher la date dans l'en-tête
        m = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
        if m:
            day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        
        if not month or not year:
            # Chercher l'en-tête "Tunis le :09/04/25" pour extrait BNA
            m = re.search(r"Tunis\s+le\s*:\s*(\d{2})/(\d{2})/(\d{2,4})", text, re.IGNORECASE)
            if m:
                day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
                # Gérer les années à 2 chiffres
                if year < 100:
                    year += 2000
        
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        current_date = None  # Garder la date courante pour les opérations multiples
        in_table = False  # Initialiser la variable pour détecter le tableau
        
        for ln in lines:
            # Détecter le début du tableau de transactions (plus strict)
            if re.search(r"Date\s+Libellé\s+Valeur\s+Débit\s+Crédit\s+Solde|Date.*Libellé.*Valeur.*Débit.*Crédit.*Solde", ln, re.IGNORECASE):
                in_table = True
                continue
            
            # Lignes de synthèse à ignorer (ne pas sortir du tableau pour TOTAL/RÉSUMÉ)
            if in_table and re.search(r"^\s*(TOTAL|R[ÉE]SUM[ÉE])\b", ln, re.IGNORECASE):
                continue

            # Fin réelle du tableau - seulement si c'est vraiment la fin
            if in_table and re.search(r"Solde\s+à\s+ce\s+jour.*sauf\s+erreur\s+ou\s+omission", ln, re.IGNORECASE):
                in_table = False
                current_date = None
                continue

            # Ignorer les pieds de page et lignes non-transactionnelles fréquentes
            if re.search(r"^Page\s+\d+|Imprimé\s+le|Imprime\s+le", ln, re.IGNORECASE):
                continue
            
            # Si on n'est pas dans le tableau, vérifier si la ligne ressemble à une transaction
            if not in_table:
                # Vérifier si la ligne contient une date ou des montants (ligne de transaction potentielle)
                has_date = re.search(r"\d{2}/\d{2}/\d{4}", ln)
                has_amount = re.search(r"\d+[.,]\d+", ln)
                has_keywords = any(word in ln.upper() for word in ['TVA', 'COM', 'COMMISSION', 'FRAIS', 'PRÉLÈVEMENT', 'VIREMENT', 'RETRAIT', 'PAIEMENT'])
                
                if has_date or has_amount or has_keywords:
                    # C'est probablement une ligne de transaction, la traiter
                    pass
                else:
                    continue
            
            # Ligne de transaction: Date | Libellé | Valeur | Débit | Crédit | Solde
            # On repère d'abord une date (DD/MM/YYYY) au début de ligne
            mdate = re.match(r"^(\d{2})/(\d{2})/(\d{4})", ln)
            
            if mdate:
                # Nouvelle date trouvée - mettre à jour la date courante
                day, month, year = int(mdate.group(1)), int(mdate.group(2)), int(mdate.group(3))
                current_date = (day, month, year)
            elif current_date is None:
                # Pas de date et pas de date courante - vérifier si c'est une ligne de transaction valide
                # Chercher une date ailleurs dans la ligne (format "Valeur")
                date_match = re.search(r"(\d{2})/(\d{2})/(\d{4})", ln)
                if date_match:
                    day, month, year = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
                    current_date = (day, month, year)
                else:
                    # Ligne sans date - vérifier si c'est une ligne de transaction valide
                    has_amount = re.search(r"\d+[.,]\d+", ln)
                    has_keywords = any(word in ln.upper() for word in ['TVA', 'COM', 'COMMISSION', 'FRAIS', 'PRÉLÈVEMENT', 'VIREMENT', 'RETRAIT', 'PAIEMENT'])
                    
                    if not has_amount and not has_keywords:
                        continue
                    # Utiliser une date par défaut si on a un mois/année de l'en-tête
                    if month and year:
                        day = 1  # Date par défaut
                        current_date = (day, month, year)
                    else:
                        continue
            else:
                # Utiliser la date courante pour les opérations multiples
                day, month, year = current_date
            
            # Extraction des montants selon la structure BNA extrait
            # Format BNA extrait: Date | Libellé | Valeur | Débit | Crédit | Solde
            
            # CORRECTION ÉQUILIBRÉE: Extraire les montants normalement mais filtrer intelligemment
            amts = re.findall(r"\d{1,3}(?:[\s\.]\d{3})*[.,]\d{2,3}", ln)
            
            # Filtrer les montants pour exclure les années et dates
            filtered_amts = []
            for amt in amts:
                # Nettoyer le montant pour vérification
                amt_clean = amt.replace(',', '.').replace(' ', '')
                
                # Exclure les années (19xx, 20xx) - plus strict
                if re.match(r'^(19|20)\d{2}$', amt_clean):
                    continue
                # Exclure les dates (DD/MM/YY ou DD/MM/YYYY)
                if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', amt):
                    continue
                # Exclure les années courtes (25, 26, etc.)
                if re.match(r'^\d{2}$', amt_clean) and int(amt_clean) <= 99:
                    continue
                # Exclure les montants qui commencent par l'année (ex: 2025xxx)
                if re.match(r'^20\d{2}', amt_clean):
                    continue
                # Ne PAS exclure les montants qui commencent par les 2 chiffres de l'année (ex: "25 717,687")
                # On les corrigera juste après l'extraction
                filtered_amts.append(amt)
            
            amts = filtered_amts
            
            # Aussi capturer les montants simples comme "0.000" ou "0,000"
            simple_amts = re.findall(r"\b0[.,]000\b", ln)
            amts.extend(simple_amts)

            # Corriger la pollution du montant par les 2 derniers chiffres de l'année (ex: "25 717,687")
            try:
                if current_date and len(current_date) == 3 and current_date[2]:
                    yy = f"{current_date[2] % 100:02d}"  # ex: 2025 -> "25"
                    fixed_amts = []
                    # Gérer avec ou sans espace entre 'YY' et le montant
                    patterns = [
                        re.compile(rf'^{yy}\s+(\d{1,3}(?:[ \.]\d{3})*[.,]\d{2,3})$'),
                        re.compile(rf'^{yy}(\d{1,3}(?:[ \.]\d{3})*[.,]\d{2,3})$')
                    ]
                    for a in amts:
                        cleaned = a
                        for pat in patterns:
                            mlead = pat.match(cleaned)
                            if mlead:
                                cleaned = mlead.group(1)
                                break
                        fixed_amts.append(cleaned)
                    amts = fixed_amts
            except Exception:
                pass
            
            # Ne pas ignorer les lignes sans montants - elles peuvent être des sous-lignes importantes
            # On les traitera avec des montants par défaut
            
            # Debug: afficher la ligne et les montants extraits
            print(f"DEBUG LIGNE: '{ln}'")
            print(f"DEBUG MONTANTS EXTRAITS: {amts}")
            print(f"DEBUG IN_TABLE: {in_table}, CURRENT_DATE: {current_date}")
            
            debit = credit = None
            
            # Analyser le libellé pour déterminer le type de transaction
            upper = ln.upper()
            
            # Mots-clés pour les CRÉDITS (priorité absolue)
            credit_keywords = [
                'DÉRÉSERVATION', 'DERESERVATION', 'DÉBLOCAGE CRÉD', 'DEBLOCAGE CRED',
                'ENCAISSEMENT', 'VIREMENT REÇU', 'VIR REÇU', 'CRÉDIT', 'CREDIT',
                'REMBOURSEMENT', 'RETOUR', 'ANNULATION', 'CANCELLATION',
                'VERSEMENT', 'APUREMENT'
            ]
            
            # Exceptions: mots qui contiennent "CREDIT" mais sont des débits
            credit_exceptions = ['CREDIT EPS', 'EPS N°']
            
            # Mots-clés pour les DÉBITS
            debit_keywords = [
                'PRÉLÈVEMENT', 'PRELEVEMENT', 'COMMISSION', 'FRAIS', 'AGIOS',
                'VIREMENT ÉMIS', 'VIR EMIS', 'DÉBIT', 'DEBIT', 'RETRAIT',
                'PAIEMENT', 'PAYMENT', 'DÉCAISSEMENT', 'DECAISSEMENT', 'TVA',
                'REJET', 'PERCEPTION', 'UTILISATION', 'COM ', 'COM.',
                'COM DE', 'COM RETRAIT', 'COM VIR', 'COM ABONNEMENT',
                'COM CAUTIONS', 'COM AUTRES', 'COM FRAIS', 'COM VIR'
            ]
            
            # Vérifier d'abord les exceptions (CREDIT EPS est un débit, pas un crédit)
            is_credit_exception = any(exception in upper for exception in credit_exceptions)
            
            if is_credit_exception:
                # C'est une exception - forcer comme débit
                is_credit = False
                is_debit = True
            else:
                # Classification normale
                is_credit = any(k in upper for k in credit_keywords)
                is_debit = any(k in upper for k in debit_keywords)
            
            # Si aucun mot-clé spécifique, considérer comme débit par défaut
            if not is_credit and not is_debit:
                is_debit = True
            
            # Classification des montants selon le type de transaction
            # Structure BNA: Date | Libellé | Valeur | Débit | Crédit | Solde
            # On ne prend que les montants qui correspondent au type de transaction
            if len(amts) >= 1:
                # Prendre le premier montant et le classifier selon le type de transaction
                first_amount = self._fmt_amount(amts[0])
                if is_credit:
                    credit = first_amount
                    debit = None  # Pas de débit pour une transaction crédit
                else:
                    debit = first_amount
                    credit = None  # Pas de crédit pour une transaction débit
            else:
                # Aucun montant détecté - ligne de sous-détail (ex: TVA, COM, etc.)
                # Vérifier si c'est une ligne de sous-détail importante
                is_subdetail = any(word in upper for word in ['TVA', 'COM', 'COMMISSION', 'FRAIS'])
                
                if is_subdetail or is_credit or is_debit:
                    # Assigner 0,000 selon le type de transaction
                    if is_credit:
                        credit = "0,000"
                        debit = None
                    else:
                        debit = "0,000"
                        credit = None
                else:
                    # Ligne sans montants - vérifier si c'est quand même une transaction valide
                    # Si on a une date courante et que la ligne contient des mots-clés, la traiter
                    if current_date and any(word in upper for word in ['TVA', 'COM', 'COMMISSION', 'FRAIS', 'PRÉLÈVEMENT', 'VIREMENT', 'RETRAIT', 'PAIEMENT']):
                        # Assigner 0,000 par défaut
                        debit = "0,000"
                        credit = None
                    else:
                        # Ligne non-transactionnelle - ignorer
                        continue
            
            # Libellé = texte entre la date et les montants
            if mdate:
                # Ligne avec date - retirer la date en tête
                lib = ln[mdate.end():].strip()
            else:
                # Ligne sans date - prendre toute la ligne
                lib = ln.strip()
            
            if amts:
                # CORRECTION: Trouver la position du premier montant dans le libellé
                # Mais éviter de mélanger avec les dates
                for amt in amts:
                    amt_pos = lib.find(amt)
                    if amt_pos != -1:
                        # Vérifier que c'est bien un montant isolé (pas dans un mot ou une date)
                        before_char = lib[amt_pos-1] if amt_pos > 0 else ' '
                        after_char = lib[amt_pos + len(amt)] if amt_pos + len(amt) < len(lib) else ' '
                        
                        # CORRECTION: Vérifier que ce n'est pas dans une date
                        # Si c'est entouré de "/", c'est probablement une date
                        if before_char == '/' or after_char == '/':
                            continue
                        
                        # Si c'est entouré d'espaces ou de ponctuation, c'est un montant
                        if before_char in ' \t\n' and after_char in ' \t\n.,':
                            lib = lib[:amt_pos].strip()
                            break
            
            # Nettoyage agressif du libellé
            # 1. Supprimer "valeur" du libellé (insensible à la casse)
            lib = re.sub(r'\s+valeur\s*', ' ', lib, flags=re.IGNORECASE)
            
            # 2. Supprimer toutes les dates (format DD/MM/YYYY ou DD/MM/YY)
            lib = re.sub(r'\s+\d{1,2}/\d{1,2}/\d{2,4}', '', lib)
            
            # 3. Supprimer tous les montants avec espaces (format: 1 234,567 ou 1 234.567)
            lib = re.sub(r'\s+[-+]?\s*\d{1,3}(?:\s\d{3})*[.,]\d{2,3}', '', lib)
            
            # 4. Supprimer tous les montants simples (format: 123,456 ou 123.456)
            lib = re.sub(r'\s+[-+]?\s*\d+[.,]\d+', '', lib)
            
            # 5. Supprimer les nombres isolés (comme -19, 0, etc.) mais pas les années
            lib = re.sub(r'\s+[-+]?\s*\d+\s*', ' ', lib)
            
            # 6. Supprimer les années isolées (25, 26, etc.) qui restent
            lib = re.sub(r'\s+\b(19|20)\d{2}\b\s*', ' ', lib)  # 2025, 2024, etc.
            lib = re.sub(r'\s+\b\d{2}\b\s*', ' ', lib)  # 25, 26, etc.
            
            # 7. Supprimer les caractères parasites (tirets, points isolés)
            lib = re.sub(r'\s+[-.]+\s*', ' ', lib)
            
            # 8. Supprimer spécifiquement les années qui pourraient être mélangées (2025, 25)
            lib = re.sub(r'\s+2025\s*', ' ', lib)
            lib = re.sub(r'\s+25\s*', ' ', lib)
            
            # Nettoyage libellé
            lib = re.sub(r"\s+", " ", lib)
            
            # Filtrer les libellés qui ne sont pas des transactions
            if not lib or len(lib) < 1:
                continue
            
            # Filtrer toutes les informations d'en-tête et non-transactionnelles
            header_words = [
                'N° DE COMPTE', 'DEVISE', 'MONTANT BLOQUÉ', 'RELATION', 'R.I.B', 'BNAPHONE', 'BNA SMS',
                'ETAT DU COMPTE', 'CONSULTEZ VOTRE COMPTE', 'BNA E-BANKING', 'EXTRAIT DU COMPTE',
                'DINARS TUNISIEN', 'VALIDE', 'COMPOSA', 'PIN CODE', 'CODE'
            ]
            
            # Supprimer les lignes contenant des mots d'en-tête
            if any(word in lib.upper() for word in header_words):
                continue
            
            # Supprimer les lignes avec des numéros de compte ou codes (format: 010 0115 008772 K)
            if re.search(r'\d{3}\s+\d{4}\s+\d{6}\s+[A-Z]', lib):
                continue
            
            # Supprimer les lignes avec des codes PIN ou numéros de téléphone
            if re.search(r'\b\d{4}\b|\b\d{5}\b|\b8840\s+0020\b|\b85230\b', lib):
                continue
            
            # Filtrer les libellés spécifiques à supprimer
            unwanted_prefixes = [
                'MONTANT BLOQUÉ :', 'CÔTE AUTORISÉE :', 'ECHÉANCE CÔTE :', 'EXTRAIT DU COMPTE',
                'N° DE COMPTE :', 'DEVISE :', 'RELATION :', 'R.I.B :', 'ETAT DU COMPTE :'
            ]
            if any(lib.upper().startswith(prefix) for prefix in unwanted_prefixes):
                continue
            
            # Construire date complète
            date_str = self._build_date(day, month, year)
            
            # Debug: afficher les montants finaux
            print(f"DEBUG FINAL - Date: {date_str}, Libellé: '{lib[:30]}...', Débit: {debit}, Crédit: {credit}")
            print(f"DEBUG CLASSIFICATION - is_credit: {is_credit}, is_debit: {is_debit}, amts: {amts}")
            print("---")
            
            # Ajouter à la liste des transactions
            self.transactions.append({
                'date': date_str,
                'libelle': lib,
                'debit': debit,
                'credit': credit
            })

    def _fmt_amount(self, amt):
        if not amt:
            return None
        
        # Nettoyer le montant - garder la structure originale
        amt_clean = amt.replace(' ', '')  # Supprimer les espaces
        
        # CORRECTION: Si le montant commence par les 2 chiffres de l'année (ex: "25 717,687" ou "25717,687"), les retirer
        try:
            year_two = datetime.now().year % 100
            patterns = [
                re.compile(rf'^{year_two:02d}\s+(\d+(?:[ \.]\d{3})*[.,]\d{2,3})$'),  # avec espace
                re.compile(rf'^{year_two:02d}(\d+(?:[ \.]\d{3})*[.,]\d{2,3})$')       # sans espace
            ]
            for pat in patterns:
                m = pat.match(amt)
                if m:
                    amt_clean = m.group(1).replace(' ', '')
                    break
        except Exception:
            pass
        
        try:
            # Si c'est déjà un format avec virgule décimale
            if ',' in amt_clean and '.' not in amt_clean:
                # Format: 123,456 -> 123.456
                val = float(amt_clean.replace(',', '.'))
            elif '.' in amt_clean and ',' not in amt_clean:
                # Format: 123.456 -> 123.456
                val = float(amt_clean)
            elif ',' in amt_clean and '.' in amt_clean:
                # Format: 1,234.567 -> 1234.567
                val = float(amt_clean.replace(',', ''))
            else:
                # Format simple: 123456 -> 123456.000
                val = float(amt_clean)
            
            # Formater avec virgule décimale et 3 décimales
            # Important: même pour 0.0, retourner "0,000"
            return f"{val:,.3f}".replace(',', ' ').replace('.', ',')
        except:
            # Si erreur, retourner le montant original
            return amt

    def _build_date(self, day, month, year):
        if not all([day, month, year]):
            return ""
        try:
            return f"{day:02d}/{month:02d}/{year}"
        except:
            return ""

    def _create_excel(self):
        # Créer le DataFrame
        df = pd.DataFrame(self.transactions)
        
        # Créer le fichier Excel dans le dossier Downloads
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'EXTRAIT_BNA'}.xlsx")
        
        # Sauvegarder avec openpyxl pour le formatage
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='J03', index=False)
            
            # Récupérer la feuille pour le formatage
            ws = writer.sheets['J03']
            
            # Style pour l'en-tête
            header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Appliquer le style à l'en-tête
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 50  # Libellé
            ws.column_dimensions['C'].width = 15  # Débit
            ws.column_dimensions['D'].width = 15  # Crédit
            
            # Ajouter des bordures
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

if __name__ == "__main__":
    root = tk.Tk()
    app = BNAExtraitConverter(root)
    root.mainloop() 
    