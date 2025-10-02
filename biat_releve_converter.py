import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
import fitz  # PyMuPDF
import cv2
import numpy as np
from PIL import Image
import io
import subprocess
import sys

class BIATReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVE BIAT vers Excel")
        self.root.geometry("600x500")

        # Variables
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar()
        self.excel_name.set("RELEVE_BIAT_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        
        self.setup_ui()
    
    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur RELEVE BIAT",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVE BIAT:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9),
                           state='readonly')
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.select_pdf_file, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        self.convert_button = tk.Button(buttons_frame, text="Convertir en Excel",
                                       command=self.convert_pdf_to_excel, 
                                       font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        self.convert_button.pack(side='left', padx=10)
        
        self.retour_button = tk.Button(buttons_frame, text="Retour page d'accueil",
                                      command=self.retour_accueil, 
                                      font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        self.retour_button.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF RELEVE BIAT",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_path.set(file_path)
    
    def is_biat_releve_pdf(self, pdf_path):
        """Vérifie si le PDF est un RELEVE BIAT avec détection du logo et mots-clés"""
        try:
            # Vérifier les mots-clés dans le texte
            if not self._check_biat_text_keywords(pdf_path):
                return False
            
            # Vérifier la structure du tableau
            if not self._check_biat_table_structure(pdf_path):
                return False
            
            # Vérifier la présence du logo BIAT
            if not self._check_biat_logo_presence(pdf_path):
                return False
            
            return True
            
        except Exception as e:
            print(f"Erreur lors de la vérification BIAT: {e}")
            return False
    
    def _check_biat_text_keywords(self, pdf_path):
        """Vérifie la présence des mots-clés spécifiques BIAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
                
                # Mots-clés spécifiques BIAT - Plus complets
                biat_keywords = [
                    "BIAT", "BANQUE INTERNATIONALE ARABE DE TUNISIE",
                    "RELEVE", "RELEVE DE COMPTE", "RELEVE DE MOUVEMENTS", "RELEVÉ DE COMPTE",
                    "TND", "DINAR TUNISIEN", "الجملة", "TOTAUX", "الرصيد", "SOLDE",
                    "بنك تونس العربي الدولي", "كشف حساب", "صاحب الحساب", "Titulaire du compte",
                    "ENCAISSEMENT", "VIREMENT", "COMMISSION", "FRAIS", "EFFETTN",
                    "VIREMENT TN MEME BQ", "VIREMENT TN AUTRE BQ", "VIR ORD CARTE PREPAYE"
                ]
                
                text_upper = text.upper()
                found_keywords = [keyword for keyword in biat_keywords if keyword.upper() in text_upper]
                
                print(f"DEBUG BIAT - Mots-clés trouvés: {found_keywords}")
                print(f"DEBUG BIAT - Texte extrait (premiers 500 caractères): {text[:500]}")
                
                # Vérifier spécifiquement les mots-clés essentiels
                essential_keywords = ["BIAT", "RELEVE", "TND", "DINAR TUNISIEN"]
                essential_found = [kw for kw in essential_keywords if kw.upper() in text_upper]
                
                print(f"DEBUG BIAT - Mots-clés essentiels trouvés: {essential_found}")
                
                # Vérifier aussi la présence de patterns BIAT spécifiques
                biat_patterns = [
                    r"VIREMENT TN MEME BQ",
                    r"VIREMENT TN AUTRE BQ",
                    r"ENCAISSEMENT EFFETTN",
                    r"COMMISSION REM EFFETS",
                    r"SOLDE AU \d{2} \d{2} \d{4}"
                ]
                
                pattern_found = False
                for pattern in biat_patterns:
                    if re.search(pattern, text_upper):
                        print(f"DEBUG BIAT - Pattern BIAT trouvé: {pattern}")
                        pattern_found = True
                        break
                
                # Retourner True si au moins 1 mot-clé essentiel OU un pattern BIAT est trouvé
                return len(essential_found) >= 1 or pattern_found
                
        except Exception as e:
            print(f"Erreur vérification mots-clés BIAT: {e}")
            return False
    
    def _check_biat_table_structure(self, pdf_path):
        """Vérifie la structure typique des tableaux BIAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if tables:
                        print(f"DEBUG BIAT - {len(tables)} tableaux trouvés")
                        for table in tables:
                            if len(table) > 1:  # Au moins header + 1 ligne
                                # Vérifier la présence de colonnes typiques BIAT
                                header_row = table[0]
                                if header_row:
                                    header_text = ' '.join([str(cell) for cell in header_row if cell])
                                    print(f"DEBUG BIAT - Header du tableau: {header_text}")
                                    if any(keyword in header_text.upper() for keyword in 
                                          ["DATE", "LIBELLE", "MONTANT", "DEBIT", "CREDIT", "التاريخ", "بيان", "عليه", "له"]):
                                        print("DEBUG BIAT - Structure de tableau BIAT détectée")
                                        return True
                    else:
                        # Si aucun tableau trouvé, vérifier le texte brut
                        text = page.extract_text()
                        if text:
                            # Chercher des patterns de transactions dans le texte
                            if re.search(r'\d{1,2}\s+\d{1,2}\s+[A-Z\s]+\s+[\d\s.,]+', text):
                                print("DEBUG BIAT - Pattern de transaction détecté dans le texte")
                                return True
            print("DEBUG BIAT - Aucune structure de tableau BIAT détectée")
            return False
        except Exception as e:
            print(f"Erreur vérification structure BIAT: {e}")
            return False
    
    def _check_biat_logo_presence(self, pdf_path):
        """Vérifie la présence du logo BIAT"""
        try:
            # Vérifier si le fichier logo existe
            logo_path = "logo/biat.png"
            if not os.path.exists(logo_path):
                print(f"DEBUG BIAT - Logo non trouvé à {logo_path}, détection basée sur le texte")
                return True  # Fallback sur la détection textuelle
            
            # Extraire les images du PDF
            doc = fitz.open(pdf_path)
            logo_found = False
            
            for page_num in range(min(3, len(doc))):  # Vérifier les 3 premières pages
                page = doc[page_num]
                image_list = page.get_images()
                
                if image_list:
                    print(f"DEBUG BIAT - {len(image_list)} images trouvées sur la page {page_num + 1}")
                    # Pour l'instant, on considère que la présence d'images indique un logo
                    logo_found = True
                    break
            
            doc.close()
            
            if logo_found:
                print("DEBUG BIAT - Images détectées dans le PDF (logo probable)")
            else:
                print("DEBUG BIAT - Aucune image détectée, mais on continue avec la détection textuelle")
            
            # Toujours retourner True car la détection textuelle est plus fiable
            return True
            
        except Exception as e:
            print(f"Erreur vérification logo BIAT: {e}")
            return True  # Fallback sur la détection textuelle
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF BIAT - chercher avant le tableau"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages[:3]):  # Chercher dans les 3 premières pages
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines[:20]:  # Chercher dans les 20 premières lignes
                            # Chercher des patterns d'année
                            year_patterns = [
                                r'(\d{4})',  # Année simple
                                r'(\d{1,2}/\d{1,2}/(\d{4}))',  # Date avec année
                                r'(\d{1,2}-\d{1,2}-(\d{4}))',  # Date avec tirets
                            ]
                            
                            for pattern in year_patterns:
                                matches = re.findall(pattern, line)
                                for match in matches:
                                    if isinstance(match, tuple):
                                        year = match[1] if len(match) > 1 else match[0]
                                    else:
                                        year = match
                                    
                                    year_int = int(year)
                                    if 2020 <= year_int <= 2030:  # Années plausibles
                                        print(f"DEBUG BIAT - Année détectée: {year_int} dans la ligne: {line}")
                                        return year_int
            
            # Si aucune année n'est trouvée, utiliser l'année actuelle
            current_year = datetime.now().year
            print(f"DEBUG BIAT - Aucune année détectée, utilisation de l'année actuelle: {current_year}")
            return current_year
            
        except Exception as e:
            print(f"Erreur détection année BIAT: {e}")
            return datetime.now().year
    
    def extract_table_data(self, pdf_path):
        """Extrait les données du tableau des transactions BIAT - APPROCHE FLEXIBLE"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG BIAT - Traitement page {page_num + 1}")
                    
                    # Essayer d'extraire les tableaux
                    tables = page.extract_tables()
                    
                    if tables:
                        print(f"DEBUG BIAT - {len(tables)} tableaux trouvés sur la page {page_num + 1}")
                        
                        # NOUVELLE MÉTHODE: Extraire les coordonnées des mots pour classification par position
                        words_data = page.extract_words()
                        print(f"DEBUG BIAT - {len(words_data)} mots extraits avec coordonnées")
                        
                        for table_idx, table in enumerate(tables):
                            print(f"DEBUG BIAT - Traitement tableau {table_idx + 1} avec données de position")
                            transactions = self.parse_biat_table(table, words_data)
                            all_transactions.extend(transactions)
                    else:
                        # Si aucun tableau trouvé, essayer d'extraire le texte brut
                        print(f"DEBUG BIAT - Aucun tableau trouvé sur la page {page_num + 1}, extraction du texte")
                        text = page.extract_text()
                        if text:
                            year = self.detect_year_from_pdf(pdf_path)
                            transactions = self.parse_biat_transactions_from_text(text, year)
                            all_transactions.extend(transactions)
                
                print(f"DEBUG BIAT - Total transactions extraites: {len(all_transactions)}")
                return all_transactions
                
        except Exception as e:
            print(f"Erreur extraction données BIAT: {e}")
            return []
    
    def parse_biat_table(self, table, words_data=None):
        """Parse un tableau BIAT extrait par pdfplumber avec données de position"""
        transactions = []
        try:
            if len(table) < 2:
                return transactions
            
            # Détecter l'année depuis le contenu du tableau
            year = None
            for row in table:
                for cell in row:
                    if cell and isinstance(cell, str):
                        year_match = re.search(r'(\d{4})', cell)
                        if year_match:
                            year_int = int(year_match.group(1))
                            if 2020 <= year_int <= 2030:
                                year = year_int
                                break
                if year:
                    break
            
            if not year:
                year = datetime.now().year
            
            # Parser les lignes de données avec données de position
            for row in table[1:]:  # Ignorer la première ligne (header)
                if row and any(cell for cell in row if cell):
                    transaction = self.parse_biat_transaction_row(row, year, words_data)
                    if transaction:
                        transactions.append(transaction)
            
        except Exception as e:
            print(f"Erreur parsing tableau BIAT: {e}")
        
        return transactions
    
    def restore_keyword_classification(self):
        """RESTAURE l'ancienne méthode de classification par mots-clés"""
        print("DEBUG BIAT - 🔄 RESTAURATION de la classification par mots-clés")
        
        # Modifier la fonction pour utiliser seulement les mots-clés
        def analyze_biat_line_structure_keywords_only(self, line):
            """Analyse la structure d'une ligne BIAT pour détecter débit/crédit par MOTS-CLÉS SEULEMENT"""
            import re
            
            libelle_upper = line.upper()
            
            # Mots-clés qui indiquent généralement un CRÉDIT (entrée d'argent)
            credit_keywords = [
                "ENCAISSEMENT", "VERSEMENT", "DEBLOCAGE", "CREDIT",
                "VIREMENT RECU", "REMBOURSEMENT", "INTERET", "DIVIDENDE","ANNUL",
                "VIREMENT VERS", "VIREMENT DE", "VIREMENT DEPUIS",
                "COMPENSATION", "AVOIR", "REMISE", "DEPOT",
                "ENCAISSEMENT CHEQUE", "ENCAISSEMENT EFFET", "ENCAISSEMENT VIR"
            ]
            
            # Mots-clés qui indiquent généralement un DÉBIT (sortie d'argent)
            debit_keywords = [
                "VIREMENT TN AUTRE BQ", "VIREMENT TN MEME BQ", "COMMISSION",
                "FRAIS", "AGIOS", "PRELEVEMENT", "RETRAIT", "CHEQUE IMPAYE",
                "EFFET IMPAYE", "COTISATION", "REGLEMENT CHEQUE", "VIR ORD CARTE",
                "VIREMENT POUR", "VIREMENT A", "VIREMENT VERS", "VIREMENT EMIS", "REDRESSEMENT",
                "ACHAT TPE", "COM ET TVA", "COM TVA", "CHEQUE PREAVISE",
                "REGLEMENT", "PAIEMENT", "SORTIE", "DEBIT"
            ]
            
            # Vérifier d'abord les patterns BIAT spécifiques (priorité haute)
            if "VIREMENT TN AUTRE BQ" in libelle_upper:
                print(f"DEBUG BIAT - Virement CRÉDIT détecté: VIREMENT TN AUTRE BQ")
                return "credit"
            
            if "VIREMENT TN MEME BQ" in libelle_upper:
                print(f"DEBUG BIAT - Virement DÉBIT détecté: VIREMENT TN MEME BQ")
                return "debit"
            
            # Vérifier les autres mots-clés spécifiques
            for keyword in credit_keywords:
                if keyword in libelle_upper:
                    print(f"DEBUG BIAT - Mot-clé crédit détecté: {keyword}")
                    return "credit"
            
            for keyword in debit_keywords:
                if keyword in libelle_upper:
                    print(f"DEBUG BIAT - Mot-clé débit détecté: {keyword}")
                    return "debit"
            
            # Par défaut: crédit
            print(f"DEBUG BIAT - Aucun mot-clé trouvé, défaut: CRÉDIT")
            return "credit"
        
        # Remplacer la méthode
        self.analyze_biat_line_structure = analyze_biat_line_structure_keywords_only.__get__(self, type(self))
        print("DEBUG BIAT - ✅ Classification par mots-clés restaurée")
    
    def parse_biat_transaction_row(self, row, year, words_data=None):
        """Parse une ligne de transaction BIAT avec classification par POSITION"""
        try:
            # Nettoyer la ligne
            clean_row = [str(cell).strip() if cell else "" for cell in row]
            clean_row = [cell for cell in clean_row if cell]
            
            if len(clean_row) < 3:
                return None
            
            # Essayer de trouver la date, le libellé et le montant
            date = None
            libelle = ""
            montant = None
            
            # Chercher une date au début
            for i, cell in enumerate(clean_row):
                if self.is_date_biat(cell):
                    date = cell
                    # Le libellé est probablement dans les cellules suivantes
                    libelle_parts = []
                    for j in range(i + 1, len(clean_row)):
                        # Inclure les identifiants numériques/alphanumériques (ex: 0833313, 0000202, S0311023)
                        token = clean_row[j]
                        ref_like = bool(re.match(r'^[A-Z]*\d{4,12}[A-Z]*$', str(token).strip(), re.IGNORECASE))
                        # Exclure uniquement les VRAIS montants (avec virgule/décimales ou séparateurs de milliers)
                        if (not self.is_probable_money(token) and not self.is_date_valeur_biat(token)) or ref_like:
                            libelle_parts.append(token)
                        elif self.is_probable_money(token):
                            montant = token
                            break
                    libelle = " ".join(libelle_parts)
                    break
            
            if not date or not montant:
                return None
            
            # NOUVELLE MÉTHODE: Classification par position physique
            debit = ""
            credit = ""
            
            if words_data:
                print(f"DEBUG BIAT - Classification par position pour: {libelle}")
                
                # Extraire les coordonnées X des montants dans cette ligne
                montant_positions = []
                for word in words_data:
                    word_text = word.get('text', '').strip()
                    # Vérifier si c'est un montant (contient des chiffres et virgule/point)
                    if re.search(r'\d+[.,]\d+', word_text):
                        x_pos = (word.get('x0', 0) + word.get('x1', 0)) / 2.0
                        montant_positions.append((x_pos, word_text))
                        print(f"DEBUG BIAT - Montant trouvé: '{word_text}' à position X: {x_pos}")
                
                if montant_positions:
                    # Trier par position X (de gauche à droite)
                    montant_positions.sort(key=lambda x: x[0])
                    
                    # Déterminer les seuils de position pour débit/crédit
                    # Seuils empiriques basés sur la structure BIAT
                    for x_pos, amount in montant_positions:
                        if x_pos < 450:  # Position à gauche = DÉBIT
                            debit = amount
                            print(f"DEBUG BIAT - Montant à gauche (X={x_pos}) -> DÉBIT: {amount}")
                        else:  # Position à droite = CRÉDIT
                            credit = amount
                            print(f"DEBUG BIAT - Montant à droite (X={x_pos}) -> CRÉDIT: {amount}")
                else:
                    # Fallback: utiliser l'ancienne méthode
                    print(f"DEBUG BIAT - Aucune position trouvée, fallback vers mots-clés")
                    classification = self.analyze_biat_line_structure(libelle)
                    if classification == "debit":
                        debit = montant
                    else:
                        credit = montant
            else:
                # Fallback: utiliser l'ancienne méthode par mots-clés
                print(f"DEBUG BIAT - Pas de données de position, fallback vers mots-clés")
                classification = self.analyze_biat_line_structure(libelle)
                if classification == "debit":
                    debit = montant
                else:
                    credit = montant
            
            # Ignorer lignes de report (libellé commençant par "Report")
            if re.match(r'(?i)^\s*report', libelle.strip()):
                return None
            
            libelle_upper = libelle.upper()
            if any(keyword in libelle_upper for keyword in ["COMMISSION", "FRAIS", "DEBIT", "RETRAIT", "VIREMENT TN AUTRE"]):
                debit = montant
            else:
                credit = montant
            
            return {
                'Date': self.format_date_biat(date, year),
                'Libellé': libelle,
                'Débit': self.format_amount(debit),
                'Crédit': self.format_amount(credit)
            }
            
        except Exception as e:
            print(f"Erreur parsing ligne BIAT: {e}")
            return None
    
    def analyze_biat_line_structure(self, line, words_data=None):
        """Analyse la structure d'une ligne BIAT pour détecter débit/crédit basé sur la POSITION"""
        import re
        
        # NOUVELLE MÉTHODE: Classification par position physique dans le PDF
        if words_data:
            print(f"DEBUG BIAT - Analyse par position pour: {line}")
            
            # Extraire les coordonnées X des montants
            montant_positions = []
            for word in words_data:
                word_text = word.get('text', '').strip()
                # Vérifier si c'est un montant (contient des chiffres et virgule/point)
                if re.search(r'\d+[.,]\d+', word_text):
                    x_pos = (word.get('x0', 0) + word.get('x1', 0)) / 2.0
                    montant_positions.append((x_pos, word_text))
                    print(f"DEBUG BIAT - Montant trouvé: '{word_text}' à position X: {x_pos}")
            
            if montant_positions:
                # Trier par position X (de gauche à droite)
                montant_positions.sort(key=lambda x: x[0])
                
                # Déterminer les seuils de position pour débit/crédit
                # Généralement: débit = position plus à gauche, crédit = position plus à droite
                if len(montant_positions) == 1:
                    # Un seul montant: analyser sa position
                    x_pos, montant = montant_positions[0]
                    # Seuil empirique: si X > 400, c'est probablement un crédit
                    if x_pos > 400:
                        print(f"DEBUG BIAT - Montant à droite (X={x_pos}) -> CRÉDIT")
                        return "credit"
                    else:
                        print(f"DEBUG BIAT - Montant à gauche (X={x_pos}) -> DÉBIT")
                        return "debit"
                elif len(montant_positions) == 2:
                    # Deux montants: le plus à gauche = débit, le plus à droite = crédit
                    print(f"DEBUG BIAT - Deux montants détectés: {montant_positions}")
                    # Retourner le type du montant le plus à droite (crédit)
                    return "credit"
                else:
                    # Plusieurs montants: prendre le plus à droite
                    print(f"DEBUG BIAT - Plusieurs montants, prendre le plus à droite")
                    return "credit"
        
        # FALLBACK: Ancienne méthode par mots-clés (gardée comme backup)
        print(f"DEBUG BIAT - Fallback vers analyse par mots-clés")
        libelle_upper = line.upper()
        
        # Mots-clés qui indiquent généralement un CRÉDIT (entrée d'argent)
        credit_keywords = [
            "ENCAISSEMENT", "VERSEMENT", "DEBLOCAGE", "CREDIT",
            "VIREMENT RECU", "REMBOURSEMENT", "INTERET", "DIVIDENDE","ANNUL",
            "VIREMENT VERS", "VIREMENT DE", "VIREMENT DEPUIS",
            # CORRECTION: Ajouter d'autres patterns de crédit
            "COMPENSATION", "AVOIR", "REMISE", "DEPOT",
            "ENCAISSEMENT CHEQUE", "ENCAISSEMENT EFFET", "ENCAISSEMENT VIR"
        ]
        
        # Mots-clés qui indiquent généralement un DÉBIT (sortie d'argent)
        debit_keywords = [
            "VIREMENT TN AUTRE BQ", "VIREMENT TN MEME BQ", "COMMISSION",
            "FRAIS", "AGIOS", "PRELEVEMENT", "RETRAIT", "CHEQUE IMPAYE",
            "EFFET IMPAYE", "COTISATION", "REGLEMENT CHEQUE", "VIR ORD CARTE",
            "VIREMENT POUR", "VIREMENT A", "VIREMENT VERS", "VIREMENT EMIS",  "REDRESSEMENT",
            # CORRECTION: Ajouter d'autres patterns de débit
            "ACHAT TPE", "COM ET TVA", "COM TVA", "CHEQUE PREAVISE",
            "REGLEMENT", "PAIEMENT", "SORTIE", "DEBIT"
        ]
        
        # CORRECTION: Logique spéciale pour les virements
        if "VIREMENT" in libelle_upper:
            print(f"DEBUG BIAT - Virement détecté, analyse du contexte: {libelle_upper}")
            
            # Virements CRÉDIT (entrée d'argent) - patterns spécifiques
            credit_virement_patterns = [
                "VIREMENT RECU", "VIREMENT DE", "VIREMENT DEPUIS", "VIREMENT VERS VOUS",
                "VIREMENT ENTRANT", "VIREMENT CREDIT", "VIREMENT BENEFICIAIRE",
                "VIREMENT TN AUTRE BQ"  # CORRECTION: Virement reçu d'une autre banque = CRÉDIT
            ]
            
            # Virements DÉBIT (sortie d'argent) - patterns spécifiques  
            debit_virement_patterns = [
                "VIREMENT POUR", "VIREMENT A", "VIREMENT VERS", "VIREMENT EMIS",
                "VIREMENT SORTANT", "VIREMENT DEBIT", "VIREMENT ORDONNATEUR",
                "VIREMENT TN MEME BQ"  # CORRECTION: Virement vers même banque = DÉBIT
            ]
            
            # CORRECTION: Vérifier d'abord les patterns BIAT spécifiques (priorité haute)
            if "VIREMENT TN AUTRE BQ" in libelle_upper:
                print(f"DEBUG BIAT - Virement CRÉDIT détecté: VIREMENT TN AUTRE BQ (reçu d'autre banque)")
                return "credit"
            
            if "VIREMENT TN MEME BQ" in libelle_upper:
                print(f"DEBUG BIAT - Virement DÉBIT détecté: VIREMENT TN MEME BQ (vers même banque)")
                return "debit"
            
            # CORRECTION: Gérer les virements avec numéros de référence
            if "VIREMENT TN AUTRE BQ" in libelle_upper and re.search(r'\d{7,}', libelle_upper):
                print(f"DEBUG BIAT - Virement CRÉDIT avec référence détecté: VIREMENT TN AUTRE BQ + numéro")
                return "credit"
            
            if "VIREMENT TN MEME BQ" in libelle_upper and re.search(r'\d{7,}', libelle_upper):
                print(f"DEBUG BIAT - Virement DÉBIT avec référence détecté: VIREMENT TN MEME BQ + numéro")
                return "debit"
            
            # Vérifier ensuite les autres patterns de virement spécifiques
            for pattern in credit_virement_patterns:
                if pattern in libelle_upper:
                    print(f"DEBUG BIAT - Virement CRÉDIT détecté: {pattern}")
                    return "credit"
            
            for pattern in debit_virement_patterns:
                if pattern in libelle_upper:
                    print(f"DEBUG BIAT - Virement DÉBIT détecté: {pattern}")
                    return "debit"
            
            # CORRECTION: Analyser tous les autres types de virements
            # Virements CRÉDIT généraux (entrée d'argent)
            if any(word in libelle_upper for word in [
                "RECU", "ENTRANT", "BENEFICIAIRE", "DE", "DEPUIS", "VERS VOUS",
                "CREDIT", "BENEFICIAIRE", "ENTRANT", "RECEPTION"
            ]):
                print(f"DEBUG BIAT - Virement classé CRÉDIT par contexte général")
                return "credit"
            
            # Virements DÉBIT généraux (sortie d'argent)
            elif any(word in libelle_upper for word in [
                "POUR", "VERS", "A", "EMIS", "ORDONNATEUR", "SORTANT", "DEBIT",
                "EMISSION", "ENVOI", "TRANSFERT", "VERSEMENT"
            ]):
                print(f"DEBUG BIAT - Virement classé DÉBIT par contexte général")
                return "debit"
            
            # CORRECTION: Analyser la structure du libellé pour les virements non classés
            else:
                print(f"DEBUG BIAT - Virement sans contexte clair, analyse avancée")
                
                # Si le libellé contient un numéro de compte ou référence longue, c'est probablement un débit
                if re.search(r'\b\d{10,}\b', libelle_upper):  # Numéro de compte long
                    print(f"DEBUG BIAT - Virement avec numéro de compte long -> DÉBIT")
                    return "debit"
                
                # Si le libellé contient des mots comme "ORDRE", "MANDAT", c'est un débit
                if any(word in libelle_upper for word in ["ORDRE", "MANDAT", "INSTRUCTION"]):
                    print(f"DEBUG BIAT - Virement avec ordre/mandat -> DÉBIT")
                    return "debit"
                
                # Si le libellé est très court (moins de 30 caractères), c'est probablement un débit
                if len(libelle_upper) < 30:
                    print(f"DEBUG BIAT - Virement avec libellé court -> DÉBIT")
                    return "debit"
                
                # Par défaut, analyser la position du montant
                print(f"DEBUG BIAT - Virement sans contexte clair, analyse de position")
                # Continuer vers l'analyse de position ci-dessous
        
        # CORRECTION: Debug pour voir les opérations non classées
        print(f"DEBUG BIAT - Analyse de l'opération: '{libelle_upper}'")
        
        # Vérifier les autres mots-clés spécifiques
        for keyword in credit_keywords:
            if keyword in libelle_upper:
                print(f"DEBUG BIAT - Mot-clé crédit détecté: {keyword}")
                return "credit"
        
        for keyword in debit_keywords:
            if keyword in libelle_upper:
                print(f"DEBUG BIAT - Mot-clé débit détecté: {keyword}")
                return "debit"
        
        # CORRECTION: Si aucun mot-clé trouvé, analyser le contenu
        print(f"DEBUG BIAT - Aucun mot-clé spécifique trouvé pour: '{libelle_upper}'")
        
        # Analyser les patterns généraux
        if any(word in libelle_upper for word in ["ENCAISSEMENT", "VERSEMENT", "CREDIT", "AVOIR", "REMBOURSEMENT"]):
            print(f"DEBUG BIAT - Pattern crédit général détecté")
            return "credit"
        elif any(word in libelle_upper for word in ["COMMISSION", "FRAIS", "DEBIT", "RETRAIT", "PAIEMENT"]):
            print(f"DEBUG BIAT - Pattern débit général détecté")
            return "debit"
        
        # Si aucun mot-clé spécifique, analyser la position du montant
        # Chercher des séquences d'espaces multiples (3+ espaces consécutifs)
        space_patterns = re.finditer(r'\s{3,}', line)
        space_positions = [match.start() for match in space_patterns]
        
        # Chercher le montant dans la ligne
        montant_match = re.search(r'([\d\s.,]+)$', line.strip())
        if not montant_match:
            return "credit"  # Par défaut
        
        montant_start = montant_match.start()
        line_before_montant = line[:montant_start].strip()
        
        # CORRECTION: Logique spéciale pour les virements sans contexte clair
        if "VIREMENT" in libelle_upper:
            print(f"DEBUG BIAT - Analyse de position pour virement: '{line_before_montant}'")
            # Si le libellé contient beaucoup d'informations (long), c'est probablement un crédit
            if len(line_before_montant) > 60:
                print(f"DEBUG BIAT - Virement avec libellé long -> CRÉDIT")
                return "credit"
            # Si le libellé est court, c'est probablement un débit
            elif len(line_before_montant) < 40:
                print(f"DEBUG BIAT - Virement avec libellé court -> DÉBIT")
                return "debit"
        
        # Analyser la position du montant par rapport aux séparations de colonnes
        if space_positions:
            # Si le montant est après la dernière séparation d'espace importante, c'est un crédit
            last_space_pos = space_positions[-1]
            if montant_start > last_space_pos:
                print(f"DEBUG BIAT - Montant après dernière séparation -> crédit")
                return "credit"
            else:
                # Le montant est avant la dernière séparation, c'est probablement un débit
                print(f"DEBUG BIAT - Montant avant dernière séparation -> débit")
                return "debit"
        
        # Méthode de fallback: analyser la longueur de la ligne avant le montant
        if len(line_before_montant) > 80:  # Ligne très longue avant le montant
            print(f"DEBUG BIAT - Ligne longue ({len(line_before_montant)}) -> crédit")
            return "credit"
        elif len(line_before_montant) > 50:  # Ligne moyennement longue
            print(f"DEBUG BIAT - Ligne moyenne ({len(line_before_montant)}) -> crédit")
            return "credit"
        else:
            # Ligne courte avant le montant, probablement un débit
            print(f"DEBUG BIAT - Ligne courte ({len(line_before_montant)}) -> débit")
            return "debit"
    
    def parse_biat_transactions_from_text(self, text, year):
        """Parse les transactions BIAT depuis le texte brut - VERSION SIMPLIFIÉE"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG BIAT TEXT - {len(lines)} lignes trouvées")
            
            for line in lines:
                line = line.strip()
                if not line or len(line) < 10:  # Ignorer les lignes trop courtes
                    continue
                
                print(f"DEBUG BIAT TEXT - Ligne: {line}")
                
                # Pattern principal: DD MM [tout le reste jusqu'au montant final]
                # Ce pattern capture tout entre la date et le montant final
                main_pattern = r'^(\d{1,2}\s+\d{1,2})\s+(.+?)\s+([\d\s.,]+)$'
                
                # Patterns spécifiques pour différents formats - AMÉLIORÉS
                patterns = [
                    # Format avec date de valeur: DD MM libellé DDMMYYYY montant
                    r'^(\d{1,2}\s+\d{1,2})\s+(.+?)\s+(\d{8})\s+([\d\s.,]+)$',
                    # Format simple: DD MM libellé montant
                    r'^(\d{1,2}\s+\d{1,2})\s+(.+?)\s+([\d\s.,]+)$',
                    # Format avec montant dans le libellé: DD MM libellé avec montant
                    r'^(\d{1,2}\s+\d{1,2})\s+(.+?)\s+([\d\s.,]+)$'
                ]
                
                # Essayer chaque pattern jusqu'à en trouver un qui fonctionne
                match = None
                for pattern in patterns:
                    match = re.match(pattern, line)
                    if match:
                        break
                
                if match:
                    # Traitement unifié pour tous les patterns
                    groups = match.groups()
                    
                    if len(groups) == 4:
                        # Format avec date de valeur: DD MM libellé DDMMYYYY montant
                        date = groups[0]  # DD MM
                        libelle = groups[1]  # Libellé complet
                        date_valeur = groups[2]  # DDMMYYYY (ignorer)
                        montant = groups[3]  # Montant
                    elif len(groups) == 3:
                        # Format simple: DD MM libellé montant
                        date = groups[0]  # DD MM
                        libelle = groups[1]  # Libellé complet
                        montant = groups[2]  # Peut être un identifiant (pas un vrai montant)
                        if not self.is_probable_money(montant):
                            # Garder ce nombre dans le libellé, pas en montant
                            libelle = f"{libelle} {montant}".strip()
                            montant = ""
                    else:
                        continue  # Pattern non reconnu
                    
                    # AMÉLIORATION: Nettoyer le libellé pour enlever les montants qui s'y trouvent
                    libelle_clean = libelle.strip()
                    
                    # Chercher si le libellé contient un montant à la fin
                    # Pattern pour détecter un montant à la fin du libellé
                    montant_in_libelle_pattern = r'(.+?)\s+([\d\s.,]+)$'
                    libelle_match = re.match(montant_in_libelle_pattern, libelle_clean)
                    
                    if libelle_match:
                        # Le libellé contient un montant à la fin
                        libelle_text = libelle_match.group(1).strip()
                        montant_in_libelle = libelle_match.group(2).strip()
                        
                        # N'accepter comme montant que si ça ressemble à un VRAI montant (avec virgule ou milliers)
                        if self.is_probable_money(montant_in_libelle):
                            if not montant or len(montant.strip()) == 0:
                                montant = montant_in_libelle
                                libelle_clean = libelle_text
                                print(f"DEBUG BIAT - Montant extrait du libellé: {montant}")
                            else:
                                print(f"DEBUG BIAT - Montant dans libellé détecté mais ignoré: {montant_in_libelle}")
                    
                    # Nettoyer le montant (enlever les points et espaces)
                    montant_clean = montant.replace('.', '').replace(' ', '') if montant else ''
                    
                    # Vérifier que le montant ne contient pas de date de valeur mélangée
                    if len(montant_clean) > 10:
                        # Chercher un pattern de date de valeur (8 chiffres) suivi d'un montant
                        date_valeur_pattern = r'^(\d{8})([\d,]+)$'
                        date_match = re.match(date_valeur_pattern, montant_clean)
                        if date_match:
                            # Extraire seulement la partie montant (après les 8 premiers chiffres)
                            montant_clean = date_match.group(2)
                            print(f"DEBUG BIAT - Date de valeur détectée et supprimée du montant: {date_match.group(1)} -> {montant_clean}")
                    
                    # Déterminer si c'est débit ou crédit selon la position du montant dans la ligne
                    debit = ""
                    credit = ""
                    
                    # Utiliser la fonction d'analyse de structure pour déterminer débit/crédit
                    transaction_type = self.analyze_biat_line_structure(line)
                    
                    if montant_clean:
                        if transaction_type == "debit":
                            debit = montant_clean
                        else:
                            credit = montant_clean
                    
                    # Utiliser la date DD MM avec l'année détectée
                    transaction = {
                        'Date': self.format_date_biat(date, year),
                        'Libellé': libelle_clean,
                        'Débit': self.format_amount(debit),
                        'Crédit': self.format_amount(credit)
                    }
                    
                    transactions.append(transaction)
                    print(f"DEBUG BIAT TEXT - Transaction créée: {transaction}")
                
                else:
                    # Si aucun pattern ne correspond, essayer une approche plus flexible
                    # Chercher des lignes qui commencent par une date
                    if re.match(r'^\d{1,2}\s+\d{1,2}\s+', line):
                        print(f"DEBUG BIAT TEXT - Ligne avec date détectée mais non parsée: {line}")
                
        except Exception as e:
            print(f"Erreur parsing texte BIAT: {e}")
        
        return transactions
    
    def is_date_biat(self, text):
        """Vérifie si le texte est une date au format BIAT DD MM"""
        if not text:
            return False
        # Pattern pour date DD MM (avec espaces)
        return bool(re.match(r'^\d{1,2}\s+\d{1,2}$', text.strip()))
    
    def is_date_valeur_biat(self, text):
        """Vérifie si le texte est une date de valeur au format BIAT DDMMYYYY"""
        if not text:
            return False
        # Pattern pour date de valeur DDMMYYYY
        return bool(re.match(r'^\d{8}$', text.strip()))
    
    def format_date_biat(self, date_str, year):
        """Formate la date BIAT au format DD/MM/YYYY"""
        try:
            if not date_str:
                return ""
            
            # Si c'est déjà au format DD/MM/YYYY
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
                return date_str
            
            # Si c'est au format DDMMYYYY
            if re.match(r'^\d{8}$', date_str):
                day = date_str[:2]
                month = date_str[2:4]
                year_part = date_str[4:8]
                return f"{day}/{month}/{year_part}"
            
            # Si c'est au format DD MM
            if re.match(r'^\d{1,2}\s+\d{1,2}$', date_str):
                parts = date_str.split()
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                return f"{day}/{month}/{year}"
            
            return date_str
            
        except Exception as e:
            print(f"Erreur formatage date BIAT: {e}")
            return date_str
    
    def is_amount(self, text):
        """Vérifie si le texte est un montant"""
        if not text:
            return False
        # Pattern pour montant avec virgule et points
        return bool(re.match(r'^[\d\s.,]+$', text.strip()))

    def is_probable_money(self, text):
        """Heuristique: vrai montant s'il a des décimales avec virgule ou des milliers séparés par espace/point."""
        if not text:
            return False
        t = str(text).strip()
        # A un séparateur décimal
        if re.search(r'\d+,\d{1,}', t):
            return True
        # A des séparateurs de milliers
        if re.search(r'\d{1,3}(?:[ \.]\d{3})+(?:,\d{1,})?$', t):
            return True
        return False
    
    def format_amount(self, amount_str):
        """Formate le montant"""
        if not amount_str or amount_str == "":
            return ""
        
        try:
            # Nettoyer le montant
            clean_amount = str(amount_str).replace(' ', '').replace('.', '')
            
            # Remplacer la virgule par un point pour la conversion
            if ',' in clean_amount:
                clean_amount = clean_amount.replace(',', '.')
            
            # Convertir en float puis formater
            amount_float = float(clean_amount)
            # 3 décimales requises
            return f"{amount_float:,.3f}".replace(',', ' ').replace('.', ',')
            
        except (ValueError, TypeError):
            return str(amount_str)
    
    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel"""
        try:
            # Créer le fichier Excel avec openpyxl pour le formatage
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "J03"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formatage des en-têtes
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Appliquer le formatage aux en-têtes
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Appliquer les bordures à toutes les cellules
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Ajuster la largeur des colonnes (libellé plus large)
            column_widths = {'A': 12, 'B': 70, 'C': 16, 'D': 16}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Sauvegarder
            wb.save(excel_path)
            print(f"Fichier Excel sauvegardé avec formatage: {excel_path}")
            
        except Exception as e:
            print(f"Erreur formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, sheet_name="J03")

    def convert_pdf_to_excel(self):
        """Fonction principale de conversion BIAT"""
        try:
            if not self.pdf_path.get():
                messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
                return
            
            if not self.excel_name.get():
                messagebox.showerror("Erreur", "Veuillez entrer un nom pour le fichier Excel")
                return
            
            # Démarrer la barre de progression
            
            self.convert_button.config(state='disabled')
            
            self.root.update()
            
            # Vérifier que c'est un PDF BIAT RELEVE
            if not self.is_biat_releve_pdf(self.pdf_path.get()):
                messagebox.showerror("Erreur", "Le fichier sélectionné n'est pas un RELEVE BIAT valide")
                return
            
            # Extraire les données
            
            self.root.update()
            
            transactions = self.extract_table_data(self.pdf_path.get())
            
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF")
                return
            
            # Créer le DataFrame
            
            self.root.update()
            
            df = pd.DataFrame(transactions)
            
            # Chemin de sauvegarde
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_filename = f"{self.excel_name.get()}.xlsx"
            excel_path = os.path.join(downloads_path, excel_filename)
            
            # Sauvegarder avec formatage
            self.save_excel_with_formatting(df, excel_path)
            
            # Succès
            
            self.convert_button.config(state='normal')

            # Message de succès plus positif
            success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
            success_msg += f"📁 Fichier: {excel_filename}\n"
            success_msg += f"📂 Emplacement: {downloads_path}\n"
            success_msg += f"📊 Nombre de transactions: {len(transactions)}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            messagebox.showinfo("✅ Conversion réussie", success_msg)
            
        except Exception as e:
            
            self.convert_button.config(state='normal')
            
            messagebox.showerror("Erreur", f"Erreur lors de la conversion: {str(e)}")

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

def main():
    root = tk.Tk()
    app = BIATReleveConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
