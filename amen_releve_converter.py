import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from datetime import datetime
import re
import unicodedata
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract  # type: ignore
    from PIL import Image
    import numpy as np
    import cv2
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class AmenReleveConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur RELEVÉ AMEN BANK vers Excel")
        self.root.geometry("600x500")

        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"RELEVE_AMEN_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        self.setup_ui()

    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur RELEVÉ AMEN",
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 8))
        subtitle_label = tk.Label(text="Conversion PDF vers Excel",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 20))
        
        # Frame principal
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=40, fill='both', expand=True)
        
        # Section PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF RELEVÉ AMEN:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        pdf_entry = tk.Entry(pdf_frame, textvariable=self.pdf_path, width=60,
                           font=("Arial", 9))
        pdf_entry.pack(pady=5, fill='x')
        
        browse_btn = tk.Button(pdf_frame, text="Parcourir", command=self.browse_pdf, font=("Segoe UI", 10, "bold"), bg="#3498db", fg="white")
        browse_btn.pack(pady=5)
        
        # Section Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:",
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_name, width=60,
                             font=("Arial", 9))
        excel_entry.pack(pady=5, fill='x')
        
        # Section boutons
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(fill='x')
        
        convert_btn = tk.Button(buttons_frame, text="Convertir en Excel",
                              command=self.convertir, 
                              font=("Segoe UI", 12, "bold"), bg="green", fg="white")
        convert_btn.pack(side='left', padx=10)
        
        retour_btn = tk.Button(buttons_frame, text="Retour page d'accueil",
                              command=self.retour_accueil, 
                              font=("Segoe UI", 12, "bold"), bg="red", fg="white")
        retour_btn.pack(side='right', padx=10)
        
        # Barre de progression
        progress_frame = tk.Frame(main_frame)
        progress_frame.pack(fill='x', pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill='x')
        
        # Zone de statut
        self.status_label = tk.Label(progress_frame, text="Prêt", 
                                   font=("Arial", 9), fg="green")
        self.status_label.pack(pady=5)

    def browse_pdf(self):
        path = filedialog.askopenfilename(title="Choisir un PDF AMEN", filetypes=[["PDF", "*.pdf"]])
        if path:
            self.pdf_path.set(path)
            # Renouveler le nom Excel pour éviter toute confusion/écrasement
            base = os.path.splitext(os.path.basename(path))[0]
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.excel_name.set(f"RELEVE_AMEN_{base}_{ts}")

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF AMEN.")
            return
        self.progress['value'] = 10; self.root.update_idletasks()
        try:
            print(f"DEBUG AMEN - PDF sélectionné: {path}")
        except Exception:
            pass

        rows = self.parse_pdf(path)
        if not rows:
            messagebox.showerror("Aucune transaction", "Impossible d'extraire des transactions du relevé AMEN.")
            return

        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'RELEVE_AMEN'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        self.progress['value'] = 100
        # Message de succès plus positif
        success_msg = f"✅ Conversion RELEVE terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n\n"
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        messagebox.showinfo("✅ Conversion réussie", success_msg)

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def is_balance_or_total_line(self, line_text):
        """CORRECTION: Détecte si une ligne est un solde, total ou footer (à bloquer)"""
        if not line_text:
            return False
        
        line_upper = line_text.upper()
        
        # BLOQUER: Lignes de soldes et totaux
        balance_keywords = [
            'SOLDE AU', 'SOLDE', 'TOTAUX', 'TOTAL', 'REPORT', 'RAPPORT',
            'SOLDE AU 31', 'SOLDE AU 30', 'SOLDE AU 29', 'SOLDE AU 28',
            'SOLDE AU 27', 'SOLDE AU 26', 'SOLDE AU 25', 'SOLDE AU 24',
            'SOLDE AU 23', 'SOLDE AU 22', 'SOLDE AU 21', 'SOLDE AU 20',
            'SOLDE AU 19', 'SOLDE AU 18', 'SOLDE AU 17', 'SOLDE AU 16',
            'SOLDE AU 15', 'SOLDE AU 14', 'SOLDE AU 13', 'SOLDE AU 12',
            'SOLDE AU 11', 'SOLDE AU 10', 'SOLDE AU 09', 'SOLDE AU 08',
            'SOLDE AU 07', 'SOLDE AU 06', 'SOLDE AU 05', 'SOLDE AU 04',
            'SOLDE AU 03', 'SOLDE AU 02', 'SOLDE AU 01'
        ]
        
        for keyword in balance_keywords:
            if keyword in line_upper:
                print(f"DEBUG AMEN - Ligne solde/total bloquée: {keyword}")
                return True
        
        # CORRECTION: BLOQUER les footers de page AMEN BANK
        footer_keywords = [
            'S.A AU CAPITAL', 'R.C :', 'AVENUE MOHAMED V', 'TÉL :', 'FAX :', 'SWIFT :',
            'E-MAIL :', 'SITE WEB :', 'CENTRE DE RELATIONS CLIENTS :',
            'AMENBANK.COM.TN', 'AMENFIRSTBANK.COM.TN', 'CFCTINTTXXX',
            'DINARS', 'TUNIS', 'CLIENTS'
        ]
        
        for keyword in footer_keywords:
            if keyword in line_upper:
                print(f"DEBUG AMEN - Ligne footer bloquée: {keyword}")
                return True
        
        # CORRECTION: BLOQUER les lignes avec des numéros de téléphone/fax
        if re.search(r'\b\d{2}\s\d{3}\s\d{3}\b', line_text):  # Format: 71 148 000
            print(f"DEBUG AMEN - Ligne numéro téléphone bloquée: {line_text}")
            return True
        
        # CORRECTION: BLOQUER les lignes avec des montants très grands (capitaux, etc.)
        if re.search(r'\b\d{3}\.\d{3}\.\d{3}\b', line_text):  # Format: 174.600.000
            print(f"DEBUG AMEN - Ligne montant très grand bloquée: {line_text}")
            return True
        
        return False
    
    def is_large_amount(self, amount):
        """CORRECTION: Détecte si un montant est probablement un solde (à bloquer)"""
        if not amount:
            return False
        
        try:
            # Convertir en float pour vérifier la valeur
            if isinstance(amount, str):
                # Nettoyer le montant
                clean_amount = amount.replace(' ', '').replace(',', '.')
                amount_value = float(clean_amount)
            else:
                amount_value = float(amount)
            
            # BLOQUER: Montants très grands (probablement des soldes)
            if amount_value > 1000000:  # Plus d'1 million
                print(f"DEBUG AMEN - Montant très grand bloqué (solde): {amount}")
                return True
            
            # BLOQUER: Montants négatifs (soldes débiteurs)
            if amount_value < 0:
                print(f"DEBUG AMEN - Montant négatif bloqué (solde): {amount}")
                return True
                
        except:
            pass
        
        return False

    def parse_pdf(self, pdf_path: str):
        # Amen: une colonne d'index précède la date (ex: "58 02/05/2025 ...")
        # Autoriser l'index optionnel avant la date
        date_line_re = re.compile(r"^\s*(?:\d+\s+)?(\d{2}/\d{2}/\d{4})")
        date_any_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
        # CORRECTION: Regex améliorée pour capturer les montants AMEN
        # Format AMEN: 5,950 (virgule = décimales) ou 1 234,567 (espace = milliers, virgule = décimales)
        amount_re = re.compile(r"-?\d{1,3}(?:[ .]\d{3})*(?:[.,]\d{2,3})?|-?\d+[.,]\d{2,3}")

        def to_float(s: str):
            if not s:
                return None
            s = s.replace('\u00a0', ' ').strip()
            neg = s.startswith('-')
            s = s.lstrip('-').replace(' ', '')
            
            # CORRECTION: Gestion améliorée des formats de montants AMEN
            # Pour AMEN, les montants sont généralement au format: 5,950 (virgule = décimales)
            if ',' in s and '.' in s:
                # Format: 1.234,567 (point = milliers, virgule = décimales)
                s = s.replace('.', '').replace(',', '.')
            elif ',' in s:
                # CORRECTION: Pour AMEN, la virgule est TOUJOURS le séparateur décimal
                # Format: 5,950 -> 5.950 (virgule = décimales)
                s = s.replace(',', '.')
            else:
                # Pas de virgule, garder tel quel
                pass
                
            try:
                val = float(s)
                return -val if neg else val
            except Exception:
                return None

        results = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not text.strip():
                    # Fallback OCR pour les PDFs scannés
                    ocr_text = self._extract_text_via_ocr(pdf_path)
                    if ocr_text.strip():
                        text = ocr_text
                        break  # Utiliser le texte OCR pour toutes les pages
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                # Préparer les mots de la page pour récupérer les positions X
                try:
                    page_words = page.extract_words() or []
                except Exception:
                    page_words = []
                # Détecter l'emplacement des en-têtes DEBIT / CREDIT si présent
                debit_x_hdr = None
                credit_x_hdr = None
                for w in page_words:
                    wt = str(w.get('text', '')).strip().upper()
                    xmid = (float(w.get('x0', 0)) + float(w.get('x1', 0))) / 2.0
                    if wt in ('DÉBIT', 'DEBIT'):
                        debit_x_hdr = xmid
                    elif wt in ('CRÉDIT', 'CREDIT'):
                        credit_x_hdr = xmid
                i = 0
                while i < len(lines):
                    line = lines[i]
                    
                    # CORRECTION: Bloquer les lignes de soldes, totaux et footers
                    if self.is_balance_or_total_line(line):
                        print(f"DEBUG AMEN - ❌ Ligne bloquée: {line[:50]}...")
                        i += 1
                        continue
                    
                    # CORRECTION: Bloquer les lignes qui contiennent des informations de contact
                    if any(keyword in line.upper() for keyword in ['AMENBANK', 'TÉLÉPHONE', 'FAX', 'EMAIL', 'SITE WEB', 'SWIFT']):
                        print(f"DEBUG AMEN - ❌ Ligne contact bloquée: {line[:50]}...")
                        i += 1
                        continue
                    
                    m = date_line_re.match(line)
                    if not m:
                        i += 1
                        continue
                    op_date = m.group(1)

                    # CORRECTION: Accumuler description jusqu'à prochaine date ou fin de page
                    j = i + 1
                    desc_lines = []
                    while j < len(lines) and not date_line_re.match(lines[j]):
                        desc_lines.append(lines[j])
                        j += 1
                    
                    # CORRECTION: Si on est à la fin de la page, traiter la ligne actuelle même sans date suivante
                    if j >= len(lines):
                        print(f"DEBUG AMEN - Fin de page détectée, traitement de la dernière ligne")

                    combined = (line + ' ' + ' '.join(desc_lines)).strip()
                    # Trouver la 2ème date (date valeur) pour cibler la zone des montants
                    dates_iter = list(date_any_re.finditer(combined))
                    tail_start = dates_iter[1].end() if len(dates_iter) >= 2 else m.end()
                    tail = combined[tail_start:]

                    # CORRECTION: Nettoyer la queue avant d'extraire les montants pour éviter les dates
                    tail_clean = tail
                    # Supprimer les dates du libellé qui peuvent interférer (format DD.MM.YYYY)
                    tail_clean = re.sub(r'\b\d{1,2}\.\d{1,2}\.\d{2,4}\b', '', tail_clean)
                    # Supprimer les dates partielles (format DD.MM)
                    tail_clean = re.sub(r'\b\d{1,2}\.\d{1,2}\b', '', tail_clean)
                    # Supprimer les années isolées
                    tail_clean = re.sub(r'\b(19|20)\d{2}\b', '', tail_clean)
                    tail_clean = re.sub(r'\s+', ' ', tail_clean).strip()
                    
                    print(f"DEBUG AMEN - Queue nettoyée: '{tail_clean}'")
                    
                    # CORRECTION: Chercher des montants après la date valeur
                    raw_amounts = amount_re.findall(tail_clean)
                    # CORRECTION: enlever le préfixe des 2 chiffres de l'année (ex: "25 1,978" ou "251,978")
                    try:
                        # Année de la date valeur si disponible, sinon année de la date d'opération
                        if len(dates_iter) >= 2:
                            y_src = dates_iter[1].group(0)
                        else:
                            y_src = op_date
                        yy = int(y_src[-4:]) % 100
                        pat_space = re.compile(rf'^{yy:02d}\s+(\d{{1,3}}(?:[ \.]\d{{3}})*[.,]\d{{2,3}})$')
                        pat_nospace = re.compile(rf'^{yy:02d}(\d{{1,3}}(?:[ \.]\d{{3}})*[.,]\d{{2,3}})$')
                        cleaned_amounts = []
                        for a in raw_amounts:
                            m1 = pat_space.match(a)
                            m2 = pat_nospace.match(a) if not m1 else None
                            if m1:
                                cleaned_amounts.append(m1.group(1))
                            elif m2:
                                cleaned_amounts.append(m2.group(1))
                            else:
                                cleaned_amounts.append(a)
                        raw_amounts = cleaned_amounts
                    except Exception:
                        pass
                    print(f"DEBUG AMEN - Queue analysée: '{tail}' -> Queue nettoyée: '{tail_clean}' -> Montants trouvés: {raw_amounts}")
                    
                    # CORRECTION: Si pas de montants trouvés dans la queue, chercher dans toute la ligne
                    if not raw_amounts and j >= len(lines):
                        print(f"DEBUG AMEN - Aucun montant trouvé dans la queue, recherche dans toute la ligne")
                        print(f"DEBUG AMEN - Ligne complète: '{combined}'")
                        raw_amounts = amount_re.findall(combined)
                        print(f"DEBUG AMEN - Montants trouvés dans toute la ligne: {raw_amounts}")
                    
                    # CORRECTION SPÉCIALE: Pour les lignes de fin de page, chercher TOUS les montants possibles
                    if j >= len(lines):
                        print(f"DEBUG AMEN - 🔍 RECHERCHE EXHAUSTIVE POUR FIN DE PAGE")
                        print(f"DEBUG AMEN - Ligne originale: '{line}'")
                        print(f"DEBUG AMEN - Ligne combinée: '{combined}'")
                        
                        # CORRECTION: Nettoyer la ligne avant de chercher les montants
                        line_for_amounts = combined
                        
                        # Supprimer les patterns qui ne sont pas des montants
                        line_for_amounts = re.sub(r'\b(?:TOTAUX|TOTAL|SOLDE|الجملة)\b.*$', '', line_for_amounts, flags=re.IGNORECASE)
                        line_for_amounts = re.sub(r'\b\d{1,2}/\d{1,2}/\d{4}\b', '', line_for_amounts)  # Supprimer les dates avec /
                        line_for_amounts = re.sub(r'\b\d{1,2}\.\d{1,2}\.\d{2,4}\b', '', line_for_amounts)  # Supprimer les dates avec .
                        line_for_amounts = re.sub(r'\b\d{1,2}\.\d{1,2}\b', '', line_for_amounts)  # Supprimer les dates partielles DD.MM
                        line_for_amounts = re.sub(r'\b(19|20)\d{2}\b', '', line_for_amounts)  # Supprimer les années
                        line_for_amounts = re.sub(r'\b[A-Z0-9]{8,}\b', '', line_for_amounts)  # Supprimer les références longues
                        
                        print(f"DEBUG AMEN - Ligne nettoyée pour montants: '{line_for_amounts}'")
                        
                        # Chercher avec une regex plus permissive
                        all_amounts = re.findall(r'-?\d+[.,]\d+', line_for_amounts)
                        print(f"DEBUG AMEN - Tous les montants trouvés: {all_amounts}")
                        
                        # Chercher aussi les montants sans décimales (mais pas trop longs)
                        all_amounts_no_decimal = re.findall(r'-?\d{1,6}', line_for_amounts)
                        print(f"DEBUG AMEN - Montants sans décimales: {all_amounts_no_decimal}")
                        
                        # Combiner tous les montants trouvés en priorisant ceux à virgule (séparateur décimal AMEN)
                        with_comma = [x for x in all_amounts if ',' in x]
                        without_comma = [x for x in all_amounts if ',' not in x]
                        all_found_amounts = with_comma + without_comma + all_amounts_no_decimal
                        print(f"DEBUG AMEN - Tous les montants combinés: {all_found_amounts}")
                        
                        if all_found_amounts:
                            raw_amounts = all_found_amounts
                            print(f"DEBUG AMEN - Utilisation de tous les montants trouvés: {raw_amounts}")
                    
                    # Filtrer les faux positifs (références longues sans séparateurs ou motifs 05.2025)
                    filtered = []
                    for a in raw_amounts:
                        if re.search(r"\d{2}\.\d{4}", a):
                            print(f"DEBUG AMEN - Montant filtré (date): {a}")
                            continue
                        if re.fullmatch(r"\d{7,}", a.replace(' ', '')):
                            print(f"DEBUG AMEN - Montant filtré (référence longue): {a}")
                            continue
                        
                        # CORRECTION: Filtrer les montants qui ressemblent à des numéros de téléphone
                        if re.match(r'^\d{2}\s\d{3}\s\d{3}$', a):  # Format: 71 148 000
                            print(f"DEBUG AMEN - Montant filtré (téléphone): {a}")
                            continue
                        
                        # CORRECTION: Filtrer les montants très grands (capitaux, etc.)
                        if re.match(r'^\d{3}\.\d{3}\.\d{3}$', a):  # Format: 174.600.000
                            print(f"DEBUG AMEN - Montant filtré (capital): {a}")
                            continue
                        
                        # CORRECTION: Filtrer les montants qui sont dans une ligne de footer
                        if any(keyword in combined.upper() for keyword in ['S.A AU CAPITAL', 'R.C :', 'AVENUE MOHAMED V', 'TÉL :', 'FAX :', 'SWIFT :', 'E-MAIL :', 'SITE WEB :']):
                            print(f"DEBUG AMEN - Montant filtré (footer): {a}")
                            continue
                        
                        # CORRECTION CRITIQUE: Filtrer les dates du libellé qui sont confondues avec des montants
                        # Pattern: 15.05.202 -> 15,050 (date mal interprétée)
                        if re.match(r'^\d{1,2}\.\d{1,2}\.\d{1,3}$', a):  # Format: 15.05.202
                            print(f"DEBUG AMEN - Montant filtré (date du libellé): {a}")
                            continue
                        
                        # CORRECTION: Filtrer les montants qui ressemblent à des dates partielles
                        if re.match(r'^\d{1,2}\.\d{1,2}$', a):  # Format: 15.05
                            print(f"DEBUG AMEN - Montant filtré (date partielle): {a}")
                            continue
                        
                        # CORRECTION: Filtrer les montants qui sont clairement des dates
                        if re.match(r'^\d{1,2}\.\d{1,2}\.\d{2,4}$', a):  # Format: 15.05.2025
                            print(f"DEBUG AMEN - Montant filtré (date complète): {a}")
                            continue
                        
                        filtered.append(a)

                    amounts_f = [to_float(x) for x in filtered if to_float(x) is not None]
                    
                    # CORRECTION: Debug pour voir les montants extraits
                    if amounts_f:
                        print(f"DEBUG AMEN - Montants extraits: {filtered} -> {amounts_f}")
                        for i, (raw, converted) in enumerate(zip(filtered, amounts_f)):
                            print(f"DEBUG AMEN - Montant {i+1}: '{raw}' -> {converted}")
                        
                        # CORRECTION: Si plusieurs montants, logique de sélection intelligente
                        if len(amounts_f) > 1:
                            print(f"DEBUG AMEN - Plusieurs montants trouvés: {amounts_f}")
                            
                            # CORRECTION SPÉCIALE: Pour les lignes de fin de page, sélectionner le montant le plus probable
                            if j >= len(lines):
                                print(f"DEBUG AMEN - 🎯 SÉLECTION INTELLIGENTE POUR FIN DE PAGE")
                                
                                # Filtrer les montants suspects (trop grands ou trop petits)
                                valid_amounts = []
                                for amount in amounts_f:
                                    if 0.1 <= amount <= 10000:  # Plage raisonnable pour une transaction
                                        valid_amounts.append(amount)
                                        print(f"DEBUG AMEN - ✅ Montant valide: {amount}")
                                    else:
                                        print(f"DEBUG AMEN - ❌ Montant rejeté (hors plage): {amount}")
                                
                                if valid_amounts:
                                    # Prendre le montant le plus petit parmi les valides
                                    amounts_f = [min(valid_amounts)]
                                    print(f"DEBUG AMEN - 🎯 Montant sélectionné (plus petit valide): {amounts_f[0]}")
                                else:
                                    # Si aucun montant valide, prendre le plus petit de tous
                                    amounts_f = [min(amounts_f)]
                                    print(f"DEBUG AMEN - ⚠️ Aucun montant valide, prise du plus petit: {amounts_f[0]}")
                            else:
                                # Logique normale pour les autres lignes
                                amounts_f_sorted = sorted(amounts_f)
                                amounts_f = [amounts_f_sorted[0]]
                                print(f"DEBUG AMEN - Montant sélectionné (normal): {amounts_f[0]}")
                    
                    # CORRECTION SPÉCIFIQUE: Gestion des lignes de fin de page
                    is_end_of_page = j >= len(lines)
                    if is_end_of_page:
                        print(f"DEBUG AMEN - ⚠️ LIGNE DE FIN DE PAGE DÉTECTÉE ⚠️")
                        print(f"DEBUG AMEN - Ligne complète: '{combined}'")
                        print(f"DEBUG AMEN - Queue actuelle: '{tail}'")
                        print(f"DEBUG AMEN - Montants trouvés dans queue: {raw_amounts}")
                    
                    # CORRECTION: Si aucune montant trouvé et c'est la fin de page, essayer une approche alternative
                    if not amounts_f and is_end_of_page:
                        print(f"DEBUG AMEN - 🔍 EXTRACTION ALTERNATIVE POUR FIN DE PAGE")
                        
                        # CORRECTION: Nettoyer la ligne avant d'extraire les montants pour éviter le mélange avec les totaux
                        line_clean = combined
                        
                        # Supprimer les patterns qui ne sont pas des montants
                        line_for_amounts = re.sub(r'\b(?:TOTAUX|TOTAL|SOLDE|الجملة)\b.*$', '', line_clean, flags=re.IGNORECASE)
                        line_for_amounts = re.sub(r'\b\d{1,2}/\d{1,2}/\d{4}\b', '', line_for_amounts)  # Supprimer les dates avec /
                        line_for_amounts = re.sub(r'\b\d{1,2}\.\d{1,2}\.\d{2,4}\b', '', line_for_amounts)  # Supprimer les dates avec .
                        line_for_amounts = re.sub(r'\b\d{1,2}\.\d{1,2}\b', '', line_for_amounts)  # Supprimer les dates partielles DD.MM
                        line_for_amounts = re.sub(r'\b(19|20)\d{2}\b', '', line_for_amounts)  # Supprimer les années
                        line_for_amounts = re.sub(r'\b[A-Z0-9]{8,}\b', '', line_for_amounts)  # Supprimer les références longues
                        
                        print(f"DEBUG AMEN - Ligne nettoyée pour montants: '{line_for_amounts}'")
                        
                        # Chercher avec une regex plus permissive
                        all_amounts = re.findall(r'-?\d+[.,]\d+', line_for_amounts)
                        print(f"DEBUG AMEN - Tous les montants trouvés: {all_amounts}")
                        
                        # Chercher aussi les montants sans décimales (mais pas trop longs)
                        all_amounts_no_decimal = re.findall(r'-?\d{1,6}', line_for_amounts)
                        print(f"DEBUG AMEN - Montants sans décimales: {all_amounts_no_decimal}")
                        
                        # Combiner tous les montants trouvés en priorisant ceux à virgule (séparateur décimal AMEN)
                        with_comma = [x for x in all_amounts if ',' in x]
                        without_comma = [x for x in all_amounts if ',' not in x]
                        all_found_amounts = with_comma + without_comma + all_amounts_no_decimal
                        print(f"DEBUG AMEN - Tous les montants combinés: {all_found_amounts}")
                        
                        if all_found_amounts:
                            raw_amounts = all_found_amounts
                            print(f"DEBUG AMEN - Utilisation de tous les montants trouvés: {raw_amounts}")
                    
                        # CORRECTION: Si toujours aucun montant, essayer une recherche plus agressive
                        if not amounts_f:
                            print(f"DEBUG AMEN - 🔍 RECHERCHE AGRESSIVE - TOUS LES NOMBRES")
                            all_numbers = re.findall(r'\d+[.,]?\d*', line_clean)
                            print(f"DEBUG AMEN - Tous les nombres trouvés: {all_numbers}")
                            
                            for number in all_numbers:
                                if len(number) >= 3:  # Au moins 3 chiffres
                                    number_val = to_float(number)
                                    if number_val and 0.1 <= number_val <= 10000:
                                        amounts_f = [number_val]
                                        print(f"DEBUG AMEN - ✅ Nombre sélectionné: {number} -> {number_val}")
                                        break
                    
                        # CORRECTION SPÉCIALE: Si on a trouvé des montants mais qu'ils semblent incorrects
                        if amounts_f and is_end_of_page:
                            print(f"DEBUG AMEN - 🔍 VÉRIFICATION DES MONTANTS TROUVÉS")
                            print(f"DEBUG AMEN - Montants actuels: {amounts_f}")
                            
                            # CORRECTION: Chercher des patterns de montants plus spécifiques
                            # Pattern 1: Montants avec virgule décimale (ex: 5,950)
                            decimal_patterns = re.findall(r'\b\d{1,4}[.,]\d{2,3}\b', combined)
                            print(f"DEBUG AMEN - Patterns décimaux trouvés: {decimal_patterns}")
                            
                            # Pattern 2: Montants entiers raisonnables (ex: 500, 1000)
                            integer_patterns = re.findall(r'\b\d{1,4}\b', combined)
                            print(f"DEBUG AMEN - Patterns entiers trouvés: {integer_patterns}")
                            
                            # Prioriser les montants décimaux (plus probables pour les transactions)
                            if decimal_patterns:
                                for pattern in decimal_patterns:
                                    pattern_val = to_float(pattern)
                                    if pattern_val and 0.1 <= pattern_val <= 10000:
                                        amounts_f = [pattern_val]
                                        print(f"DEBUG AMEN - ✅ Montant décimal sélectionné: {pattern} -> {pattern_val}")
                                        break
                            elif integer_patterns:
                                # Si pas de décimaux, prendre le plus petit entier raisonnable
                                for pattern in sorted(integer_patterns, key=int):
                                    pattern_val = int(pattern)
                                    if 1 <= pattern_val <= 10000:
                                        amounts_f = [float(pattern_val)]
                                        print(f"DEBUG AMEN - ✅ Montant entier sélectionné: {pattern} -> {pattern_val}")
                                        break
                            
                            # CORRECTION: Si aucun pattern raisonnable, garder le plus petit montant trouvé
                            if not amounts_f or (amounts_f and min(amounts_f) > 10000):
                                if amounts_f:
                                    smallest = min(amounts_f)
                                    if smallest > 10000:
                                        # Diviser par 1000 si c'est trop grand (erreur de formatage)
                                        corrected = smallest / 1000
                                        amounts_f = [corrected]
                                        print(f"DEBUG AMEN - 🔧 Montant corrigé (divisé par 1000): {smallest} -> {corrected}")
                                    else:
                                        amounts_f = [smallest]
                                        print(f"DEBUG AMEN - ✅ Plus petit montant gardé: {smallest}")
                    
                    # CORRECTION: Si c'est une ligne de fin de page et qu'on a des montants, validation spéciale
                    if is_end_of_page and amounts_f:
                        print(f"DEBUG AMEN - 🎯 VALIDATION SPÉCIALE FIN DE PAGE")
                        print(f"DEBUG AMEN - Montants avant validation: {amounts_f}")
                        
                        # Filtrer les montants suspects (trop grands pour être des transactions)
                        valid_amounts = []
                        for amount in amounts_f:
                            if not self.is_large_amount(amount) and amount > 0:
                                valid_amounts.append(amount)
                                print(f"DEBUG AMEN - ✅ Montant valide: {amount}")
                            else:
                                print(f"DEBUG AMEN - ❌ Montant rejeté (trop grand): {amount}")
                        
                        if valid_amounts:
                            amounts_f = valid_amounts
                            print(f"DEBUG AMEN - Montants finaux après validation: {amounts_f}")
                        else:
                            print(f"DEBUG AMEN - ⚠️ Aucun montant valide trouvé pour la ligne de fin de page")

                    # Positions X des montants extraits (quand on peut les retrouver dans les mots)
                    pos_list = []  # (xmid, value)
                    if filtered and page_words:
                        norm_filtered = [f.replace(' ', '') for f in filtered]
                        for w in page_words:
                            wt = str(w.get('text', ''))
                            wt_norm = wt.replace(' ', '')
                            if wt_norm in norm_filtered:
                                val = to_float(wt)
                                if val is not None:
                                    xmid = (float(w.get('x0', 0)) + float(w.get('x1', 0))) / 2.0
                                    pos_list.append((xmid, val))

                    # Construire libellé (entre date opération et valeurs)
                    lib = combined[m.end():tail_start].strip()
                    
                    # CORRECTION CRITIQUE: Supprimer TOUTES les dates du libellé pour éviter la confusion
                    # Supprimer les dates au format DD/MM/YYYY
                    lib = re.sub(r"\b\d{2}/\d{2}/\d{4}\b", "", lib)
                    # Supprimer les dates au format DD.MM.YYYY (dans les libellés)
                    lib = re.sub(r"\b\d{1,2}\.\d{1,2}\.\d{2,4}\b", "", lib)
                    # Supprimer les dates partielles DD.MM
                    lib = re.sub(r"\b\d{1,2}\.\d{1,2}\b", "", lib)
                    # Supprimer les années isolées
                    lib = re.sub(r"\b(19|20)\d{2}\b", "", lib)
                    
                    lib = re.sub(r"\s+", " ", lib).strip()
                    
                    print(f"DEBUG AMEN - Libellé nettoyé: '{lib}'")

                    debit = credit = None
                    # CORRECTION: Filtrer les montants très grands (soldes) avant classification
                    filtered_amounts = []
                    for amount in amounts_f:
                        if not self.is_large_amount(amount):
                            filtered_amounts.append(amount)
                        else:
                            print(f"DEBUG AMEN - Montant filtré (solde): {amount}")
                    
                    # CORRECTION: Utiliser les montants filtrés
                    amounts_f = filtered_amounts
                    
                    # 1) Si deux montants trouvés et positions disponibles, trier par X (gauche=Débit, droite=Crédit)
                    if len(pos_list) >= 2:
                        pos_sorted = sorted(pos_list, key=lambda z: z[0])
                        # CORRECTION: Vérifier que les montants ne sont pas des soldes
                        if not self.is_large_amount(pos_sorted[0][1]):
                            debit = abs(pos_sorted[0][1])
                        if not self.is_large_amount(pos_sorted[-1][1]):
                            credit = abs(pos_sorted[-1][1])
                    elif len(amounts_f) >= 2:
                        # fallback si positions manquantes
                        if not self.is_large_amount(amounts_f[0]):
                            debit = abs(amounts_f[0])
                        if not self.is_large_amount(amounts_f[1]):
                            credit = abs(amounts_f[1])
                    elif len(amounts_f) == 1:
                        only = amounts_f[0]
                        # CORRECTION: Vérifier que le montant n'est pas un solde
                        if not self.is_large_amount(only):
                            # 1bis) si une position et des en-têtes, départager par distance aux en-têtes
                            if len(pos_list) == 1 and (debit_x_hdr is not None and credit_x_hdr is not None):
                                xonly = pos_list[0][0]
                                if abs(xonly - debit_x_hdr) <= abs(xonly - credit_x_hdr):
                                    debit = abs(only)
                                else:
                                    credit = abs(only)
                            else:
                                # 2) Heuristiques par mots-clés enrichies (basées sur vos exemples)
                                upper_lib = lib.upper()
                                credit_kw = [
                                # Encaissements / dépôts / remises
                                'VERSEMENT', 'DEPOT', 'ENCAISSEMENT', 'REM COMMER', 'REM COM',
                                'REM COMMERCANT', 'REM COMMERÇANT', 'REM CHEQ', 'REMISE CHEQUE',
                                'CHEQUE DEPOSIT', 'CHQ DEPOSIT','PLV'
                                # Avoir TPE
                                'AV.TPE', 'AV TPE', 'AVTPE',
                                # Virements reçus
                                'VIREMENT RECU', 'VIR RECU', 'CREDIT'
                                ]
                                debit_kw = [
                                # Commissions / frais / taxes / intérêts
                                'COMMISSION', 'COMM SUR REMISE', 'FRAIS', 'AGIOS', 'TVA', 'INTERET',
                                # Retraits / paiements
                                'RETRAIT', 'PAIEMENT',
                                # Effets / cion
                                'CION', 'EFFET', 'CION/EFFET', 'CION EFFET', 'REG EFFET',
                                # Prélèvements / règlements
                                'PRELEVEMENT', 'REGLEMENT', 'PLV REGL', 'PLV REGLE',
                                # Virements émis
                                'VIREMENT EMIS', 'DEBIT',
                                # CORRECTION: Ajouter les mots-clés manquants
                                'COTIS', 'COTIS.TPE', 'COTIS TPE', 'COTISATION'
                                ]
                                # Règles spécifiques (très fréquentes)
                                credit_strong = ['VERSEMENT', 'AV.TPE', 'AV TPE', 'AVTPE', 'CHEQUE DEPOSIT', 'CHQ DEPOSIT']
                                debit_strong = ['COMMISSION SUR REMISE', 'TVA SUR', 'PLV REGLE', 'PLV REGL', 'CION/EFFET', 'CION EFFET', 'COTIS', 'COTIS.TPE']

                                if any(k in upper_lib for k in debit_strong):
                                    debit = abs(only)
                                elif any(k in upper_lib for k in credit_strong):
                                    credit = abs(only)
                                else:
                                    is_credit_kw = any(k in upper_lib for k in credit_kw)
                                    is_debit_kw = any(k in upper_lib for k in debit_kw)
                                    if is_debit_kw and not is_credit_kw:
                                        debit = abs(only)
                                    elif is_credit_kw and not is_debit_kw:
                                        credit = abs(only)
                                    else:
                                        # CORRECTION: Logique améliorée pour les petites transactions
                                        # Si le montant est très petit (< 1), c'est probablement un débit (frais/commissions)
                                        if only < 1.0:
                                            debit = abs(only)
                                            print(f"DEBUG AMEN - Petite transaction classée DÉBIT: {only}")
                                        # 3) Fallback: signe de la donnée extraite
                                        elif (filtered and filtered[0].strip().startswith('-')) or only < 0:
                                            debit = abs(only)
                                        else:
                                            credit = abs(only)

                    # Ignorer lignes de soldes et totaux
                    upper = lib.upper()
                    if re.search(r"(?i)solde|totaux|report", upper):
                        i = j
                        continue

                    # CORRECTION: Logique spéciale pour les dernières lignes de page
                    if is_end_of_page and (debit or credit):
                        print(f"DEBUG AMEN - 🎯 DERNIÈRE LIGNE DE PAGE TRAITÉE")
                        print(f"DEBUG AMEN - Libellé: {lib}")
                        if debit:
                            print(f"DEBUG AMEN - DÉBIT détecté: {debit}")
                        if credit:
                            print(f"DEBUG AMEN - CRÉDIT détecté: {credit}")
                    
                    # CORRECTION: Validation supplémentaire pour les lignes de fin de page
                    if is_end_of_page:
                        # Vérifier que le libellé ne contient pas de mots de totaux
                        lib_upper = lib.upper()
                        if any(word in lib_upper for word in ['TOTAUX', 'TOTAL', 'SOLDE', 'الجملة']):
                            print(f"DEBUG AMEN - ❌ Ligne de fin de page ignorée (contient totaux): {lib}")
                            i = j
                            continue
                        
                        # CORRECTION: Validation spéciale pour les montants de fin de page
                        if debit and debit > 10000:  # Montant suspect pour une transaction normale
                            print(f"DEBUG AMEN - ⚠️ DÉBIT suspect (trop grand): {debit}")
                            # Essayer de corriger en divisant par 1000 (erreur de formatage)
                            if debit > 1000000:
                                debit_corrected = debit / 1000
                                print(f"DEBUG AMEN - 🔧 DÉBIT corrigé: {debit} -> {debit_corrected}")
                                debit = debit_corrected
                        
                        if credit and credit > 10000:  # Montant suspect pour une transaction normale
                            print(f"DEBUG AMEN - ⚠️ CRÉDIT suspect (trop grand): {credit}")
                            # Essayer de corriger en divisant par 1000 (erreur de formatage)
                            if credit > 1000000:
                                credit_corrected = credit / 1000
                                print(f"DEBUG AMEN - 🔧 CRÉDIT corrigé: {credit} -> {credit_corrected}")
                                credit = credit_corrected

                    # CORRECTION: Éviter d'avoir des montants dans les deux colonnes
                    if debit and credit:
                        # Si les montants sont identiques, garder seulement le crédit (versements)
                        if abs(debit - credit) < 0.001:
                            if any(kw in lib.upper() for kw in ["VERSEMENT", "DEPOT", "ENCAISSEMENT", "REM COM", "AV.TPE", "VIREMENT RECU"]):
                                debit = None
                                print(f"DEBUG AMEN - Montant identique, gardé CRÉDIT: {credit}")
                            else:
                                credit = None
                                print(f"DEBUG AMEN - Montant identique, gardé DÉBIT: {debit}")
                        elif debit < 0.01 or credit < 0.01:
                            # Si l'un des montants est quasi-nul, le supprimer
                            if debit < 0.01:
                                debit = None
                                print(f"DEBUG AMEN - Montant quasi-nul supprimé du DÉBIT")
                            else:
                                credit = None
                                print(f"DEBUG AMEN - Montant quasi-nul supprimé du CRÉDIT")
                        else:
                            # CORRECTION: Si deux montants différents, classifier selon le libellé
                            upper_lib = lib.upper()
                            credit_keywords = [
                                'VERSEMENT', 'DEPOT', 'ENCAISSEMENT', 'REM COM', 'AV.TPE', 
                                'VIREMENT RECU', 'VIR RECU', 'CREDIT', 'REMISE'
                            ]
                            debit_keywords = [
                                'COMMISSION', 'FRAIS', 'AGIOS', 'TVA', 'RETRAIT', 'PAIEMENT',
                                'CION', 'EFFET', 'COTIS'
                            ]
                            
                            if any(kw in upper_lib for kw in credit_keywords):
                                debit = None
                                print(f"DEBUG AMEN - Classification CRÉDIT selon libellé: {credit}")
                            elif any(kw in upper_lib for kw in debit_keywords):
                                credit = None
                                print(f"DEBUG AMEN - Classification DÉBIT selon libellé: {debit}")
                            else:
                                # Par défaut, garder le plus petit montant comme crédit
                                if debit < credit:
                                    debit = None
                                    print(f"DEBUG AMEN - Par défaut, gardé CRÉDIT (plus grand): {credit}")
                                else:
                                    credit = None
                                    print(f"DEBUG AMEN - Par défaut, gardé DÉBIT (plus grand): {debit}")
                    
                    # CORRECTION: Préserver les montants originaux avant formatage
                    debit_original = debit
                    credit_original = credit
                    
                    # Formater les montants selon le standard BIAT
                    if debit:
                        debit_formatted = self._format_amount_biats_style(debit)
                        print(f"DEBUG AMEN - DÉBIT formaté: {debit} -> {debit_formatted}")
                    else:
                        debit_formatted = None
                        
                    if credit:
                        credit_formatted = self._format_amount_biats_style(credit)
                        print(f"DEBUG AMEN - CRÉDIT formaté: {credit} -> {credit_formatted}")
                    else:
                        credit_formatted = None
                    
                    results.append({
                        "date": op_date,
                        "libelle": lib,
                        "debit": debit_formatted,
                        "credit": credit_formatted
                    })

                    i = j
        return results

    def _format_amount_biats_style(self, amount_float):
        """Formate un montant selon le style BIAT: espace pour milliers, virgule pour décimales"""
        if amount_float is None:
            return None
        
        try:
            # Formater avec 3 décimales exactement
            formatted = f"{amount_float:,.3f}"
            # Remplacer les virgules par des espaces (milliers) et le point par une virgule (décimales)
            formatted = formatted.replace(',', ' ').replace('.', ',')
            return formatted
        except (ValueError, TypeError):
            return str(amount_float) if amount_float is not None else None

    def _format_excel(self, path: str):
        wb = load_workbook(path)
        ws = wb.active
        ws.title = "J03"
        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        # CORRECTION: Les montants sont déjà formatés en format BIAT dans le DataFrame
        # Pas besoin de reformater dans Excel
        max_row = ws.max_row
        print(f"DEBUG AMEN - Formatage Excel: {max_row} lignes à traiter")
        # Bordures fines sur tout le tableau
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
        wb.save(path)

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

def main():
    root = tk.Tk(); AmenReleveConverter(root); root.mainloop()

if __name__ == '__main__':
    main()

