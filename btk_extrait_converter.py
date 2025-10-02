import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import subprocess
import sys
import pdfplumber
import pandas as pd
from datetime import datetime
import re
try:
    import pytesseract  # type: ignore
    from PIL import Image
    import fitz
    import numpy as np
    import cv2
    TESS = True
except Exception:
    TESS = False


class BTKExtraitConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAİT BTK vers Excel")
        self.root.geometry("600x480")
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar(value=f"EXTRAIT_BTK_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        self._ui()

    def _ui(self):
        tk.Label(self.root, text="Convertisseur EXTRAİT BTK", font=("Arial", 18, "bold")).pack(pady=(20, 4))
        frm = tk.Frame(self.root); frm.pack(pady=10, padx=40, fill='x')
        tk.Label(frm, text="Fichier PDF EXTRAİT BTK:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(frm, width=48, textvariable=self.pdf_path).grid(row=1, column=0, padx=(0,10))
        tk.Button(frm, text="Parcourir", command=self.browse, bg="#3498db", fg='white').grid(row=1, column=1)
        tk.Label(frm, text="Nom du fichier Excel:", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky='w', pady=(15,5))
        tk.Entry(frm, width=48, textvariable=self.excel_name).grid(row=3, column=0, padx=(0,10))
        self.progress = ttk.Progressbar(self.root, mode='indeterminate', length=520)
        self.progress.pack(pady=16)
        btns = tk.Frame(self.root); btns.pack()
        tk.Button(btns, text="Convertir en Excel", command=self.convertir, bg="#16A34A", fg='white', font=("Arial", 13, "bold")).pack(side='left', padx=6)
        tk.Button(btns, text="Retour", command=self.back, bg="#e74c3c", fg='white', font=("Arial", 11, "bold")).pack(side='left', padx=6)

    def browse(self):
        f = filedialog.askopenfilename(title="Choisir un PDF BTK EXTRAIT", filetypes=[["PDF", "*.pdf"]])
        if f: self.pdf_path.set(f)

    def back(self):
        try:
            self.root.destroy(); subprocess.Popen([sys.executable, "lancer_convertisseur.py"])  # noqa
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def convertir(self):
        path = self.pdf_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("PDF manquant", "Veuillez choisir un fichier PDF BTK extrait.")
            return
        self.progress.start(); self.root.update_idletasks()
        rows = self._extract_rows(path)
        if not rows:
            self.progress.stop(); messagebox.showerror("Aucune transaction", "Impossible d'extraire des transactions BTK."); return
        df = pd.DataFrame(rows, columns=["date", "libelle", "debit", "credit"])
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out = os.path.join(downloads, f"{self.excel_name.get().strip() or 'EXTRAIT_BTK'}.xlsx")
        df.to_excel(out, index=False)
        self._format_excel(out)
        self.progress.stop()
        
        # Message de succès plus positif
        success_msg = f"✅ Conversion EXTRAT terminée avec succès !\n\n"
        success_msg += f"📁 Fichier enregistré: {out}\n\n"
        success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
        
        messagebox.showinfo("✅ Conversion réussie", success_msg)

    def _extract_rows(self, pdf_path: str):
        # 1) tentative texte
        try:
            with pdfplumber.open(pdf_path) as pdf:
                consolidated = []
                cur = None
                for page in pdf.pages:
                    words = page.extract_words(use_text_flow=True)
                    if not words:
                        continue
                    # colonnes approximatives: Date | Opération | Débit | Crédit | Solde
                    page_width = page.width
                    edges = [0, 120, 520, 660, 760, page_width]
                    # grouper par ligne
                    line_map = {}
                    for w in words:
                        y = int(round((w['top'] + w['bottom'])/2)); line_map.setdefault(y, []).append(w)
                    for _, ws in sorted(line_map.items()):
                        ws.sort(key=lambda t: t['x0'])
                        text_line = ' '.join(w['text'] for w in ws)
                        up = text_line.upper()
                        if any(k in up for k in ['BTK@DIRECT', 'EXTRAIT DE COMPTE', 'SOLDE', 'PAGE ', 'IMPRIMER', 'HTTPS://']):
                            continue
                        buckets = [[] for _ in range(5)]
                        for w in ws:
                            x = w['x0']
                            for i in range(5):
                                if edges[i] <= x < edges[i+1]:
                                    buckets[i].append(w['text']); break
                        date_raw = ' '.join(buckets[0]).strip()
                        date = self._norm_date(date_raw)
                        lib = (' '.join(buckets[1]).strip() + ' ' + ' '.join(buckets[2]).strip()).strip()
                        debit = self._fmt_amount(' '.join(buckets[3]).strip())
                        credit = self._fmt_amount(' '.join(buckets[4]).strip())
                        if date:
                            if cur: consolidated.append(cur)
                            cur = {'date': date, 'libelle': lib, 'debit': debit, 'credit': credit}
                        else:
                            # continuation
                            if cur:
                                if lib:
                                    cur['libelle'] = (cur['libelle'] + ' ' + lib).strip()
                                if not cur['debit'] and debit: cur['debit'] = debit
                                if not cur['credit'] and credit: cur['credit'] = credit
                if cur: consolidated.append(cur)
                if consolidated:
                    return consolidated
        except Exception:
            pass
        # 2) OCR fallback
        if not TESS:
            # 2bis) Fallback texte brut si available
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    txt = "\n".join([p.extract_text() or '' for p in pdf.pages])
                rows = self._extract_rows_from_text(txt)
                if rows:
                    return rows
            except Exception:
                pass
            return []
        try:
            doc = fitz.open(pdf_path)
            text_all = ""
            for p in doc:
                mat = fitz.Matrix(2,2); pix = p.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
                text_all += pytesseract.image_to_string(thr, lang='fra') + "\n"
            return self._parse_text_ocr(text_all)
        except Exception:
            return []

    def _parse_text_ocr(self, text: str):
        rows = []
        date_re = re.compile(r"(\d{2}[./-]\d{2}[./-]\d{4})")
        amount_re = re.compile(r"\d{1,3}(?:\s\d{3})*(?:[.,]\d{3})")
        cur = None
        for line in [ln.strip() for ln in text.splitlines() if ln.strip()]:
            d = date_re.search(line)
            if d:
                if cur: rows.append(cur)
                cur = {'date': self._norm_date(d.group(1)), 'libelle': line[d.end():].strip(), 'debit': None, 'credit': None}
            else:
                if cur:
                    cur['libelle'] = (cur['libelle'] + ' ' + line).strip()
            amts = amount_re.findall(line)
            if cur and amts:
                if len(amts) == 2:
                    cur['debit'] = self._fmt_amount(amts[0]); cur['credit'] = self._fmt_amount(amts[1])
                elif len(amts) == 1 and cur['debit'] is None and cur['credit'] is None:
                    # heuristique: petits = débit
                    val = self._fmt_amount(amts[0]);
                    cur['debit'] = val if self._as_float(val) and self._as_float(val) < 1000 else None
                    cur['credit'] = None if cur['debit'] else val
        if cur: rows.append(cur)
        return rows

    def _fmt_amount(self, s: str):
        if not s: return None
        v = s.replace('\u00A0', ' ').replace(' ', '').replace('.', '').replace(',', '.')
        try:
            f = float(v); return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
        except Exception:
            return None

    def _as_float(self, s: str):
        try:
            return float(s.replace(' ', '').replace(',', '.'))
        except Exception:
            return None

    def _norm_date(self, raw: str):
        if not raw: return ''
        try:
            m = re.search(r"(\d{2})[./-](\d{2})[./-](\d{4})", raw)
            if not m: return ''
            d, mth, y = m.groups()
            return f"{d}/{mth}/{y}"
        except Exception:
            return ''

    def _format_excel(self, path: str):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = load_workbook(path)
            ws = wb.active
            ws.title = "Extrait BTK"
            # tailles colonnes
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 16
            # header style
            yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
            bold = Font(bold=True)
            center = Alignment(horizontal='center', vertical='center')
            for cell in ws[1]:
                cell.fill = yellow; cell.font = bold; cell.alignment = center
            # number formats
            for r in range(2, ws.max_row + 1):
                ws[f'C{r}'].number_format = '# ##0,000'
                ws[f'D{r}'].number_format = '# ##0,000'
            # borders
            thin = Side(style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
                for c in row:
                    c.border = border
            wb.save(path)
        except Exception:
            pass

    def _extract_rows_from_text(self, text: str):
        rows = []
        if not text:
            return rows
        # 0) Tentative stricte par regex ligne complète (Date, Date valeur, Libellé, Débit, Crédit, Solde)
        try:
            norm = '\n'.join(re.sub(r"\s+", " ", ln.strip()) for ln in text.splitlines())
            amt = r"\d{1,3}(?:\s\d{3})*,\d{3}"
            row_re = re.compile(
                rf"^(?P<date>\d{{2}}/\d{{2}}/\d{{4}})\s+(?:\d{{2}}/\d{{2}}/\d{{4}})\s+"
                rf"(?P<libelle>.+?)\s+"
                rf"(?:(?P<debit>{amt}))?\s*"
                rf"(?:(?P<credit>{amt}))?\s+{amt}$",
                re.MULTILINE)
            for m in row_re.finditer(norm):
                date = self._norm_date(m.group('date'))
                lib = m.group('libelle').strip()
                debit = self._fmt_amount(m.group('debit') or '')
                credit = self._fmt_amount(m.group('credit') or '')
                rows.append({'date': date, 'libelle': lib, 'debit': debit or None, 'credit': credit or None})
            if rows:
                return rows
        except Exception:
            pass
        # Normalisations basiques
        lines = [re.sub(r"\s+", " ", ln.strip()) for ln in text.splitlines() if ln.strip()]
        date_re = re.compile(r"^(\d{2}/\d{2}/\d{4})\b")
        amount_re = re.compile(r"\d{1,3}(?:\s\d{3})*(?:[.,]\d{3})")
        headers = ['EXTRAIT DE COMPTE', 'SOLDE ', 'DATE DE VALEUR', 'DATE VALEUR', 'OPÉRATION', 'OPERATION', 'DÉBIT', 'DEBIT', 'CRÉDIT', 'CREDIT', 'BTK@DIRECT']
        cur = None
        for raw in lines:
            up = raw.upper()
            if any(h in up for h in headers):
                continue
            m = date_re.match(raw)
            if m:
                # Commit previous row
                if cur:
                    rows.append(cur)
                date = self._norm_date(m.group(1))
                rest = raw[m.end():].strip()
                # Certains PDFs ont la date de valeur comme 2e date
                rest = re.sub(r"^\d{2}/\d{2}/\d{4}\s+", "", rest)
                # Chercher les montants sur la ligne (débit, crédit, solde)
                amts = amount_re.findall(rest)
                debit = credit = None
                lib = rest
                if amts:
                    # Le dernier est souvent le solde -> ignorer
                    core = amts[:-1] if len(amts) >= 2 else amts
                    # Déterminer la limite de libellé avant le premier montant
                    pos_first = rest.find(core[0]) if core else -1
                    if pos_first > 0:
                        lib = rest[:pos_first].strip()
                    # Classification
                    if len(core) >= 2:
                        debit = self._fmt_amount(core[0])
                        credit = self._fmt_amount(core[1])
                    elif len(core) == 1:
                        single = self._fmt_amount(core[0])
                        hint = lib.upper()
                        if any(k in hint for k in ['RETRAIT', 'PAIEMENT', 'PRÉL', 'PRELEV', 'FRAIS', 'AGIOS', 'COM ', ' TVA', 'VIR. EMIS']):
                            debit = single
                        else:
                            credit = single
                cur = {'date': date, 'libelle': lib, 'debit': debit, 'credit': credit}
            else:
                # Ligne de continuation du libellé
                if cur:
                    # éviter d'ajouter les URL/mentions techniques
                    if not re.search(r"https?://|IMPRIMER|PAGE \d+/\d+", raw, re.IGNORECASE):
                        cur['libelle'] = (cur['libelle'] + ' ' + raw).strip()
        if cur:
            rows.append(cur)
        return rows


def main():
    root = tk.Tk(); BTKExtraitConverter(root); root.mainloop()


if __name__ == '__main__':
    main()


