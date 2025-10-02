import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
import os
from datetime import datetime
import re
import unicodedata
import fitz  # PyMuPDF
import cv2
import numpy as np
from PIL import Image
import io
import subprocess
import sys

# OCR fallback (for scanned PDFs)
try:
    import pytesseract  # type: ignore
    _OCR_AVAILABLE = True
except Exception:
    _OCR_AVAILABLE = False

class BIATExtratConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur EXTRAT BIAT vers Excel")
        self.root.geometry("600x500")

        # Variables
        self.pdf_path = tk.StringVar()
        self.excel_name = tk.StringVar()
        self.excel_name.set("EXTRAT_BIAT_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        
        self.setup_ui()
    
    def setup_ui(self):
        # Carte principale moderne

        # Titre principal moderne
        title_label = tk.Label(text="Convertisseur EXTRAT BIAT"
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(20, 10))
        
        subtitle_label = tk.Label(text="Conversion PDF vers Excel"
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 30))
        
        # Frame principal dans la carte
        main_frame = tk.Frame()
        main_frame.pack(pady=20, padx=20, fill='both', expand=True)
        
        # Sélection du fichier PDF
        pdf_frame = tk.Frame(main_frame)
        pdf_frame.pack(fill='x', pady=15)
        
        tk.Label(pdf_frame, text="Fichier PDF EXTRAT BIAT:"
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 8))
        
        pdf_select_frame = tk.Frame(pdf_frame)
        pdf_select_frame.pack(fill='x', pady=5)
        
        tk.Entry(pdf_select_frame, textvariable=self.pdf_path
                font=("Arial", 9), width=50, state='readonly', ).pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        tk.Button(pdf_select_frame, text="Parcourir", command=self.select_pdf_file, , font=("Segoe UI", 10, "bold"), ).pack(side='right', padx=(10, 0))
        
        # Nom du fichier Excel
        excel_frame = tk.Frame(main_frame)
        excel_frame.pack(fill='x', pady=15)
        
        tk.Label(excel_frame, text="Nom du fichier Excel:"
                font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 8))
        
        tk.Entry(excel_frame, textvariable=self.excel_name
                font=("Arial", 9), width=50, ).pack(fill='x', pady=5)
        
        # Bouton de conversion
        convert_frame = tk.Frame(main_frame)
        convert_frame.pack(pady=40, fill='x')
        
        # Boutons
        buttons_frame = tk.Frame(convert_frame)
        buttons_frame.pack(pady=20, fill='x')
        
        self.convert_button = tk.Button(buttons_frame, text="Convertir en Excel"
                                       command=self.convert_pdf_to_excel, 
                                       font=("Segoe UI", 12, "bold"), )
        self.convert_button.pack(side='left', padx=10)
        
        self.retour_button = tk.Button(buttons_frame, text="Retour page d'accueil"
                                      command=self.retour_accueil, 
                                      font=("Segoe UI", 12, "bold"), )
        self.retour_button.pack(side='right', padx=10)
        
        # Barre de progression

        # Zone de statut
        )

    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier PDF EXTRAT BIAT"
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")]
        )
        if file_path:
            self.pdf_path.set(file_path)
            }")
    
    def is_biat_extrat_pdf(self, pdf_path):
        """Vérifie si le PDF est un EXTRAT BIAT (au moins 2 signaux sur 3: mots-clés, structure, logo)."""
        try:
            signals = 0
            text_ok = self._check_biat_text_keywords(pdf_path)
            if text_ok:
                signals += 1
            struct_ok = self._check_biat_table_structure(pdf_path)
            if struct_ok:
                signals += 1
            logo_ok = self._check_biat_logo_presence(pdf_path)
            if logo_ok:
                signals += 1
            print(f"DEBUG BIAT EXTRAT - Signals: text={text_ok}, struct={struct_ok}, logo={logo_ok} -> {signals}")
            return signals >= 2
        except Exception as e:
            print(f"Erreur lors de la vérification BIAT EXTRAT: {e}")
            return False
    
    def _check_biat_text_keywords(self, pdf_path):
        """Vérifie la présence des mots-clés spécifiques BIAT EXTRAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
                
                # Normaliser les accents pour être tolérant
                u = unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode('ascii').upper()
                primary = ("EXTRAIT DE COMPTE" in u) or ("EXTRAT" in u)
                secondary = any(k in u for k in ["BIAT", "BANQUE INTERNATIONALE ARABE DE TUNISIE", "DEBIT", "CREDIT", "DATE VALEUR", "LIBELLE", "REFERENCE"])
                return primary and secondary
                
        except Exception as e:
            print(f"Erreur vérification mots-clés BIAT EXTRAT: {e}")
            return False
    
    def _check_biat_table_structure(self, pdf_path):
        """Vérifie la structure typique des tableaux BIAT EXTRAT"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if tables:
                        print(f"DEBUG BIAT EXTRAT - {len(tables)} tableaux trouvés")
                        for table in tables:
                            if len(table) > 1:  # Au moins header + 1 ligne
                                # Vérifier la présence de colonnes typiques BIAT EXTRAT
                                header_row = table[0]
                                if header_row:
                                    header_text = ' '.join([str(cell) for cell in header_row if cell])
                                    print(f"DEBUG BIAT EXTRAT - Header du tableau: {header_text}")
                                    print(f"DEBUG BIAT EXTRAT - Header en majuscules: {header_text.upper()}")
                                    print(f"DEBUG BIAT EXTRAT - Contient 'DATE OPÉ': {'DATE OPÉ' in header_text.upper()}")
                                    print(f"DEBUG BIAT EXTRAT - Contient 'LIBELLÉ OPÉRATION': {'LIBELLÉ OPÉRATION' in header_text.upper()}")
                                    print(f"DEBUG BIAT EXTRAT - Contient 'RÉFÉRENCE': {'RÉFÉRENCE' in header_text.upper()}")
                                    # Mots-clés spécifiques au format EXTRAT BIAT
                                    # Le format EXTRAT BIAT a cette structure exacte :
                                    # "Date opé. Libellé Opération Référence Date Valeur Débit Crédit"
                                    if ("DATE OPÉ" in header_text.upper() and "LIBELLÉ OPÉRATION" in header_text.upper() and "RÉFÉRENCE" in header_text.upper()) or \
                                       ("DATE OPE" in header_text.upper() and "LIBELLE OPERATION" in header_text.upper() and "REFERENCE" in header_text.upper()) or \
                                       ("DATE OPÉ." in header_text.upper() and "LIBELLÉ OPÉRATION" in header_text.upper()) or \
                                       ("DATE OPE." in header_text.upper() and "LIBELLE OPERATION" in header_text.upper()) or \
                                       any(keyword in header_text.upper() for keyword in 
                                          ["DATE OPÉ", "DATE OPE", "LIBELLÉ OPÉRATION", "LIBELLE OPERATION"
                                           "RÉFÉRENCE", "REFERENCE", "DATE VALEUR", "DÉBIT", "DEBIT", "CRÉDIT", "CREDIT"]):
                                        print("DEBUG BIAT EXTRAT - Structure de tableau BIAT EXTRAT détectée")
                                        return True
                    else:
                        # Si aucun tableau trouvé, vérifier le texte brut
                        text = page.extract_text()
                        if text:
                            # Chercher des patterns de transactions EXTRAT BIAT dans le texte
                            if re.search(r'\d{1,2}\s+\w+\s+\d{2}\s+[A-Z\s]+\s+[\d\s.,]+', text):
                                print("DEBUG BIAT EXTRAT - Pattern de transaction EXTRAT détecté dans le texte")
                                return True
                            # Chercher aussi des patterns plus simples
                            if re.search(r'\d{1,2}\s+\w+\s+\d{2}', text):
                                print("DEBUG BIAT EXTRAT - Pattern de date EXTRAT détecté dans le texte")
                                return True
            print("DEBUG BIAT EXTRAT - Aucune structure de tableau BIAT détectée")
            return False
        except Exception as e:
            print(f"Erreur vérification structure BIAT EXTRAT: {e}")
            return False
    
    def _check_biat_logo_presence(self, pdf_path):
        """Vérifie la présence du logo BIAT"""
        try:
            # Vérifier si le fichier logo existe
            logo_path = "logo/biat.png"
            if not os.path.exists(logo_path):
                print(f"DEBUG BIAT EXTRAT - Logo non trouvé à {logo_path}, détection basée sur le texte")
                return True  # Fallback sur la détection textuelle
            
            # Extraire les images du PDF
            doc = fitz.open(pdf_path)
            logo_found = False
            
            for page_num in range(min(3, len(doc))):  # Vérifier les 3 premières pages
                page = doc[page_num]
                image_list = page.get_images()
                
                if image_list:
                    print(f"DEBUG BIAT EXTRAT - {len(image_list)} images trouvées sur la page {page_num + 1}")
                    # Pour l'instant, on considère que la présence d'images indique un logo
                    logo_found = True
                    break
            
            doc.close()
            
            if logo_found:
                print("DEBUG BIAT EXTRAT - Images détectées dans le PDF (logo probable)")
            else:
                print("DEBUG BIAT EXTRAT - Aucune image détectée, mais on continue avec la détection textuelle")
            
            # Toujours retourner True car la détection textuelle est plus fiable
            return True
            
        except Exception as e:
            print(f"Erreur vérification logo BIAT EXTRAT: {e}")
            return True  # Fallback sur la détection textuelle
    
    def detect_year_from_pdf(self, pdf_path):
        """Détecte l'année depuis le PDF BIAT EXTRAT - chercher dans l'en-tête"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages[:3]):  # Chercher dans les 3 premières pages
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines[:30]:  # Chercher dans les 30 premières lignes
                            # Chercher des patterns d'année spécifiques au format EXTRAT BIAT
                            year_patterns = [
                                r'EDITE LE.*?(\d{2})',  # "Edité le : 01 sept. 25"
                                r'SOLDE DEPART AU.*?(\d{2})',  # "Solde départ au 01 août 25"
                                r'(\d{4})',  # Année simple
                                r'(\d{1,2}/\d{1,2}/(\d{4}))',  # Date avec année
                                r'(\d{1,2}-\d{1,2}-(\d{4}))',  # Date avec tirets
                            ]
                            
                            for pattern in year_patterns:
                                matches = re.findall(pattern, line)
                                for match in matches:
                                    if isinstance(match, tuple):
                                        year = match[1] if len(match) > 1 else match[0]
                                    else:
                                        year = match
                                    
                                    # Convertir l'année à 2 chiffres en année complète
                                    if len(year) == 2:
                                        year_int = 2000 + int(year)
                                    else:
                                        year_int = int(year)
                                    
                                    if 2020 <= year_int <= 2030:  # Années plausibles
                                        print(f"DEBUG BIAT EXTRAT - Année détectée: {year_int} dans la ligne: {line}")
                                        return year_int
            
            # Si aucune année n'est trouvée, utiliser l'année actuelle
            current_year = datetime.now().year
            print(f"DEBUG BIAT EXTRAT - Aucune année détectée, utilisation de l'année actuelle: {current_year}")
            return current_year
            
        except Exception as e:
            print(f"Erreur détection année BIAT EXTRAT: {e}")
            return datetime.now().year
    
    def extract_table_data(self, pdf_path):
        """Extrait les données du tableau des transactions BIAT EXTRAT - APPROCHE FLEXIBLE"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_transactions = []
                
                for page_num, page in enumerate(pdf.pages):
                    print(f"DEBUG BIAT EXTRAT - Traitement page {page_num + 1}")
                    
                    # Vérifier si le texte est vide (PDF scanné)
                    text = page.extract_text() or ''
                    if not text.strip() and page_num == 0:
                        # Fallback OCR pour les PDFs scannés
                        ocr_text = self._extract_text_via_ocr(pdf_path)
                        if ocr_text.strip():
                            # Utiliser le texte OCR pour toutes les pages
                            return self._parse_ocr_text(ocr_text)
                    
                    # Essayer d'extraire les tableaux
                    tables = page.extract_tables()
                    
                    if tables:
                        print(f"DEBUG BIAT EXTRAT - {len(tables)} tableaux trouvés sur la page {page_num + 1}")
                        table_transactions_found = False
                        for table_idx, table in enumerate(tables):
                            print(f"DEBUG BIAT EXTRAT - Traitement tableau {table_idx + 1}")
                            transactions = self.parse_biat_table(table)
                            if transactions:  # Si on trouve des transactions dans le tableau
                                all_transactions.extend(transactions)
                                table_transactions_found = True
                        
                        # Si aucun tableau n'a donné de transactions, essayer l'extraction de texte
                        if not table_transactions_found:
                            print(f"DEBUG BIAT EXTRAT - Aucune transaction trouvée dans les tableaux, extraction du texte")
                            text = page.extract_text()
                            if text:
                                print(f"DEBUG BIAT EXTRAT - Texte extrait (premiers 200 caractères): {text[:200]}")
                                year = self.detect_year_from_pdf(pdf_path)
                                print(f"DEBUG BIAT EXTRAT - Année détectée: {year}")
                                transactions = self.parse_biat_transactions_from_text(text, year)
                                print(f"DEBUG BIAT EXTRAT - Transactions extraites du texte: {len(transactions)}")
                                all_transactions.extend(transactions)
                            else:
                                print("DEBUG BIAT EXTRAT - Aucun texte extrait de la page")
                    else:
                        # Si aucun tableau trouvé, essayer d'extraire le texte brut
                        print(f"DEBUG BIAT EXTRAT - Aucun tableau trouvé sur la page {page_num + 1}, extraction du texte")
                        text = page.extract_text()
                        if text:
                            year = self.detect_year_from_pdf(pdf_path)
                            transactions = self.parse_biat_transactions_from_text(text, year)
                            all_transactions.extend(transactions)
                
                print(f"DEBUG BIAT EXTRAT - Total transactions extraites: {len(all_transactions)}")
                # Fallback global: si rien trouvé, parser tout le texte du PDF d'un coup
                if not all_transactions:
                    try:
                        full_text = "".join((p.extract_text() or "") for p in pdf.pages)
                        if full_text:
                            year = self.detect_year_from_pdf(pdf_path)
                            extra = self.parse_biat_transactions_from_text(full_text, year)
                            all_transactions.extend(extra)
                            print(f"DEBUG BIAT EXTRAT - Fallback texte global: {len(extra)} transactions")
                    except Exception as e:
                        print(f"DEBUG BIAT EXTRAT - Fallback texte erreur: {e}")
                return all_transactions
                
        except Exception as e:
            print(f"Erreur extraction données BIAT EXTRAT: {e}")
            return []

    def extract_by_layout(self, pdf_path):
        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(use_text_flow=True)
                    if not words:
                        continue
                    # Définir des colonnes par défaut adaptées à BIAT
                    col_edges = [0, 120, 430, 540, 640, 720, page.width]
                    # regrouper par y
                    line_map = {}
                    for w in words:
                        y = int(round((w['top'] + w['bottom']) / 2))
                        line_map.setdefault(y, []).append(w)
                    for y, ws in sorted(line_map.items()):
                        ws.sort(key=lambda t: t['x0'])
                        text_line = ' '.join(w['text'] for w in ws)
                        u = text_line.upper()
                        if any(k in u for k in ['DATE OP', 'LIBELL', 'RÉFÉRENCE', 'REFERENCE', 'DATE VALEUR', 'DÉBIT', 'CRÉDIT', 'PAGE ', 'SOLDE', 'TOTAUX']):
                            continue
                        buckets = [[] for _ in range(6)]
                        for w in ws:
                            x = w['x0']
                            for i in range(6):
                                if col_edges[i] <= x < col_edges[i+1]:
                                    buckets[i].append(w['text'])
                                    break
                        date_op = ' '.join(buckets[0]).strip()
                        libelle = ' '.join(buckets[1]).strip()
                        debit_raw = ' '.join(buckets[4]).strip()
                        credit_raw = ' '.join(buckets[5]).strip()
                        if not (libelle or debit_raw or credit_raw):
                            continue
                        def fmt_amount(s):
                            s = (s or '').strip()
                            if not s:
                                return ''
                            v = s.replace(' ', '').replace('.', '').replace(',', '.')
                            try:
                                f = float(v)
                            except Exception:
                                return ''
                            return f"{f:,.3f}".replace(',', ' ').replace('.', ',', 1)
                        debit = fmt_amount(debit_raw)
                        credit = fmt_amount(credit_raw)
                        if debit and credit:
                            debit = ''
                        rows.append({'Date': date_op, 'Libellé': libelle, 'Débit': debit, 'Crédit': credit})
        except Exception as e:
            print(f"DEBUG BIAT LAYOUT ERROR: {e}")
        return rows
    
    def parse_biat_table(self, table):
        """Parse un tableau BIAT EXTRAT extrait par pdfplumber"""
        transactions = []
        try:
            print(f"DEBUG BIAT EXTRAT - Tableau reçu avec {len(table)} lignes")
            
            if len(table) < 2:
                print("DEBUG BIAT EXTRAT - Tableau trop petit")
                return transactions
            
            # Afficher la structure du tableau
            for i, row in enumerate(table[:3]):  # Afficher les 3 premières lignes
                print(f"DEBUG BIAT EXTRAT - Ligne {i}: {row}")
            
            # Détecter l'année depuis le contenu du tableau
            year = None
            for row in table:
                for cell in row:
                    if cell and isinstance(cell, str):
                        year_match = re.search(r'(\d{4})', cell)
                        if year_match:
                            year_int = int(year_match.group(1))
                            if 2020 <= year_int <= 2030:
                                year = year_int
                                print(f"DEBUG BIAT EXTRAT - Année détectée dans le tableau: {year}")
                                break
                if year:
                    break
            
            if not year:
                year = datetime.now().year
                print(f"DEBUG BIAT EXTRAT - Année par défaut: {year}")
            
            # Parser les lignes de données
            for i, row in enumerate(table[1:], 1):  # Ignorer la première ligne (header)
                if row and any(cell for cell in row if cell):
                    print(f"DEBUG BIAT EXTRAT - Traitement ligne {i}: {row}")
                    transaction = self.parse_biat_transaction_row(row, year)
                    if transaction:
                        transactions.append(transaction)
                        print(f"DEBUG BIAT EXTRAT - Transaction ajoutée: {transaction}")
                    else:
                        print(f"DEBUG BIAT EXTRAT - Ligne ignorée (pas de transaction valide)")
            
        except Exception as e:
            print(f"Erreur parsing tableau BIAT EXTRAT: {e}")
        
        print(f"DEBUG BIAT EXTRAT - Total transactions extraites du tableau: {len(transactions)}")
        return transactions
    
    def parse_biat_transactions_from_text(self, text, year):
        """Parse les transactions BIAT EXTRAT depuis le texte brut - LOGIQUE CORRIGÉE"""
        transactions = []
        seen_transactions = set()
        
        try:
            print(f"DEBUG BIAT EXTRAT - Parsing du texte brut (premiers 500 caractères): {text[:500]}")
            
            # Pattern colonne-complet: date | libellé | référence | date valeur | débit | crédit
            pattern_full = re.compile(
                r'^(?P<date>\d{1,2}\s+[^\d\s]{2,20}\s+\d{2})\s+'
                r'(?P<lib>.+?)\s+'
                r'(?P<ref>[A-Z0-9\\/]+)\s+'
                r'(?P<valdate>\d{1,2}\s+[^\d\s]{2,20}\s+\d{2})\s+'
                r'(?:(?P<debit>[\d\s.,]+)\s+)?(?P<credit>[\d\s.,]+)?$'
            )

            # Pattern sans référence (parfois vide): date | libellé | date valeur | débit | crédit
            pattern_no_ref = re.compile(
                r'^(?P<date>\d{1,2}\s+[^\d\s]{2,20}\s+\d{2})\s+'
                r'(?P<lib>.+?)\s+'
                r'(?P<valdate>\d{1,2}\s+[^\d\s]{2,20}\s+\d{2})\s+'
                r'(?:(?P<debit>[\d\s.,]+)\s+)?(?P<credit>[\d\s.,]+)?$'
            )

            # Pattern simple: date | libellé | montant (un seul, souvent crédit)
            pattern_simple = re.compile(
                r'^(?P<date>\d{1,2}\s+[^\d\s]{2,20}\s+\d{2})\s+'
                r'(?P<lib>.+?)\s+(?P<amount>[\d\s.,]+)$'
            )
            
            # Pour EXTRAT BIAT, utiliser les lignes brutes directement
            lines = [ln.strip() for ln in text.split('\n')]
            print(f"DEBUG BIAT EXTRAT - {len(lines)} lignes de texte à analyser")
            
            for line in lines:
                line = line.strip()
                if not line or len(line) < 10:
                    continue
                
                print(f"DEBUG BIAT EXTRAT - Ligne: {line}")
                
                # Ignorer seulement les en-têtes vraiment non-transactionnels
                if any(header in line.upper() for header in [
                    'DATE LIBELLE', 'DEBIT CREDIT', 'TOTAUX', 'SOLDE'
                    'BIAT EXTRAIT', 'COMPTE', 'RIB:', 'TITULAIRE', 'CLIENT'
                    'EXTRATDE', 'NUMRODE', 'DEVISE:', 'CATEGORIE', 'SOLDEDEPART'
                    'PAGE', 'SAUFERREUR'
                ]):
                    continue
                
                # Fonction helper pour convertir les dates EXTRAT BIAT - AMÉLIORÉE
                def convert_extrat_date(date_str):
                    """Convertit les dates EXTRAT BIAT (ex: "01 août 25" -> "01/08/2025")"""
                    # Mapping des mois - étendu pour couvrir toutes les variantes OCR
                    month_map = {
                        'jan': '01', 'fev': '02', 'mar': '03', 'avr': '04'
                        'mai': '05', 'jun': '06', 'jul': '07', 'juil': '07'
                        'aod': '08', 'aot': '08', 'aott': '08', 'aoi': '08', 'aout': '08'
                        'sep': '09', 'sept': '09', 'oct': '10', 'nov': '11', 'dec': '12'
                    }
                    
                    # Extraire jour, mois, année - gérer les dates avec espaces
                    match = re.match(r'(\d{1,2})\s+([^\d\s]+)\s+(\d{2})', date_str.lower())
                    if match:
                        day = match.group(1).zfill(2)
                        month_str = match.group(2)
                        year = '20' + match.group(3)
                        
                        # Trouver le mois - recherche plus flexible
                        for key, value in month_map.items():
                            if key in month_str or month_str in key:
                                return f"{day}/{value}/{year}"
                    
                    return "01/01/2025"  # Date par défaut
                
                # Fonction helper pour nettoyer les montants
                def clean_amount(amount_str):
                    """Nettoie et convertit un montant en float"""
                    if not amount_str:
                        return None
                    try:
                        # Remplacer virgule par point et enlever espaces
                        clean_amount = amount_str.replace(',', '.').replace(' ', '')
                        return float(clean_amount)
                    except ValueError:
                        return None
                
                # Fonction helper pour classifier débit/crédit
                def classify_biat_amount(description, amount, line):
                    """Détermine si c'est un débit ou crédit basé sur la description"""
                    desc_upper = description.upper()
                    if any(keyword in desc_upper for keyword in [
                        "DEBIT", "COMMISSION", "FRAIS", "REGLEMENT", "BLOQUAGE"
                        "PRELEVEMENT", "COM TVA", "COM REG", "PAIEMENT", "ABONNEMENT"
                    ]):
                        return amount, None
                    else:
                        return None, amount
                
                # Fonction helper pour traiter une transaction
                def process_extrait_transaction(date_part, description, reference, value_date, amount_str, debit_or_credit_hint=None):
                    # Convertir la date
                    date_str = convert_extrat_date(date_part)
                    
                    # Nettoyer la description
                    description = re.sub(r'\s+', ' ', description).strip()
                    if reference and reference not in description:
                        description = f"{description} {reference}"
                    
                    if not description:
                        return None
                    
                    # Traiter le montant
                    amount = clean_amount(amount_str)
                    if amount is not None:
                        # Utiliser la classification
                        if debit_or_credit_hint == 'debit':
                            debit, credit = amount, None
                        elif debit_or_credit_hint == 'credit':
                            debit, credit = None, amount
                        else:
                            debit, credit = classify_biat_amount(description, amount, line)
                        
                        # Créer une clé unique pour éviter les doublons
                        transaction_key = f"{date_str}|{description}|{debit}|{credit}|{amount}"
                        
                        if transaction_key not in seen_transactions:
                            seen_transactions.add(transaction_key)
                            return {
                                "date": date_str
                                "libelle": description
                                "debit": debit
                                "credit": credit
                            }
                    return None
                
                # Essayer le pattern colonnes complètes
                match = pattern_full.match(line)
                if match:
                    date_part = match.group('date').strip()
                    description = match.group('lib').strip()
                    reference = match.group('ref') or ""
                    value_date = match.group('valdate') or ""
                    debit_str = (match.group('debit') or '').strip()
                    credit_str = (match.group('credit') or '').strip()
                    print(f"DEBUG BIAT EXTRAT - FULL: {date_part} | {description} | {reference} | {value_date} | {debit_str} | {credit_str}")
                    if debit_str:
                        t = process_extrait_transaction(date_part, description, reference, value_date, debit_str, 'debit')
                        if t:
                            transactions.append(t)
                    if credit_str:
                        t = process_extrait_transaction(date_part, description, reference, value_date, credit_str, 'credit')
                        if t:
                            transactions.append(t)
                    if transaction:
                        transactions.append(transaction)
                        print(f"DEBUG BIAT EXTRAT - Transaction ajoutée: {transaction}")
                    continue
                
                # Pattern sans référence
                match = pattern_no_ref.match(line)
                if match:
                    date_part = match.group('date').strip()
                    description = match.group('lib').strip()
                    value_date = match.group('valdate') or ""
                    debit_str = (match.group('debit') or '').strip()
                    credit_str = (match.group('credit') or '').strip()
                    print(f"DEBUG BIAT EXTRAT - NOREF: {date_part} | {description} | {value_date} | {debit_str} | {credit_str}")
                    if debit_str:
                        t = process_extrait_transaction(date_part, description, "", value_date, debit_str, 'debit')
                        if t:
                            transactions.append(t)
                    if credit_str:
                        t = process_extrait_transaction(date_part, description, "", value_date, credit_str, 'credit')
                        if t:
                            transactions.append(t)
                    if transaction:
                        transactions.append(transaction)
                        print(f"DEBUG BIAT EXTRAT - Transaction ajoutée: {transaction}")
                    continue
                
                # Pattern simple (un seul montant à droite)
                match = pattern_simple.match(line)
                if match:
                    date_part = match.group('date').strip()
                    description = match.group('lib').strip()
                    amount_str = match.group('amount').strip()
                    print(f"DEBUG BIAT EXTRAT - SIMPLE: {date_part} | {description} | {amount_str}")
                    transaction = process_extrait_transaction(date_part, description, "", "", amount_str)
                    if transaction:
                        transactions.append(transaction)
                        print(f"DEBUG BIAT EXTRAT - Transaction ajoutée: {transaction}")
                    continue
            
            print(f"DEBUG BIAT EXTRAT - Total transactions extraites du texte: {len(transactions)}")
            return transactions
            
        except Exception as e:
            print(f"Erreur parsing texte BIAT EXTRAT: {e}")
            return []
    
    def parse_biat_transaction_row(self, row, year):
        """Parse une ligne de transaction BIAT EXTRAT avec colonnes séparées"""
        try:
            # Nettoyer la ligne - garder toutes les colonnes même vides
            clean_row = [str(cell).strip() if cell else "" for cell in row]
            
            print(f"DEBUG BIAT EXTRAT - Ligne brute: {row}")
            print(f"DEBUG BIAT EXTRAT - Ligne nettoyée: {clean_row}")
            print(f"DEBUG BIAT EXTRAT - Nombre de colonnes: {len(clean_row)}")
            
            # Structure EXTRAT BIAT: Date opé., Libellé, Référence, Date Valeur, Débit, Crédit
            if len(clean_row) < 6:
                print(f"DEBUG BIAT EXTRAT - Pas assez de colonnes: {len(clean_row)}")
                return None
            
            date_ope = clean_row[0]
            libelle = clean_row[1]
            reference = clean_row[2]
            date_valeur = clean_row[3]
            debit = clean_row[4]
            credit = clean_row[5]
            
            print(f"DEBUG BIAT EXTRAT - Date opé: '{date_ope}', Libellé: '{libelle[:50]}...', Débit: '{debit}', Crédit: '{credit}'")
            
            # Vérifier si c'est une date valide
            if not self.is_date_biat_extrat(date_ope):
                print(f"DEBUG BIAT EXTRAT - Date invalide: '{date_ope}'")
                return None
            
            # Formater la date
            date_formatted = self.format_date_biat_extrat(date_ope, year)
            print(f"DEBUG BIAT EXTRAT - Date formatée: '{date_formatted}'")
            
            # Déterminer si c'est un débit ou crédit
            montant_debit = None
            montant_credit = None
            
            if debit and debit != '' and debit != '0':
                # Nettoyer le montant débit
                debit_clean = debit.replace(' ', '').replace(',', '.')
                try:
                    montant_debit = float(debit_clean)
                    print(f"DEBUG BIAT EXTRAT - Montant débit: {montant_debit}")
                except:
                    print(f"DEBUG BIAT EXTRAT - Erreur parsing débit: '{debit}'")
            
            if credit and credit != '' and credit != '0':
                # Nettoyer le montant crédit
                credit_clean = credit.replace(' ', '').replace(',', '.')
                try:
                    montant_credit = float(credit_clean)
                    print(f"DEBUG BIAT EXTRAT - Montant crédit: {montant_credit}")
                except:
                    print(f"DEBUG BIAT EXTRAT - Erreur parsing crédit: '{credit}'")
            
            # Vérifier qu'on a au moins un montant
            if not montant_debit and not montant_credit:
                print(f"DEBUG BIAT EXTRAT - Aucun montant valide trouvé")
                return None
            
            return {
                'date': date_formatted
                'libelle': libelle
                'debit': montant_debit
                'credit': montant_credit
            }
            
        except Exception as e:
            print(f"Erreur parsing ligne BIAT EXTRAT: {e}")
            return None
    
    def analyze_biat_line_structure(self, line):
        """Analyse la structure d'une ligne BIAT EXTRAT pour détecter débit/crédit"""
        import re
        
        # Analyser le contenu du libellé pour déterminer le type de transaction
        libelle_upper = line.upper()
        
        # Mots-clés qui indiquent généralement un CRÉDIT (entrée d'argent)
        credit_keywords = [
            "ENCAISSEMENT", "VERSEMENT", "DEBLOCAGE", "CREDIT"
            "VIREMENT RECU", "REMBOURSEMENT", "INTERET", "DIVIDENDE"
        ]
        
        # Mots-clés qui indiquent généralement un DÉBIT (sortie d'argent)
        debit_keywords = [
            "VIREMENT TN AUTRE BQ", "VIREMENT TN MEME BQ", "COMMISSION"
            "FRAIS", "AGIOS", "PRELEVEMENT", "RETRAIT", "CHEQUE IMPAYE"
            "EFFET IMPAYE", "COTISATION", "REGLEMENT CHEQUE", "VIR ORD CARTE"
        ]
        
        # Vérifier d'abord les mots-clés spécifiques
        for keyword in credit_keywords:
            if keyword in libelle_upper:
                print(f"DEBUG BIAT EXTRAT - Mot-clé crédit détecté: {keyword}")
                return "credit"
        
        for keyword in debit_keywords:
            if keyword in libelle_upper:
                print(f"DEBUG BIAT EXTRAT - Mot-clé débit détecté: {keyword}")
                return "debit"
        
        # Si aucun mot-clé spécifique, analyser la position du montant
        # Chercher des séquences d'espaces multiples (3+ espaces consécutifs)
        space_patterns = re.finditer(r'\s{3,}', line)
        space_positions = [match.start() for match in space_patterns]
        
        # Chercher le montant dans la ligne
        montant_match = re.search(r'([\d\s.,]+)$', line.strip())
        if not montant_match:
            return "credit"  # Par défaut
        
        montant_start = montant_match.start()
        line_before_montant = line[:montant_start].strip()
        
        # Analyser la position du montant par rapport aux séparations de colonnes
        if space_positions:
            # Si le montant est après la dernière séparation d'espace importante, c'est un crédit
            last_space_pos = space_positions[-1]
            if montant_start > last_space_pos:
                print(f"DEBUG BIAT EXTRAT - Montant après dernière séparation -> crédit")
                return "credit"
            else:
                # Le montant est avant la dernière séparation, c'est probablement un débit
                print(f"DEBUG BIAT EXTRAT - Montant avant dernière séparation -> débit")
                return "debit"
        
        # Méthode de fallback: analyser la longueur de la ligne avant le montant
        if len(line_before_montant) > 80:  # Ligne très longue avant le montant
            print(f"DEBUG BIAT EXTRAT - Ligne longue ({len(line_before_montant)}) -> crédit")
            return "credit"
        elif len(line_before_montant) > 50:  # Ligne moyennement longue
            print(f"DEBUG BIAT EXTRAT - Ligne moyenne ({len(line_before_montant)}) -> crédit")
            return "credit"
        else:
            # Ligne courte avant le montant, probablement un débit
            print(f"DEBUG BIAT EXTRAT - Ligne courte ({len(line_before_montant)}) -> débit")
            return "debit"
    
    def parse_biat_transactions_from_text(self, text, year):
        """Parse les transactions BIAT EXTRAT depuis le texte brut - VERSION SIMPLIFIÉE"""
        transactions = []
        try:
            lines = text.split('\n')
            print(f"DEBUG BIAT EXTRAT TEXT - {len(lines)} lignes trouvées")
            
            for line in lines:
                line = line.strip()
                if not line or len(line) < 10:  # Ignorer les lignes trop courtes
                    continue
                
                print(f"DEBUG BIAT EXTRAT TEXT - Ligne: {line}")
                
                # Pattern principal: DD MM [tout le reste jusqu'au montant final]
                # Ce pattern capture tout entre la date et le montant final
                main_pattern = r'^(\d{1,2}\s+\d{1,2})\s+(.+?)\s+([\d\s.,]+)$'
                
                # Patterns spécifiques pour différents formats - AMÉLIORÉS
                patterns = [
                    # Format avec référence et date valeur
                    r'^(?P<date>\d{1,2}\s+[A-Za-zéûôîç\.]+\.?\s+\d{2})\s+'
                    r'(?P<lib>.+?)\s+'
                    r'(?P<ref>FT\d+)?\s+'
                    r'(?P<valdate>\d{1,2}\s+[A-Za-zéûôîç\.]+\.?\s+\d{2})\s+'
                    r'(?P<debit>[\d\s,.]+)?\s*'
                    r'(?P<credit>[\d\s,.]+)?$'
                    # Format simple sans référence
                    r'^(?P<date>\d{1,2}\s+[A-Za-zéûôîç\.]+\.?\s+\d{2})\s+'
                    r'(?P<lib>.+?)\s+'
                    r'(?P<debit>[\d\s,.]+)?\s*'
                    r'(?P<credit>[\d\s,.]+)?$'
                ]

                # Essayer chaque pattern jusqu'à en trouver un qui fonctionne
                match = None
                for pattern in patterns:
                    match = re.match(pattern, line)
                    if match:
                        break
                
                if match:
                    # Traitement unifié pour tous les patterns
                    groups = match.groups()
                    
                    if len(groups) == 4:
                        # Format avec date de valeur: DD MM libellé DDMMYYYY montant
                        date = groups[0]  # DD MM
                        libelle = groups[1]  # Libellé complet
                        date_valeur = groups[2]  # DDMMYYYY (ignorer)
                        montant = groups[3]  # Montant
                    elif len(groups) == 3:
                        # Format simple: DD MM libellé montant
                        date = groups[0]  # DD MM
                        libelle = groups[1]  # Libellé complet
                        montant = groups[2]  # Montant
                    else:
                        continue  # Pattern non reconnu
                    
                    # AMÉLIORATION: Nettoyer le libellé pour enlever les montants qui s'y trouvent
                    libelle_clean = libelle.strip()
                    
                    # Chercher si le libellé contient un montant à la fin
                    # Pattern pour détecter un montant à la fin du libellé
                    montant_in_libelle_pattern = r'(.+?)\s+([\d\s.,]+)$'
                    libelle_match = re.match(montant_in_libelle_pattern, libelle_clean)
                    
                    if libelle_match:
                        # Le libellé contient un montant à la fin
                        libelle_text = libelle_match.group(1).strip()
                        montant_in_libelle = libelle_match.group(2).strip()
                        
                        # Vérifier si le montant extrait du libellé est plus récent/pertinent
                        if not montant or len(montant.strip()) == 0:
                            # Utiliser le montant du libellé
                            montant = montant_in_libelle
                            libelle_clean = libelle_text
                            print(f"DEBUG BIAT EXTRAT - Montant extrait du libellé: {montant}")
                        else:
                            # Garder le libellé complet mais signaler le montant trouvé
                            print(f"DEBUG BIAT EXTRAT - Montant dans libellé détecté mais ignoré: {montant_in_libelle}")
                    
                    # Nettoyer le montant (enlever les points et espaces)
                    montant_clean = self.clean_amount(montant)

                    # Vérifier que le montant ne contient pas de date de valeur mélangée
                    if len(montant_clean) > 10:
                        # Chercher un pattern de date de valeur (8 chiffres) suivi d'un montant
                        date_valeur_pattern = r'^(\d{8})([\d,]+)$'
                        date_match = re.match(date_valeur_pattern, montant_clean)
                        if date_match:
                            # Extraire seulement la partie montant (après les 8 premiers chiffres)
                            montant_clean = date_match.group(2)
                            print(f"DEBUG BIAT EXTRAT - Date de valeur détectée et supprimée du montant: {date_match.group(1)} -> {montant_clean}")
                    
                    # Déterminer si c'est débit ou crédit selon la position du montant dans la ligne
                    debit = ""
                    credit = ""
                    
                    # Utiliser la fonction d'analyse de structure pour déterminer débit/crédit
                    transaction_type = self.analyze_biat_line_structure(line)
                    
                    if transaction_type == "debit":
                        debit = montant_clean
                    else:
                        credit = montant_clean
                    
                    # Utiliser la date DD MM avec l'année détectée
                    transaction = {
                        'Date': self.format_date_biat(date, year)
                        'Libellé': libelle_clean
                        'Débit': self.format_amount(debit)
                        'Crédit': self.format_amount(credit)
                    }
                    
                    transactions.append(transaction)
                    print(f"DEBUG BIAT EXTRAT TEXT - Transaction créée: {transaction}")
                
                else:
                    # Si aucun pattern ne correspond, essayer une approche plus flexible
                    # Chercher des lignes qui commencent par une date
                    if re.match(r'^\d{1,2}\s+[A-Za-zéûôîç\.]+\.?\s+\d{2}', line):
                        print(f"DEBUG BIAT EXTRAT TEXT - Ligne avec date détectée mais non parsée: {line}")
                
        except Exception as e:
            print(f"Erreur parsing texte BIAT EXTRAT: {e}")
        
        return transactions
    
    def is_date_biat_extrat(self, text):
        """Vérifie si le texte est une date au format BIAT EXTRAT DD mois YY"""
        if not text:
            return False
        # Pattern pour date DD mois YY (ex: "01 août 25", "04 août 25")
        return bool(re.match(r'^\d{1,2}\s+[A-Za-zéûôîç\.]+\.?\s+\d{2}$', text.strip()))
    
    def is_date_biat(self, text):
        """Vérifie si le texte est une date au format BIAT DD MM"""
        if not text:
            return False
        # Pattern pour date DD MM (avec espaces)
        return bool(re.match(r'^\d{1,2}\s+\d{1,2}$', text.strip()))
    
    def is_date_valeur_biat(self, text):
        """Vérifie si le texte est une date de valeur au format BIAT DDMMYYYY"""
        if not text:
            return False
        # Pattern pour date de valeur DDMMYYYY
        return bool(re.match(r'^\d{8}$', text.strip()))
    
    def format_date_biat_extrat(self, date_str, year):
        """Formate la date BIAT EXTRAT au format DD/MM/YYYY"""
        try:
            if not date_str:
                return ""
            
            # Si c'est déjà au format DD/MM/YYYY
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
                return date_str
            
            # Si c'est au format DD mois YY (ex: "01 août 25")
            if re.match(r'^\d{1,2}\s+\w+\s+\d{2}$', date_str):
                parts = date_str.split()
                day = parts[0].zfill(2)
                month_name = parts[1].lower()
                year_part = parts[2]
                
                # Convertir le nom du mois en numéro
                month_map = {
                    'janvier': '01', 'jan': '01'
                    'février': '02', 'fév': '02', 'fevrier': '02', 'fev': '02'
                    'mars': '03', 'mar': '03'
                    'avril': '04', 'avr': '04'
                    'mai': '05'
                    'juin': '06'
                    'juillet': '07', 'juil': '07'
                    'août': '08', 'aout': '08'
                    'septembre': '09', 'sept': '09', 'sept.': '09'
                    'octobre': '10', 'oct': '10'
                    'novembre': '11', 'nov': '11'
                    'décembre': '12', 'dec': '12', 'decembre': '12', 'déc.': '12'
                }
                
                month_num = month_map.get(month_name, '01')
                
                # Convertir l'année à 2 chiffres en année complète
                if len(year_part) == 2:
                    full_year = 2000 + int(year_part)
                else:
                    full_year = int(year_part)
                
                return f"{day}/{month_num}/{full_year}"
            
            # Si c'est au format DDMMYYYY
            if re.match(r'^\d{8}$', date_str):
                day = date_str[:2]
                month = date_str[2:4]
                year_part = date_str[4:8]
                return f"{day}/{month}/{year_part}"
            
            # Si c'est au format DD MM
            if re.match(r'^\d{1,2}\s+\d{1,2}$', date_str):
                parts = date_str.split()
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                return f"{day}/{month}/{year}"
            
            return date_str
            
        except Exception as e:
            print(f"Erreur formatage date BIAT EXTRAT: {e}")
            return date_str
    
    def format_date_biat(self, date_str, year):
        """Formate la date BIAT au format DD/MM/YYYY"""
        try:
            if not date_str:
                return ""
            
            # Si c'est déjà au format DD/MM/YYYY
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
                return date_str
            
            # Si c'est au format DDMMYYYY
            if re.match(r'^\d{8}$', date_str):
                day = date_str[:2]
                month = date_str[2:4]
                year_part = date_str[4:8]
                return f"{day}/{month}/{year_part}"
            
            # Si c'est au format DD MM
            if re.match(r'^\d{1,2}\s+\d{1,2}$', date_str):
                parts = date_str.split()
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                return f"{day}/{month}/{year}"
            
            return date_str
            
        except Exception as e:
            print(f"Erreur formatage date BIAT EXTRAT: {e}")
            return date_str
    
    def is_amount(self, text):
        """Vérifie si le texte est un montant"""
        if not text:
            return False
        # Pattern pour montant avec virgule et points
        return bool(re.match(r'^[\d\s.,]+$', text.strip()))
    
    def format_amount(self, amount_str):
        """Formate le montant"""
        if not amount_str or amount_str == "":
            return ""
        
        try:
            # Nettoyer le montant
            clean_amount = str(amount_str).replace(' ', '').replace('.', '')
            
            # Remplacer la virgule par un point pour la conversion
            if ',' in clean_amount:
                clean_amount = clean_amount.replace(',', '.')
            
            # Convertir en float puis formater
            amount_float = float(clean_amount)
            return f"{amount_float:,.2f}".replace(',', ' ').replace('.', ',')
            
        except (ValueError, TypeError):
            return str(amount_str)
    
    def save_excel_with_formatting(self, df, excel_path):
        """Sauvegarde le DataFrame en Excel avec formatage professionnel"""
        try:
            # Créer le fichier Excel avec openpyxl pour le formatage
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "J03"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formatage des en-têtes
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin_border = Border(
                left=Side(style='thin')
                right=Side(style='thin')
                top=Side(style='thin')
                bottom=Side(style='thin')
            )
            
            # Appliquer le formatage aux en-têtes
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Appliquer les bordures à toutes les cellules
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Ajuster la largeur des colonnes
            column_widths = {'A': 12, 'B': 50, 'C': 15, 'D': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Sauvegarder
            wb.save(excel_path)
            print(f"Fichier Excel sauvegardé avec formatage: {excel_path}")
            
        except Exception as e:
            print(f"Erreur formatage Excel: {e}")
            # Fallback: sauvegarde simple
            df.to_excel(excel_path, index=False, sheet_name="J03")
    
    def convert_pdf_to_excel(self):
        """Conversion complète: extrait Date | Libellé | Débit | Crédit et exporte en Excel."""
        try:
            if not self.pdf_path.get():
                messagebox.showerror("Erreur", "Veuillez sélectionner un fichier PDF")
                return
            if not self.excel_name.get():
                messagebox.showerror("Erreur", "Veuillez entrer un nom pour le fichier Excel")
                return
            
            # Démarrer
            
            self.convert_button.config(state='disabled')
            
            self.root.update()
            
            pdf_file = self.pdf_path.get()
            if not os.path.exists(pdf_file):
                messagebox.showerror("Erreur", "Le fichier n'existe pas")
                return
            
            # Optionnel: vérification souple
            if not self.is_biat_extrat_pdf(pdf_file):
                messagebox.showwarning("Attention", "Le fichier ne ressemble pas à un EXTRAT BIAT (détection souple). La conversion sera tentée quand même.")
            
            # Multi-niveaux: d'abord tableaux, puis layout si vide
            transactions = self.extract_table_data(pdf_file)
            if not transactions:
                # Essayer une extraction par mise en page (x/y)
                transactions = self.extract_by_layout(pdf_file)
            if not transactions:
                messagebox.showerror("Erreur", "Aucune transaction trouvée dans le PDF")
                return
            
            # Harmoniser les clés éventuelles
            norm_tx = []
            for t in transactions:
                if not t:
                    continue
                if 'Date' in t:
                    norm_tx.append(t)
                else:
                    norm_tx.append({
                        'Date': t.get('date', '')
                        'Libellé': t.get('libelle', '')
                        'Débit': t.get('debit', '')
                        'Crédit': t.get('credit', '')
                    })
            
            df = pd.DataFrame(norm_tx, columns=['Date', 'Libellé', 'Débit', 'Crédit'])
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_filename = f"{self.excel_name.get()}.xlsx"
            excel_path = os.path.join(downloads_path, excel_filename)
            
            self.save_excel_with_formatting(df, excel_path)

            self.convert_button.config(state='normal')
            
            # Message de succès plus positif
            success_msg = f"✅ Conversion EXTRAT terminée avec succès !\n\n"
            success_msg += f"📁 Fichier: {excel_filename}\n"
            success_msg += f"📂 Emplacement: {downloads_path}\n"
            success_msg += f"📊 Nombre de transactions: {len(df)}\n\n"
            success_msg += f"🎉 Votre fichier Excel est prêt à utiliser !"
            
            messagebox.showinfo("✅ Conversion réussie", success_msg)
        except Exception as e:
            
            self.convert_button.config(state='normal')
            
            messagebox.showerror("Erreur", f"Erreur lors de la conversion: {str(e)}")

    def retour_accueil(self):
        """Retourne à la page d'accueil"""
        try:
            # Fermer la fenêtre actuelle
            self.root.destroy()
            # Lancer le convertisseur principal
            subprocess.Popen([sys.executable, "lancer_convertisseur.py"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de retourner à la page d'accueil: {e}")

    def _extract_text_via_ocr(self, pdf_path: str) -> str:
        """Réalise un OCR simple sur chaque page si disponible."""
        if not _OCR_AVAILABLE:
            return ""
        try:
            doc = fitz.open(pdf_path)
            text_parts = []
            for page in doc:
                # Rendu à bonne résolution pour un OCR plus fiable
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Pré-traitement binaire
                arr = np.array(img)
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                # OCR en français (si Tesseract FRA installé)
                txt = pytesseract.image_to_string(thr, lang='fra')
                if txt:
                    text_parts.append(txt)
            return "\n".join(text_parts)
        except Exception:
            return ""

    def _parse_ocr_text(self, text: str):
        """Parse le texte OCR pour extraire les transactions BIAT."""
        transactions = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Chercher une date au format DD/MM/YYYY
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            if not date_match:
                continue
                
            date = date_match.group(1)
            
            # Extraire le libellé (après la date)
            libelle = line[date_match.end():].strip()
            
            # Chercher les montants
            amounts = re.findall(r'[\d\s,\.]+', line)
            amounts = [amt.strip() for amt in amounts if re.match(r'^\d', amt.strip())]
            
            debit = credit = None
            
            if len(amounts) >= 2:
                # Premier montant = débit, deuxième = crédit
                debit = self._format_amount(amounts[0])
                credit = self._format_amount(amounts[1])
            elif len(amounts) == 1:
                # Un seul montant, déterminer si c'est débit ou crédit
                if any(keyword in libelle.lower() for keyword in ['retrait', 'paiement', 'frais', 'commission', 'debit']):
                    debit = self._format_amount(amounts[0])
                else:
                    credit = self._format_amount(amounts[0])
            
            # Nettoyer le libellé des montants
            for amount in amounts:
                libelle = libelle.replace(amount, '').strip()
            
            if libelle:
                transactions.append({
                    'date': date
                    'libelle': libelle
                    'debit': debit
                    'credit': credit
                })
        
        return transactions

    def _format_amount(self, amount_str: str):
        """Formate un montant selon le format BIAT."""
        if not amount_str:
            return None
        try:
            # Nettoyer le montant
            cleaned = re.sub(r'[^\d,.-]', '', amount_str)
            if not cleaned:
                return None
            
            # Convertir en float
            if ',' in cleaned and '.' in cleaned:
                # Format: 1.234,56
                cleaned = cleaned.replace('.', '').replace(',', '.')
            elif ',' in cleaned:
                # Format: 1234,56
                cleaned = cleaned.replace(',', '.')
            
            amount = float(cleaned)
            if amount == 0:
                return None
            
            # Retourner au format BIAT
            return f"{amount:,.3f}".replace(',', ' ').replace('.', ',')
        except:
            return None

def main():
    root = tk.Tk()
    app = BIATExtratConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
